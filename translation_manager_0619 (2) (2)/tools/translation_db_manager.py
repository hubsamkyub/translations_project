# translation_db_manager.py에 추가할 내용

import sqlite3
import os
import json
import time
import pandas as pd
import re
from openpyxl import load_workbook
import gc
from openpyxl.styles import PatternFill
from datetime import datetime

class TranslationDBManager:
    def __init__(self, parent_window=None):
        self.parent = parent_window
        self.excluded_count = 0
    
    def find_string_id_position(self, worksheet):
        """최적화된 STRING_ID 위치 탐색 (2~6행 → 1행 → 실패 시 None)"""
        for row in worksheet.iter_rows(min_row=2, max_row=6, max_col=5):
            for cell in row:
                if isinstance(cell.value, str) and "STRING_ID" in cell.value.upper():
                    return cell.column, cell.row

        for row in worksheet.iter_rows(min_row=1, max_row=1, max_col=5):
            for cell in row:
                if isinstance(cell.value, str) and "STRING_ID" in cell.value.upper():
                    return cell.column, cell.row

        return None, None

    def find_language_columns(self, worksheet, header_row, langs, language_mapping=None):
        """STRING_ID 행 기준으로 언어 컬럼 위치 탐색 (ZH/CN 등 매핑 포함)"""
        if not header_row:
            return {}

        lang_cols = {}
        
        # 역방향 매핑 준비 (예: CN → [ZH])
        reverse_mapping = {}
        if language_mapping:
            for alt, main in language_mapping.items():
                reverse_mapping.setdefault(main, []).append(alt)

        # 지정한 헤더 행에서만 검색 (빠르게)
        for row in worksheet.iter_rows(min_row=header_row, max_row=header_row):
            for cell in row:
                if not cell.value:
                    continue

                header_text = str(cell.value).strip()

                # 직접 매칭
                if header_text in langs:
                    lang_cols[header_text] = cell.column
                    continue

                # ZH 같은 대체 키 → CN 등으로 매핑
                if language_mapping and header_text in language_mapping:
                    mapped_lang = language_mapping[header_text]
                    if mapped_lang in langs and mapped_lang not in lang_cols:
                        lang_cols[mapped_lang] = cell.column

        return lang_cols
    
    def create_translation_table(self, cursor, table_name="translation_data"):
        """번역 테이블 생성 (상태 및 업데이트 날짜 포함)"""
        cursor.execute(f'''
        CREATE TABLE IF NOT EXISTS {table_name} (
            id INTEGER PRIMARY KEY,
            file_name TEXT,
            sheet_name TEXT,
            string_id TEXT UNIQUE,
            kr TEXT,
            en TEXT, 
            cn TEXT,
            tw TEXT,
            th TEXT,
            status TEXT DEFAULT 'active',
            update_date TEXT
        )
        ''')
        
        # 인덱스 생성
        cursor.execute(f'CREATE INDEX IF NOT EXISTS idx_{table_name}_string_id ON {table_name}(string_id)')
        cursor.execute(f'CREATE INDEX IF NOT EXISTS idx_{table_name}_file_sheet ON {table_name}(file_name, sheet_name)')
        cursor.execute(f'CREATE INDEX IF NOT EXISTS idx_{table_name}_status ON {table_name}(status)')

    def update_translation_db(self, excel_files, db_path, language_list, batch_size=2000, use_read_only=True, progress_callback=None):
        """번역 DB 업데이트 함수 (A열 # 체크 기능 추가)"""
        if not excel_files:
            return {"status": "error", "message": "번역 파일이 선택되지 않았습니다."}
            
        if not db_path:
            return {"status": "error", "message": "DB 파일 경로가 지정되지 않았습니다."}
        
        if not os.path.exists(db_path):
            return {"status": "error", "message": "기존 DB 파일이 존재하지 않습니다. 먼저 DB를 구축해주세요."}
            
        selected_langs = language_list
        if not selected_langs:
            return {"status": "error", "message": "하나 이상의 언어를 선택하세요."}
        
        language_mapping = {
            "ZH": "CN",
        }
        
        try:
            conn = sqlite3.connect(db_path)
            conn.execute("PRAGMA journal_mode = WAL")
            conn.execute("PRAGMA synchronous = NORMAL")
            conn.execute("PRAGMA cache_size = 10000")
            cursor = conn.cursor()
            
            self._update_table_schema(cursor)
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            existing_data = self._load_existing_data_by_string_id(cursor)
            
            total_rows = 0
            processed_count = 0
            error_count = 0
            updated_rows = 0
            new_rows = 0
            batch_data = []
            
            for idx, (file_name, file_path) in enumerate(excel_files):
                if progress_callback:
                    progress_callback(f"파일 ({idx+1}/{len(excel_files)}) {file_name} 처리 중...", 
                                    idx+1, len(excel_files))
                
                try:
                    gc.collect()
                    workbook = load_workbook(file_path, read_only=use_read_only, data_only=True)
                    
                    string_sheets = [sheet for sheet in workbook.sheetnames 
                                if sheet.lower().startswith("string") and not sheet.startswith("#")]
                    
                    for sheet_name in string_sheets:
                        worksheet = workbook[sheet_name]
                        
                        string_id_col, header_row = self.find_string_id_position(worksheet)
                        if not string_id_col or not header_row:
                            continue
                        
                        lang_cols = self.find_language_columns(worksheet, header_row, selected_langs, language_mapping)
                        if not lang_cols:
                            continue
                        
                        max_row = worksheet.max_row
                        
                        for row_cells in worksheet.iter_rows(min_row=header_row + 1, max_row=max_row, values_only=True):                            
                            if not row_cells or string_id_col - 1 >= len(row_cells):
                                continue
                            string_id = row_cells[string_id_col - 1]
                            if not string_id:
                                continue
                            
                            first_col_val = str(row_cells[0] or '').strip()
                            status = '비활성' if first_col_val.startswith('#') else 'active'
                            
                            values = {"string_id": string_id}
                            has_translation = False
                            
                            for lang, col in lang_cols.items():
                                if col - 1 < len(row_cells):
                                    cell_value = row_cells[col - 1]
                                    mapped_lang = language_mapping.get(lang.upper(), lang).lower()
                                    values[mapped_lang] = cell_value
                                    if cell_value:
                                        has_translation = True
                            
                            if string_id in existing_data:
                                existing_row = existing_data[string_id]
                                update_needed = False
                                
                                if existing_row.get('status') != status:
                                    update_needed = True

                                if not update_needed:
                                    for lang in selected_langs:
                                        lang_key = lang.lower()
                                        if values.get(lang_key) is not None and existing_row.get(lang_key) != values.get(lang_key):
                                            update_needed = True
                                            break
                                
                                if update_needed:
                                    update_data = {
                                        'string_id': string_id,
                                        'status': status,
                                        'update_date': current_time,
                                        'file_name': file_name,
                                        'sheet_name': sheet_name
                                    }
                                    for lang in ['kr', 'en', 'cn', 'tw', 'th']:
                                        update_data[lang] = values.get(lang, existing_row.get(lang))
                                    batch_data.append(('update', update_data))
                                    updated_rows += 1
                            else:
                                if has_translation:
                                    insert_data = {
                                        'string_id': string_id,
                                        'status': status,
                                        'update_date': current_time,
                                        'file_name': file_name,
                                        'sheet_name': sheet_name
                                    }
                                    for lang in ['kr', 'en', 'cn', 'tw', 'th']:
                                        insert_data[lang] = values.get(lang)
                                    batch_data.append(('insert', insert_data))
                                    new_rows += 1
                            
                            if len(batch_data) >= batch_size:
                                self._process_batch_data(cursor, batch_data)
                                total_rows += len(batch_data)
                                batch_data = []
                                gc.collect()
                    
                    workbook.close()
                    del workbook
                    processed_count += 1
                    
                except Exception as e:
                    if progress_callback:
                        progress_callback(f"파일 처리 오류: {e}", idx+1, len(excel_files))
                    error_count += 1
            
            if batch_data:
                self._process_batch_data(cursor, batch_data)
                total_rows += len(batch_data)
            
            if progress_callback:
                progress_callback("DB 최적화 중...", len(excel_files), len(excel_files))
            
            cursor.execute("PRAGMA optimize")
            conn.commit()
            conn.close()
            
            return {
                "status": "success",
                "processed_count": processed_count,
                "error_count": error_count,
                "total_rows": total_rows,
                "updated_rows": updated_rows,
                "new_rows": new_rows,
                "deleted_rows": 0
            }
            
        except Exception as e:
            try:
                conn.close()
            except:
                pass
                
            return {
                "status": "error",
                "message": str(e)
            }

    def _update_table_schema(self, cursor):
        """테이블 스키마 업데이트 (status, update_date 컬럼 추가)"""
        try:
            # 기존 테이블 구조 확인
            cursor.execute("PRAGMA table_info(translation_data)")
            columns = [row[1] for row in cursor.fetchall()]
            
            # status 컬럼이 없으면 추가
            if 'status' not in columns:
                cursor.execute("ALTER TABLE translation_data ADD COLUMN status TEXT DEFAULT 'active'")
                # 기존 데이터의 status를 'active'로 업데이트
                cursor.execute("UPDATE translation_data SET status = 'active' WHERE status IS NULL")
            
            # update_date 컬럼이 없으면 추가
            if 'update_date' not in columns:
                cursor.execute("ALTER TABLE translation_data ADD COLUMN update_date TEXT")
                # 기존 데이터의 update_date를 현재 시간으로 설정
                current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                cursor.execute("UPDATE translation_data SET update_date = ? WHERE update_date IS NULL", (current_time,))
            
            # 기존 UNIQUE 제약조건 확인 및 수정 (STRING_ID만 유일하도록)
            cursor.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='translation_data'")
            table_schema = cursor.fetchone()
            if table_schema and 'UNIQUE(file_name, sheet_name, string_id)' in table_schema[0]:
                # 기존 테이블을 백업하고 새로운 구조로 재생성
                cursor.execute('''
                CREATE TABLE translation_data_new (
                    id INTEGER PRIMARY KEY,
                    file_name TEXT,
                    sheet_name TEXT,
                    string_id TEXT UNIQUE,
                    kr TEXT,
                    en TEXT, 
                    cn TEXT,
                    tw TEXT,
                    th TEXT,
                    status TEXT DEFAULT 'active',
                    update_date TEXT
                )
                ''')
                
                # 데이터 이전 (중복 STRING_ID는 최신 것만 유지)
                cursor.execute('''
                INSERT OR REPLACE INTO translation_data_new 
                SELECT * FROM translation_data
                ''')
                
                # 기존 테이블 삭제 후 이름 변경
                cursor.execute('DROP TABLE translation_data')
                cursor.execute('ALTER TABLE translation_data_new RENAME TO translation_data')
            
            # 인덱스 생성
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_string_id ON translation_data(string_id)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_file_sheet ON translation_data(file_name, sheet_name)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_status ON translation_data(status)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_update_date ON translation_data(update_date)')
            
        except Exception as e:
            print(f"스키마 업데이트 오류: {e}")

    def _load_existing_data_by_string_id(self, cursor):
        """기존 데이터 로드 (STRING_ID만을 키로 사용)"""
        cursor.execute("""
            SELECT file_name, sheet_name, string_id, kr, en, cn, tw, th, status, update_date
            FROM translation_data
        """)
        
        existing_data = {}
        for row in cursor.fetchall():
            string_id = row[2]  # string_id만을 키로 사용
            existing_data[string_id] = {
                'file_name': row[0],
                'sheet_name': row[1],
                'kr': row[3],
                'en': row[4],
                'cn': row[5],
                'tw': row[6],
                'th': row[7],
                'status': row[8],
                'update_date': row[9]
            }
        
        return existing_data

    def _process_batch_data(self, cursor, batch_data):
        """배치 데이터 처리"""
        for operation, data in batch_data:
            try:
                if operation == 'insert':
                    cursor.execute('''
                        INSERT OR REPLACE INTO translation_data 
                        (file_name, sheet_name, string_id, kr, en, cn, tw, th, status, update_date)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        data['file_name'], data['sheet_name'], data['string_id'],
                        data.get('kr'), data.get('en'), data.get('cn'), 
                        data.get('tw'), data.get('th'), data['status'], data['update_date']
                    ))
                    
                elif operation == 'update':
                    cursor.execute('''
                        UPDATE translation_data 
                        SET file_name = ?, sheet_name = ?, kr = ?, en = ?, cn = ?, tw = ?, th = ?, status = ?, update_date = ?
                        WHERE string_id = ?
                    ''', (
                        data['file_name'], data['sheet_name'],
                        data.get('kr'), data.get('en'), data.get('cn'), 
                        data.get('tw'), data.get('th'), 
                        data['status'], data['update_date'], data['string_id']
                    ))
                    
                elif operation == 'status':
                    cursor.execute('''
                        UPDATE translation_data 
                        SET status = ?, update_date = ?
                        WHERE string_id = ?
                    ''', (
                        data['status'], data['update_date'], data['string_id']
                    ))
                    
            except Exception as e:
                print(f"배치 처리 오류: {e}")
                continue

    # 기존 build_translation_db 메서드는 그대로 유지
# build_translation_db 함수를 아래 코드로 교체합니다.

    def build_translation_db(self, excel_files, output_db_path, language_list, batch_size=2000, use_read_only=True, progress_callback=None):
        """번역 DB 구축 함수 (A열 # 체크 기능 추가)"""
        if not excel_files:
            return {"status": "error", "message": "번역 파일이 선택되지 않았습니다."}
            
        if not output_db_path:
            return {"status": "error", "message": "DB 파일 경로가 지정되지 않았습니다."}
            
        selected_langs = language_list
        if not selected_langs:
            return {"status": "error", "message": "하나 이상의 언어를 선택하세요."}
        
        language_mapping = {"ZH": "CN"}
        
        if os.path.exists(output_db_path):
            os.remove(output_db_path)
            
        try:
            conn = sqlite3.connect(output_db_path)
            conn.execute("PRAGMA journal_mode = WAL")
            conn.execute("PRAGMA synchronous = NORMAL")
            cursor = conn.cursor()
            
            self.create_translation_table(cursor) # 테이블 생성 헬퍼 함수 호출
            conn.commit()
 
            total_rows = 0
            processed_count = 0
            error_count = 0
            batch_data = []
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            for idx, (file_name, file_path) in enumerate(excel_files):
                if progress_callback:
                    progress_callback(f"파일 ({idx+1}/{len(excel_files)}) {file_name} 처리 중...", idx+1, len(excel_files))
                
                try:
                    gc.collect()
                    workbook = load_workbook(file_path, read_only=use_read_only, data_only=True)
                    
                    string_sheets = [sheet for sheet in workbook.sheetnames if sheet.lower().startswith("string") and not sheet.startswith("#")]
                    
                    for sheet_name in string_sheets:
                        worksheet = workbook[sheet_name]
                        string_id_col, header_row = self.find_string_id_position(worksheet)
                        if not string_id_col or not header_row: continue
                        
                        lang_cols = self.find_language_columns(worksheet, header_row, selected_langs, language_mapping)
                        if not lang_cols: continue
                        
                        for row_cells in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
                            if not row_cells or not row_cells[string_id_col - 1]:
                                continue

                            # [수정] A열 값을 확인하여 status 결정
                            first_col_val = str(row_cells[0] or '').strip()
                            status = '비활성' if first_col_val.startswith('#') else 'active'

                            string_id = row_cells[string_id_col - 1]
                            values = {"string_id": string_id}
                            has_translation = False
                            
                            for lang, col in lang_cols.items():
                                if col - 1 < len(row_cells):
                                    cell_value = row_cells[col - 1]
                                    mapped_lang = language_mapping.get(lang.upper(), lang).lower()
                                    values[mapped_lang] = cell_value
                                    if cell_value: has_translation = True
                            
                            if has_translation:
                                batch_data.append((
                                    file_name, sheet_name, values["string_id"],
                                    values.get("kr"), values.get("en"), values.get("cn"), 
                                    values.get("tw"), values.get("th"),
                                    status, # [수정] 결정된 status 값 사용
                                    current_time
                                ))
                                
                                if len(batch_data) >= batch_size:
                                    cursor.executemany('''
                                    INSERT OR REPLACE INTO translation_data (file_name, sheet_name, string_id, kr, en, cn, tw, th, status, update_date)
                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', batch_data)
                                    total_rows += len(batch_data)
                                    batch_data = []
                    workbook.close()
                    processed_count += 1
                except Exception as e:
                    if progress_callback: progress_callback(f"파일 처리 오류: {e}", idx+1, len(excel_files))
                    error_count += 1
            
            if batch_data:
                cursor.executemany('''
                INSERT OR REPLACE INTO translation_data (file_name, sheet_name, string_id, kr, en, cn, tw, th, status, update_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', batch_data)
                total_rows += len(batch_data)

            conn.commit()
            conn.close()
            
            return {"status": "success", "processed_count": processed_count, "error_count": error_count, "total_rows": total_rows}
        except Exception as e:
            if 'conn' in locals() and conn: conn.close()
            return {"status": "error", "message": str(e)}
        
    # 기존 load_translation_cache 메서드도 유지하되 status 필터링 추가
    def load_translation_cache(self, db_path):
        """번역 DB를 메모리에 캐싱 (활성 상태만)"""
        try:
            # DB 연결
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            # 활성 상태의 데이터만 로드
            cursor.execute("SELECT * FROM translation_data WHERE status = 'active'")
            rows = cursor.fetchall()
            
            # 3단계 캐싱 구조 구성
            translation_cache = {}              # STRING_ID만 (3순위)
            translation_file_cache = {}         # 파일명 + STRING_ID (1순위)
            translation_sheet_cache = {}        # 시트명 + STRING_ID (2순위)
            
            # 중복 STRING_ID 추적용
            duplicate_ids = {}
            
            for row in rows:
                file_name = row["file_name"]
                sheet_name = row["sheet_name"]
                string_id = row["string_id"]
                
                # 중복 STRING_ID 추적
                if string_id not in duplicate_ids:
                    duplicate_ids[string_id] = []
                duplicate_ids[string_id].append(file_name)
                
                # 데이터 딕셔너리 생성
                data = {
                    "kr": row["kr"],
                    "en": row["en"],
                    "cn": row["cn"],
                    "tw": row["tw"],
                    "th": row["th"],
                    "file_name": file_name,
                    "sheet_name": sheet_name
                }
                
                # 1. 파일명 + STRING_ID 캐싱 (1순위)
                norm_file_name = file_name.lower()
                if norm_file_name not in translation_file_cache:
                    translation_file_cache[norm_file_name] = {}
                
                if string_id not in translation_file_cache[norm_file_name]:
                    translation_file_cache[norm_file_name][string_id] = data
                
                # 2. 시트명 + STRING_ID 캐싱 (2순위)
                norm_sheet_name = sheet_name.lower()
                if norm_sheet_name not in translation_sheet_cache:
                    translation_sheet_cache[norm_sheet_name] = {}
                
                if string_id not in translation_sheet_cache[norm_sheet_name]:
                    translation_sheet_cache[norm_sheet_name][string_id] = data
                
                # 3. STRING_ID만 캐싱 (3순위)
                translation_cache[string_id] = data
            
            conn.close()
            
            # 결과 반환
            return {
                "translation_cache": translation_cache,
                "translation_file_cache": translation_file_cache,
                "translation_sheet_cache": translation_sheet_cache,
                "duplicate_ids": duplicate_ids,
                "file_count": len(translation_file_cache),
                "sheet_count": len(translation_sheet_cache),
                "id_count": len(translation_cache)
            }
            
        except Exception as e:
            return {
                "status": "error",
                "message": str(e)
            }