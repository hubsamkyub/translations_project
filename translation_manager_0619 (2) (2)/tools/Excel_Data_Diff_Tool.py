# 상단 import 문
import os
import sys
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import logging
import threading
import hashlib
import time
import warnings
import sqlite3
import concurrent.futures  # 파일 상단에 추가
import json
from datetime import datetime


# 현재 스크립트 파일이 있는 디렉토리 경로
current_dir = os.path.dirname(os.path.abspath(__file__))
# 루트 디렉토리 경로 (tools 상위 폴더)
root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# 루트 디렉토리를 모듈 검색 경로에 추가
if root_dir not in sys.path:
    sys.path.append(root_dir)

# 내부 유틸리티 모듈 임포트
from utils.config_utils import load_config, save_config
from utils.cache_utils import (load_cached_data, save_cache, hash_paths, 
                              get_file_mtime, update_excel_cache)
from utils.excel_utils import ExcelFileManager
from utils.common_utils import PathUtils, HashUtils, FileUtils, DBUtils, logger

from ui.common_components import ScrollableCheckList


# 경고 무시 (openpyxl 관련)
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# 로깅 설정
logging.basicConfig(
    filename=os.path.join(root_dir, 'excel_diff_tool.log'),
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    filemode='w'
)

# 안전한 DB 접근 함수 (독립적 구현)
def get_columns_from_db_safe(db_path, table_name):
    """안전하게 DB 컬럼 정보를 가져오는 함수 (특수 문자 처리)"""
    if not os.path.exists(db_path):
        return []
    
    try:
        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        
        # 방법 1: 따옴표로 감싸기
        try:
            escaped_table_name = f'"{table_name}"'
            cur.execute(f"PRAGMA table_info({escaped_table_name})")
            columns = [row[1] for row in cur.fetchall()]
            conn.close()
            return columns
        except Exception as e1:
            logging.warning(f"첫 번째 이스케이프 방식 실패: {e1}")
            
            # 방법 2: 대괄호로 감싸기
            try:
                conn = sqlite3.connect(db_path)
                cur = conn.cursor()
                
                escaped_table_name = f'[{table_name}]'
                cur.execute(f"PRAGMA table_info({escaped_table_name})")
                columns = [row[1] for row in cur.fetchall()]
                conn.close()
                return columns
            except Exception as e2:
                logging.warning(f"두 번째 이스케이프 방식 실패: {e2}")
                
                # 모든 방법 실패 시 직접 쿼리 시도
                try:
                    conn = sqlite3.connect(db_path)
                    cur = conn.cursor()
                    
                    # 테이블 목록 가져와서 시작하는 부분만 비교
                    cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
                    all_tables = [row[0] for row in cur.fetchall()]
                    
                    # 테이블 이름의 시작 부분이 같은 테이블 찾기
                    matching_tables = [t for t in all_tables if t.startswith(table_name.split('@')[0])]
                    
                    if matching_tables:
                        # 첫 번째 매칭 테이블 사용
                        safe_table = matching_tables[0]
                        cur.execute(f"PRAGMA table_info(\"{safe_table}\")")
                        columns = [row[1] for row in cur.fetchall()]
                        conn.close()
                        return columns
                except Exception as e3:
                    logging.error(f"모든 DB 접근 방식 실패: {e3}")
    except Exception as e:
        logging.error(f"DB 연결 실패: {db_path} - {e}")
    
    return []  # 모든 시도 실패 시 빈 목록 반환

class LoadingPopup:
    """비동기 안전하게 상태를 업데이트하는 로딩 팝업 클래스"""
    def __init__(self, parent, title="처리 중...", message="⏳ 준비 중입니다..."):
        self.top = tk.Toplevel(parent)
        self.top.title(title)
        self.top.geometry("400x150")
        self.top.grab_set()
        self.label = tk.Label(self.top, text=message, font=("Arial", 12))
        self.label.pack(expand=True, pady=20)
        
        # 진행 상황 표시용
        self.progress_label = tk.Label(self.top, text="0%", font=("Arial", 10))
        self.progress_label.pack()

    def update_message(self, text):
        self.label.config(text=text)
        self.top.update_idletasks()

    def update_progress(self, current, total):
        percent = int((current / total) * 100) if total else 0
        self.progress_label.config(text=f"{percent}% ({current}/{total})")
        self.top.update_idletasks()

    def close(self):
        self.top.destroy()

class ImprovedFileList(tk.Frame):
    """향상된 파일 목록 클래스 - 기존 UI 연결 유지"""
    def __init__(self, parent, width=300, height=200):
        super().__init__(parent)
        
        # 메인 프레임
        self.main_frame = tk.Frame(self)
        self.main_frame.pack(fill="both", expand=True)
        
        # 검색 프레임
        self.search_frame = tk.Frame(self.main_frame)
        self.search_frame.pack(fill="x", padx=2, pady=2)
        
        # 검색 레이블
        tk.Label(self.search_frame, text="검색:", anchor="w").pack(side="left", padx=(0, 2))
        
        # 검색창
        self.search_var = tk.StringVar()
        self.search_entry = tk.Entry(self.search_frame, textvariable=self.search_var)
        self.search_entry.pack(side="left", fill="x", expand=True)
        self.search_entry.bind("<KeyRelease>", self.filter_list)
        
        # 리스트와 스크롤바 프레임
        self.list_frame = tk.Frame(self.main_frame)
        self.list_frame.pack(fill="both", expand=True, pady=2)
        
        # 스크롤바
        self.scrollbar_y = tk.Scrollbar(self.list_frame, orient="vertical")
        self.scrollbar_y.pack(side="right", fill="y")
        
        # 수평 스크롤바
        self.scrollbar_x = tk.Scrollbar(self.list_frame, orient="horizontal")
        self.scrollbar_x.pack(side="bottom", fill="x")
        
        # 리스트박스 (체크박스 텍스트 포함)
        self.listbox = tk.Listbox(self.list_frame, 
                                 width=width, 
                                 height=height,
                                 selectmode="extended",
                                 xscrollcommand=self.scrollbar_x.set,
                                 yscrollcommand=self.scrollbar_y.set)
        self.listbox.pack(side="left", fill="both", expand=True)
        
        # 스크롤바 연결
        self.scrollbar_y.config(command=self.listbox.yview)
        self.scrollbar_x.config(command=self.listbox.xview)
        
        # 버튼 프레임 (선택/해제 버튼)
        self.button_frame = tk.Frame(self.main_frame)
        self.button_frame.pack(fill="x", pady=2)
        
        # 전체 선택 버튼
        self.select_all_btn = tk.Button(self.button_frame, text="모두 선택", 
                                      command=self.select_all)
        self.select_all_btn.pack(side="left", padx=2)
        
        # 전체 해제 버튼
        self.deselect_all_btn = tk.Button(self.button_frame, text="모두 해제", 
                                        command=self.deselect_all)
        self.deselect_all_btn.pack(side="left", padx=2)
        
        # 선택 반전 버튼
        self.invert_btn = tk.Button(self.button_frame, text="선택 반전", 
                                   command=self.invert_selection)
        self.invert_btn.pack(side="left", padx=2)
        
        # 항목 목록과 체크 상태 저장
        self.items = []  # 모든 항목
        self.checked = {}  # 체크 상태 {항목: 체크여부}
        
        # 마지막 선택 인덱스 (쉬프트 클릭용)
        self.last_selected_index = None
        
        # 이벤트 바인딩
        self.listbox.bind("<Button-1>", self.on_click)
        self.listbox.bind("<Shift-Button-1>", self.on_shift_click)
        self.listbox.bind("<space>", self.on_space)
        self.listbox.bind("<<ListboxSelect>>", self.on_selection_change)
    
    def add_item(self, text, checked=False):
        """항목 추가 - 기존 인터페이스 유지"""
        self.items.append(text)
        self.checked[text] = checked
        
        # 체크 표시와 함께 리스트박스에 추가
        self._update_display()
    
    def clear(self):
        """모든 항목 삭제 - 기존 인터페이스 유지"""
        self.listbox.delete(0, "end")
        self.items.clear()
        self.checked.clear()
    
    def get_checked_items(self):
        """체크된 항목 목록 반환 - 기존 인터페이스 유지"""
        return [item for item in self.items if self.checked.get(item, False)]
    
    def select_all(self):
        """모든 항목 선택"""
        for item in self.items:
            self.checked[item] = True
        self._update_display()
    
    def deselect_all(self):
        """모든 항목 선택 해제"""
        for item in self.items:
            self.checked[item] = False
        self._update_display()
    
    def invert_selection(self):
        """선택 상태 반전"""
        for item in self.items:
            self.checked[item] = not self.checked.get(item, False)
        self._update_display()
    
    def on_click(self, event):
        """클릭 이벤트 처리 - 체크박스 토글"""
        index = self.listbox.nearest(event.y)
        if index < 0 or index >= self.listbox.size():
            return
        
        # 체크박스 영역 클릭 여부 확인 (왼쪽 25픽셀)
        bbox = self.listbox.bbox(index)
        if bbox and event.x < 25:
            self._toggle_item_without_refresh(index)
        
        # 마지막 선택 인덱스 저장
        self.last_selected_index = index
    
    def on_shift_click(self, event):
        """쉬프트 클릭 이벤트 처리 - 다중 선택"""
        current_index = self.listbox.nearest(event.y)
        if current_index < 0 or current_index >= self.listbox.size():
            return
        
        # 마지막 선택 인덱스가 있는 경우
        if self.last_selected_index is not None:
            # 체크박스 영역 클릭 여부 확인
            bbox = self.listbox.bbox(current_index)
            if bbox and event.x < 25:
                # 범위 설정
                start = min(self.last_selected_index, current_index)
                end = max(self.last_selected_index, current_index)
                
                # 범위 내 모든 항목 토글
                for i in range(start, end + 1):
                    self._toggle_item_without_refresh(i)
        
        # 마지막 선택 인덱스 갱신
        self.last_selected_index = current_index
    
    def on_space(self, event):
        """스페이스바 이벤트 처리 - 선택 항목 토글, 포커스 유지"""
        selections = self.listbox.curselection()
        if selections:
            # 선택된 항목들의 원래 텍스트 저장
            selected_texts = [self.listbox.get(i) for i in selections]
            
            # 상태 토글
            for index in selections:
                self._toggle_item_without_refresh(index, update_display=False)  # 화면 갱신 없이 상태만 변경
            
            # 한 번만 화면 갱신
            self._update_display()
            
            # 원본 텍스트로 새 인덱스 찾기 및 선택 상태 복원
            for item_text in selected_texts:
                # 체크 표시를 제거한 원본 텍스트
                if item_text.startswith("☑ "):
                    original_text = item_text[2:]
                elif item_text.startswith("☐ "):
                    original_text = item_text[2:]
                else:
                    original_text = item_text
                
                # 새로운 텍스트 (토글 후 체크 상태)
                checked = self.checked.get(original_text, False)
                new_prefix = "☑ " if checked else "☐ "
                new_text = new_prefix + original_text
                
                # 새 목록에서 위치 찾기
                for i in range(self.listbox.size()):
                    if self.listbox.get(i) == new_text:
                        self.listbox.selection_set(i)
                        # 첫 번째 선택된 항목으로 보기 이동
                        if i == selections[0]:
                            self.listbox.see(i)
                        break

    # def _toggle_item(self, index, update_display=True):
    #     """항목 체크 상태 토글"""
    #     if index < 0 or index >= self.listbox.size():
    #         return
        
    #     # 리스트박스 항목 텍스트 가져오기
    #     text = self.listbox.get(index)
        
    #     # 체크 표시 제거하고 원본 텍스트 추출
    #     if text.startswith("☑ "):
    #         item_text = text[2:]
    #     elif text.startswith("☐ "):
    #         item_text = text[2:]
    #     else:
    #         item_text = text
        
    #     # 체크 상태 토글
    #     if item_text in self.checked:
    #         self.checked[item_text] = not self.checked.get(item_text, False)
        
    #     # 화면 갱신 (필요한 경우만)
    #     if update_display:
    #         current_text = text  # 현재 텍스트 저장
            
    #         self._update_display()
            
    #         # 토글 후 새 텍스트 형식
    #         new_prefix = "☑ " if self.checked.get(item_text, False) else "☐ "
    #         new_text = new_prefix + item_text
            
    #         # 새 목록에서 위치 찾기
    #         for i in range(self.listbox.size()):
    #             if self.listbox.get(i) == new_text:
    #                 self.listbox.selection_set(i)
    #                 self.listbox.see(i)  # 보이는 영역으로 스크롤
    #                 break

    def on_selection_change(self, event):
        """선택 변경 이벤트 처리"""
        pass  # 필요한 경우 구현
    
    def filter_list(self, event=None):
        """검색어로 목록 필터링"""
        self._update_display()


    def _toggle_item_without_refresh(self, index):
        """개별 항목의 체크 상태만 토글하고 전체 목록 갱신 없이 해당 항목만 업데이트"""
        if index < 0 or index >= self.listbox.size():
            return
        
        # 리스트박스 항목 텍스트 가져오기
        text = self.listbox.get(index)
        
        # 체크 표시 제거하고 원본 텍스트 추출
        if text.startswith("☑ "):
            item_text = text[2:]
            is_checked = True
        elif text.startswith("☐ "):
            item_text = text[2:]
            is_checked = False
        else:
            item_text = text
            is_checked = False
        
        # 체크 상태 토글
        if item_text in self.checked:
            self.checked[item_text] = not is_checked
        
        # 새 체크 상태에 따라 프리픽스 변경
        new_prefix = "☑ " if self.checked.get(item_text, False) else "☐ "
        new_text = new_prefix + item_text
        
        # 개별 항목만 업데이트
        self.listbox.delete(index)
        self.listbox.insert(index, new_text)
        
        # 선택 상태 유지
        self.listbox.selection_set(index)
        self.listbox.see(index)  # 해당 항목이 보이도록 스크롤 조정

    def on_space(self, event):
        """스페이스바 이벤트 처리 - 선택 항목 토글, 포커스 유지"""
        selections = list(self.listbox.curselection())
        for index in selections:
            self._toggle_item_without_refresh(index)
        
        # 마지막 선택 인덱스 갱신
        if selections:
            self.last_selected_index = selections[-1]

    def on_click(self, event):
        """클릭 이벤트 처리 - 체크박스 토글"""
        index = self.listbox.nearest(event.y)
        if index < 0 or index >= self.listbox.size():
            return
        
        # 체크박스 영역 클릭 여부 확인 (왼쪽 25픽셀)
        bbox = self.listbox.bbox(index)
        if bbox and event.x < 25:
            self._toggle_item_without_refresh(index)
        
        # 마지막 선택 인덱스 저장
        self.last_selected_index = index

    def on_shift_click(self, event):
        """쉬프트 클릭 이벤트 처리 - 다중 선택"""
        current_index = self.listbox.nearest(event.y)
        if current_index < 0 or current_index >= self.listbox.size():
            return
        
        # 마지막 선택 인덱스가 있는 경우
        if self.last_selected_index is not None:
            # 체크박스 영역 클릭 여부 확인
            bbox = self.listbox.bbox(current_index)
            if bbox and event.x < 25:
                # 범위 설정
                start = min(self.last_selected_index, current_index)
                end = max(self.last_selected_index, current_index)
                
                # 범위 내 모든 항목 토글
                for i in range(start, end + 1):
                    self._toggle_item_without_refresh(i)
        
        # 마지막 선택 인덱스 갱신
        self.last_selected_index = current_index

    def _update_display(self):
        """리스트박스 표시 업데이트"""
        # 현재 스크롤 위치 저장
        current_selection = self.listbox.curselection()
        yview = self.listbox.yview()
        
        # 검색어 가져오기
        search_text = self.search_var.get().lower()
        
        # 리스트박스 초기화
        self.listbox.delete(0, "end")
        
        # 필터링하여 항목 추가
        for item in self.items:
            if not search_text or search_text in item.lower():
                checked = self.checked.get(item, False)
                prefix = "☑ " if checked else "☐ "
                self.listbox.insert("end", prefix + item)
        
        # 스크롤 위치 복원
        if yview:
            self.listbox.yview_moveto(yview[0])
        
        # 선택 상태 복원
        for i in current_selection:
            if i < self.listbox.size():
                self.listbox.selection_set(i)

class ExcelDiffTool:
    def __init__(self, root):
        self.root = root
        self.root.title("엑셀 데이터 Diff 툴")
        self.root.geometry("900x700")
            
        # 설정 로드 - 절대 경로 사용
        config_path = os.path.join(root_dir, "config.json")
        self.config = load_config(config_path)
        
        # 변수 초기화
        self.source_path = tk.StringVar(value=self.config.get("source_path", ""))
        self.target_path = tk.StringVar(value=self.config.get("target_path", ""))
        self.source_cache = {}
        self.target_cache = {}
        self.diff_results = []
        
        # UI 빌드
        self.build_ui()

        # 저장된 경로가 있으면 파일 목록 자동 업데이트
        if self.source_path.get() and os.path.exists(self.source_path.get()):
            self.root.after(100, lambda: self.update_file_list(self.source_path.get()))


    def build_ui(self):
        """UI 구성 - 좌/우 분할 레이아웃"""
        # 상단 프레임 - 경로 선택
        self.build_path_ui()
        
        # 중간 프레임 - 파일 목록과 옵션 (좌/우 분할)
        self.build_file_and_option_ui()
        
        # 하단 프레임 - 결과 표시
        self.build_result_ui()


    def build_path_ui(self):
        """경로 선택 UI 구성"""
        path_frame = tk.LabelFrame(self.root, text="파일/폴더 선택")
        path_frame.pack(fill="x", padx=10, pady=5)
        
        # 원본 경로
        source_frame = tk.Frame(path_frame)
        source_frame.pack(fill="x", padx=5, pady=5)
        
        tk.Label(source_frame, text="원본 경로:").grid(row=0, column=0, padx=5, pady=5)
        source_entry = tk.Entry(source_frame, textvariable=self.source_path, width=50)
        source_entry.grid(row=0, column=1, padx=5, pady=5)
        
        # 엔터 키 이벤트 추가
        source_entry.bind('<Return>', lambda event: self.update_file_list(self.source_path.get()))
        
        tk.Button(source_frame, text="폴더 선택", command=lambda: self.select_folder("source")).grid(row=0, column=2, padx=2, pady=5)
        tk.Button(source_frame, text="파일 선택", command=lambda: self.select_file("source")).grid(row=0, column=3, padx=2, pady=5)
        
        # 직접 업데이트 버튼 추가
        tk.Button(source_frame, text="목록 새로고침", 
                command=lambda: self.update_file_list(self.source_path.get())).grid(row=0, column=4, padx=2, pady=5)
        
        # 비교본 경로
        target_frame = tk.Frame(path_frame)
        target_frame.pack(fill="x", padx=5, pady=5)
        
        tk.Label(target_frame, text="비교본 경로:").grid(row=0, column=0, padx=5, pady=5)
        tk.Entry(target_frame, textvariable=self.target_path, width=50).grid(row=0, column=1, padx=5, pady=5)
        tk.Button(target_frame, text="폴더 선택", command=lambda: self.select_folder("target")).grid(row=0, column=2, padx=2, pady=5)
        tk.Button(target_frame, text="파일 선택", command=lambda: self.select_file("target")).grid(row=0, column=3, padx=2, pady=5)
        
        # # 파일 목록 선택 프레임 추가
        # self.build_file_list_ui()


    def build_file_and_option_ui(self):
        """파일 목록과 옵션 UI를 좌/우로 분할"""
        # 메인 프레임
        main_frame = tk.Frame(self.root)
        main_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        # 좌측 프레임 - 파일 목록
        left_frame = tk.LabelFrame(main_frame, text="비교 대상 파일")
        left_frame.pack(side="left", fill="both", expand=True, padx=(0, 5))
        
        # 소스 파일 목록 프레임
        source_list_frame = tk.Frame(left_frame)
        source_list_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # 소스 파일 목록 레이블
        tk.Label(source_list_frame, text="원본 파일 목록:").pack(anchor="w")
        
        # 개선된 파일 목록 사용
        self.source_file_list = ImprovedFileList(source_list_frame, width=30, height=15)
        self.source_file_list.pack(fill="both", expand=True)
        
        # 도움말 추가
        help_text = "팁: 좌측 체크박스 클릭, 스페이스바, 쉬프트+클릭으로 여러 항목 선택 가능"
        tk.Label(left_frame, text=help_text, fg="gray", anchor="w", 
                font=("Arial", 8)).pack(fill="x", padx=5, pady=(0, 5))
        
        # 우측 프레임 - 옵션 및 버튼
        right_frame = tk.Frame(main_frame)
        right_frame.pack(side="right", fill="both", padx=(5, 0))
        
        # 비교 옵션 프레임
        option_frame = tk.LabelFrame(right_frame, text="비교 옵션")
        option_frame.pack(fill="x", pady=(0, 5))
        
        # 옵션들
        self.include_added_var = tk.BooleanVar(value=True)
        self.include_removed_var = tk.BooleanVar(value=True)
        self.include_modified_var = tk.BooleanVar(value=True)
        
        options_frame = tk.Frame(option_frame)
        options_frame.pack(fill="x", padx=5, pady=5)
        
        # 옵션 체크박스를 세로로 배치
        tk.Checkbutton(options_frame, text="추가된 항목 포함", 
                    variable=self.include_added_var).pack(anchor="w", pady=2)
        tk.Checkbutton(options_frame, text="삭제된 항목 포함", 
                    variable=self.include_removed_var).pack(anchor="w", pady=2)
        tk.Checkbutton(options_frame, text="변경된 항목 포함", 
                    variable=self.include_modified_var).pack(anchor="w", pady=2)
        
        # 실행 버튼 프레임
        button_frame = tk.LabelFrame(right_frame, text="실행")
        button_frame.pack(fill="x", pady=5)
        
        # 버튼을 세로로 배치
        tk.Button(button_frame, text="비교 시작", 
                command=self.start_comparison, 
                width=15, height=2).pack(pady=(5, 2), padx=10)
        
        tk.Button(button_frame, text="결과 내보내기", 
                command=self.export_results, 
                width=15, height=2).pack(pady=(2, 5), padx=10)
        
        # 추가 도움말 프레임
        help_frame = tk.LabelFrame(right_frame, text="도움말")
        help_frame.pack(fill="both", expand=True, pady=5)
        
        # 도움말 내용
        help_info = (
            "1. 원본 및 비교본 경로를 선택하세요.\n"
            "2. 비교할 파일을 체크하세요.\n"
            "3. 비교 옵션을 설정하세요.\n"
            "4. '비교 시작' 버튼을 클릭하세요.\n"
            "5. 결과를 엑셀로 내보낼 수 있습니다."
        )
        tk.Label(help_frame, text=help_info, justify="left", 
                padx=5, pady=5).pack(fill="both", expand=True)

   
    # def build_file_list_ui(self):
    #     """파일 목록 선택 UI 구성 - 기존 UI 연결 유지"""
    #     files_frame = tk.LabelFrame(self.root, text="비교 대상 파일")
    #     files_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
    #     # 소스 파일 목록 프레임
    #     source_list_frame = tk.Frame(files_frame)
    #     source_list_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
    #     # 소스 파일 목록 레이블
    #     tk.Label(source_list_frame, text="원본 파일 목록:").pack(anchor="w")
        
    #     # 개선된 파일 목록 사용 - 기존 인터페이스 유지
    #     self.source_file_list = ImprovedFileList(source_list_frame, width=40, height=15)
    #     self.source_file_list.pack(fill="both", expand=True)
        
    #     # 도움말 추가
    #     help_text = "팁: 좌측 체크박스 클릭, 스페이스바, 쉬프트+클릭으로 여러 항목 선택 가능"
    #     tk.Label(files_frame, text=help_text, fg="gray", anchor="w", 
    #             font=("Arial", 8)).pack(fill="x", padx=5, pady=(0, 5))


    def build_result_ui(self):
        """결과 표시 UI 구성"""
        result_frame = tk.LabelFrame(self.root, text="비교 결과")
        result_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        # 요약 결과 트리뷰
        self.summary_tree = ttk.Treeview(result_frame, columns=("파일", "신규", "수정", "삭제"), show="headings")
        
        # 스크롤바
        vsb = ttk.Scrollbar(result_frame, orient="vertical", command=self.summary_tree.yview)
        hsb = ttk.Scrollbar(result_frame, orient="horizontal", command=self.summary_tree.xview)
        self.summary_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # 컬럼 설정
        self.summary_tree.heading("파일", text="파일")
        self.summary_tree.heading("신규", text="신규")
        self.summary_tree.heading("수정", text="수정")
        self.summary_tree.heading("삭제", text="삭제")
        
        # 컬럼 너비
        self.summary_tree.column("파일", width=250)
        self.summary_tree.column("신규", width=80)
        self.summary_tree.column("수정", width=80)
        self.summary_tree.column("삭제", width=80)
        
        # 스크롤바 배치
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.summary_tree.pack(side="left", fill="both", expand=True)
        
        # 더블 클릭 이벤트 바인딩
        self.summary_tree.bind("<Double-1>", self.show_detail_popup)
        
        # 상태 표시
        self.status_var = tk.StringVar(value="준비 완료")
        tk.Label(self.root, textvariable=self.status_var).pack(fill="x", padx=10, pady=5)


    def select_folder(self, target):
        """폴더 선택 다이얼로그"""
        folder = filedialog.askdirectory()
        if folder:
            if target == "source":
                self.source_path.set(folder)
                self.update_file_list(folder)  # 3번째 인수 제거
                
                # 비교본 경로도 같은 경로로 설정
                self.target_path.set(folder)
            else:
                self.target_path.set(folder)
            
            # 설정 저장
            self.save_paths()


    def select_file(self, target):
        """파일 선택 다이얼로그"""
        file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
        if file:
            if target == "source":
                self.source_path.set(file)
                # 파일 선택 시에도 파일 목록 업데이트
                self.update_file_list(file)  # 3번째 인수 제거
                
                # 비교본 경로도 같은 파일로 설정
                self.target_path.set(file)
            else:
                self.target_path.set(file)
            
            # 설정 저장
            self.save_paths()

#1
    def update_file_list(self, folder_path):
        """선택된 폴더의 파일 목록 업데이트 - Excel_ 시작 폴더 필터링 적용, String_ 시작 파일 및 Excel_String 폴더 제외"""
        print(f"파일 목록 업데이트 시작: {folder_path}")
        
        if not folder_path or not os.path.exists(folder_path):
            print(f"경로가 존재하지 않음: {folder_path}")
            return
        
        try:
            # ScrollableCheckList 객체가 있는지 확인
            if not hasattr(self, 'source_file_list'):
                print("source_file_list 속성이 없음")
                return
                
            # 파일 목록 클리어
            self.source_file_list.clear()
            
            # 단일 파일인 경우
            if os.path.isfile(folder_path):
                file_name = os.path.basename(folder_path)
                
                # String으로 시작하는 파일 제외
                if file_name.startswith("String_"):
                    print(f"String_ 시작 파일 제외: {file_name}")
                    self.status_var.set(f"String_ 시작 파일은 제외됩니다: {file_name}")
                    return
                    
                print(f"단일 파일 발견: {file_name}")
                self.source_file_list.add_item(file_name, checked=True)
                self.status_var.set(f"파일 목록 업데이트 완료: {file_name}")
                return
            
            # Excel_ 시작하는 폴더 필터링 및 파일 수집
            excel_files = []
            excluded_string_files = 0
            excluded_string_folders = 0
            
            # 루트 폴더 자체에서 파일 검색
            for file in os.listdir(folder_path):
                file_path = os.path.join(folder_path, file)
                # 폴더인 경우 Excel_ 시작 여부 확인
                if os.path.isdir(file_path):
                    # Excel_String 폴더는 제외
                    if file.startswith("Excel_String"):
                        excluded_string_folders += 1
                        print(f"Excel_String 폴더 제외: {file}")
                        continue
                        
                    # 일반 Excel_ 폴더 처리    
                    elif file.startswith("Excel_"):
                        # Excel_ 폴더 내 엑셀 파일 검색
                        for root, _, files in os.walk(file_path):
                            for excel_file in files:
                                # String으로 시작하는 파일 제외
                                if excel_file.startswith("String_"):
                                    excluded_string_files += 1
                                    continue
                                    
                                if excel_file.endswith(('.xlsx', '.xls')) and not excel_file.startswith('~$'):
                                    file_full_path = os.path.join(root, excel_file)
                                    rel_path = os.path.relpath(file_full_path, folder_path).replace('\\', '/')
                                    excel_files.append(rel_path)
                # 파일인 경우 엑셀 파일인지 확인하여 추가
                elif file.endswith(('.xlsx', '.xls')) and not file.startswith('~$'):
                    # String으로 시작하는 파일 제외
                    if file.startswith("String_"):
                        excluded_string_files += 1
                        continue
                        
                    rel_path = file
                    excel_files.append(rel_path)
            
            print(f"발견된 엑셀 파일 수: {len(excel_files)}")
            print(f"제외된 항목: String_ 시작 파일 {excluded_string_files}개, Excel_String 폴더 {excluded_string_folders}개")
            
            # 발견된 파일이 없을 경우 메시지 표시
            if not excel_files:
                if excluded_string_files > 0 or excluded_string_folders > 0:
                    self.status_var.set(f"모든 파일이 제외되었습니다. (String_ 파일: {excluded_string_files}개, Excel_String 폴더: {excluded_string_folders}개)")
                else:
                    self.status_var.set("Excel_ 시작하는 폴더가 없거나 엑셀 파일이 없습니다.")
                return
            
            # 파일명 정렬 후 추가
            for file_path in sorted(excel_files):
                self.source_file_list.add_item(file_path, checked=True)
            
            # 상태 업데이트
            exclude_msg = []
            if excluded_string_files > 0:
                exclude_msg.append(f"{excluded_string_files}개 String_ 파일")
            if excluded_string_folders > 0:
                exclude_msg.append(f"{excluded_string_folders}개 Excel_String 폴더")
                
            if exclude_msg:
                self.status_var.set(f"파일 목록 업데이트 완료: {len(excel_files)}개 파일 발견, {' 및 '.join(exclude_msg)} 제외됨")
            else:
                self.status_var.set(f"파일 목록 업데이트 완료: {len(excel_files)}개 파일 발견")
        except Exception as e:
            error_msg = f"파일 목록 업데이트 오류: {str(e)}"
            print(error_msg)
            import traceback
            traceback.print_exc()
            self.status_var.set(error_msg)
            messagebox.showerror("파일 목록 오류", f"파일 목록 업데이트 중 오류가 발생했습니다:\n{str(e)}")


    def save_paths(self):
        """경로 설정 저장"""
        # 현재 설정에 경로 업데이트
        self.config["source_path"] = self.source_path.get()
        self.config["target_path"] = self.target_path.get()
        
        # 설정 저장 - 절대 경로 사용
        config_path = os.path.join(root_dir, "config.json")
        save_config(config_path, self.config)
        
        # 디버깅 메시지
        print(f"설정 저장 완료: {config_path}")
        self.status_var.set(f"설정 저장 완료: {os.path.basename(config_path)}")

    
    def show_detail_popup(self, event):
        """파일 상세 정보 팝업 - 컬럼 정렬 순서 및 PK 표시 개선"""
        item = self.summary_tree.focus()
        if not item:
            return
        
        # 선택된 파일 정보 가져오기
        values = self.summary_tree.item(item, "values")
        file_name = values[0]
        
        # 해당 파일의 상세 결과 필터링
        file_results = [r for r in self.diff_results if r[0] == file_name]
        if not file_results:
            messagebox.showinfo("정보", "해당 파일의 상세 정보가 없습니다.")
            return
        
        # 팝업 창 생성
        popup = tk.Toplevel(self.root)
        popup.title(f"상세 비교 결과 - {file_name}")
        popup.geometry("1200x700")
        
        # ----- 1. 시트별 PK 컬럼 식별 -----
        sheet_pk_columns = {}
        for result in file_results:
            _, sheet_name, _, _, _, _, _ = result
            if sheet_name not in sheet_pk_columns:
                # DB 이름 추출 (시트 이름에서)
                db_name = sheet_name.split('@')[0].replace(".xlsx", "")
                # PK 정보 조회
                pk_columns = self.get_pk_from_excel_cache(db_name, sheet_name=sheet_name)
                sheet_pk_columns[sheet_name] = pk_columns
                print(f"[{sheet_name}] PK 컬럼: {pk_columns}")
        
        # ----- 2. 시트별 원본 컬럼 순서 가져오기 -----
        sheet_columns_order = {}
        for sheet_name in set(result[1] for result in file_results):
            column_order = []
            
            # 2.1 원본 Excel 파일에서 컬럼 순서 가져오기
            try:
                # 원본 파일 경로
                source_file_path = self.source_path.get()
                if os.path.isdir(source_file_path):
                    # 폴더인 경우 해당 파일 찾기
                    for root, _, files in os.walk(source_file_path):
                        for file in files:
                            if file == file_name or file.endswith(file_name):
                                source_file_path = os.path.join(root, file)
                                break
                
                # 엑셀 파일 열기
                if os.path.isfile(source_file_path):
                    with pd.ExcelFile(source_file_path) as xls:
                        if sheet_name in xls.sheet_names:
                            # 캐시에서 헤더 행 정보 가져오기
                            header_row = self.get_sheet_header_row_from_cache(source_file_path, sheet_name, self.source_cache)
                            # 엑셀에서 컬럼 순서 가져오기
                            df = pd.read_excel(xls, sheet_name=sheet_name, header=header_row)
                            column_order = df.columns.tolist()
                            print(f"[{sheet_name}] 엑셀에서 컬럼 순서 가져옴: {column_order[:5]}...")
            except Exception as e:
                print(f"[{file_name} - {sheet_name}] 원본 컬럼 순서 가져오기 실패: {e}")
            
            # 2.2 Excel_Indent.json에서 컬럼 순서 가져오기 (원본 파일에서 가져오지 못한 경우)
            if not column_order:
                try:
                    indent_path = os.path.join(root_dir, "Excel_Indent.json")
                    if os.path.exists(indent_path):
                        with open(indent_path, 'r', encoding='utf-8') as f:
                            indent_data = json.load(f)
                        
                        # DB 이름 추출 
                        db_name = sheet_name.split('@')[0].replace(".xlsx", "")
                        
                        # Excel_Indent.json에서 컬럼 순서 확인
                        if db_name in indent_data and "Columns" in indent_data[db_name]:
                            column_order = indent_data[db_name]["Columns"]
                            print(f"[{sheet_name}] Excel_Indent.json에서 컬럼 순서 가져옴: {column_order[:5]}...")
                except Exception as e:
                    print(f"[{file_name} - {sheet_name}] Excel_Indent.json에서 컬럼 순서 가져오기 실패: {e}")
            
            # 2.3 DB에서 컬럼 순서 가져오기 (다른 방법이 실패한 경우)
            if not column_order:
                try:
                    # DB 이름 추출
                    db_name = sheet_name.split('@')[0].replace(".xlsx", "")
                    # 데이터베이스 파일 경로 추정
                    source_dir = os.path.dirname(source_file_path) if os.path.isfile(source_file_path) else source_file_path
                    db_path = os.path.join(source_dir, f"{db_name}.db")
                    if os.path.exists(db_path):
                        columns = get_columns_from_db_safe(db_path, sheet_name)
                        if columns:
                            column_order = columns
                            print(f"[{sheet_name}] DB에서 컬럼 순서 가져옴: {column_order[:5]}...")
                except Exception as e:
                    print(f"[{file_name} - {sheet_name}] DB에서 컬럼 순서 가져오기 실패: {e}")
            
            # 컬럼 순서 저장
            sheet_columns_order[sheet_name] = column_order
        
        # ----- 3. 데이터 그룹화 -----
        sheets_data = {}
        for result in file_results:
            _, sheet, row, col, source_val, target_val, status = result
            
            if sheet not in sheets_data:
                sheets_data[sheet] = {}
            
            # 행 ID 사용
            row_key = row
            
            # 각 시트마다 행별로 데이터 정리
            if row_key not in sheets_data[sheet]:
                sheets_data[sheet][row_key] = {}
                
            # 각 행의 컬럼 데이터 저장 (원본/비교본 값 별도 저장)
            sheets_data[sheet][row_key][col] = {
                "원본": source_val,
                "비교본": target_val,
                "상태": status
            }
        
        # ----- 4. 모든 컬럼 수집 및 정렬 -----
        # 4.1 모든 시트의 모든 컬럼 목록 수집
        all_columns = set()
        for sheet_name, sheet_data in sheets_data.items():
            # PK 컬럼 추가
            pk_cols = sheet_pk_columns.get(sheet_name, [])
            all_columns.update(pk_cols)
            
            # 모든 행의 모든 컬럼 추가
            for row_data in sheet_data.values():
                all_columns.update(row_data.keys())
        
        # 4.2 컬럼 순서 정렬
        sorted_columns = []
        
        # 우선 모든 시트의 원본 컬럼 순서를 종합
        all_ordered_columns = []
        for sheet_name, column_order in sheet_columns_order.items():
            for col in column_order:
                if col in all_columns and col not in all_ordered_columns:
                    all_ordered_columns.append(col)
        
        # PK 컬럼 추가 (이미 추가되지 않은 경우)
        all_pk_columns = []
        for pk_cols in sheet_pk_columns.values():
            for pk_col in pk_cols:
                if pk_col not in all_pk_columns:
                    all_pk_columns.append(pk_col)
        
        # PK 컬럼 먼저 추가
        for pk_col in all_pk_columns:
            if pk_col in all_columns:
                if pk_col not in sorted_columns:
                    sorted_columns.append(pk_col)
                    all_columns.discard(pk_col)
        
        # 원본 순서 컬럼 추가
        for col in all_ordered_columns:
            if col in all_columns:
                sorted_columns.append(col)
                all_columns.discard(col)
        
        # 나머지 컬럼은 알파벳순 추가
        sorted_columns.extend(sorted(all_columns))
        
        # ----- 5. 트리뷰 생성 -----
        # 5.1 트리뷰 컬럼 구성: 파일|시트|행 + 모든 컬럼(각 컬럼은 원본/비교본 분리)
        tree_columns = ["파일", "시트", "행ID"]  # 기본 컬럼
        
        # 각 데이터 컬럼에 대해 원본/비교본 컬럼 추가
        for col in sorted_columns:
            tree_columns.append(f"{col} (원본)")
            tree_columns.append(f"{col} (비교본)")
        
        # 5.2 메인 프레임 생성
        main_frame = tk.Frame(popup)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # 5.3 트리뷰 생성
        tree = ttk.Treeview(main_frame, columns=tree_columns, show="headings")
        
        # 스크롤바
        vsb = ttk.Scrollbar(main_frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(main_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # 5.4 컬럼 헤더 설정
        for col in tree_columns:
            tree.heading(col, text=col)
            # 컬럼 너비 설정
            width = 80  # 기본 너비
            if col == "파일":
                width = 150
            elif col == "시트":
                width = 150
            elif col == "행ID":
                width = 60
            tree.column(col, width=width, stretch=True if "원본" in col or "비교본" in col else False)
        
        # ----- 6. 데이터 표시 -----
        # 6.1 숫자 형식 처리 함수
        def format_value(val):
            """값을 적절한 형식으로 변환 (소수점 0 제거)"""
            if pd.isna(val):
                return ""
            
            # 소수점 0 제거
            if isinstance(val, (int, float)):
                if val == int(val):
                    return str(int(val))
            
            # 문자열에서 .0 패턴 제거
            if isinstance(val, str) and val.endswith('.0') and val[:-2].isdigit():
                return val[:-2]
            
            return str(val)
        
        # 6.2 각 시트의 데이터를 하나의 트리뷰에 추가
        row_id_counter = 0
        for sheet_name, sheet_data in sorted(sheets_data.items()):
            # 행 데이터를 정렬하여 추가
            for row_key, row_data in sorted(sheet_data.items(), key=lambda x: x[0]):
                # 행 기본 정보 (파일, 시트, 행ID)
                row_values = [file_name, sheet_name, row_key]
                
                # 각 컬럼의 원본/비교본 값 별도로 추가
                for col in sorted_columns:
                    source_val = ""
                    target_val = ""
                    
                    if col in row_data:
                        cell_data = row_data[col]
                        source_val = format_value(cell_data['원본'])
                        target_val = format_value(cell_data['비교본'])
                    
                    row_values.append(source_val)  # 원본 값
                    row_values.append(target_val)  # 비교본 값
                
                # 트리에 행 추가
                row_id = tree.insert("", "end", values=row_values)
                row_id_counter += 1
                
                # 상태에 따라 태그 적용
                status_tags = set(cell_data["상태"] for cell_data in row_data.values() if "상태" in cell_data)
                if "추가됨" in status_tags:
                    tree.item(row_id, tags=("added",))
                elif "삭제됨" in status_tags:
                    tree.item(row_id, tags=("removed",))
                elif "변경됨" in status_tags:
                    tree.item(row_id, tags=("modified",))
        
        # ----- 7. UI 완성 -----
        # 7.1 스크롤바 배치
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        tree.pack(side="left", fill="both", expand=True)
        
        # 7.2 태그 색상 설정
        tree.tag_configure("added", background="#e6ffec")
        tree.tag_configure("removed", background="#ffebe9")
        tree.tag_configure("modified", background="#fcf3dc")
        
        # 7.3 필터 프레임 추가
        filter_frame = tk.Frame(popup)
        filter_frame.pack(fill="x", padx=10, pady=(5, 10))
        
        # 컬럼 자동 조정 버튼
        ttk.Button(filter_frame, text="컬럼 크기 자동 조정", 
                command=lambda: self.auto_adjust_columns(tree)).pack(side="left", padx=5)
        
        # 필터 입력창
        tk.Label(filter_frame, text="필터:").pack(side="left", padx=(10, 5))
        filter_var = tk.StringVar()
        filter_entry = ttk.Entry(filter_frame, textvariable=filter_var, width=30)
        filter_entry.pack(side="left", padx=5)
        
        # 필터 적용 버튼
        ttk.Button(filter_frame, text="적용", 
                command=lambda: self._filter_detail_tree(tree, filter_var.get())).pack(side="left", padx=5)
        
        # 필터 초기화 버튼
        ttk.Button(filter_frame, text="초기화", 
                command=lambda: self._reset_detail_filter(tree, filter_var)).pack(side="left", padx=5)
        
        # 엑셀 내보내기 버튼
        ttk.Button(filter_frame, text="엑셀로 내보내기", 
                command=lambda: self.export_sheet_results(file_name, [sheet for sheet in sheets_data.keys()])).pack(side="right", padx=5)
        
        # 7.4 상태 정보 표시
        status_label = tk.Label(popup, text=f"총 {row_id_counter}개 행이 표시됨", anchor="w")
        status_label.pack(fill="x", padx=10, pady=5)


    def get_excel_column_order(self, file_path, sheet_name):
        """엑셀 파일에서 컬럼 순서 가져오기"""
        try:
            if os.path.isfile(file_path):
                with pd.ExcelFile(file_path) as xls:
                    if sheet_name in xls.sheet_names:
                        # 캐시에서 헤더 행 정보 가져오기
                        header_row = self.get_sheet_header_row_from_cache(file_path, sheet_name, self.source_cache)
                        # 엑셀에서 컬럼 순서 가져오기 
                        df = pd.read_excel(xls, sheet_name=sheet_name, header=header_row)
                        return df.columns.tolist()
        except Exception as e:
            print(f"엑셀 컬럼 순서 가져오기 실패: {file_path} - {sheet_name} - {e}")
        return []

    def get_db_column_order(self, db_path, table_name):
        """DB에서 컬럼 순서 가져오기"""
        try:
            if os.path.exists(db_path):
                return get_columns_from_db_safe(db_path, table_name)
        except Exception as e:
            print(f"DB 컬럼 순서 가져오기 실패: {db_path} - {table_name} - {e}")
        return []

    def get_indent_column_order(self, db_name):
        """Excel_Indent.json에서 컬럼 순서 가져오기"""
        try:
            indent_path = os.path.join(root_dir, "Excel_Indent.json")
            if os.path.exists(indent_path):
                with open(indent_path, 'r', encoding='utf-8') as f:
                    indent_data = json.load(f)
                
                if db_name in indent_data and "Columns" in indent_data[db_name]:
                    return indent_data[db_name]["Columns"]
        except Exception as e:
            print(f"Excel_Indent 컬럼 순서 가져오기 실패: {db_name} - {e}")
        return []

    def _filter_detail_tree(self, tree, filter_text):
        """상세 트리뷰 필터링"""
        if not filter_text:
            # 필터 텍스트가 없으면 모든 항목 표시
            for item in tree.get_children():
                tree.item(item, open=True)
                tree.item(item, tags=tree.item(item, "tags"))  # 원래 태그 유지
            return
        
        # 모든 항목을 순회하며 필터링
        filter_text = filter_text.lower()
        count = 0
        for item in tree.get_children():
            # 항목의 모든 컬럼 값을 가져와 검색
            values = tree.item(item, "values")
            item_text = " ".join(str(v) for v in values).lower()
            
            # 필터 텍스트가 포함되어 있는지 확인
            if filter_text in item_text:
                tree.item(item, open=True)
                tree.item(item, tags=(tree.item(item, "tags") or ()) + ("filtered",))  # 필터 태그 추가
                count += 1
            else:
                tree.detach(item)  # 항목 숨기기
        
        # 필터 태그 스타일 설정
        tree.tag_configure("filtered", background="#f0f8ff")  # 연한 파란색 배경
        
        # 상태 업데이트
        status_label = tree.master.master.winfo_children()[-1]  # 상태 라벨은 마지막 위젯
        status_label.config(text=f"필터 적용됨: {count}개 항목 표시")

    def _reset_detail_filter(self, tree, filter_var):
        """상세 트리뷰 필터 초기화"""
        # 필터 입력 필드 초기화
        filter_var.set("")
        
        # 숨겨진 모든 항목 다시 표시
        for item in tree.detached():
            tree.reattach(item, "", 0)  # 루트에 다시 부착
        
        # 원래 태그 복원
        for item in tree.get_children():
            tags = tree.item(item, "tags")
            if tags and "filtered" in tags:
                new_tags = tuple(t for t in tags if t != "filtered")
                tree.item(item, tags=new_tags)
        
        # 상태 업데이트
        status_label = tree.master.master.winfo_children()[-1]
        status_label.config(text=f"총 {len(tree.get_children())}개 행이 표시됨")


    def auto_adjust_columns(self, tree):
        """트리뷰 컬럼 크기 자동 조정"""
        for column in tree["columns"]:
            tree.column(column, width=None)  # 자동 크기 조정 위해 초기화
            
            # 컬럼 헤더 너비 계산
            header_width = len(tree.heading(column)["text"]) * 10
            
            # 각 행의 해당 컬럼 값 너비 계산
            max_width = header_width
            for item in tree.get_children():
                idx = tree["columns"].index(column)
                if idx < len(tree.item(item)["values"]):
                    cell_value = str(tree.item(item)["values"][idx])
                    cell_width = len(cell_value) * 8
                    max_width = max(max_width, cell_width)
            
            # 최대 300px, 최소 80px로 제한
            max_width = min(max(max_width, 80), 300)
            tree.column(column, width=max_width)


    def export_sheet_results(self, file_name, sheets):
        """선택한 파일의 시트별 결과를 엑셀로 내보내기"""
        try:
            # 저장할 파일 경로 선택
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel 파일", "*.xlsx")],
                initialfile=f"{file_name}_비교결과.xlsx",
                title="비교 결과 저장"
            )
            
            if not save_path:
                return  # 사용자가 취소함
            
            # 해당 파일 결과 필터링
            file_results = [r for r in self.diff_results if r[0] == file_name]
            
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment
            
            # 워크북 생성
            wb = Workbook()
            # 기본 시트 제거
            wb.remove(wb.active)
            
            # 변경 상태별 셀 스타일
            added_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 녹색
            removed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 빨간색
            modified_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 노란색
            
            # 각 시트별 처리
            for sheet_name in sorted(sheets):
                # 시트 생성 (이름 길이 제한)
                safe_sheet_name = sheet_name[:31]  # Excel 시트 이름 길이 제한
                ws = wb.create_sheet(title=safe_sheet_name)
                
                # 시트별 결과 필터링
                sheet_results = [r for r in file_results if r[1] == sheet_name]
                
                # 데이터 구조화
                sheet_data = {}
                columns = set()
                
                for result in sheet_results:
                    # 결과 형식: (file_name, sheet_name, row_id, col, source_val, target_val, status)
                    _, _, row_id, col, source_val, target_val, status = result
                    
                    if row_id not in sheet_data:
                        sheet_data[row_id] = {}
                        
                    columns.add(col)
                    
                    sheet_data[row_id][col] = {
                        "원본": source_val,
                        "비교본": target_val,
                        "상태": status
                    }
                
                # 헤더 추가
                headers = ["PK/ID"] + sorted(columns)
                ws.append(headers)
                
                # 헤더 스타일 적용
                for col_idx, _ in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                # 데이터 추가
                for row_id, row_data in sorted(sheet_data.items(), key=lambda x: x[0]):
                    row_values = [row_id]  # PK/ID 값
                    
                    # 각 컬럼의 값 추가
                    for col in sorted(columns):
                        if col in row_data:
                            source_val = row_data[col]["원본"]
                            target_val = row_data[col]["비교본"]
                            status = row_data[col]["상태"]
                            
                            # null 값 처리
                            if pd.isna(source_val):
                                source_val = ""
                            if pd.isna(target_val):
                                target_val = ""
                            
                            # 변경 상태에 따라 값 표시
                            if str(source_val) == str(target_val):
                                cell_value = source_val
                                cell_fill = None
                            else:
                                if "추가" in status:
                                    cell_value = f"[추가] {target_val}"
                                    cell_fill = added_fill
                                elif "삭제" in status:
                                    cell_value = f"[삭제] {source_val}"
                                    cell_fill = removed_fill
                                else:
                                    cell_value = f"{source_val} → {target_val}"
                                    cell_fill = modified_fill
                            
                            row_values.append(cell_value)
                        else:
                            row_values.append("")
                    
                    # 엑셀에 행 추가
                    row_idx = ws.max_row + 1
                    for col_idx, value in enumerate(row_values, 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = value
                        
                        # 변경된 컬럼에 스타일 적용
                        col_name = headers[col_idx-1] if col_idx > 0 and col_idx <= len(headers) else None
                        if col_name and col_name in row_data and str(row_data[col_name]["원본"]) != str(row_data[col_name]["비교본"]):
                            if "추가" in row_data[col_name]["상태"]:
                                cell.fill = added_fill
                            elif "삭제" in row_data[col_name]["상태"]:
                                cell.fill = removed_fill
                            else:
                                cell.fill = modified_fill
                
                # 컬럼 너비 자동 조정
                for col_idx, column in enumerate(headers, 1):
                    max_length = len(str(column)) + 2
                    for row_idx in range(2, ws.max_row + 1):
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if cell_value:
                            max_length = max(max_length, len(str(cell_value)) + 2)
                    
                    # 최대 너비 제한
                    col_width = min(max_length, 50)
                    column_letter = ws.cell(row=1, column=col_idx).column_letter
                    ws.column_dimensions[column_letter].width = col_width
            
            # 파일 저장
            wb.save(save_path)
            
            messagebox.showinfo("내보내기 완료", f"비교 결과가 '{save_path}'에 저장되었습니다.")
            
        except Exception as e:
            messagebox.showerror("내보내기 오류", f"결과 내보내기 중 오류가 발생했습니다:\n{str(e)}")
            logging.error(f"시트별 결과 내보내기 오류: {e}")
            import traceback
            traceback.print_exc()


    def export_results(self):
        """비교 결과를 엑셀 파일로 내보내기"""
        if not self.diff_results:
            messagebox.showinfo("알림", "내보낼 비교 결과가 없습니다.")
            return
        
        # 저장할 파일 경로 선택
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            title="비교 결과 저장"
        )
        
        if not file_path:
            return  # 사용자가 취소함
        
        try:
            # 로딩 팝업 표시
            loading_popup = LoadingPopup(self.root, title="내보내기 중", message="결과 파일 생성 중...")
            
            # 엑셀 파일 생성 준비
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font
            
            # 워크북 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "요약"
            
            # 요약 시트 헤더
            ws.append(["파일", "시트", "추가", "수정", "삭제", "총 변경"])
            
            # 파일별 시트별 변경사항 집계
            file_sheet_changes = {}
            for result in self.diff_results:
                # 결과 형식: (file_name, sheet_name, row_id, col, source_val, target_val, status)
                file_name, sheet, row_id, col, source_val, target_val, status = result
                key = (file_name, sheet)
                
                if key not in file_sheet_changes:
                    file_sheet_changes[key] = {"추가": 0, "수정": 0, "삭제": 0}
                
                if "추가" in status:
                    file_sheet_changes[key]["추가"] += 1
                elif "삭제" in status:
                    file_sheet_changes[key]["삭제"] += 1
                elif "변경" in status:
                    file_sheet_changes[key]["수정"] += 1
            
            # 요약 정보 추가
            for (file_name, sheet), counts in file_sheet_changes.items():
                total = sum(counts.values())
                ws.append([
                    file_name, sheet, 
                    counts["추가"], counts["수정"], counts["삭제"], 
                    total
                ])
            
            # 각 파일-시트 조합에 대한 상세 시트 생성
            for (file_name, sheet), _ in file_sheet_changes.items():
                # 시트명 생성 (30자 제한)
                sheet_name = f"{file_name[:15]}_{sheet[:10]}"
                if len(sheet_name) > 30:
                    sheet_name = sheet_name[:30]
                
                detail_ws = wb.create_sheet(title=sheet_name)
                
                # 상세 시트 헤더
                detail_ws.append(["파일", "시트", "행", "열", "원본값", "변경값", "상태"])
                
                # 해당 파일-시트의 변경사항 추가
                sheet_results = [r for r in self.diff_results if r[0] == file_name and r[1] == sheet]
                for result in sheet_results:
                    detail_ws.append(result)
                
                # 상태별 서식 적용
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                
                for row_idx, result in enumerate(sheet_results, start=2):  # 헤더 다음부터 시작
                    status = result[6]
                    if "추가" in status:
                        detail_ws.cell(row=row_idx, column=7).fill = green_fill
                    elif "삭제" in status:
                        detail_ws.cell(row=row_idx, column=7).fill = red_fill
                    elif "변경" in status:
                        detail_ws.cell(row=row_idx, column=7).fill = yellow_fill
            
            # 파일 저장
            wb.save(file_path)
            
            # 로딩 팝업 닫기
            loading_popup.close()
            
            # 완료 메시지
            messagebox.showinfo("내보내기 완료", f"비교 결과가 '{file_path}'에 저장되었습니다.")
            
        except Exception as e:
            # 오류 처리
            if 'loading_popup' in locals():
                loading_popup.close()
            messagebox.showerror("내보내기 오류", f"결과 내보내기 중 오류가 발생했습니다:\n{str(e)}")
            logging.error(f"결과 내보내기 오류: {e}")
            import traceback
            traceback.print_exc()


    def start_comparison(self):
        """비교 시작"""
        source_path = self.source_path.get()
        target_path = self.target_path.get()

        if not source_path or not target_path:
            messagebox.showwarning("경로 오류", "원본과 비교본 경로를 모두 지정해주세요.")
            return

        if not os.path.exists(source_path) or not os.path.exists(target_path):
            messagebox.showwarning("경로 오류", "지정된 경로가 존재하지 않습니다.")
            return

        # 선택된 파일만 비교
        selected_files = None
        if hasattr(self, 'source_file_list'):
            checked_items = self.source_file_list.get_checked_items()

            if checked_items:  # 선택된 항목이 있는 경우에만 필터링
                selected_files = checked_items
                # 콘솔에 선택된 파일 출력 (디버깅용)
                print(f"선택된 파일: {selected_files}")

        # 로딩 팝업 생성
        loading_popup = LoadingPopup(self.root, title="비교 진행 중", message="준비 중...")

        # 비교 스레드 시작
        thread = threading.Thread(
            target=self.compare_files,
            args=(source_path, target_path, loading_popup, selected_files)
        )
        thread.daemon = True
        thread.start()

    def prepare_cache(self, path):
        """경로에 대한 캐시 준비 - cache_utils 활용"""
        loading_popup = LoadingPopup(self.root, title="캐시 준비 중", message="파일 스캔 중...")
        
        try:
            if os.path.isfile(path):
                # 단일 파일인 경우
                folder_path = os.path.dirname(path)
            else:
                # 폴더인 경우
                folder_path = path
            
            # 캐시 ID 및 경로 생성 - root_dir 사용하여 상위 폴더에 캐시 생성
            cache_id = hash_paths(folder_path)
            
            # 변경된 부분: root_dir(상위 디렉토리)에 .cache 폴더 생성
            cache_dir = os.path.join(root_dir, ".cache", cache_id)
            os.makedirs(cache_dir, exist_ok=True)
            cache_path = os.path.join(cache_dir, "excel_cache.json")
            
            # 캐시 데이터 로드 또는 생성
            cache_data = load_cached_data(cache_path)
            if not cache_data:
                # 캐시 없으면 생성
                loading_popup.update_message("캐시 생성 중...")
                
                # 직접 캐시 생성하기
                try:
                    # 캐시 생성 함수 직접 호출
                    cache_data = {}
                    # 파일 스캔
                    loading_popup.update_message("엑셀 파일 검색 중...")
                    
                    for root, _, files in os.walk(folder_path):
                        excel_files = [f for f in files if f.endswith(".xlsx") and not f.startswith("~$")]
                        total_files = len(excel_files)
                        
                        for idx, file in enumerate(excel_files):
                            path = os.path.join(root, file)
                            rel_path = os.path.relpath(path, folder_path).replace("\\", "/")
                            mtime = get_file_mtime(path)
                            
                            loading_popup.update_message(f"파일 분석 중 ({idx+1}/{total_files}): {file}")
                            loading_popup.update_progress(idx+1, total_files)
                            
                            # 간단한 파일 정보만 저장
                            cache_data[rel_path] = {
                                "path": path,
                                "mtime": mtime,
                                "sheets": {}  # 빈 시트 정보
                            }
                    
                    # 캐시 저장
                    save_cache(cache_path, cache_data)
                except Exception as e:
                    logging.error(f"캐시 생성 오류: {e}")
                    cache_data = {}
            else:
                loading_popup.update_message("기존 캐시를 사용합니다.")
                time.sleep(1)  # 메시지를 볼 수 있게 잠시 대기
                        
            loading_popup.close()
            return cache_data
        except Exception as e:
            loading_popup.close()
            messagebox.showerror("캐시 준비 오류", f"캐시 준비 중 오류가 발생했습니다:\n{str(e)}")
            return {}


    def create_excel_cache(self, folder_path, cache_path, loading_popup=None):
        """엑셀 캐시 생성"""
        # 기존 캐시 로드
        old_cache = load_cached_data(cache_path)
        new_cache = {}
        files_to_scan = []
        
        # 파일 스캔
        if loading_popup:
            loading_popup.update_message("엑셀 파일 검색 중...")
        
        for root, _, files in os.walk(folder_path):
            excel_files = [f for f in files if f.endswith(".xlsx") and not f.startswith("~$")]
            for file in excel_files:
                path = os.path.join(root, file)
                rel_path = os.path.relpath(path, folder_path).replace("\\", "/")
                mtime = get_file_mtime(path)
                
                # 캐시 상태 확인
                if rel_path not in old_cache or old_cache[rel_path]["mtime"] != mtime:
                    files_to_scan.append((rel_path, path, mtime))
                else:
                    new_cache[rel_path] = old_cache[rel_path]
        
        # 진행 상황 표시 개선
        total_files = len(files_to_scan)
        if loading_popup:
            loading_popup.update_message(f"변경된 {total_files}개 파일 분석 시작")
            loading_popup.update_progress(0, total_files)
        
        # 변경된 파일만 분석
        for idx, (rel_path, path, mtime) in enumerate(files_to_scan):
            try:
                if loading_popup:
                    loading_popup.update_message(f"파일 분석 중 ({idx+1}/{total_files}): {os.path.basename(path)}")
                    loading_popup.update_progress(idx+1, total_files)
                
                sheets_result = self.analyze_excel_file(path, folder_path)
                if sheets_result:
                    new_cache[rel_path] = {
                        "path": path,
                        "mtime": mtime,
                        "sheets": sheets_result
                    }
            except Exception as e:
                logging.error(f"파일 분석 실패: {path} - {e}")
        
        # 캐시 저장
        if loading_popup:
            loading_popup.update_message("캐시 저장 중...")
        
        save_cache(cache_path, new_cache)
        return new_cache
    
    def analyze_excel_file(self, file_path, db_folder=None):
        """엑셀 파일 분석 (ExcelFileManager 함수 대체)"""
        result = {}
        
        try:
            with pd.ExcelFile(file_path, engine="openpyxl") as xls:
                for sheet_name in xls.sheet_names:
                    try:
                        # #으로 시작하는 시트 제외
                        if sheet_name.startswith("#"):
                            continue
                        
                        # 헤더 행 찾기 (기본값 사용)
                        header_row = 2  # 기본 헤더 행 (0-based 인덱스)
                        
                        # DB 컬럼 정보 얻기 (안전한 방식)
                        if db_folder:
                            table_name = sheet_name.split("@")[0].replace(".xlsx", "")
                            db_path = os.path.join(db_folder, f"{table_name}.db")
                            
                            # 안전한 DB 접근 함수 사용
                            db_columns = get_columns_from_db_safe(db_path, sheet_name)
                            
                            # DB 컬럼 정보가 있으면 헤더 행 찾기 시도
                            if db_columns and len(db_columns) > 0:
                                try:
                                    # 헤더 행 찾기 로직 (간소화)
                                    for row_idx in [2, 3]:  # 주로 3, 4번째 행이 헤더
                                        df_row = pd.read_excel(
                                            xls, 
                                            sheet_name=sheet_name, 
                                            header=None, 
                                            nrows=1,
                                            skiprows=row_idx
                                        )
                                        row = df_row.iloc[0].tolist()
                                        
                                        # DB 컬럼과 일치하는지 확인
                                        if isinstance(row, list) and db_columns:
                                            if db_columns[0] in row:
                                                header_row = row_idx
                                                break
                                except Exception as e:
                                    logging.warning(f"헤더 행 검색 실패: {sheet_name} - {e}")
                        
                        # 헤더가 있는 위치부터 데이터 읽기
                        df = pd.read_excel(xls, sheet_name=sheet_name, header=header_row)
                        
                        result[sheet_name] = {
                            "columns": df.columns.tolist(),
                            "rows": len(df),
                            "header_row": header_row
                        }
                    except Exception as e:
                        logging.warning(f"시트 분석 실패: {sheet_name} - {e}")
            
            return result
        except Exception as e:
            logging.error(f"엑셀 파일 분석 실패: {file_path} - {e}")
            return {}
    

    def get_excel_files(self, path, selected_files=None):
        """경로에서 엑셀 파일 목록 가져오기"""
        result = []
        
        if os.path.isfile(path):
            # 파일 경로인 경우
            if path.endswith(('.xlsx', '.xls')) and not os.path.basename(path).startswith('~$'):
                return [path]
            return []
        
        # 폴더 경로인 경우
        for root, _, files in os.walk(path):
            for file in files:
                if file.endswith(('.xlsx', '.xls')) and not file.startswith('~$'):
                    file_path = os.path.join(root, file)
                    # 상대 경로 계산
                    rel_path = os.path.relpath(file_path, path).replace('\\', '/')
                    
                    # 선택된 파일이 지정된 경우 필터링
                    if selected_files:
                        # 파일명만으로도 비교
                        file_name = os.path.basename(file_path)
                        if rel_path not in selected_files and file_name not in selected_files:
                            continue
                        
                    result.append(file_path)
        
        return result


    def get_columns_order_from_cache(self, file_path, sheet_name):
        """캐시에서 원본 컬럼 순서를 가져오는 함수"""
        try:
            print(f"[{sheet_name}] 원본 컬럼 순서 검색")
            
            # DB 이름 정규화 (소문자 변환 및 '@' 기호 이후 제거)
            db_name = sheet_name.split('@')[0].replace(".xlsx", "").lower()
            
            # 1. 먼저 메모리에 로드된 캐시 객체 확인
            for cache_name, current_cache in [("원본", self.source_cache), ("비교본", self.target_cache)]:
                # 1.1 정확한 시트 이름으로 검색
                for cache_file_path, file_info in current_cache.items():
                    if "sheets" not in file_info:
                        continue
                    
                    for sheet_name_in_cache, sheet_info in file_info["sheets"].items():
                        # 정확한 시트 이름 일치
                        if sheet_name.lower() == sheet_name_in_cache.lower():
                            if "columns" in sheet_info:
                                columns = sheet_info["columns"]
                                print(f"[{sheet_name}] {cache_name} 캐시에서 컬럼 순서 발견: {columns}")
                                return columns
                        
                        # DB 이름 일치
                        sheet_db_name = sheet_name_in_cache.lower().split('@')[0].replace(".xlsx", "")
                        if sheet_db_name == db_name:
                            if "columns" in sheet_info:
                                columns = sheet_info["columns"]
                                print(f"[{sheet_name}] {cache_name} 캐시에서 DB 이름으로 컬럼 순서 발견: {columns}")
                                return columns
            
            # 2. 모든 캐시 폴더 검색 (.cache 폴더 내 모든 하위 폴더)
            cache_base = os.path.join(root_dir, ".cache")
            if os.path.exists(cache_base):
                for folder in os.listdir(cache_base):
                    cache_path = os.path.join(cache_base, folder, "excel_cache.json")
                    if os.path.exists(cache_path):
                        print(f"컬럼 순서 캐시 파일 확인: {folder} - {cache_path}")
                        try:
                            with open(cache_path, 'r', encoding='utf-8') as f:
                                cache_data = json.load(f)
                            
                            for cache_file_path, file_info in cache_data.items():
                                if "sheets" not in file_info:
                                    continue
                                
                                for sheet_name_in_cache, sheet_info in file_info["sheets"].items():
                                    # 정확한 시트 이름 일치
                                    if sheet_name.lower() == sheet_name_in_cache.lower():
                                        if "columns" in sheet_info:
                                            columns = sheet_info["columns"]
                                            print(f"[{sheet_name}] 캐시 {folder}에서 컬럼 순서 발견: {columns}")
                                            return columns
                                    
                                    # DB 이름으로 일치 확인
                                    sheet_db_name = sheet_name_in_cache.lower().split('@')[0].replace(".xlsx", "")
                                    if sheet_db_name == db_name:
                                        if "columns" in sheet_info:
                                            columns = sheet_info["columns"]
                                            print(f"[{sheet_name}] 캐시 {folder}에서 DB 이름으로 컬럼 순서 발견: {columns}")
                                            return columns
                        except Exception as e:
                            print(f"캐시 파일 {cache_path} 로드 중 오류: {e}")
            
            # 3. Excel_Indent.json 파일에서 직접 컬럼 순서 확인
            indent_path = os.path.join(root_dir, "Excel_Indent.json")
            if os.path.exists(indent_path):
                try:
                    with open(indent_path, 'r', encoding='utf-8') as f:
                        indent_data = json.load(f)
                    
                    # 정확한 DB 이름으로 검색
                    if db_name in indent_data and "Columns" in indent_data[db_name]:
                        columns = indent_data[db_name]["Columns"]
                        print(f"[{sheet_name}] Excel_Indent.json에서 컬럼 순서 발견: {columns}")
                        return columns if isinstance(columns, list) else columns.split(',')
                    
                    # 대소문자 무관 검색
                    for key, value in indent_data.items():
                        if key.lower() == db_name and "Columns" in value:
                            columns = value["Columns"]
                            print(f"[{sheet_name}] Excel_Indent.json에서 소문자 키로 컬럼 순서 발견: {columns}")
                            return columns if isinstance(columns, list) else columns.split(',')
                except Exception as e:
                    print(f"Excel_Indent.json 로드 오류: {e}")
            
            # 4. DB에서 컬럼 순서 가져오기
            try:
                db_folder = os.path.dirname(file_path)
                db_path = os.path.join(db_folder, f"{db_name}.db")
                if os.path.exists(db_path):
                    columns = self.get_columns_from_db_safe(db_path, sheet_name)
                    if columns:
                        print(f"[{sheet_name}] DB에서 컬럼 순서 발견: {columns}")
                        return columns
            except Exception as e:
                print(f"DB에서 컬럼 순서 조회 오류: {e}")
            
            # 캐시에 정보가 없으면 None 반환
            print(f"[{sheet_name}] 컬럼 순서 정보를 찾을 수 없음")
            return None
        
        except Exception as e:
            logging.error(f"컬럼 순서 검색 오류: {file_path}/{sheet_name} - {e}")
            return None


    def compare_files(self, source_path, target_path, loading_popup, selected_files=None):
        """파일 비교 작업 수행 - 엑셀 파일 직접 비교 방식으로 변경"""
        try:
            # 1. 캐시 준비 (기존 코드 유지)
            loading_popup.update_message("원본 폴더 캐시 준비 중...")
            self.source_cache = self.prepare_cache(source_path)
            
            loading_popup.update_message("비교본 폴더 캐시 준비 중...")
            self.target_cache = self.prepare_cache(target_path)
            
            # 2. 파일 목록 가져오기
            loading_popup.update_message("파일 목록 준비 중...")
            source_files = self.get_excel_files(source_path, selected_files)
            target_files = self.get_excel_files(target_path, selected_files)
            
            # 공통 파일 찾기
            common_files = []
            if os.path.isfile(source_path) and os.path.isfile(target_path):
                common_files = [(os.path.basename(source_path), source_path, target_path)]
            else:
                source_file_dict = {os.path.basename(f): f for f in source_files}
                target_file_dict = {os.path.basename(f): f for f in target_files}
                common_file_names = set(source_file_dict.keys()) & set(target_file_dict.keys())
                common_files = [(name, source_file_dict[name], target_file_dict[name]) 
                            for name in common_file_names]
            
            # 3. 각 파일 직접 비교
            loading_popup.update_message(f"총 {len(common_files)}개 파일 비교 중...")
            self.diff_results = []
            
            for idx, (file_name, source_file, target_file) in enumerate(common_files):
                loading_popup.update_message(f"파일 비교 중 ({idx+1}/{len(common_files)}): {file_name}")
                loading_popup.update_progress(idx+1, len(common_files))
                
                # 직접 엑셀 파일 비교 함수 호출
                file_diff = self.compare_excel_file(file_name, source_file, target_file)
                self.diff_results.extend(file_diff)
            
            # 4. 결과 UI 업데이트
            self.root.after(0, lambda: self.update_results_ui(loading_popup))
            
        except Exception as e:
            self.root.after(0, lambda: self.show_error(str(e), loading_popup))


    def compare_excel_file(self, file_name, source_file, target_file):
        """엑셀 파일 직접 비교 - DataFrame 비교 방식 사용, A열에 #이 있는 행과 #으로 시작하는 컬럼 제외"""
        file_diff = []
        import time

        try:
            start_time = time.time()
            print(f"[{file_name}] 엑셀 비교 시작: {time.strftime('%H:%M:%S')}")
            self.root.after(0, lambda: self.status_var.set(f"파일 비교 중: {file_name}"))

            with pd.ExcelFile(source_file) as source_xls, pd.ExcelFile(target_file) as target_xls:
                source_sheets = [s for s in source_xls.sheet_names if not s.startswith('#')]
                target_sheets = [s for s in target_xls.sheet_names if not s.startswith('#')]
                
                common_sheets = set(source_sheets) & set(target_sheets)
                print(f"[{file_name}] 소스 시트 수: {len(source_sheets)}, 타겟 시트 수: {len(target_sheets)}, 공통 시트 수: {len(common_sheets)}")
                
                sheet_count = 0
                total_sheets = len(common_sheets)
                
                for sheet in common_sheets:
                    sheet_count += 1
                    sheet_start_time = time.time()
                    print(f"\n[{file_name}] 시트 {sheet_count}/{total_sheets} - '{sheet}' 비교 시작: {time.strftime('%H:%M:%S')}")
                    
                    # 엑셀 데이터 로드
                    excel_load_start = time.time()
                    print(f"[{file_name} - {sheet}] 헤더 행 정보 조회 중...")
                    source_header_row = self.get_sheet_header_row_from_cache(source_file, sheet, self.source_cache)
                    target_header_row = self.get_sheet_header_row_from_cache(target_file, sheet, self.target_cache)
                    print(f"[{file_name} - {sheet}] 헤더 행: 소스={source_header_row}, 타겟={target_header_row}")
                    
                    try:
                        # 엑셀 데이터 로드
                        print(f"[{file_name} - {sheet}] 엑셀 데이터 로드 중...")
                        source_df = pd.read_excel(source_xls, sheet_name=sheet, header=source_header_row, engine='openpyxl')
                        target_df = pd.read_excel(target_xls, sheet_name=sheet, header=target_header_row, engine='openpyxl')
                        
                        # #으로 시작하는 컬럼 제외
                        original_source_cols = len(source_df.columns)
                        source_df = source_df[[col for col in source_df.columns if not str(col).startswith('#')]]
                        filtered_source_cols = original_source_cols - len(source_df.columns)
                        if filtered_source_cols > 0:
                            print(f"[{file_name} - {sheet}] 소스에서 #으로 시작하는 {filtered_source_cols}개 컬럼 제외됨")
                        
                        original_target_cols = len(target_df.columns)
                        target_df = target_df[[col for col in target_df.columns if not str(col).startswith('#')]]
                        filtered_target_cols = original_target_cols - len(target_df.columns)
                        if filtered_target_cols > 0:
                            print(f"[{file_name} - {sheet}] 타겟에서 #으로 시작하는 {filtered_target_cols}개 컬럼 제외됨")
                        
                        # A열에 #이 있는 행 제외
                        if len(source_df.columns) > 0:
                            first_col = source_df.columns[0]
                            original_source_rows = len(source_df)
                            # 문자열 타입 확인 및 변환
                            if source_df[first_col].dtype != 'object':
                                source_df[first_col] = source_df[first_col].astype(str)
                            # #으로 시작하는 행 필터링
                            source_df = source_df[~source_df[first_col].str.startswith('#', na=False)]
                            filtered_source_rows = original_source_rows - len(source_df)
                            if filtered_source_rows > 0:
                                print(f"[{file_name} - {sheet}] 소스에서 A열에 #으로 시작하는 {filtered_source_rows}개 행 제외됨")
                        
                        if len(target_df.columns) > 0:
                            first_col = target_df.columns[0]
                            original_target_rows = len(target_df)
                            # 문자열 타입 확인 및 변환
                            if target_df[first_col].dtype != 'object':
                                target_df[first_col] = target_df[first_col].astype(str)
                            # #으로 시작하는 행 필터링
                            target_df = target_df[~target_df[first_col].str.startswith('#', na=False)]
                            filtered_target_rows = original_target_rows - len(target_df)
                            if filtered_target_rows > 0:
                                print(f"[{file_name} - {sheet}] 타겟에서 A열에 #으로 시작하는 {filtered_target_rows}개 행 제외됨")
                        
                        excel_load_end = time.time()
                        print(f"[{file_name} - {sheet}] DataFrame 로드 완료: 소스={source_df.shape}, 타겟={target_df.shape}, 소요 시간: {excel_load_end - excel_load_start:.2f}초")
                        
                        # PK 정보 가져오기
                        pk_start_time = time.time()
                        print(f"[{file_name} - {sheet}] PK 정보 조회 중...")
                        # 시트 이름에서 DB 이름 추출하여 PK 정보 조회 (기존 함수 활용)
                        db_name = sheet.split('@')[0].replace(".xlsx", "")
                        pk_columns = self.get_pk_from_excel_cache(db_name, sheet_name=sheet)
                        
                        # PK 컬럼이 #으로 시작하는 컬럼 제외 후에도 존재하는지 확인
                        if pk_columns:
                            pk_columns = [col for col in pk_columns if col in source_df.columns and col in target_df.columns]
                            
                        if not pk_columns:
                            print(f"[{file_name} - {sheet}] PK 정보 없음, 기본 PK 검색 중...")
                            # 간단한 기본 PK 설정
                            for col in source_df.columns:
                                if 'id' in col.lower() or 'key' in col.lower() or 'index' in col.lower():
                                    pk_columns = [col]
                                    print(f"[{file_name} - {sheet}] ID 포함 컬럼을 PK로 사용: {col}")
                                    break
                            if not pk_columns and len(source_df.columns) > 0:
                                pk_columns = [source_df.columns[0]]  # 첫 번째 컬럼을 PK로 사용
                                print(f"[{file_name} - {sheet}] 첫 번째 컬럼을 PK로 사용: {pk_columns[0]}")
                        
                        pk_end_time = time.time()
                        print(f"[{file_name} - {sheet}] 사용 PK: {pk_columns}, 소요 시간: {pk_end_time - pk_start_time:.2f}초")
                        
                        # 데이터프레임 비교 수행
                        compare_start_time = time.time()
                        
                        # PK 컬럼 존재 확인
                        pk_valid = False
                        if pk_columns:
                            # 모든 PK 컬럼이 양쪽 DataFrame에 존재하는지 확인
                            pk_in_source = all(col in source_df.columns for col in pk_columns)
                            pk_in_target = all(col in target_df.columns for col in pk_columns)
                            
                            if pk_in_source and pk_in_target:
                                print(f"[{file_name} - {sheet}] 모든 PK 컬럼이 소스와 타겟에 존재함: {pk_columns}")
                                pk_valid = True
                            else:
                                missing_in_source = [col for col in pk_columns if col not in source_df.columns]
                                missing_in_target = [col for col in pk_columns if col not in target_df.columns]
                                if missing_in_source:
                                    print(f"[{file_name} - {sheet}] 소스에 없는 PK 컬럼: {missing_in_source}")
                                if missing_in_target:
                                    print(f"[{file_name} - {sheet}] 타겟에 없는 PK 컬럼: {missing_in_target}")

                        if pk_valid:
                            print(f"[{file_name} - {sheet}] PK 기반 비교 수행")
                            # PK 기반 비교
                            sheet_diff = self.compare_dataframes_pk(source_df, target_df, pk_columns, file_name, sheet)
                        else:
                            print(f"[{file_name} - {sheet}] 위치 기반 비교 수행 (유효한 PK 없음)")
                            # 행 위치 기반 비교
                            sheet_diff = self.compare_dataframes_position(source_df, target_df, file_name, sheet)
                        
                        compare_end_time = time.time()
                        print(f"[{file_name} - {sheet}] 비교 완료: {len(sheet_diff)}개 변경 사항, 소요 시간: {compare_end_time - compare_start_time:.2f}초")
                        
                        # 변경 사항 추가
                        file_diff.extend(sheet_diff)
                        
                        sheet_end_time = time.time()
                        print(f"[{file_name} - {sheet}] 시트 처리 완료, 총 소요 시간: {sheet_end_time - sheet_start_time:.2f}초")
                        
                    except Exception as e:
                        logging.error(f"시트 처리 오류 ({file_name}/{sheet}): {e}")
                        print(f"[{file_name} - {sheet}] 처리 중 오류 발생: {str(e)}")
                        import traceback
                        traceback.print_exc()
                
                end_time = time.time()
                print(f"[{file_name}] 엑셀 비교 완료: 총 {len(file_diff)}개 변경 사항, 소요 시간: {end_time - start_time:.2f}초")
                return file_diff
                    
        except Exception as e:
            logging.error(f"파일 처리 오류 ({file_name}): {e}")
            print(f"[{file_name}] 처리 중 오류 발생: {str(e)}")
            import traceback
            traceback.print_exc()
            return []

    def compare_dataframes_pk(self, df_source, df_target, pk_columns, file_name, sheet_name):
        """PK를 기준으로 DataFrame 비교 - PK 처리 수정"""
        diff_results = []
        
        try:
            print(f"[{file_name} - {sheet_name}] PK 기반 비교 시작: PK={pk_columns}")
            
            # 원본 데이터프레임 복사 (인덱스 설정 전)
            df_source_original = df_source.copy()
            df_target_original = df_target.copy()
            
            # 1. PK 열을 인덱스로 설정
            # NaN 값 처리
            for col in pk_columns:
                df_source[col] = df_source[col].fillna('').astype(str)
                df_target[col] = df_target[col].fillna('').astype(str)
            
            # 인덱스 설정
            if len(pk_columns) == 1:
                # 단일 PK
                pk_col = pk_columns[0]
                df_source.set_index(pk_col, inplace=True)
                df_target.set_index(pk_col, inplace=True)
            else:
                # 복합 PK
                df_source['_temp_pk_'] = df_source[pk_columns].apply(lambda row: '_'.join(row.values.astype(str)), axis=1)
                df_target['_temp_pk_'] = df_target[pk_columns].apply(lambda row: '_'.join(row.values.astype(str)), axis=1)
                df_source.set_index('_temp_pk_', inplace=True)
                df_target.set_index('_temp_pk_', inplace=True)
            
            # 2. 인덱스(PK) 비교
            source_indices = set(df_source.index)
            target_indices = set(df_target.index)
            
            added_indices = target_indices - source_indices
            removed_indices = source_indices - target_indices
            common_indices = source_indices & target_indices
            
            print(f"[{file_name} - {sheet_name}] 인덱스 비교 결과: 추가={len(added_indices)}, 삭제={len(removed_indices)}, 공통={len(common_indices)}")
            
            # 행 카운터 - 실제 행 번호 관리 (PK가 아닌 순차적 ID)
            row_counter = 1
            
            # 3. 추가된 행 처리
            for idx in added_indices:
                # 행 ID 생성 (순차적 번호)
                row_id = f"행 {row_counter}"
                row_counter += 1
                
                # 해당 인덱스의 타겟 데이터 찾기
                if len(pk_columns) == 1:
                    # 단일 PK
                    pk_col = pk_columns[0]
                    target_row = df_target_original[df_target_original[pk_col] == idx]
                    if not target_row.empty:
                        target_row = target_row.iloc[0]
                        
                        # PK 컬럼도 포함하여 모든 컬럼 처리
                        for col in df_target_original.columns:
                            val = target_row[col]
                            if not pd.isna(val):
                                diff_results.append((file_name, sheet_name, row_id, col, None, val, "추가됨"))
                else:
                    # 복합 PK
                    idx_parts = idx.split('_')
                    if len(idx_parts) == len(pk_columns):
                        # 복합 PK 조건으로 데이터 찾기
                        mask = pd.Series(True, index=df_target_original.index)
                        for i, col in enumerate(pk_columns):
                            mask &= df_target_original[col].astype(str) == idx_parts[i]
                        
                        target_row = df_target_original[mask]
                        if not target_row.empty:
                            target_row = target_row.iloc[0]
                            
                            # PK 컬럼도 포함하여 모든 컬럼 처리
                            for col in df_target_original.columns:
                                val = target_row[col]
                                if not pd.isna(val):
                                    diff_results.append((file_name, sheet_name, row_id, col, None, val, "추가됨"))
            
            # 4. 삭제된 행 처리
            for idx in removed_indices:
                # 행 ID 생성 (순차적 번호)
                row_id = f"행 {row_counter}"
                row_counter += 1
                
                # 해당 인덱스의 소스 데이터 찾기
                if len(pk_columns) == 1:
                    # 단일 PK
                    pk_col = pk_columns[0]
                    source_row = df_source_original[df_source_original[pk_col] == idx]
                    if not source_row.empty:
                        source_row = source_row.iloc[0]
                        
                        # PK 컬럼도 포함하여 모든 컬럼 처리
                        for col in df_source_original.columns:
                            val = source_row[col]
                            if not pd.isna(val):
                                diff_results.append((file_name, sheet_name, row_id, col, val, None, "삭제됨"))
                else:
                    # 복합 PK
                    idx_parts = idx.split('_')
                    if len(idx_parts) == len(pk_columns):
                        # 복합 PK 조건으로 데이터 찾기
                        mask = pd.Series(True, index=df_source_original.index)
                        for i, col in enumerate(pk_columns):
                            mask &= df_source_original[col].astype(str) == idx_parts[i]
                        
                        source_row = df_source_original[mask]
                        if not source_row.empty:
                            source_row = source_row.iloc[0]
                            
                            # PK 컬럼도 포함하여 모든 컬럼 처리
                            for col in df_source_original.columns:
                                val = source_row[col]
                                if not pd.isna(val):
                                    diff_results.append((file_name, sheet_name, row_id, col, val, None, "삭제됨"))
            
            # 5. 공통 행의 값 비교
            for idx in common_indices:
                # 행 ID 생성 (순차적 번호)
                row_id = f"행 {row_counter}"
                row_counter += 1
                
                # 해당 인덱스의 소스/타겟 데이터 찾기
                if len(pk_columns) == 1:
                    # 단일 PK
                    pk_col = pk_columns[0]
                    source_row = df_source_original[df_source_original[pk_col] == idx]
                    target_row = df_target_original[df_target_original[pk_col] == idx]
                    
                    if not source_row.empty and not target_row.empty:
                        source_row = source_row.iloc[0]
                        target_row = target_row.iloc[0]
                        
                        # 모든 컬럼 비교 (변경된 것만 결과에 포함)
                        all_columns = set(df_source_original.columns) | set(df_target_original.columns)
                        
                        for col in all_columns:
                            if col in df_source_original.columns and col in df_target_original.columns:
                                # 양쪽에 컬럼이 있는 경우
                                source_val = source_row[col]
                                target_val = target_row[col]
                                
                                # NaN 값 처리
                                if pd.isna(source_val) and pd.isna(target_val):
                                    continue
                                
                                if pd.isna(source_val):
                                    source_val = None
                                if pd.isna(target_val):
                                    target_val = None
                                
                                # 값이 다른 경우만 결과에 추가
                                if str(source_val) != str(target_val):
                                    diff_results.append((file_name, sheet_name, row_id, col, source_val, target_val, "변경됨"))
                            elif col in df_source_original.columns:
                                # 소스에만 있는 컬럼
                                val = source_row[col]
                                if not pd.isna(val):
                                    diff_results.append((file_name, sheet_name, row_id, col, val, None, "컬럼 삭제됨"))
                            elif col in df_target_original.columns:
                                # 타겟에만 있는 컬럼
                                val = target_row[col]
                                if not pd.isna(val):
                                    diff_results.append((file_name, sheet_name, row_id, col, None, val, "컬럼 추가됨"))
                else:
                    # 복합 PK
                    idx_parts = idx.split('_')
                    if len(idx_parts) == len(pk_columns):
                        # 소스 데이터 찾기
                        source_mask = pd.Series(True, index=df_source_original.index)
                        for i, col in enumerate(pk_columns):
                            source_mask &= df_source_original[col].astype(str) == idx_parts[i]
                        
                        # 타겟 데이터 찾기
                        target_mask = pd.Series(True, index=df_target_original.index)
                        for i, col in enumerate(pk_columns):
                            target_mask &= df_target_original[col].astype(str) == idx_parts[i]
                        
                        source_row = df_source_original[source_mask]
                        target_row = df_target_original[target_mask]
                        
                        if not source_row.empty and not target_row.empty:
                            source_row = source_row.iloc[0]
                            target_row = target_row.iloc[0]
                            
                            # 모든 컬럼 비교 (변경된 것만 결과에 포함)
                            all_columns = set(df_source_original.columns) | set(df_target_original.columns)
                            
                            for col in all_columns:
                                if col in df_source_original.columns and col in df_target_original.columns:
                                    # 양쪽에 컬럼이 있는 경우
                                    source_val = source_row[col]
                                    target_val = target_row[col]
                                    
                                    # NaN 값 처리
                                    if pd.isna(source_val) and pd.isna(target_val):
                                        continue
                                    
                                    if pd.isna(source_val):
                                        source_val = None
                                    if pd.isna(target_val):
                                        target_val = None
                                    
                                    # 값이 다른 경우만 결과에 추가
                                    if str(source_val) != str(target_val):
                                        diff_results.append((file_name, sheet_name, row_id, col, source_val, target_val, "변경됨"))
                                elif col in df_source_original.columns:
                                    # 소스에만 있는 컬럼
                                    val = source_row[col]
                                    if not pd.isna(val):
                                        diff_results.append((file_name, sheet_name, row_id, col, val, None, "컬럼 삭제됨"))
                                elif col in df_target_original.columns:
                                    # 타겟에만 있는 컬럼
                                    val = target_row[col]
                                    if not pd.isna(val):
                                        diff_results.append((file_name, sheet_name, row_id, col, None, val, "컬럼 추가됨"))
            
            return diff_results
            
        except Exception as e:
            logging.error(f"PK 기반 비교 오류 ({file_name}/{sheet_name}): {e}")
            print(f"[{file_name} - {sheet_name}] PK 기반 비교 중 오류 발생: {str(e)}")
            import traceback
            traceback.print_exc()
            return []

    def compare_dataframes_position(self, df_source, df_target, file_name, sheet_name):
        """행 위치를 기준으로 DataFrame 비교 - 표시 개선"""
        diff_results = []
        
        try:
            print(f"[{file_name} - {sheet_name}] 위치 기반 비교 시작")
            
            # 1. 행 수 비교
            source_rows = len(df_source)
            target_rows = len(df_target)
            
            print(f"[{file_name} - {sheet_name}] 행 수: 소스={source_rows}, 타겟={target_rows}")
            
            # 2. 컬럼 비교
            source_cols = set(df_source.columns)
            target_cols = set(df_target.columns)
            
            # 3. 행 단위 비교 (변경된 부분만 결과에 포함)
            # 최대 행 수
            max_rows = max(source_rows, target_rows)
            
            for row_idx in range(max_rows):
                # 행 ID 생성 (순차적 번호)
                row_id = f"행 {row_idx+1}"
                
                # 추가된 행
                if row_idx >= source_rows and row_idx < target_rows:
                    target_row = df_target.iloc[row_idx]
                    
                    # 추가된 행의 모든 컬럼을 결과에 추가
                    for col in df_target.columns:
                        val = target_row[col]
                        if not pd.isna(val):
                            diff_results.append((file_name, sheet_name, row_id, col, None, val, "추가됨"))
                
                # 삭제된 행
                elif row_idx < source_rows and row_idx >= target_rows:
                    source_row = df_source.iloc[row_idx]
                    
                    # 삭제된 행의 모든 컬럼을 결과에 추가
                    for col in df_source.columns:
                        val = source_row[col]
                        if not pd.isna(val):
                            diff_results.append((file_name, sheet_name, row_id, col, val, None, "삭제됨"))
                
                # 공통 행 - 값 비교 (변경된 값만 결과에 포함)
                elif row_idx < source_rows and row_idx < target_rows:
                    source_row = df_source.iloc[row_idx]
                    target_row = df_target.iloc[row_idx]
                    
                    # 모든 컬럼 확인
                    all_columns = set(df_source.columns) | set(df_target.columns)
                    
                    for col in all_columns:
                        if col in df_source.columns and col in df_target.columns:
                            # 양쪽에 컬럼이 있는 경우
                            source_val = source_row[col]
                            target_val = target_row[col]
                            
                            # NaN 값 처리
                            if pd.isna(source_val) and pd.isna(target_val):
                                continue
                            
                            if pd.isna(source_val):
                                source_val = None
                            if pd.isna(target_val):
                                target_val = None
                            
                            # 값이 다른 경우만 결과에 추가
                            if str(source_val) != str(target_val):
                                diff_results.append((file_name, sheet_name, row_id, col, source_val, target_val, "변경됨"))
                        elif col in df_source.columns:
                            # 소스에만 있는 컬럼
                            val = source_row[col]
                            if not pd.isna(val):
                                diff_results.append((file_name, sheet_name, row_id, col, val, None, "컬럼 삭제됨"))
                        elif col in df_target.columns:
                            # 타겟에만 있는 컬럼
                            val = target_row[col]
                            if not pd.isna(val):
                                diff_results.append((file_name, sheet_name, row_id, col, None, val, "컬럼 추가됨"))
            
            return diff_results
            
        except Exception as e:
            logging.error(f"위치 기반 비교 오류 ({file_name}/{sheet_name}): {e}")
            print(f"[{file_name} - {sheet_name}] 위치 기반 비교 중 오류 발생: {str(e)}")
            import traceback
            traceback.print_exc()
            return []

    def get_pk_from_excel_cache(self, db_name, sheet_name=None):
        """excel_cache에서 DB에 해당하는 PK 정보 추출 (모든 캐시 폴더 검색)"""
        try:
            # DB 이름 정규화 (소문자 변환 및 '@' 기호 이후 제거)
            db_name_lower = db_name.lower().split('@')[0].replace(".xlsx", "")
            
            # 시트명을 이용한 검색도 준비
            sheet_key = None
            if sheet_name:
                sheet_key = sheet_name.lower().split('@')[0].replace(".xlsx", "")
                print(f"[DEBUG] 검색 키: DB={db_name_lower}, 시트={sheet_key}")
            
            # 1. 먼저 메모리에 로드된 캐시 객체 확인
            for cache_name, cache in [("원본", self.source_cache), ("비교본", self.target_cache)]:
                for file_path, file_info in cache.items():
                    if "sheets" not in file_info:
                        continue
                    
                    for sheet_name_in_cache, sheet_info in file_info["sheets"].items():
                        # DB 이름과 시트 이름 모두 비교
                        sheet_db_name = sheet_name_in_cache.lower().split('@')[0].replace(".xlsx", "")
                        
                        # 1.1 정확한 시트 이름 일치 확인
                        if sheet_name and sheet_name.lower() == sheet_name_in_cache.lower():
                            if "pk" in sheet_info and sheet_info["pk"]:
                                print(f"[{db_name}] {cache_name} 캐시에서 정확한 시트명으로 PK 발견: {sheet_info['pk']}")
                                return sheet_info["pk"]
                        
                        # 1.2 DB 이름 일치 확인
                        if sheet_db_name == db_name_lower:
                            if "pk" in sheet_info and sheet_info["pk"]:
                                print(f"[{db_name}] {cache_name} 캐시에서 DB 이름으로 PK 발견: {sheet_info['pk']}")
                                return sheet_info["pk"]
            
            # 2. 모든 캐시 폴더 검색 (.cache 폴더 내 모든 하위 폴더)
            cache_base = os.path.join(root_dir, ".cache")
            if os.path.exists(cache_base):
                for folder in os.listdir(cache_base):
                    cache_path = os.path.join(cache_base, folder, "excel_cache.json")
                    if os.path.exists(cache_path):
                        print(f"캐시 파일 확인: {folder} - {cache_path}")
                        try:
                            with open(cache_path, 'r', encoding='utf-8') as f:
                                cache_data = json.load(f)
                            
                            for file_path, file_info in cache_data.items():
                                if "sheets" not in file_info:
                                    continue
                                
                                for sheet_name_in_cache, sheet_info in file_info["sheets"].items():
                                    # DB 이름과 시트 이름 모두 비교
                                    sheet_db_name = sheet_name_in_cache.lower().split('@')[0].replace(".xlsx", "")
                                    
                                    # 2.1 정확한 시트 이름 일치 확인
                                    if sheet_name and sheet_name.lower() == sheet_name_in_cache.lower():
                                        if "pk" in sheet_info and sheet_info["pk"]:
                                            print(f"[{db_name}] 캐시 {folder}에서 정확한 시트명으로 PK 발견: {sheet_info['pk']}")
                                            return sheet_info["pk"]
                                    
                                    # 2.2 DB 이름 일치 확인
                                    if sheet_db_name == db_name_lower:
                                        if "pk" in sheet_info and sheet_info["pk"]:
                                            print(f"[{db_name}] 캐시 {folder}에서 DB 이름으로 PK 발견: {sheet_info['pk']}")
                                            return sheet_info["pk"]
                        except Exception as e:
                            print(f"캐시 파일 {cache_path} 로드 중 오류: {e}")
            
            # 3. Excel_Indent.json 파일에서 직접 PK 정보 확인
            indent_path = os.path.join(root_dir, "Excel_Indent.json")
            if os.path.exists(indent_path):
                try:
                    with open(indent_path, 'r', encoding='utf-8') as f:
                        indent_data = json.load(f)
                    
                    # 3.1 정확한 DB 이름으로 검색
                    if db_name in indent_data and "OrderBy" in indent_data[db_name]:
                        pk_columns = [col.strip() for col in indent_data[db_name]["OrderBy"].split(",")]
                        print(f"[{db_name}] Excel_Indent.json에서 PK 발견: {pk_columns}")
                        return pk_columns
                    
                    # 3.2 소문자로 변환한 DB 이름으로 검색
                    for key, value in indent_data.items():
                        if key.lower() == db_name_lower and "OrderBy" in value:
                            pk_columns = [col.strip() for col in value["OrderBy"].split(",")]
                            print(f"[{db_name}] Excel_Indent.json에서 소문자 키로 PK 발견: {pk_columns}")
                            return pk_columns
                except Exception as e:
                    print(f"Excel_Indent.json 로드 오류: {e}")
            
            # 4. 특정 경로에 있는 캐시 확인 (Excel_Data_Search_47의 캐시)
            specific_cache_path = os.path.join(root_dir, ".cache", "excel_cache.json")
            if os.path.exists(specific_cache_path):
                print(f"알려진 특정 캐시 확인: {specific_cache_path}")
                try:
                    with open(specific_cache_path, 'r', encoding='utf-8') as f:
                        cache_data = json.load(f)
                    
                    for file_path, file_info in cache_data.items():
                        if "sheets" not in file_info:
                            continue
                        
                        for sheet_name_in_cache, sheet_info in file_info["sheets"].items():
                            # DB 이름과 시트 이름 모두 비교
                            sheet_db_name = sheet_name_in_cache.lower().split('@')[0].replace(".xlsx", "")
                            
                            # 4.1 정확한 시트 이름 일치 확인
                            if sheet_name and sheet_name.lower() == sheet_name_in_cache.lower():
                                if "pk" in sheet_info and sheet_info["pk"]:
                                    print(f"[{db_name}] 특정 캐시에서 정확한 시트명으로 PK 발견: {sheet_info['pk']}")
                                    return sheet_info["pk"]
                            
                            # 4.2 DB 이름 일치 확인
                            if sheet_db_name == db_name_lower:
                                if "pk" in sheet_info and sheet_info["pk"]:
                                    print(f"[{db_name}] 특정 캐시에서 DB 이름으로 PK 발견: {sheet_info['pk']}")
                                    return sheet_info["pk"]
                except Exception as e:
                    print(f"특정 캐시 로드 중 오류: {e}")
            
            print(f"[{db_name}] 캐시에서 PK 정보를 찾을 수 없습니다.")
            return []
        except Exception as e:
            logging.error(f"PK 정보 추출 오류: {db_name} - {e}")
            import traceback
            traceback.print_exc()
            return []


    def get_sheet_header_row_from_cache(self, file_path, sheet_name, cache):
        """캐시에서 시트의 헤더 행 번호를 가져오는 함수"""
        try:
            # DB 이름 정규화 (소문자 변환 및 '@' 기호 이후 제거)
            db_name = sheet_name.split('@')[0].replace(".xlsx", "").lower()
            print(f"[{sheet_name}] 헤더 행 검색 (DB 이름: {db_name})")
            
            # 1. 먼저 메모리에 로드된 캐시 객체 확인
            for cache_name, current_cache in [("현재", cache)]:
                # 1.1 정확한 시트 이름으로 검색
                for file_path, file_info in current_cache.items():
                    if "sheets" not in file_info:
                        continue
                    
                    for sheet_name_in_cache, sheet_info in file_info["sheets"].items():
                        # 정확한 시트 이름 일치
                        if sheet_name.lower() == sheet_name_in_cache.lower():
                            if "header_row" in sheet_info:
                                header_row = sheet_info["header_row"]
                                print(f"[{sheet_name}] {cache_name} 캐시에서 정확한 시트명으로 헤더 행 발견: {header_row}")
                                return header_row
                        
                        # DB 이름 일치
                        sheet_db_name = sheet_name_in_cache.lower().split('@')[0].replace(".xlsx", "")
                        if sheet_db_name == db_name:
                            if "header_row" in sheet_info:
                                header_row = sheet_info["header_row"]
                                print(f"[{sheet_name}] {cache_name} 캐시에서 DB 이름으로 헤더 행 발견: {header_row}")
                                return header_row
            
            # 2. 모든 캐시 폴더 검색 (.cache 폴더 내 모든 하위 폴더)
            cache_base = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".cache")
            if os.path.exists(cache_base):
                for folder in os.listdir(cache_base):
                    cache_path = os.path.join(cache_base, folder, "excel_cache.json")
                    if os.path.exists(cache_path):
                        print(f"헤더 행 캐시 파일 확인: {folder} - {cache_path}")
                        try:
                            with open(cache_path, 'r', encoding='utf-8') as f:
                                cache_data = json.load(f)
                            
                            for file_path, file_info in cache_data.items():
                                if "sheets" not in file_info:
                                    continue
                                
                                for sheet_name_in_cache, sheet_info in file_info["sheets"].items():
                                    # 정확한 시트 이름 일치
                                    if sheet_name.lower() == sheet_name_in_cache.lower():
                                        if "header_row" in sheet_info:
                                            header_row = sheet_info["header_row"]
                                            print(f"[{sheet_name}] 캐시 {folder}에서 정확한 시트명으로 헤더 행 발견: {header_row}")
                                            return header_row
                                    
                                    # DB 이름으로 일치 확인
                                    sheet_db_name = sheet_name_in_cache.lower().split('@')[0].replace(".xlsx", "")
                                    if sheet_db_name == db_name:
                                        if "header_row" in sheet_info:
                                            header_row = sheet_info["header_row"]
                                            print(f"[{sheet_name}] 캐시 {folder}에서 DB 이름으로 헤더 행 발견: {header_row}")
                                            return header_row
                        except Exception as e:
                            print(f"캐시 파일 {cache_path} 로드 중 오류: {e}")
            
            # 3. Excel_Indent.json 파일에서 직접 헤더 행 정보 확인
            indent_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Excel_Indent.json")
            if os.path.exists(indent_path):
                try:
                    with open(indent_path, 'r', encoding='utf-8') as f:
                        indent_data = json.load(f)
                    
                    # 정확한 DB 이름으로 검색
                    if db_name in indent_data and "HeaderRow" in indent_data[db_name]:
                        header_row = indent_data[db_name]["HeaderRow"]
                        print(f"[{sheet_name}] Excel_Indent.json에서 헤더 행 발견: {header_row}")
                        return header_row
                    
                    # 대소문자 무관 검색
                    for key, value in indent_data.items():
                        if key.lower() == db_name and "HeaderRow" in value:
                            header_row = value["HeaderRow"]
                            print(f"[{sheet_name}] Excel_Indent.json에서 소문자 키로 헤더 행 발견: {header_row}")
                            return header_row
                except Exception as e:
                    print(f"Excel_Indent.json 로드 오류: {e}")
            
            # 캐시에 정보가 없으면 기본값 반환
            print(f"[{sheet_name}] 헤더 행 정보를 캐시에서 찾을 수 없어 기본값 사용")
            return 2  # 기본값을 0 대신 2로 변경 (통상적인 엑셀 헤더 위치)
        
        except Exception as e:
            logging.error(f"헤더 행 검색 오류: {file_path}/{sheet_name} - {e}")
            return 2  # 오류 발생 시 기본값 반환


    def update_results_ui(self, loading_popup=None):
        """비교 결과를 UI에 표시 - 행 단위로 변경 사항 집계"""
        try:
            # 로딩 팝업 닫기
            if loading_popup:
                loading_popup.close()
            
            # 요약 트리뷰 항목 초기화
            self.summary_tree.delete(*self.summary_tree.get_children())
            
            # 파일별로 결과 집계
            file_summary = {}
            
            # 파일-시트-행 단위로 그룹화
            row_changes = {}
            
            for result in self.diff_results:
                # 결과 형식: (file_name, sheet_name, row_id, col, source_val, target_val, status)
                file_name, sheet_name, row_id, col_name, source_val, target_val, status = result
                
                # 파일-시트-행 복합 키
                key = (file_name, sheet_name, row_id)
                
                if key not in row_changes:
                    row_changes[key] = {"추가": False, "수정": False, "삭제": False}
                
                # 상태 표시
                if "추가" in status:
                    row_changes[key]["추가"] = True
                elif "삭제" in status:
                    row_changes[key]["삭제"] = True
                elif "변경" in status:
                    row_changes[key]["수정"] = True
            
            # 파일별로 행 단위 변경 집계
            for (file_name, sheet_name, row_id), status in row_changes.items():
                if file_name not in file_summary:
                    file_summary[file_name] = {"추가": 0, "수정": 0, "삭제": 0}
                
                # 행 단위로 한 번만 카운트 (우선순위: 추가 > 삭제 > 수정)
                if status["추가"]:
                    file_summary[file_name]["추가"] += 1
                elif status["삭제"]:
                    file_summary[file_name]["삭제"] += 1
                elif status["수정"]:
                    file_summary[file_name]["수정"] += 1
            
            # 요약 트리뷰에 결과 추가
            for file_name, counts in file_summary.items():
                self.summary_tree.insert(
                    "", "end", 
                    values=(
                        file_name, 
                        counts["추가"], 
                        counts["수정"], 
                        counts["삭제"]
                    )
                )
            
            # 결과가 없는 경우
            if not file_summary:
                self.status_var.set("변경 사항이 없습니다.")
            else:
                # 행 단위 변경 사항 수
                total_row_changes = sum(sum(counts.values()) for counts in file_summary.values())
                self.status_var.set(f"총 {total_row_changes}개 행에 변경 사항이 발견되었습니다.")
                    
            # 컬럼 자동 조정
            for column in self.summary_tree["columns"]:
                self.summary_tree.column(column, width=None)  # 자동 조정 준비
                
                # 헤더 너비 기준
                header_width = len(self.summary_tree.heading(column)["text"]) * 10
                
                # 데이터 너비 확인
                max_width = header_width
                for item_id in self.summary_tree.get_children():
                    item = self.summary_tree.item(item_id)
                    idx = self.summary_tree["columns"].index(column)
                    if idx < len(item["values"]):
                        cell_width = len(str(item["values"][idx])) * 8
                        max_width = max(max_width, cell_width)
                
                # 최대 너비 설정
                self.summary_tree.column(column, width=min(max_width, 300))
                
        except Exception as e:
            if loading_popup:
                loading_popup.close()
            
            self.status_var.set(f"결과 표시 중 오류 발생: {str(e)}")
            logging.error(f"결과 업데이트 오류: {e}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("결과 표시 오류", f"결과를 UI에 표시하는 중 오류가 발생했습니다:\n{str(e)}")

def apply_pk_from_cache(sheet_name, sheet_meta):
    """
    캐시에 저장된 PK 정보를 반환하는 유틸리티 함수.
    PK 정보가 없을 경우 기본값 설정.
    """
    pk_cols = sheet_meta.get("pk", [])
    if not pk_cols:
        # 컬럼 정보 확인
        columns = sheet_meta.get("columns", [])
        
        # ID 관련 컬럼 이름들 (우선순위 순)
        id_columns = ["ID", "UniqueID", "RewardID", "ItemID", "QuestID", 
                      "StringID", "STRING_ID", "HeroID", "CollectionID", 
                      "GroupID", "Index", "Idx"]
        
        # 컬럼 목록에서 ID 관련 컬럼 찾기
        for id_col in id_columns:
            if id_col in columns:
                print(f"[{sheet_name}] 자동 선택된 PK: {id_col}")
                return [id_col]
        
        # 'ID'를 포함하는 컬럼명 찾기
        for col in columns:
            if 'ID' in col or 'Id' in col or 'id' in col:
                print(f"[{sheet_name}] ID 포함 컬럼 자동 선택: {col}")
                return [col]
        
        # 첫 번째 컬럼을 기본값으로 사용
        if columns:
            print(f"[{sheet_name}] 첫 번째 컬럼을 PK로 사용: {columns[0]}")
            return [columns[0]]
            
        # 컬럼 정보도 없는 경우 'ID' 사용
        print(f"[{sheet_name}] 기본 PK 'ID' 사용")
        return ["ID"]
    
    return pk_cols

# 메인 함수
def main():
    root = tk.Tk()
    app = ExcelDiffTool(root)
    root.mainloop()

if __name__ == "__main__":
    main()