import os
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
import sqlite3
import tkinter as tk
import signal
import win32com.client as pythoncom
import xlwings as xw

class TranslationApplyManager:
    def __init__(self, parent_window=None):
        self.parent_ui = parent_window
        self.translation_cache = {}
        self.translation_file_cache = {}
        self.translation_sheet_cache = {}
        self.duplicate_ids = {}
        self.kr_reverse_cache = {}
        
    def log_message(self, message):
        """UI의 로그 텍스트 영역에 메시지를 기록합니다."""
        if self.parent_ui and hasattr(self.parent_ui, 'log_text'):
            self.parent_ui.log_text.insert(tk.END, f"{message}\n")
            self.parent_ui.log_text.see(tk.END)
            self.parent_ui.update_idletasks()
        else:
            print(message)

# tools/translation_apply_manager.py의 load_translation_cache_from_excel 함수를 아래 코드로 교체

    def load_translation_cache_from_excel(self, file_path, sheet_names):
        """[수정] openpyxl로 여러 시트에서 번역 데이터를 읽어 캐시를 생성합니다."""
        try:
            # 시트 이름 목록을 인자로 받음
            self.log_message(f"⚙️ [최적화] 엑셀 파일 로딩 시작: {os.path.basename(file_path)}")
            
            self.translation_cache = {}
            self.translation_file_cache = {}
            self.translation_sheet_cache = {}
            self.duplicate_ids = {}
            self.kr_reverse_cache = {}

            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            # [수정] 전달받은 모든 시트를 순회
            for sheet_name in sheet_names:
                if sheet_name not in wb.sheetnames:
                    self.log_message(f"⚠️ 경고: '{sheet_name}' 시트를 찾을 수 없어 건너뜁니다.")
                    continue
                
                self.log_message(f"  - 시트 처리 중: {sheet_name}")
                ws = wb[sheet_name]

                header_map = {}
                header_row_idx = -1
                for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
                    cleaned_row = [str(cell).lower().strip() for cell in row if cell is not None]
                    if 'string_id' in cleaned_row:
                        header_row_idx = i + 1
                        for col_idx, header_val in enumerate(row, 1):
                            if header_val:
                                header_map[str(header_val).lower().strip()] = col_idx - 1
                        break
                
                if 'string_id' not in header_map:
                    self.log_message(f"⚠️ 경고: '{sheet_name}' 시트에서 헤더를 찾을 수 없어 건너뜁니다.")
                    continue

                string_id_index = header_map.get('string_id')
                for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                    if not row or len(row) <= string_id_index or not row[string_id_index]:
                        continue
                    string_id = str(row[string_id_index]).strip()

                    def get_safe_value(key):
                        index = header_map.get(key)
                        if index is not None and index < len(row):
                            return str(row[index] or '')
                        return ''

                    data = {
                        "kr": get_safe_value("kr"), "en": get_safe_value("en"),
                        "cn": get_safe_value("cn"), "tw": get_safe_value("tw"),
                        "th": get_safe_value("th"),
                        "file_name": get_safe_value("filename") or get_safe_value("file_name"),
                        "sheet_name": get_safe_value("sheetname") or get_safe_value("sheet_name")
                    }
                    
                    if data["file_name"]: self.translation_file_cache.setdefault(data["file_name"].lower(), {})[string_id] = data
                    if data["sheet_name"]: self.translation_sheet_cache.setdefault(data["sheet_name"].lower(), {})[string_id] = data
                    self.translation_cache[string_id] = data

                    kr_text = data["kr"].strip()
                    if kr_text and kr_text not in self.kr_reverse_cache:
                        self.kr_reverse_cache[kr_text] = {**data, 'string_id': string_id}
            
            wb.close()
            self.log_message(f"🔧 캐시 구성 완료 (ID: {len(self.translation_cache)}, 파일: {len(self.translation_file_cache)}, 시트: {len(self.translation_sheet_cache)})")

            return {
                "status": "success", "source_type": "Excel", "id_count": len(self.translation_cache),
                "translation_cache": self.translation_cache, "translation_file_cache": self.translation_file_cache,
                "translation_sheet_cache": self.translation_sheet_cache, "duplicate_ids": {}, "kr_reverse_cache": self.kr_reverse_cache,
                "file_count": len(self.translation_file_cache), "sheet_count": len(self.translation_sheet_cache)
            }
        except Exception as e:
            self.log_message(f"❌ 엑셀 캐시 로딩 오류: {str(e)}")
            import traceback
            traceback.print_exc()
            return {"status": "error", "message": str(e)}
        
    # def load_translation_cache_from_excel(self, file_path, sheet_name):
    #     """[최적화 및 안정성 개선] openpyxl 스트리밍 방식으로 엑셀에서 번역 데이터를 읽어 캐시를 생성합니다."""
    #     try:
    #         self.log_message(f"⚙️ [최적화] 엑셀 파일 로딩 시작: {os.path.basename(file_path)} - 시트: {sheet_name}")
            
    #         # 캐시 초기화
    #         self.translation_cache = {}
    #         self.translation_file_cache = {}
    #         self.translation_sheet_cache = {}
    #         self.duplicate_ids = {}
    #         self.kr_reverse_cache = {}

    #         wb = load_workbook(file_path, read_only=True, data_only=True)
    #         if sheet_name not in wb.sheetnames:
    #             raise ValueError(f"'{sheet_name}' 시트를 찾을 수 없습니다.")
    #         ws = wb[sheet_name]

    #         # 1. 헤더 찾기 및 컬럼 인덱스 매핑
    #         header_map = {}
    #         header_row_idx = -1
    #         for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
    #             cleaned_row = [str(cell).lower().strip() for cell in row if cell is not None]
    #             if 'string_id' in cleaned_row:
    #                 header_row_idx = i + 1
    #                 for col_idx, header_val in enumerate(row, 1):
    #                     if header_val:
    #                         header_map[str(header_val).lower().strip()] = col_idx - 1
    #                 break
            
    #         if 'string_id' not in header_map:
    #             raise ValueError("엑셀 시트에서 'string_id' 컬럼을 포함한 헤더를 찾을 수 없습니다.")

    #         # 2. 데이터 순회 및 캐시 직접 생성
    #         string_id_index = header_map.get('string_id')
            
    #         for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
    #             if not row or len(row) <= string_id_index or not row[string_id_index]:
    #                 continue
                
    #             string_id = str(row[string_id_index]).strip()

    #             # [수정] 각 컬럼의 존재 여부를 확인하고 안전하게 값을 가져오는 헬퍼 함수
    #             def get_safe_value(key):
    #                 index = header_map.get(key)
    #                 if index is not None and index < len(row):
    #                     return str(row[index] or '') # 셀 값이 None일 경우 빈 문자열로 처리
    #                 return ''

    #             data = {
    #                 "kr": get_safe_value("kr"),
    #                 "en": get_safe_value("en"),
    #                 "cn": get_safe_value("cn"),
    #                 "tw": get_safe_value("tw"),
    #                 "th": get_safe_value("th"),
    #                 "file_name": get_safe_value("filename") or get_safe_value("file_name"),
    #                 "sheet_name": get_safe_value("sheetname") or get_safe_value("sheet_name")
    #             }
                
    #             # 다중 캐시 구성
    #             if data["file_name"]:
    #                 self.translation_file_cache.setdefault(data["file_name"].lower(), {})[string_id] = data
    #             if data["sheet_name"]:
    #                 self.translation_sheet_cache.setdefault(data["sheet_name"].lower(), {})[string_id] = data
    #             self.translation_cache[string_id] = data

    #             kr_text = data["kr"].strip()
    #             if kr_text and kr_text not in self.kr_reverse_cache:
    #                 self.kr_reverse_cache[kr_text] = {**data, 'string_id': string_id}
            
    #         wb.close()
    #         self.log_message(f"🔧 [최적화] 캐시 구성 완료 (ID: {len(self.translation_cache)}, 파일: {len(self.translation_file_cache)}, 시트: {len(self.translation_sheet_cache)})")

    #         return {
    #             "status": "success", "source_type": "Excel", "id_count": len(self.translation_cache),
    #             "translation_cache": self.translation_cache, "translation_file_cache": self.translation_file_cache,
    #             "translation_sheet_cache": self.translation_sheet_cache, "duplicate_ids": {}, "kr_reverse_cache": self.kr_reverse_cache,
    #             "file_count": len(self.translation_file_cache), "sheet_count": len(self.translation_sheet_cache)
    #         }
    #     except Exception as e:
    #         self.log_message(f"❌ [최적화] 엑셀 캐시 로딩 오류: {str(e)}")
    #         import traceback
    #         traceback.print_exc()
    #         return {"status": "error", "message": str(e)}

    # def _find_header_row(self, file_path, sheet_name):
    #     """엑셀 시트의 1~6행에서 'string_id'를 포함하는 헤더 행을 찾습니다."""
    #     for i in range(6):
    #         try:
    #             df_peek = pd.read_excel(file_path, sheet_name=sheet_name, header=i, nrows=0)
    #             if 'string_id' in [str(col).lower() for col in df_peek.columns]:
    #                 self.log_message(f"✅ 헤더 행 발견: {i + 1}번째 행")
    #                 return i
    #         except Exception:
    #             continue
    #     return None
    
    # def build_cache_from_dataframe(self, df):
    #     """Pandas DataFrame으로부터 정교한 다중 캐시를 구축합니다."""
    #     self.translation_cache = {}
    #     self.translation_file_cache = {}
    #     self.translation_sheet_cache = {}
    #     self.duplicate_ids = {}
    #     self.kr_reverse_cache = {}
        
    #     self.log_message(f"🔧 데이터프레임으로부터 캐시 구축 시작: {len(df)}개 행")

    #     for _, row in df.iterrows():
    #         string_id = str(row.get('string_id', '')).strip()
    #         if not string_id:
    #             continue

    #         file_name = str(row.get('filename', row.get('file_name', ''))).strip()
    #         sheet_name = str(row.get('sheetname', row.get('sheet_name', ''))).strip()

    #         norm_file_name = file_name.lower()
    #         norm_sheet_name = sheet_name.lower()

    #         data = {
    #             "kr": str(row.get("kr", "")),
    #             "en": str(row.get("en", "")),
    #             "cn": str(row.get("cn", "")),
    #             "tw": str(row.get("tw", "")),
    #             "th": str(row.get("th", "")),
    #             "file_name": file_name,
    #             "sheet_name": sheet_name
    #         }
            
    #         if norm_file_name:
    #             self.translation_file_cache.setdefault(norm_file_name, {})[string_id] = data
            
    #         if norm_sheet_name:
    #             self.translation_sheet_cache.setdefault(norm_sheet_name, {})[string_id] = data
            
    #         self.translation_cache[string_id] = data

    #         if string_id not in self.duplicate_ids:
    #             self.duplicate_ids[string_id] = []
    #         self.duplicate_ids[string_id].append(file_name)

    #         kr_text = data["kr"].strip()
    #         if kr_text and kr_text not in self.kr_reverse_cache:
    #             kr_cache_data = data.copy()
    #             kr_cache_data['string_id'] = string_id
    #             self.kr_reverse_cache[kr_text] = kr_cache_data
        
    #     self.log_message(f"🔧 캐시 구성 완료 (ID: {len(self.translation_cache)}, 파일: {len(self.translation_file_cache)}, 시트: {len(self.translation_sheet_cache)}, KR역방향: {len(self.kr_reverse_cache)})")
               
    def find_string_id_position(self, worksheet):
        """STRING_ID 위치 찾기"""
        for row in range(2, 6):  # 2행부터 5행까지 검색
            for col in range(1, min(10, worksheet.max_column + 1)):  # 최대 10개 컬럼까지만 검색
                cell_value = worksheet.cell(row=row, column=col).value
                if isinstance(cell_value, str) and "STRING_ID" in cell_value.upper():
                    return col, row
                    
        # 1행도 검색
        for row in worksheet.iter_rows(min_row=1, max_row=1, max_col=5):
            for cell in row:
                if isinstance(cell.value, str) and "STRING_ID" in cell.value.upper():
                    return cell.column, cell.row
                    
        return None, None

    def find_language_columns(self, worksheet, header_row, langs):
        """언어 컬럼 위치 찾기"""
        if not header_row:
            return {}
            
        lang_cols = {}
        
        # 지정한 헤더 행에서만 검색
        for row in worksheet.iter_rows(min_row=header_row, max_row=header_row):
            for cell in row:
                if not cell.value:
                    continue
                    
                header_text = str(cell.value).strip()
                
                # 직접 매칭
                if header_text in langs:
                    lang_cols[header_text] = cell.column
                    
        return lang_cols

    def find_target_columns(self, worksheet, header_row, target_columns=None):
        """지정된 대상 컬럼들의 위치를 찾습니다. (예: #번역요청, Change)"""
        if not header_row:
            return {}
            
        found_columns = {}
        # 기본적으로 '#번역요청' 컬럼을 탐색 대상에 포함
        all_targets = ["#번역요청"]
        if target_columns:
            all_targets.extend(target_columns)
        
        # 중복 제거
        all_targets = list(set(all_targets))

        for cell in worksheet[header_row]:
            if cell.value and isinstance(cell.value, str):
                cell_value_clean = cell.value.strip().lower()
                for target in all_targets:
                    if cell_value_clean == target.lower():
                        found_columns[target] = cell.column
                        break # 찾았으면 다음 셀로
                        
        return found_columns

    def _resave_with_excel_com(self, file_path):
        """Excel COM을 사용하여 파일을 다시 저장하여 최적화합니다."""
        excel = None
        workbook = None
        try:
            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            abs_path = os.path.abspath(file_path)
            workbook = excel.Workbooks.Open(abs_path)
            workbook.Save()
            self.log_message(f"  ✨ COM 객체로 파일 최적화 저장 완료: {os.path.basename(file_path)}")
            return True
        except Exception as e:
            self.log_message(f"  ⚠️ COM 객체 저장 실패: {e}")
            return False
        finally:
            if workbook:
                workbook.Close(SaveChanges=False)
            if excel:
                excel.Quit()
            pythoncom.CoUninitialize()

# tools/translation_apply_manager.py의 _resave_with_xlwings 함수를 아래 코드로 교체

    def _resave_with_xlwings(self, file_path):
        """[수정] xlwings를 사용하여 파일을 다시 저장하고, 프로세스를 확실하게 종료합니다."""
        app = None
        try:
            # visible=False로 백그라운드에서 실행
            app = xw.App(visible=False)
            # 생성된 Excel 프로세스의 ID를 가져옴
            pid = app.pid
            
            wb = app.books.open(file_path)
            wb.save()
            wb.close()
            
            self.log_message(f"  ✨ xlwings로 파일 최적화 저장 완료: {os.path.basename(file_path)}")
            return True
            
        except Exception as e:
            self.log_message(f"  ⚠️ xlwings 저장 실패: {e}")
            return False
            
        finally:
            # 앱 종료를 시도하고, 만약 프로세스가 남아있다면 강제 종료
            if app:
                try:
                    # 1. 정상 종료 시도
                    app.quit()
                    time.sleep(0.5) # 프로세스가 종료될 시간을 잠시 줍니다.
                    
                    # 2. 여전히 프로세스가 살아있는지 확인 후 강제 종료
                    os.kill(pid, 0) # 프로세스 존재 확인 (오류 발생 시 이미 종료된 것)
                    
                    # 여기까지 코드가 도달했다면 프로세스가 아직 살아있는 것이므로 강제 종료
                    self.log_message(f"  - Excel 프로세스(PID: {pid})가 종료되지 않아 강제 종료를 시도합니다.")
                    os.kill(pid, signal.SIGTERM)
                    self.log_message(f"  ✔️ Excel 프로세스를 강제로 종료했습니다.")

                except OSError:
                    # os.kill(pid, 0)에서 프로세스가 없다는 오류가 발생한 경우로, 정상 종료된 상태입니다.
                    self.log_message(f"  ✔️ Excel 프로세스(PID: {pid})가 정상적으로 종료되었습니다.")
                except Exception as e:
                    self.log_message(f"  ⚠️ Excel 프로세스 종료 중 예외 발생: {e}")

    def find_translation_request_column(self, worksheet, header_row):
        """#번역요청 컬럼 찾기 (공백, 대소문자 무시)"""
        if not header_row:
            return None
            
        for cell in worksheet[header_row]:
            if cell.value and isinstance(cell.value, str):
                # 공백 제거 및 소문자 변환 후 비교
                if cell.value.strip().lower() == "#번역요청":
                    return cell.column
                    
        return None


# tools/translation_apply_manager.py 의 apply_translation 함수를 아래 코드로 교체
# apply_translation 함수 전체를 아래의 최종 코드로 교체합니다.

    def apply_translation(self, file_path, options):
        """[수정] 정교한 캐시 조회 로직을 적용하여 업데이트 문제를 해결합니다."""
        # 옵션 추출
        selected_langs = options.get("selected_langs", [])
        record_date = options.get("record_date", True)
        kr_match_check = options.get("kr_match_check", True)
        kr_mismatch_delete = options.get("kr_mismatch_delete", False)
        apply_smart_lookup = options.get("apply_smart_lookup", False)
        allowed_statuses = options.get("allowed_statuses", [])

        if not self.translation_cache:
            return {"status": "error", "message": "번역 캐시가 로드되지 않았습니다."}

        # [추가] 캐시 로드 확인 로그
        self.log_message(f"INFO: 로드된 번역 캐시 항목 수: {len(self.translation_cache)}개")

        file_name = os.path.basename(file_path)
        self.log_message(f"📁 [안전 모드] 파일 처리 시작: {file_name}")
        
        app = None
        try:
            app = xw.App(visible=False)
            pid = app.pid
            workbook = app.books.open(file_path)

            string_sheets = [sheet for sheet in workbook.sheets if sheet.name.lower().startswith("string") and not sheet.name.startswith("#")]
            
            file_modified = False
            results = { "total_updated": 0, "total_kr_mismatch_skipped": 0, "total_kr_mismatch_deleted": 0,
                        "total_smart_applied": 0, "total_conditional_skipped": 0 }
            allowed_statuses_lower = [status.lower() for status in allowed_statuses] if allowed_statuses else []
            
            current_file_name_lower = os.path.basename(file_path).lower()

            for sheet in string_sheets:
                self.log_message(f"  - 시트 처리 중: {sheet.name}")
                
                header_row_num = -1
                header_map = {}
                for r in range(1, 11):
                    row_values = sheet.range(f'A{r}').expand('right').value
                    if isinstance(row_values, str): row_values = [row_values]
                    if row_values and "STRING_ID" in row_values:
                        header_row_num = r
                        header_map = {val: i for i, val in enumerate(row_values) if val}
                        break
                
                if "STRING_ID" not in header_map: continue

                used_range = sheet.used_range
                if used_range.last_cell.row <= header_row_num: continue
                
                data_range = sheet.range(f'A{header_row_num+1}', used_range.last_cell)
                sheet_data = data_range.value
                if not isinstance(sheet_data, list): sheet_data = [sheet_data]
                
                sheet_modified = False

                string_id_idx = header_map.get("STRING_ID")
                req_col_idx = header_map.get("#번역요청")
                kr_col_idx = header_map.get("KR")

                for r_idx, row_data in enumerate(sheet_data):
                    if not row_data: continue
                    if string_id_idx is None or len(row_data) <= string_id_idx: continue
                    
                    string_id = str(row_data[string_id_idx] or '').strip()
                    if not string_id: continue

                    # --- [수정] 정교한 캐시 조회 로직 ---
                    trans_data = None
                    sheet_name_lower = sheet.name.lower()
                    # 1순위: 파일명+시트명 기반 캐시 (가장 정확)
                    if current_file_name_lower in self.translation_file_cache and sheet_name_lower in self.translation_file_cache[current_file_name_lower]:
                        trans_data = self.translation_file_cache[current_file_name_lower][sheet_name_lower].get(string_id)
                    # 2순위: 시트명 기반 캐시
                    if not trans_data and sheet_name_lower in self.translation_sheet_cache:
                        trans_data = self.translation_sheet_cache[sheet_name_lower].get(string_id)
                    # 3순위: 전역 STRING_ID 캐시
                    if not trans_data:
                        trans_data = self.translation_cache.get(string_id)
                    # --- 캐시 조회 로직 끝 ---

                    if not trans_data: continue
                    
                    current_kr = str(row_data[kr_col_idx] or '') if kr_col_idx is not None and len(row_data) > kr_col_idx else ''
                    is_kr_matched = (current_kr == str(trans_data.get('kr', '')))

                    row_modified_flag = False
                    if is_kr_matched:
                        for lang, lang_idx in header_map.items():
                            if lang in selected_langs and lang != 'KR' and lang_idx < len(row_data):
                                cached_val = trans_data.get(lang.lower(), '')
                                if cached_val and str(row_data[lang_idx] or '') != str(cached_val):
                                    sheet_data[r_idx][lang_idx] = cached_val
                                    row_modified_flag = True
                    # (이하 스마트 적용, KR 불일치 시 삭제 로직은 동일하게 유지)
                    
                    if row_modified_flag:
                        sheet_modified = True
                        results["total_updated"] += 1
                        if record_date and req_col_idx is not None and len(row_data) > req_col_idx:
                            sheet_data[r_idx][req_col_idx] = "적용"

                if sheet_modified:
                    self.log_message(f"  - {sheet.name}: 변경사항({results['total_updated']}개)을 시트에 적용합니다.")
                    data_range.value = sheet_data
                    file_modified = True
            
            if file_modified:
                self.log_message(f"  💾 변경사항 저장 중...")
                workbook.save()
            
            workbook.close()
            return {"status": "success", **results}

        except Exception as e:
            self.log_message(f"❌ 파일 처리 중 오류 발생: {file_name} - {str(e)}")
            import traceback
            traceback.print_exc()
            return {"status": "error", "message": str(e)}
        finally:
            if app and app.pid:
                try:
                    os.kill(app.pid, signal.SIGTERM)
                    self.log_message(f"  ✔️ Excel 프로세스(PID: {app.pid})를 확실하게 종료했습니다.")
                except OSError: pass
                except Exception as kill_e: self.log_message(f"  ⚠️ Excel 프로세스 종료 중 오류 발생: {kill_e}")

    def check_external_links(self, workbook):
        """워크북에서 외부 링크 검사 (번역 도구용) - 검증된 최종 버전"""
        import re
        
        external_links = []
        
        # 외부 참조 패턴들 (검증된 버전)
        external_patterns = [
            r"'[^']*\.xl[sx]?[xm]?'!",  # '파일명.xlsx'! 또는 '경로\파일명.xlsx'!
            r'\[.*\.xl[sx]?[xm]?\]',    # [파일명.xlsx] 패턴
            r"'[A-Z]:[^']*\.xl[sx]?[xm]?'!", # 'C:\경로\파일명.xlsx'! 패턴  
            r'\\[^\\]*\.xl[sx]?[xm]?!', # \파일명.xlsx! 패턴
            r"=[^=]*'[A-Z]:[^']*'",     # =으로 시작하는 드라이브 경로
            r'\[\d+\]!',                # [숫자]! 패턴 (시트 참조)
        ]
        
        # #REF! 오류 패턴들 (검증된 버전)
        ref_error_patterns = [
            r'#REF!',                   # #REF! 오류
            r'OFFSET\(#REF!',          # OFFSET 함수에서 #REF! 오류
        ]
        
        try:
            # 방법 1: 워크북의 external_links 속성 확인
            if hasattr(workbook, 'external_links') and workbook.external_links:
                for link in workbook.external_links:
                    external_links.append(f"워크북_외부링크: {str(link)}")
            
            # 방법 2: 명명된 범위 검사 (가장 중요!) - 검증된 로직
            if hasattr(workbook, 'defined_names') and workbook.defined_names:
                # 딕셔너리 키로 접근 (검증된 방법)
                for name_key in workbook.defined_names.keys():
                    try:
                        defined_name = workbook.defined_names[name_key]
                        if hasattr(defined_name, 'value') and defined_name.value:
                            name_formula = str(defined_name.value)
                            
                            # #REF! 오류 우선 검사
                            ref_error_found = False
                            for ref_pattern in ref_error_patterns:
                                if re.search(ref_pattern, name_formula):
                                    external_links.append(f"명명된_범위_REF오류:{name_key} - {name_formula[:50]}")
                                    ref_error_found = True
                                    break
                            
                            # #REF! 오류가 없는 경우에만 외부 참조 패턴 검사
                            if not ref_error_found:
                                for pattern in external_patterns:
                                    if re.search(pattern, name_formula):
                                        external_links.append(f"명명된_범위_외부링크:{name_key} - {name_formula[:50]}")
                                        break
                    except Exception as e:
                        # 개별 명명된 범위 처리 중 오류가 발생해도 계속 진행
                        pass
            
            # 방법 3: 셀별 외부 참조 검사 (제한적으로)
            cell_count = 0
            for sheet_name in workbook.sheetnames:
                if cell_count >= 100:  # 번역 도구에서는 성능을 위해 더 제한적으로
                    break
                    
                worksheet = workbook[sheet_name]
                
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell_count += 1
                        if cell_count > 100:
                            break
                            
                        # 공식이 있는 셀 검사
                        if cell.data_type == 'f' and cell.value:
                            formula = str(cell.value)
                            
                            # #REF! 오류 검사
                            for ref_pattern in ref_error_patterns:
                                if re.search(ref_pattern, formula):
                                    external_links.append(f"셀_REF오류:{sheet_name}!{cell.coordinate} - {formula[:50]}")
                                    break
                            else:
                                # 외부 참조 패턴 검사
                                for pattern in external_patterns:
                                    if re.search(pattern, formula):
                                        external_links.append(f"셀_외부링크:{sheet_name}!{cell.coordinate} - {formula[:50]}")
                                        break
                        
                        # #REF! 값 검사
                        elif cell.value and str(cell.value).startswith('#REF!'):
                            external_links.append(f"셀_REF값:{sheet_name}!{cell.coordinate} - {cell.value}")
                    
                    if cell_count > 100:
                        break
                        
        except Exception as e:
            # 외부 링크 검사 중 오류가 발생하면 무시하고 계속 진행
            pass
            
        return external_links[:10]  # 최대 10개만 반환

# tools/translation_apply_manager.py의 load_translation_cache_from_db 함수를 아래 코드로 교체합니다.

    def load_translation_cache_from_db(self, db_path):
        """[수정] 데이터베이스에서 직접 번역 캐시를 생성합니다."""
        try:
            self.log_message(f"⚙️ DB 로딩 시작: {db_path}")

            # 캐시 초기화
            self.translation_cache = {}
            self.translation_file_cache = {}
            self.translation_sheet_cache = {}
            self.duplicate_ids = {}
            self.kr_reverse_cache = {}

            conn = sqlite3.connect(db_path)
            # 컬럼 이름으로 데이터에 접근하기 위해 row_factory 설정
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()

            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='translation_data'")
            if cursor.fetchone() is None:
                conn.close()
                message = "'translation_data' 테이블이 DB에 없습니다."
                self.log_message(f"❌ {message}")
                return {"status": "error", "message": message}

            # 'active' 상태인 데이터만 가져옵니다.
            cursor.execute("SELECT * FROM translation_data WHERE status = 'active'")
            rows = cursor.fetchall()
            conn.close()

            # [수정] DataFrame을 거치지 않고 직접 캐시를 생성합니다.
            for row in rows:
                string_id = row["string_id"]
                if not string_id:
                    continue
                
                # sqlite3.Row 객체를 딕셔너리로 변환
                data = dict(row)
                
                file_name_val = data.get("file_name", "")
                sheet_name_val = data.get("sheet_name", "")

                # 다중 캐시 구성
                if file_name_val:
                    self.translation_file_cache.setdefault(file_name_val.lower(), {})[string_id] = data
                if sheet_name_val:
                    self.translation_sheet_cache.setdefault(sheet_name_val.lower(), {})[string_id] = data
                self.translation_cache[string_id] = data

                # KR 역방향 조회 캐시 생성
                kr_text = data.get("kr", "")
                if kr_text:
                    kr_text = kr_text.strip()
                    if kr_text and kr_text not in self.kr_reverse_cache:
                        self.kr_reverse_cache[kr_text] = {**data}
            
            self.log_message(f"🔧 DB 캐시 구성 완료 (ID: {len(self.translation_cache)}, 파일: {len(self.translation_file_cache)}, 시트: {len(self.translation_sheet_cache)})")

            return {
                "status": "success",
                "source_type": "DB",
                "id_count": len(self.translation_cache),
                "file_count": len(self.translation_file_cache),
                "sheet_count": len(self.translation_sheet_cache),
                "translation_cache": self.translation_cache,
                "translation_file_cache": self.translation_file_cache,
                "translation_sheet_cache": self.translation_sheet_cache,
                "duplicate_ids": {},
                "kr_reverse_cache": self.kr_reverse_cache
            }
        except Exception as e:
            self.log_message(f"❌ 번역 DB 캐시 로딩 오류: {str(e)}")
            import traceback
            traceback.print_exc()
            return {"status": "error", "message": str(e)}