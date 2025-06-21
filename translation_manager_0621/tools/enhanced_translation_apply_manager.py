import os
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
import sqlite3
import tkinter as tk
import signal
import win32com.client as pythoncom
import xlwings as xw
from collections import defaultdict
import hashlib

class EnhancedTranslationApplyManager:
    def __init__(self, parent_window=None):
        self.parent_ui = parent_window
        
        # 기본 캐시
        self.translation_cache = {}
        self.translation_file_cache = {}
        self.translation_sheet_cache = {}
        self.duplicate_ids = {} # 이 부분은 DB 기반으로 변경되면서 사용성이 줄어들 수 있음
        self.kr_reverse_cache = {}
        
        # [신규] 특수 컬럼 관련 캐시
        self.special_filtered_cache = {}  # 특수 컬럼 필터링된 데이터
        self.detected_special_columns = {}  # 감지된 특수 컬럼 정보
        self.current_filter_conditions = {}  # 현재 적용된 필터 조건
        
        # 지원 언어 제한
        self.supported_languages = ["KR", "CN", "TW"]
        
        # [신규] DB 캐시 폴더
        self.cache_dir = os.path.join(os.getcwd(), "temp_db")
        os.makedirs(self.cache_dir, exist_ok=True)
            
        self.SPECIAL_COLUMN_NAMES = [
            "번역요청", 
            "수정요청",
            "Help",
            "원문",
            "번역요청2",
            "번역추가",
            "번역적용"
            # 필요한 다른 특수 컬럼 이름을 여기에 추가
        ]

    ### safe_strip: [변경 없음]
    def safe_strip(self, value):
        if value is None:
            return ""
        return str(value).strip()

    ### safe_lower: [변경 없음]
    def safe_lower(self, value):
        if value is None:
            return ""
        return str(value).lower().strip()

    ### log_message: [변경 없음]
    def log_message(self, message):
        if self.parent_ui and hasattr(self.parent_ui, 'log_text'):
            self.parent_ui.log_text.insert(tk.END, f"{message}\n")
            self.parent_ui.log_text.see(tk.END)
            self.parent_ui.update_idletasks()
        else:
            print(message)

    ### _get_db_path: [신규]
    def _get_db_path(self, excel_path):
        """[신규] 엑셀 파일 경로를 기반으로 고유한 DB 캐시 파일 경로를 생성합니다."""
        # 파일 경로를 해싱하여 고유하고 안전한 파일 이름을 만듭니다.
        file_hash = hashlib.md5(excel_path.encode()).hexdigest()
        return os.path.join(self.cache_dir, f"cache_{file_hash}.db")

    ### _is_db_cache_valid: [신규]
    def _is_db_cache_valid(self, excel_path, db_path):
        """[신규] DB 캐시가 최신 상태인지 확인합니다."""
        if not os.path.exists(db_path):
            return False
            
        try:
            # 원본 엑셀 파일의 최종 수정 시간
            excel_mod_time = os.path.getmtime(excel_path)
            
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT value FROM metadata WHERE key = 'source_mod_time'")
            stored_mod_time = cursor.fetchone()
            conn.close()
            
            if stored_mod_time and float(stored_mod_time[0]) == excel_mod_time:
                self.log_message("✅ 유효한 DB 캐시를 발견했습니다. 캐시를 재사용합니다.")
                return True
            else:
                self.log_message("⚠️ 원본 파일이 변경되었습니다. DB 캐시를 재구성합니다.")
                return False
        except Exception as e:
            self.log_message(f"DB 캐시 유효성 검사 오류: {e}. 캐시를 재구성합니다.")
            return False

    ### _build_db_from_excel: [신규]
    def _build_db_from_excel(self, excel_path, db_path, progress_callback=None):
        """[신규] 엑셀 파일의 모든 시트를 읽어 SQLite DB 캐시를 구축합니다."""
        try:
            self.log_message(f"⚙️ DB 캐시 구축 시작: {os.path.basename(excel_path)}")
            
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()

            cursor.execute('DROP TABLE IF EXISTS metadata')
            cursor.execute('DROP TABLE IF EXISTS translation_data')
            conn.commit()

            cursor.execute('''
                CREATE TABLE metadata (key TEXT PRIMARY KEY, value TEXT)
            ''')
            cursor.execute('''
                CREATE TABLE translation_data (
                    string_id TEXT PRIMARY KEY, kr TEXT, cn TEXT, tw TEXT,
                    file_name TEXT, sheet_name TEXT, special_columns TEXT 
                )
            ''')
            
            wb = load_workbook(excel_path, read_only=True, data_only=True)
            all_sheets = wb.sheetnames

            for idx, sheet_name in enumerate(all_sheets):
                if progress_callback:
                    progress_callback((idx / len(all_sheets)) * 100, f"시트 처리 중 ({idx+1}/{len(all_sheets)}): {sheet_name}")

                if not sheet_name.lower().startswith("string") or sheet_name.startswith("#"):
                    continue
                
                ws = wb[sheet_name]

                header_map = {}
                header_row_idx = -1
                for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
                    cleaned_row = [self.safe_lower(str(cell)) for cell in row if cell is not None]
                    if 'string_id' in cleaned_row:
                        header_row_idx = i + 1
                        for col_idx, header_val in enumerate(row, 1):
                            if header_val:
                                header_map[self.safe_strip(str(header_val))] = col_idx
                        break
                
                if 'string_id' not in [self.safe_lower(k) for k in header_map.keys()]:
                    continue

                string_id_col_key = [k for k in header_map if self.safe_lower(k) == 'string_id'][0]
                string_id_col = header_map[string_id_col_key]

                rows_to_insert = []
                for row_data in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                    if not row_data or len(row_data) < string_id_col or not row_data[string_id_col-1]:
                        continue
                    
                    string_id = self.safe_strip(str(row_data[string_id_col-1]))
                    if not string_id:
                        continue
                        
                    data = {}
                    special_data = {}
                    # ▼▼▼ 헤더 정규화 로직 적용 ▼▼▼
                    for header, col in header_map.items():
                        value = self.safe_strip(row_data[col-1]) if len(row_data) >= col else ""

                        # 헤더 이름에서 # 제거, 모든 공백 제거하여 정규화
                        normalized_header = header.lstrip('#').replace(' ', '')

                        # 정규화된 이름이 SPECIAL_COLUMN_NAMES 목록에 있는지 확인
                        if normalized_header in self.SPECIAL_COLUMN_NAMES:
                            # DB에 저장할 때도 정규화된 이름(예: '수정요청')을 키로 사용
                            special_data[normalized_header] = value
                        else:
                            # 일반 컬럼은 소문자로 통일
                            data[self.safe_lower(header)] = value

                    # special_columns를 JSON 문자열로 변환하여 저장
                    import json
                    special_columns_json = json.dumps(special_data, ensure_ascii=False)

                    kr_val = data.get("kr", "")
                    cn_val = data.get("cn", "")
                    tw_val = data.get("tw", "")
                    file_val = data.get("filename", data.get("file_name", os.path.basename(excel_path)))
                    sheet_val = data.get("sheetname", data.get("sheet_name", sheet_name))

                    rows_to_insert.append((
                        string_id, kr_val, cn_val, tw_val, file_val, sheet_val, special_columns_json
                    ))

                cursor.executemany('''
                    INSERT OR REPLACE INTO translation_data 
                    (string_id, kr, cn, tw, file_name, sheet_name, special_columns)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', rows_to_insert)

            excel_mod_time = os.path.getmtime(excel_path)
            cursor.execute("INSERT INTO metadata (key, value) VALUES (?, ?)", ('source_mod_time', str(excel_mod_time)))
            
            conn.commit()
            conn.close()
            self.log_message("✅ DB 캐시 구축 완료.")
            return {"status": "success"}
        except Exception as e:
            self.log_message(f"❌ DB 캐시 구축 중 오류 발생: {e}")
            import traceback
            traceback.print_exc()
            return {"status": "error", "message": str(e)}

    ### initiate_excel_caching: [신규]
    def initiate_excel_caching(self, excel_path, force_rebuild=False, progress_callback=None):
        """[수정] force_rebuild 플래그를 추가하여 DB 캐시 강제 재구축을 지원합니다."""
        db_path = self._get_db_path(excel_path)

        # ▼▼▼ force_rebuild 로직 추가 ▼▼▼
        if force_rebuild:
            self.log_message("ℹ️ 사용자의 요청으로 DB 캐시를 강제 재구성합니다.")
        elif self._is_db_cache_valid(excel_path, db_path):
            # 캐시가 유효하면 시트 목록만 DB에서 가져와서 반환 (강제 재구축이 아닐 때만)
            try:
                conn = sqlite3.connect(db_path)
                cursor = conn.cursor()
                cursor.execute("SELECT DISTINCT sheet_name FROM translation_data ORDER BY sheet_name")
                sheets = [row[0] for row in cursor.fetchall()]
                conn.close()
                return {"status": "success", "source_type": "DB_Cache", "sheets": sheets}
            except Exception as e:
                self.log_message(f"캐시된 시트 목록 로드 오류: {e}")

        # 캐시가 유효하지 않거나 강제 재구축이 요청된 경우 새로 구축
        build_result = self._build_db_from_excel(excel_path, db_path, progress_callback)
        if build_result["status"] == "success":
            # 구축 성공 후, 재귀 호출 대신 직접 시트 목록을 가져와서 반환하도록 최적화
            try:
                conn = sqlite3.connect(db_path)
                cursor = conn.cursor()
                cursor.execute("SELECT DISTINCT sheet_name FROM translation_data ORDER BY sheet_name")
                sheets = [row[0] for row in cursor.fetchall()]
                conn.close()
                return {"status": "success", "source_type": "DB_Cache", "sheets": sheets}
            except Exception as e:
                self.log_message(f"재구축 후 시트 목록 로드 오류: {e}")
                return {"status": "error", "message": "DB 재구축 후 시트 목록을 가져오는 데 실패했습니다."}
        else:
            return build_result
        
    ### load_translation_cache_from_excel_with_filter: [변경]
    def load_translation_cache_from_excel_with_filter(self, excel_path, sheet_names, special_column_filter=None):
        """[변경] DB 캐시에서 데이터를 쿼리하여 메모리 캐시를 구성합니다."""
        self.log_message("⚙️ DB 캐시로부터 메모리 캐시 로딩 시작...")
        db_path = self._get_db_path(excel_path)
        if not os.path.exists(db_path):
            return {"status": "error", "message": "DB 캐시 파일이 없습니다. 먼저 파일을 선택하여 캐시를 생성하세요."}

        try:
            # 캐시 초기화
            self.translation_cache = {}
            self.kr_reverse_cache = {}
            self.special_filtered_cache = {}
            # file/sheet cache는 DB기반에서는 다른 방식으로 관리되어야 하므로 여기서는 초기화만 함
            self.translation_file_cache = {} 
            self.translation_sheet_cache = {}
            
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()

            # 기본 쿼리문
            query = "SELECT * FROM translation_data"
            params = []
            
            # 시트 이름 조건 추가
            if sheet_names:
                query += f" WHERE sheet_name IN ({','.join('?' for _ in sheet_names)})"
                params.extend(sheet_names)

            cursor.execute(query, params)
            rows = cursor.fetchall()

            total_loaded = 0
            total_filtered = 0

            for row in rows:
                data = dict(row)
                string_id = data["string_id"]
                
                # 메모리 캐시 구성
                self.translation_cache[string_id] = data
                kr_text = data.get("kr", "")
                if kr_text and kr_text not in self.kr_reverse_cache:
                    self.kr_reverse_cache[kr_text] = {**data}
                
                total_loaded += 1

                # 특수 컬럼 필터링 로직
                if special_column_filter:
                    filter_column = special_column_filter['column_name']
                    filter_value = special_column_filter['condition_value']
                    
                    import json
                    special_columns = json.loads(data.get('special_columns', '{}'))
                    
                    if filter_column in special_columns:
                        special_cell_value = special_columns[filter_column]
                        if self.safe_lower(filter_value) in self.safe_lower(special_cell_value):
                            self.special_filtered_cache[string_id] = {**data, **special_columns}
                            total_filtered += 1
            
            conn.close()

            # 결과 로그
            self.log_message(f"🔧 메모리 캐시 구성 완료:")
            self.log_message(f"  - 선택된 시트로부터 로드: {total_loaded}개")
            if special_column_filter:
                self.log_message(f"  - 필터링된 데이터: {total_filtered}개 ({filter_column} = '{special_column_filter['condition_value']}')")

            return {
                "status": "success", "source_type": "Excel_DB_Cache",
                "id_count": len(self.translation_cache),
                "filtered_count": len(self.special_filtered_cache),
                "translation_cache": self.translation_cache,
                "special_filtered_cache": self.special_filtered_cache,
                "kr_reverse_cache": self.kr_reverse_cache,
                "duplicate_ids": {} # 이 로직은 단순화됨
            }
        except Exception as e:
            self.log_message(f"❌ DB 캐시 로딩 오류: {str(e)}")
            import traceback
            traceback.print_exc()
            return {"status": "error", "message": str(e)}

    ### load_translation_cache_from_excel: [변경]
    def load_translation_cache_from_excel(self, file_path, sheet_names):
        """[호환성] 기존 메서드는 필터링 없는 새 메서드를 호출합니다."""
        return self.load_translation_cache_from_excel_with_filter(file_path, sheet_names, None)

    def apply_translation_with_filter_option(self, file_path, options):
        """[수정] 중복 로그 제거 및 덮어쓰기된 항목 상세 정보 반환 기능 추가"""

        mode = options.get("mode", "id")
        selected_langs = options.get("selected_langs", [])
        record_date = options.get("record_date", True)
        use_filtered_data = options.get("use_filtered_data", False)

        kr_match_check = options.get("kr_match_check", True)
        kr_mismatch_delete = options.get("kr_mismatch_delete", False)
        kr_overwrite = options.get("kr_overwrite", False)

        kr_overwrite_on_kr_mode = options.get("kr_overwrite_on_kr_mode", False)

        allowed_statuses = options.get("allowed_statuses", [])
        allowed_statuses_lower = [status.lower() for status in allowed_statuses] if allowed_statuses else []

        if use_filtered_data and self.special_filtered_cache:
            active_cache = self.special_filtered_cache
            cache_type = "특수필터링"
            # ▼▼▼ [요청 1] 이 위치의 로그 메시지 제거 ▼▼▼
            # self.log_message(f"🔍 특수 필터링된 캐시 사용: {len(active_cache)}개 항목")
        else:
            active_cache = self.translation_cache
            cache_type = "전체"

        if not active_cache:
            return {"status": "error", "message": f"{cache_type} 번역 캐시가 로드되지 않았습니다."}

        if mode == 'kr' and not self.kr_reverse_cache:
            return {"status": "error", "message": "KR 기반 적용을 위한 역방향 캐시가 없습니다."}

        file_name = os.path.basename(file_path)

        option_summary = []
        option_summary.append(f"{mode.upper()} 기반")
        option_summary.append(f"{cache_type} 캐시")

        if mode == 'id' and kr_match_check:
            option_summary.append("KR일치검사")
            if kr_mismatch_delete:
                option_summary.append("불일치시삭제")
            if kr_overwrite:
                option_summary.append("덮어쓰기")
        elif mode == 'kr' and kr_overwrite_on_kr_mode:
            option_summary.append("덮어쓰기")

        if allowed_statuses:
            option_summary.append(f"조건:{','.join(allowed_statuses)}")

        self.log_message(f"📁 {file_name} 처리시작 [{' | '.join(option_summary)}]")

        workbook = None
        try:
            workbook = load_workbook(file_path)

            string_sheets = [sheet for sheet in workbook.sheetnames if sheet.lower().startswith("string") and not sheet.startswith("#")]

            if not string_sheets:
                self.log_message(f"   ⚠️ String 시트 없음")
                return {"status": "info", "message": "파일에 String 시트가 없습니다"}

            file_modified = False
            results = {
                "total_updated": 0, "total_overwritten": 0, "total_kr_mismatch_skipped": 0,
                "total_kr_mismatch_deleted": 0, "total_conditional_skipped": 0
            }

            # ▼▼▼ [요청 3] 덮어쓴 항목을 기록할 리스트 초기화 ▼▼▼
            overwritten_items = []

            sheet_details = {}

            fill_green = PatternFill(start_color="DAF2D0", end_color="DAF2D0", fill_type="solid")
            fill_orange = PatternFill(start_color="FFDDC1", end_color="FFDDC1", fill_type="solid")
            fill_blue = PatternFill(start_color="D0E7FF", end_color="D0E7FF", fill_type="solid")

            for sheet_name in string_sheets:
                worksheet = workbook[sheet_name]
                string_id_col, header_row = self.find_string_id_position(worksheet)
                if not string_id_col or not header_row:
                    self.log_message(f"   ⚠️ {sheet_name}: STRING_ID 컬럼 없음")
                    continue

                supported_langs = [lang for lang in selected_langs if lang in self.supported_languages]
                lang_cols = self.find_language_columns(worksheet, header_row, supported_langs + ['KR'])
                request_col_idx = self.find_target_columns(worksheet, header_row, ["#번역요청"]).get("#번역요청")

                sheet_stats = {
                    "updated": 0, "overwritten": 0, "conditional_skipped": 0,
                    "kr_mismatch_skipped": 0, "kr_mismatch_deleted": 0,
                    "total_rows": 0, "processed_rows": 0
                }

                lang_apply_count = {lang: 0 for lang in supported_langs if lang != 'KR'}
                sheet_stats["total_rows"] = worksheet.max_row - header_row

                for row_idx in range(header_row + 1, worksheet.max_row + 1):
                    sheet_stats["processed_rows"] += 1

                    if allowed_statuses_lower and request_col_idx:
                        request_val = self.safe_lower(str(worksheet.cell(row=row_idx, column=request_col_idx).value or ''))
                        if request_val not in allowed_statuses_lower:
                            sheet_stats["conditional_skipped"] += 1
                            continue

                    trans_data = None
                    key_value = ''
                    if mode == 'id':
                        key_value = self.safe_strip(str(worksheet.cell(row=row_idx, column=string_id_col).value or ''))
                        if key_value:
                            trans_data = active_cache.get(key_value)
                    else: 
                        if 'KR' in lang_cols:
                            key_value = self.safe_strip(str(worksheet.cell(row=row_idx, column=lang_cols['KR']).value or ''))
                            if key_value:
                                trans_data = self.kr_reverse_cache.get(key_value)

                    if not key_value or not trans_data:
                        continue

                    row_modified_this_iteration = False

                    if mode == 'id' and kr_match_check:
                        current_kr_val = self.safe_strip(str(worksheet.cell(row=row_idx, column=lang_cols['KR']).value or ''))
                        cache_kr_val = self.safe_strip(str(trans_data.get('kr', '')))
                        if current_kr_val != cache_kr_val:
                            if kr_mismatch_delete:
                                deleted_count = 0
                                for lang, col_idx in lang_cols.items():
                                    if lang != 'KR' and worksheet.cell(row=row_idx, column=col_idx).value:
                                        worksheet.cell(row=row_idx, column=col_idx).value = ""
                                        deleted_count += 1
                                        row_modified_this_iteration = True
                                if deleted_count > 0:
                                    sheet_stats["kr_mismatch_deleted"] += 1
                            else:
                                sheet_stats["kr_mismatch_skipped"] += 1
                            continue 

                    for lang in supported_langs:
                        if lang == 'KR': continue

                        lang_lower = lang.lower()
                        col_idx = lang_cols.get(lang)
                        if not col_idx: continue

                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        current_val = self.safe_strip(str(cell.value or ''))
                        cached_val = self.safe_strip(str(trans_data.get(lang_lower, '')))

                        if cached_val and current_val != cached_val:
                            should_overwrite = (mode == 'id' and kr_match_check and kr_overwrite) or \
                                            (mode == 'kr' and kr_overwrite_on_kr_mode)

                            if should_overwrite:
                                cell.value = cached_val
                                cell.fill = fill_orange
                                sheet_stats["overwritten"] += 1
                                lang_apply_count[lang] += 1
                                row_modified_this_iteration = True
                                # ▼▼▼ [요청 3] 덮어쓰기 정보 수집 ▼▼▼
                                overwritten_items.append({
                                    "file_name": file_name,
                                    "sheet_name": sheet_name,
                                    "string_id": key_value,
                                    "language": lang,
                                    "kr_text": trans_data.get('kr', ''),
                                    "overwritten_text": cached_val
                                })
                                # ▲▲▲ 여기까지 추가 ▲▲▲
                            elif not current_val: # 빈 칸에만 적용
                                cell.value = cached_val
                                cell.fill = fill_blue if use_filtered_data else fill_green
                                sheet_stats["updated"] += 1
                                lang_apply_count[lang] += 1
                                row_modified_this_iteration = True

                    if row_modified_this_iteration:
                        file_modified = True
                        if record_date and request_col_idx:
                            worksheet.cell(row=row_idx, column=request_col_idx).value = "특수필터적용" if use_filtered_data else "적용"

                if sheet_stats["updated"] > 0 or sheet_stats["overwritten"] > 0:
                    lang_details = [f"{lang}:{count}" for lang, count in lang_apply_count.items() if count > 0]
                    log_parts = []
                    if sheet_stats["updated"] > 0: log_parts.append(f"신규:{sheet_stats['updated']}")
                    if sheet_stats["overwritten"] > 0: log_parts.append(f"덮어씀:{sheet_stats['overwritten']}")
                    if lang_details: log_parts.append(f"[{', '.join(lang_details)}]")
                    self.log_message(f"   ✅ {sheet_name}: {' | '.join(log_parts)}")
                else:
                    skip_reasons = []
                    if sheet_stats["conditional_skipped"] > 0: skip_reasons.append(f"조건불일치:{sheet_stats['conditional_skipped']}")
                    if sheet_stats["kr_mismatch_skipped"] > 0: skip_reasons.append(f"KR불일치:{sheet_stats['kr_mismatch_skipped']}")
                    if sheet_stats["kr_mismatch_deleted"] > 0: skip_reasons.append(f"KR불일치삭제:{sheet_stats['kr_mismatch_deleted']}")

                    if skip_reasons: self.log_message(f"   ⚠️ {sheet_name}: 적용없음 ({' | '.join(skip_reasons)})")
                    else: self.log_message(f"   ⚠️ {sheet_name}: 적용없음 (번역데이터 없음)")

                for key in results:
                    if key.startswith("total_"):
                        results[key] += sheet_stats.get(key[6:], 0)

                sheet_details[sheet_name] = sheet_stats

            if file_modified:
                self.log_message(f"   💾 변경사항 저장 중...")
                workbook.save(file_path)

                summary_parts = []
                if results["total_updated"] > 0: summary_parts.append(f"신규 {results['total_updated']}개")
                if results["total_overwritten"] > 0: summary_parts.append(f"덮어씀 {results['total_overwritten']}개")

                total_applied = results["total_updated"] + results["total_overwritten"]
                cache_info = f"({cache_type}캐시)" if use_filtered_data else ""
                self.log_message(f"   ✅ {file_name} 완료: {' | '.join(summary_parts)} (총 {total_applied}개 적용) {cache_info}")
            else:
                skip_summary = []
                if results["total_conditional_skipped"] > 0: skip_summary.append(f"조건 {results['total_conditional_skipped']}개")
                if results["total_kr_mismatch_skipped"] > 0: skip_summary.append(f"KR불일치 {results['total_kr_mismatch_skipped']}개")

                if skip_summary: self.log_message(f"   ⚠️ {file_name} 완료: 변경없음 ({' | '.join(skip_summary)} 건너뜀)")
                else: self.log_message(f"   ⚠️ {file_name} 완료: 변경없음 (번역 데이터 없음)")

            # ▼▼▼ [요청 3] 수집한 덮어쓰기 목록을 반환값에 추가 ▼▼▼
            return {"status": "success", **results, "overwritten_items": overwritten_items}

        except Exception as e:
            self.log_message(f"   ❌ {file_name} 오류: {str(e)}")
            import traceback
            traceback.print_exc()
            return {"status": "error", "message": str(e), "error_type": "processing_error"}
        finally:
            if workbook:
                workbook.close()
 
    ### apply_translation: [변경 없음]
    def apply_translation(self, file_path, options):
        return self.apply_translation_with_filter_option(file_path, options)

    ### find_string_id_position: [변경 없음]
    def find_string_id_position(self, worksheet):
        for row in range(2, 6):
            for col in range(1, min(10, worksheet.max_column + 1)):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    if "string_id" in self.safe_lower(cell_value):
                        return col, row
                        
        for row in worksheet.iter_rows(min_row=1, max_row=1, max_col=5):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if "string_id" in self.safe_lower(cell.value):
                        return cell.column, cell.row
                        
        return None, None

    ### find_language_columns: [변경 없음]
    def find_language_columns(self, worksheet, header_row, langs):
        if not header_row:
            return {}
            
        lang_cols = {}
        
        langs_upper = [lang.upper() for lang in langs]
        
        for row in worksheet.iter_rows(min_row=header_row, max_row=header_row):
            for cell in row:
                if not cell.value:
                    continue
                    
                header_text = self.safe_strip(cell.value).upper()
                
                if header_text in langs_upper:
                    lang_cols[header_text] = cell.column
                    
        return lang_cols

    ### find_target_columns: [변경 없음]
    def find_target_columns(self, worksheet, header_row, target_columns=None):
        if not header_row:
            return {}
            
        found_columns = {}
        all_targets = ["#번역요청"]
        if target_columns:
            all_targets.extend(target_columns)
        
        all_targets = list(set(all_targets))

        for cell in worksheet[header_row]:
            if cell.value and isinstance(cell.value, str):
                cell_value_clean = self.safe_strip(cell.value).lower()
                for target in all_targets:
                    if cell_value_clean == target.lower():
                        found_columns[target] = cell.column
                        break
                        
        return found_columns

    ### get_filter_statistics: [삭제]
    # 이 함수는 새로운 DB 기반 로직과 맞지 않아 삭제하고, 필요시 DB 쿼리를 통해 직접 정보를 가져오는 방식으로 대체합니다.

    ### load_translation_cache_from_db: [변경 없음]
    # 이 함수는 사용자가 직접 제공하는 별도의 DB 파일을 로드하는 기능으로, 그대로 유지합니다.
    def load_translation_cache_from_db(self, db_path):
        """[기존] 데이터베이스에서 직접 번역 캐시를 생성합니다."""
        try:
            self.log_message(f"⚙️ DB 로딩 시작: {db_path}")

            # 캐시 초기화
            self.translation_cache = {}
            self.translation_file_cache = {}
            self.translation_sheet_cache = {}
            self.duplicate_ids = {}
            self.kr_reverse_cache = {}

            conn = sqlite3.connect(db_path)
            # 컬럼 이름으로 데이터에 접근하기 위해 row_factory 설정
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()

            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='translation_data'")
            if cursor.fetchone() is None:
                conn.close()
                message = "'translation_data' 테이블이 DB에 없습니다."
                self.log_message(f"❌ {message}")
                return {"status": "error", "message": message}

            # 'active' 상태인 데이터만 가져옵니다.
            cursor.execute("SELECT * FROM translation_data WHERE status = 'active'")
            rows = cursor.fetchall()
            conn.close()

            # DataFrame을 거치지 않고 직접 캐시를 생성합니다.
            for row in rows:
                string_id = row["string_id"]
                if not string_id:
                    continue
                
                # sqlite3.Row 객체를 딕셔너리로 변환
                data = dict(row)
                
                # [개선] 모든 텍스트 값에 TRIM 적용
                cleaned_data = {}
                for key, value in data.items():
                    if isinstance(value, str):
                        cleaned_data[key] = self.safe_strip(value)
                    else:
                        cleaned_data[key] = value
                
                file_name_val = cleaned_data.get("file_name", "")
                sheet_name_val = cleaned_data.get("sheet_name", "")

                # 다중 캐시 구성
                if file_name_val:
                    self.translation_file_cache.setdefault(file_name_val.lower(), {})[string_id] = cleaned_data
                if sheet_name_val:
                    self.translation_sheet_cache.setdefault(sheet_name_val.lower(), {})[string_id] = cleaned_data
                self.translation_cache[string_id] = cleaned_data

                # KR 역방향 조회 캐시 생성
                kr_text = cleaned_data.get("kr", "")
                if kr_text:
                    kr_text = self.safe_strip(kr_text)
                    if kr_text and kr_text not in self.kr_reverse_cache:
                        self.kr_reverse_cache[kr_text] = {**cleaned_data}
            
            self.log_message(f"🔧 DB 캐시 구성 완료 (ID: {len(self.translation_cache)}, 파일: {len(self.translation_file_cache)}, 시트: {len(self.translation_sheet_cache)})")

            return {
                "status": "success",
                "source_type": "DB",
                "id_count": len(self.translation_cache),
                "file_count": len(self.translation_file_cache),
                "sheet_count": len(self.translation_sheet_cache),
                "translation_cache": self.translation_cache,
                "translation_file_cache": self.translation_file_cache,
                "translation_sheet_cache": self.translation_sheet_cache,
                "duplicate_ids": {},
                "kr_reverse_cache": self.kr_reverse_cache
            }
        except Exception as e:
            self.log_message(f"❌ 번역 DB 캐시 로딩 오류: {str(e)}")
            import traceback
            traceback.print_exc()
            return {"status": "error", "message": str(e)}

    def detect_special_column_in_excel(self, excel_path, target_column_name):
        """[변경] DB 캐시에서 특수 컬럼 정보를 분석합니다."""
        db_path = self._get_db_path(excel_path)
        if not os.path.exists(db_path):
            return {"status": "error", "message": "DB 캐시가 생성되지 않았습니다."}

        self.log_message(f"⚙️ [특수컬럼감지] DB 캐시에서 '{target_column_name}' 분석 시작...")

        # ▼▼▼ 사용자 입력값 정규화 ▼▼▼
        normalized_target_column = target_column_name.lstrip('#').replace(' ', '')
        self.log_message(f"   (정규화된 검색어: '{normalized_target_column}')")
        
        try:
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT special_columns, sheet_name FROM translation_data")
            rows = cursor.fetchall()
            conn.close()
            
            import json
            values_count = defaultdict(int)
            found_in_sheets = set()
            non_empty_rows = 0

            for special_json, sheet_name in rows:
                if not special_json: continue
                
                special_data = json.loads(special_json)
                if normalized_target_column in special_data:
                    value = self.safe_strip(special_data[normalized_target_column])
                    if value:
                        values_count[value] += 1
                        found_in_sheets.add(sheet_name)
                        non_empty_rows += 1
            
            if not values_count:
                self.log_message(f"⚠️ 특수 컬럼 '{target_column_name}'에 대한 데이터를 찾을 수 없습니다.")
                return {"status": "success", "detected_info": {}}

            most_common = sorted(values_count.items(), key=lambda x: x[1], reverse=True)[:5]
            
            analysis_result = {
                'non_empty_rows': non_empty_rows,
                'unique_values': dict(values_count),
                'most_common': most_common,
                'found_in_sheets': list(found_in_sheets)
            }

            self.log_message(f"✅ 특수 컬럼 '{target_column_name}' 분석 완료:")
            self.log_message(f"  - 발견된 시트: {len(found_in_sheets)}개")
            self.log_message(f"  - 데이터 항목: {non_empty_rows}개 / 고유값: {len(values_count)}개")
            self.log_message(f"  - 최빈값: {most_common}")

            return {"status": "success", "detected_info": analysis_result}

        except Exception as e:
            self.log_message(f"❌ 특수 컬럼 분석 오류: {e}")
            import traceback
            traceback.print_exc()
            return {"status": "error", "message": str(e)}