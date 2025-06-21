# enhanced_integrated_translation_manager.py

import os
import sqlite3
import pandas as pd
from collections import defaultdict
from datetime import datetime
from openpyxl import load_workbook
import gc
import threading
import time
import re

class EnhancedIntegratedTranslationManager:
    def __init__(self, parent_window=None):
        self.parent = parent_window
        
        # 기본 데이터 저장소
        self.master_data = {}  # 메모리 내 마스터 데이터
        self.target_data = {}  # 메모리 내 타겟 데이터
        self.duplicate_data = {}  # 중복 데이터 추적
        self.comparison_results = []  # 비교 결과
        
        # [신규] 특수 컬럼 관련 데이터
        self.special_column_data = {}  # 특수 컬럼 필터링된 데이터
        self.detected_special_columns = {}  # 감지된 특수 컬럼 정보
        self.special_column_stats = {}  # 특수 컬럼 통계
        
        # [수정] 언어 목록을 KR, CN, TW로 제한
        self.supported_languages = ["KR", "CN", "TW"]
        
    def safe_strip(self, value):
        """안전한 TRIM 기능 - None 값과 빈 문자열 처리"""
        if value is None:
            return ""
        return str(value).strip()
    
    def safe_lower(self, value):
        """안전한 소문자 변환 - None 값 처리"""
        if value is None:
            return ""
        return str(value).lower().strip()

    def find_string_id_position(self, worksheet):
        """STRING_ID 컬럼 위치 찾기 (lower 적용)"""
        for row in worksheet.iter_rows(min_row=1, max_row=6, max_col=5):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # [개선] lower() 적용하여 대소문자 무관하게 검색
                    if "string_id" in self.safe_lower(cell.value):
                        return cell.column, cell.row
        return None, None

    def find_language_columns(self, worksheet, header_row, langs, language_mapping=None):
        """언어 컬럼 매핑 찾기 (lower 적용)"""
        if not header_row: 
            return {}
        
        lang_cols = {}
        extended_mapping = {}
        
        # [수정] 언어 매핑도 소문자로 처리
        if language_mapping:
            for alt, main in language_mapping.items():
                extended_mapping[alt.upper()] = main.upper()
        
        # [수정] 언어 목록을 대문자로 변환
        langs_upper = [lang.upper() for lang in langs]
        
        for row in worksheet.iter_rows(min_row=header_row, max_row=header_row):
            for cell in row:
                if not cell.value: 
                    continue
                    
                # [개선] lower() 적용 후 비교
                header_text = self.safe_strip(cell.value).upper()
                
                if header_text in langs_upper:
                    lang_cols[header_text] = cell.column
                elif header_text in extended_mapping:
                    mapped_lang = extended_mapping[header_text]
                    if mapped_lang in langs_upper:
                        lang_cols[mapped_lang] = cell.column
        
        return lang_cols

    def detect_special_columns(self, worksheet, header_row):
        """[신규] 특수 컬럼 감지 (#으로 시작하는 컬럼들)"""
        special_cols = {}
        
        if not header_row:
            return special_cols
        
        for row in worksheet.iter_rows(min_row=header_row, max_row=header_row):
            for cell in row:
                if not cell.value:
                    continue
                    
                header_text = self.safe_strip(cell.value)
                
                # # 으로 시작하는 컬럼 감지
                if header_text.startswith('#'):
                    special_cols[header_text] = cell.column
        
        return special_cols

    def analyze_special_column_values(self, worksheet, special_cols, header_row, max_rows=1000):
        """[신규] 특수 컬럼의 값들을 분석하여 통계 정보 수집"""
        special_analysis = {}
        
        for col_name, col_idx in special_cols.items():
            values_count = defaultdict(int)
            total_rows = 0
            non_empty_rows = 0
            
            # 최대 max_rows까지만 분석 (성능 고려)
            end_row = min(header_row + max_rows, worksheet.max_row + 1)
            
            for row_idx in range(header_row + 1, end_row):
                total_rows += 1
                cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                
                if cell_value:
                    cleaned_value = self.safe_strip(cell_value)
                    if cleaned_value:
                        values_count[cleaned_value] += 1
                        non_empty_rows += 1
            
            special_analysis[col_name] = {
                'column_index': col_idx,
                'total_rows': total_rows,
                'non_empty_rows': non_empty_rows,
                'unique_values': dict(values_count),
                'most_common': sorted(values_count.items(), key=lambda x: x[1], reverse=True)[:5]
            }
        
        return special_analysis

    def filter_data_by_special_column(self, data_dict, special_column_name, condition_value):
        """[신규] 특수 컬럼 조건에 따라 데이터 필터링"""
        filtered_data = {}
        
        for string_id, data in data_dict.items():
            # 특수 컬럼 값 확인
            special_value = data.get(special_column_name, "")
            special_value_clean = self.safe_strip(special_value)
            condition_value_clean = self.safe_strip(condition_value)
            
            # [개선] 조건 매칭 시 TRIM 적용
            if special_value_clean and condition_value_clean:
                if condition_value_clean.lower() in special_value_clean.lower():
                    filtered_data[string_id] = data
        
        return filtered_data

    def load_excel_data_to_memory(self, excel_files, language_list, special_column_filter=None, progress_callback=None):
        """
        [개선] 엑셀 파일들을 메모리로 로드하여 데이터와 중복 정보 반환
        special_column_filter = {"column_name": "#수정요청", "condition_value": "요청"}
        """
        language_mapping = {"ZH": "CN"}
        unique_data = {}
        duplicate_data = defaultdict(list)
        special_filtered_data = {}  # [신규] 특수 컬럼 필터링된 데이터
        all_special_columns = {}  # [신규] 모든 파일에서 발견된 특수 컬럼
        
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        processed_count = 0
        error_count = 0

        for idx, (file_name, file_path) in enumerate(excel_files):
            if progress_callback:
                progress_callback(f"파일 로드 중 ({idx+1}/{len(excel_files)}) {file_name}...", idx, len(excel_files))
            
            try:
                gc.collect()
                workbook = load_workbook(file_path, read_only=True, data_only=True)
                
                file_special_columns = {}  # 현재 파일의 특수 컬럼
                
                for sheet_name in workbook.sheetnames:
                    if not sheet_name.lower().startswith("string") or sheet_name.startswith("#"):
                        continue
                    
                    worksheet = workbook[sheet_name]
                    string_id_col, header_row = self.find_string_id_position(worksheet)
                    if not string_id_col or not header_row:
                        continue
                    
                    # [개선] 지원되는 언어만 사용
                    lang_cols = self.find_language_columns(worksheet, header_row, self.supported_languages, language_mapping)
                    if not lang_cols:
                        continue
                    
                    # [신규] 특수 컬럼 감지
                    special_cols = {}
                    if special_column_filter and special_column_filter.get("column_name"):
                        target_column = special_column_filter["column_name"]
                        special_cols = self.detect_special_columns_fast(worksheet, header_row, target_column)
                        
                        if special_cols:
                            # 간단한 통계만 수집 (전체 스캔 생략)
                            col_name = list(special_cols.keys())[0]
                            col_idx = special_cols[col_name]
                            
                            if col_name not in all_special_columns:
                                all_special_columns[col_name] = {
                                    'files': [],
                                    'total_occurrences': 0,
                                    'unique_values': set()
                                }
                            
                            all_special_columns[col_name]['files'].append(f"{file_name}:{sheet_name}")
                            # 대략적인 데이터 행 수만 계산
                            data_row_count = max(0, worksheet.max_row - header_row)
                            all_special_columns[col_name]['total_occurrences'] += data_row_count
                    
                    for row_cells in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
                        if not row_cells or string_id_col - 1 >= len(row_cells):
                            continue
                        
                        string_id = self.safe_strip(row_cells[string_id_col - 1]) if string_id_col - 1 < len(row_cells) else ""
                        if not string_id:
                            continue
                        
                        # [개선] 상태 판단 시 TRIM 적용
                        first_cell = self.safe_strip(row_cells[0]) if row_cells else ""
                        status = '비활성' if first_cell.startswith('#') else 'active'
                        
                        # 데이터 구조 생성
                        data_dict = {
                            'string_id': string_id,
                            'file_name': file_name,
                            'sheet_name': sheet_name,
                            'status': status,
                            'update_date': current_time
                        }
                        
                        # [개선] 언어 데이터 추가 시 TRIM 적용
                        for lang, col in lang_cols.items():
                            if col - 1 < len(row_cells):
                                raw_value = row_cells[col - 1]
                                data_dict[lang.lower()] = self.safe_strip(raw_value)
                        
                        # [신규] 특수 컬럼 데이터 추가
                        for special_col_name, special_col_idx in special_cols.items():
                            if special_col_idx - 1 < len(row_cells):
                                special_value = row_cells[special_col_idx - 1]
                                data_dict[special_col_name] = self.safe_strip(special_value)
                        
                        # 중복 검사 및 데이터 저장
                        if string_id not in unique_data and string_id not in duplicate_data:
                            unique_data[string_id] = data_dict
                        elif string_id in unique_data:
                            # 기존 데이터를 중복 리스트로 이동
                            original_data = unique_data.pop(string_id)
                            duplicate_data[string_id].extend([original_data, data_dict])
                        else:
                            # 이미 중복 리스트에 있으면 추가
                            duplicate_data[string_id].append(data_dict)

                workbook.close()
                processed_count += 1
                
            except Exception as e:
                if progress_callback:
                    progress_callback(f"파일 처리 오류: {e}", idx + 1, len(excel_files))
                error_count += 1
        
        # [신규] 특수 컬럼 필터링 적용
        if special_column_filter and special_column_filter.get("column_name") and special_column_filter.get("condition_value"):
            special_filtered_data = self.filter_data_by_special_column(
                unique_data, 
                special_column_filter["column_name"], 
                special_column_filter["condition_value"]
            )
        
        # 특수 컬럼 통계 정리
        special_columns_summary = {}
        for col_name, info in all_special_columns.items():
            special_columns_summary[col_name] = {
                'files_count': len(info['files']),
                'files': info['files'],
                'total_occurrences': info['total_occurrences'],
                'unique_values': list(info['unique_values'])
            }
        
        return {
            "unique_data": unique_data,
            "duplicate_data": dict(duplicate_data),
            "special_filtered_data": special_filtered_data,
            "detected_special_columns": special_columns_summary,
            "processed_count": processed_count,
            "error_count": error_count,
            "total_unique": len(unique_data),
            "special_filtered_count": len(special_filtered_data)
        }

    def load_db_data_to_memory(self, db_path):
        """[개선] 기존 DB 파일을 메모리로 로드 (TRIM 적용)"""
        if not os.path.exists(db_path):
            return {}
        
        try:
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            cursor.execute("SELECT * FROM translation_data WHERE status = 'active'")
            rows = cursor.fetchall()
            
            # 컬럼 정보 가져오기
            column_names = [description[0] for description in cursor.description]
            
            db_data = {}
            for row in rows:
                row_dict = dict(zip(column_names, row))
                string_id = self.safe_strip(row_dict.get('string_id'))
                if string_id:
                    # [개선] 모든 텍스트 값에 TRIM 적용
                    cleaned_dict = {}
                    for key, value in row_dict.items():
                        if isinstance(value, str):
                            cleaned_dict[key] = self.safe_strip(value)
                        else:
                            cleaned_dict[key] = value
                    db_data[string_id] = cleaned_dict
            
            conn.close()
            return db_data
            
        except Exception as e:
            print(f"DB 로드 중 오류: {e}")
            return {}

    def compare_data_in_memory(self, master_data, target_data, language_list, comparison_options):
        """[개선] 메모리 내 데이터 비교 (TRIM 적용)"""
        results = {
            "new_in_target": [],      # 타겟에만 있는 항목 (신규)
            "new_in_master": [],      # 마스터에만 있는 항목 (삭제됨)
            "modified": [],           # 변경된 항목
            "unchanged": []           # 변경되지 않은 항목
        }
        
        # 타겟에만 있는 항목 (신규)
        if comparison_options.get("include_new", True):
            for string_id, data in target_data.items():
                if string_id not in master_data:
                    results["new_in_target"].append(data)
        
        # 마스터에만 있는 항목 (삭제됨)
        if comparison_options.get("include_deleted", True):
            for string_id, data in master_data.items():
                if string_id not in target_data:
                    results["new_in_master"].append(data)
        
        # 양쪽에 모두 있는 항목 비교
        if comparison_options.get("include_modified", True):
            common_ids = set(master_data.keys()) & set(target_data.keys())
            
            for string_id in common_ids:
                master_item = master_data[string_id]
                target_item = target_data[string_id]
                
                is_changed = False
                changes_detail = {}
                
                # [개선] 언어별 비교 시 TRIM 적용
                for lang in language_list:
                    lang_key = lang.lower()
                    master_val = self.safe_strip(master_item.get(lang_key, ''))
                    target_val = self.safe_strip(target_item.get(lang_key, ''))
                    
                    if master_val != target_val:
                        is_changed = True
                        changes_detail[f'{lang_key}_master'] = master_val
                        changes_detail[f'{lang_key}_target'] = target_val
                
                if is_changed:
                    # 변경된 항목 추가
                    modified_item = target_item.copy()
                    modified_item.update(changes_detail)
                    results["modified"].append(modified_item)
                else:
                    results["unchanged"].append(target_item)
        
        return results

    def integrated_process(self, excel_files, language_list, comparison_options=None, 
                            master_db_path=None, master_excel_path=None, special_column_filter=None, progress_callback=None):
        """[개선] 통합 프로세스: 로드 → 비교 → 결과 반환 (특수 컬럼 필터링 추가)"""
        
        if comparison_options is None:
            comparison_options = {
                "include_new": True,
                "include_deleted": True, 
                "include_modified": True
            }
        
        results = {
            "status": "success",
            "load_results": None,
            "comparison_results": None,
            "duplicate_data": None,
            "special_column_results": None,  # [신규]
            "detected_special_columns": None,  # [신규]
            "master_count": 0,
            "target_count": 0,
            "special_filtered_count": 0  # [신규]
        }
        
        try:
            # 1단계: 엑셀 파일들을 메모리로 로드 (특수 컬럼 필터링 포함)
            if progress_callback:
                progress_callback("엑셀 파일 로드 중...", 1, 5)
            
            load_results = self.load_excel_data_to_memory(
                excel_files, language_list, special_column_filter, progress_callback
            )
            results["load_results"] = load_results
            results["duplicate_data"] = load_results["duplicate_data"]
            results["detected_special_columns"] = load_results["detected_special_columns"]
            results["special_filtered_count"] = load_results["special_filtered_count"]
            
            self.target_data = load_results["unique_data"]
            self.special_column_data = load_results["special_filtered_data"]  # [신규]
            results["target_count"] = len(self.target_data)
            
            # 2단계: 마스터 DB 로드 (있는 경우)
            if master_db_path and os.path.exists(master_db_path):
                if progress_callback:
                    progress_callback("마스터 DB 로드 중...", 2, 5)
                
                self.master_data = self.load_db_data_to_memory(master_db_path)
                results["master_count"] = len(self.master_data)
            elif master_excel_path and os.path.exists(master_excel_path):
                if progress_callback:
                    progress_callback("마스터 엑셀 로드 중...", 2, 5)
                
                # 마스터 엑셀 로드 (모든 String 시트)
                master_result = self.load_excel_data_to_memory(
                    [(os.path.basename(master_excel_path), master_excel_path)], 
                    language_list, None, None
                )
                self.master_data = master_result.get("unique_data", {})
                results["master_count"] = len(self.master_data)
            else:
                self.master_data = {}
                results["master_count"] = 0
            
            # 3단계: 데이터 비교
            if progress_callback:
                progress_callback("데이터 비교 중...", 3, 5)
            
            # 기본 비교
            comparison_results = self.compare_data_in_memory(
                self.master_data, self.target_data, language_list, comparison_options
            )
            results["comparison_results"] = comparison_results
            self.comparison_results = comparison_results
            
            # [신규] 특수 컬럼 필터링 데이터 비교 (필터링된 데이터가 있는 경우)
            if self.special_column_data:
                special_comparison_results = self.compare_data_in_memory(
                    self.master_data, self.special_column_data, language_list, comparison_options
                )
                results["special_column_results"] = special_comparison_results
            
            # 4단계: 결과 정리
            if progress_callback:
                progress_callback("결과 정리 중...", 4, 5)
            
            results["summary"] = {
                "new_items": len(comparison_results["new_in_target"]),
                "deleted_items": len(comparison_results["new_in_master"]),
                "modified_items": len(comparison_results["modified"]),
                "unchanged_items": len(comparison_results["unchanged"]),
                "duplicate_ids": len(results["duplicate_data"]),
                "special_filtered_items": len(self.special_column_data),
                "detected_special_columns_count": len(results["detected_special_columns"])
            }
            
            if progress_callback:
                progress_callback("통합 프로세스 완료!", 5, 5)
            
        except Exception as e:
            results["status"] = "error"
            results["message"] = str(e)
        
        return results

    def export_results_to_excel(self, output_path, export_options=None):
        """[개선] 비교 결과를 엑셀로 내보내기 (특수 컬럼 결과 포함)"""
        if export_options is None:
            export_options = {
                "export_new": True,
                "export_deleted": True,
                "export_modified": True,
                "export_duplicates": True,
                "export_special_filtered": True  # [신규]
            }
        
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                
                # 신규 항목 시트
                if export_options.get("export_new", True) and self.comparison_results:
                    new_items = self.comparison_results.get("new_in_target", [])
                    if new_items:
                        df_new = pd.DataFrame(new_items)
                        df_new.to_excel(writer, sheet_name='신규항목', index=False)
                
                # 삭제된 항목 시트
                if export_options.get("export_deleted", True) and self.comparison_results:
                    deleted_items = self.comparison_results.get("new_in_master", [])
                    if deleted_items:
                        df_deleted = pd.DataFrame(deleted_items)
                        df_deleted.to_excel(writer, sheet_name='삭제된항목', index=False)
                
                # 변경된 항목 시트
                if export_options.get("export_modified", True) and self.comparison_results:
                    modified_items = self.comparison_results.get("modified", [])
                    if modified_items:
                        df_modified = pd.DataFrame(modified_items)
                        df_modified.to_excel(writer, sheet_name='변경된항목', index=False)
                
                # [신규] 특수 컬럼 필터링 항목 시트
                if export_options.get("export_special_filtered", True) and self.special_column_data:
                    special_items = list(self.special_column_data.values())
                    if special_items:
                        df_special = pd.DataFrame(special_items)
                        df_special.to_excel(writer, sheet_name='특수컬럼필터링', index=False)
                
                # 중복 항목 시트
                if export_options.get("export_duplicates", True) and self.duplicate_data:
                    duplicate_flat = []
                    for string_id, items in self.duplicate_data.items():
                        for item in items:
                            duplicate_flat.append(item)
                    
                    if duplicate_flat:
                        df_duplicates = pd.DataFrame(duplicate_flat)
                        df_duplicates.to_excel(writer, sheet_name='중복항목', index=False)
                
                # [신규] 특수 컬럼 통계 시트
                if self.detected_special_columns:
                    stats_data = []
                    for col_name, info in self.detected_special_columns.items():
                        stats_data.append({
                            '특수컬럼명': col_name,
                            '발견된파일수': info['files_count'],
                            '총발생횟수': info['total_occurrences'],
                            '고유값개수': len(info['unique_values']),
                            '고유값목록': ', '.join(info['unique_values'][:10]),  # 최대 10개만
                            '발견된파일': ', '.join(info['files'][:5])  # 최대 5개만
                        })
                    
                    if stats_data:
                        df_stats = pd.DataFrame(stats_data)
                        df_stats.to_excel(writer, sheet_name='특수컬럼통계', index=False)
            
            return {"status": "success", "path": output_path}
            
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def get_statistics(self):
        """[개선] 현재 데이터 통계 반환 (특수 컬럼 정보 포함)"""
        stats = {
            "master_data_count": len(self.master_data),
            "target_data_count": len(self.target_data),
            "duplicate_count": len(self.duplicate_data),
            "special_filtered_count": len(self.special_column_data),  # [신규]
            "detected_special_columns_count": len(self.detected_special_columns),  # [신규]
            "comparison_executed": bool(self.comparison_results)
        }
        
        if self.comparison_results:
            stats.update({
                "new_items": len(self.comparison_results.get("new_in_target", [])),
                "deleted_items": len(self.comparison_results.get("new_in_master", [])),
                "modified_items": len(self.comparison_results.get("modified", [])),
                "unchanged_items": len(self.comparison_results.get("unchanged", []))
            })
        
        return stats

    def clear_data(self):
        """[개선] 메모리 데이터 초기화 (특수 컬럼 데이터 포함)"""
        self.master_data.clear()
        self.target_data.clear()
        self.duplicate_data.clear()
        self.comparison_results.clear()
        self.special_column_data.clear()  # [신규]
        self.detected_special_columns.clear()  # [신규]
        self.special_column_stats.clear()  # [신규]
        gc.collect()

    
    def detect_special_columns_fast(self, worksheet, header_row, target_column_name=None):
        """[최적화] 빠른 특수 컬럼 감지 (지정된 컬럼명만 검색)"""
        special_cols = {}
        
        if not header_row or not target_column_name:
            return special_cols
        
        target_clean = target_column_name.strip().lower()
        
        # 지정된 헤더 행에서만 검색
        for row in worksheet.iter_rows(min_row=header_row, max_row=header_row):
            for cell in row:
                if not cell.value:
                    continue
                    
                header_text = self.safe_strip(cell.value)
                header_clean = header_text.strip().lower()
                
                # 지정된 컬럼명과 정확히 일치하는지 확인
                if header_clean == target_clean:
                    special_cols[header_text] = cell.column
                    break  # 찾았으면 즉시 중단
        
        return special_cols
    
# enhanced_integrated_translation_manager.py에 추가할 새로운 메서드들

    def execute_file_comparison(self, translation_files, master_files, comparison_options, progress_callback=None):
        """파일 비교 실행"""
        try:
            if progress_callback:
                progress_callback("번역 파일 로드 중...", 1, 6)
            
            # 1단계: 번역 파일들 로드
            translation_data = self.load_files_to_memory(translation_files, comparison_options["languages"], progress_callback)
            
            if progress_callback:
                progress_callback("마스터 파일 로드 중...", 2, 6)
            
            # 2단계: 마스터 파일들 로드
            master_data = self.load_files_to_memory(master_files, comparison_options["languages"], progress_callback)
            
            if progress_callback:
                progress_callback("비교 키 생성 중...", 3, 6)
            
            # 3단계: 비교 기준에 따른 키 생성 및 데이터 매핑
            criteria = comparison_options["comparison_criteria"]
            translation_mapped = self.create_comparison_keys(translation_data, criteria)
            master_mapped = self.create_comparison_keys(master_data, criteria)
            
            if progress_callback:
                progress_callback("데이터 비교 중...", 4, 6)
            
            # 4단계: 실제 비교 수행
            comparison_results = self.perform_data_comparison(
                translation_mapped, master_mapped, comparison_options
            )
            
            if progress_callback:
                progress_callback("결과 정리 중...", 5, 6)
            
            # 5단계: 결과 정리
            summary = self.create_comparison_summary(comparison_results)
            
            if progress_callback:
                progress_callback("비교 완료!", 6, 6)
            
            return {
                "status": "success",
                "comparison_results": comparison_results,
                "summary": summary,
                "translation_count": len(translation_data),
                "master_count": len(master_data),
                "comparison_criteria": criteria
            }
            
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def load_files_to_memory(self, file_list, languages, progress_callback=None):
        """여러 파일을 메모리로 로드"""
        all_data = {}
        
        for idx, (file_name, file_path) in enumerate(file_list):
            if progress_callback:
                progress_callback(f"파일 로드: {file_name}", idx, len(file_list))
            
            try:
                workbook = load_workbook(file_path, read_only=True, data_only=True)
                
                for sheet_name in workbook.sheetnames:
                    if not sheet_name.lower().startswith("string") or sheet_name.startswith("#"):
                        continue
                    
                    worksheet = workbook[sheet_name]
                    string_id_col, header_row = self.find_string_id_position(worksheet)
                    if not string_id_col or not header_row:
                        continue
                    
                    lang_cols = self.find_language_columns(worksheet, header_row, languages + ['KR'])
                    if not lang_cols:
                        continue
                    
                    for row_cells in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
                        if not row_cells or string_id_col - 1 >= len(row_cells):
                            continue
                        
                        string_id = self.safe_strip(row_cells[string_id_col - 1]) if string_id_col - 1 < len(row_cells) else ""
                        if not string_id:
                            continue
                        
                        # 상태 판단
                        first_cell = self.safe_strip(row_cells[0]) if row_cells else ""
                        status = '비활성' if first_cell.startswith('#') else 'active'
                        
                        if status == '비활성':
                            continue  # 비활성 항목은 제외
                        
                        # 데이터 구조 생성
                        data_dict = {
                            'string_id': string_id,
                            'file_name': file_name,
                            'sheet_name': sheet_name,
                            'file_path': file_path,
                            'status': status
                        }
                        
                        # 언어 데이터 추가
                        for lang, col in lang_cols.items():
                            if col - 1 < len(row_cells):
                                raw_value = row_cells[col - 1]
                                data_dict[lang.lower()] = self.safe_strip(raw_value)
                        
                        # 키 생성 (파일별로 고유한 키)
                        unique_key = f"{file_name}:{sheet_name}:{string_id}"
                        all_data[unique_key] = data_dict
                
                workbook.close()
                
            except Exception as e:
                if progress_callback:
                    progress_callback(f"파일 오류: {file_name} - {str(e)}", idx + 1, len(file_list))
                continue
        
        return all_data

    def create_comparison_keys(self, data_dict, criteria):
        """비교 기준에 따른 키 생성"""
        mapped_data = {}
        
        for unique_key, data in data_dict.items():
            string_id = data.get('string_id', '')
            file_name = data.get('file_name', '')
            sheet_name = data.get('sheet_name', '')
            kr_text = data.get('kr', '')
            cn_text = data.get('cn', '')
            tw_text = data.get('tw', '')
            
            # 비교 기준에 따른 키 생성
            comparison_key = None
            
            if criteria == "file_id":
                comparison_key = f"{file_name}:{string_id}"
            elif criteria == "sheet_id":
                comparison_key = f"{sheet_name}:{string_id}"
            elif criteria == "id_only":
                comparison_key = string_id
            elif criteria == "id_kr":
                comparison_key = f"{string_id}:{kr_text}"
            elif criteria == "kr_only":
                comparison_key = kr_text
            elif criteria == "id_cn":
                comparison_key = f"{string_id}:{cn_text}"
            elif criteria == "id_tw":
                comparison_key = f"{string_id}:{tw_text}"
            
            if comparison_key:
                # 같은 키에 여러 데이터가 있으면 리스트로 관리
                if comparison_key not in mapped_data:
                    mapped_data[comparison_key] = []
                mapped_data[comparison_key].append(data)
        
        return mapped_data

    def perform_data_comparison(self, translation_mapped, master_mapped, comparison_options):
        """실제 데이터 비교 수행"""
        results = {
            "new_in_target": [],      # 번역에만 있는 항목 (신규)
            "new_in_master": [],      # 마스터에만 있는 항목 (삭제됨)
            "modified": [],           # 변경된 항목
            "unchanged": []           # 변경되지 않은 항목
        }
        
        languages = comparison_options.get("languages", [])
        include_new = comparison_options.get("include_new", True)
        include_deleted = comparison_options.get("include_deleted", True)
        include_modified = comparison_options.get("include_modified", True)
        
        # 번역에만 있는 항목 (신규)
        if include_new:
            for key in translation_mapped:
                if key not in master_mapped:
                    for item in translation_mapped[key]:
                        results["new_in_target"].append(item)
        
        # 마스터에만 있는 항목 (삭제됨)
        if include_deleted:
            for key in master_mapped:
                if key not in translation_mapped:
                    for item in master_mapped[key]:
                        results["new_in_master"].append(item)
        
        # 양쪽에 모두 있는 항목 비교
        if include_modified:
            common_keys = set(translation_mapped.keys()) & set(master_mapped.keys())
            
            for key in common_keys:
                translation_items = translation_mapped[key]
                master_items = master_mapped[key]
                
                # 각 조합에 대해 비교 (보통은 1:1이지만 중복 키가 있을 수 있음)
                for trans_item in translation_items:
                    best_match = None
                    is_changed = False
                    
                    # 마스터 아이템 중에서 가장 적합한 것 찾기
                    for master_item in master_items:
                        current_changed = False
                        change_details = {}
                        
                        # 언어별 비교
                        for lang in languages:
                            lang_key = lang.lower()
                            trans_val = self.safe_strip(trans_item.get(lang_key, ''))
                            master_val = self.safe_strip(master_item.get(lang_key, ''))
                            
                            if trans_val != master_val:
                                current_changed = True
                                change_details[f'{lang_key}_master'] = master_val
                                change_details[f'{lang_key}_target'] = trans_val
                        
                        # 첫 번째 매치를 사용 (더 정교한 매칭 로직 추가 가능)
                        if best_match is None:
                            best_match = master_item
                            is_changed = current_changed
                            if is_changed:
                                # 변경된 항목 추가
                                modified_item = trans_item.copy()
                                modified_item.update(change_details)
                                results["modified"].append(modified_item)
                            else:
                                results["unchanged"].append(trans_item)
                            break
        
        return results

    def create_comparison_summary(self, comparison_results):
        """비교 결과 요약 생성"""
        return {
            "new_items": len(comparison_results.get("new_in_target", [])),
            "deleted_items": len(comparison_results.get("new_in_master", [])),
            "modified_items": len(comparison_results.get("modified", [])),
            "unchanged_items": len(comparison_results.get("unchanged", []))
        }

    def export_comparison_results_to_excel(self, output_path, comparison_results):
        """비교 결과를 엑셀로 내보내기"""
        try:
            comparison_data = comparison_results.get("comparison_results", {})
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                
                # 신규 항목 시트
                new_items = comparison_data.get("new_in_target", [])
                if new_items:
                    df_new = pd.DataFrame(new_items)
                    df_new.to_excel(writer, sheet_name='신규항목', index=False)
                
                # 삭제된 항목 시트
                deleted_items = comparison_data.get("new_in_master", [])
                if deleted_items:
                    df_deleted = pd.DataFrame(deleted_items)
                    df_deleted.to_excel(writer, sheet_name='삭제된항목', index=False)
                
                # 변경된 항목 시트
                modified_items = comparison_data.get("modified", [])
                if modified_items:
                    df_modified = pd.DataFrame(modified_items)
                    df_modified.to_excel(writer, sheet_name='변경된항목', index=False)
                
                # 동일한 항목 시트
                unchanged_items = comparison_data.get("unchanged", [])
                if unchanged_items:
                    df_unchanged = pd.DataFrame(unchanged_items)
                    df_unchanged.to_excel(writer, sheet_name='동일한항목', index=False)
                
                # 요약 시트
                summary = comparison_results.get("summary", {})
                summary_data = [
                    {"구분": "신규 항목", "개수": summary.get("new_items", 0)},
                    {"구분": "삭제된 항목", "개수": summary.get("deleted_items", 0)},
                    {"구분": "변경된 항목", "개수": summary.get("modified_items", 0)},
                    {"구분": "동일한 항목", "개수": summary.get("unchanged_items", 0)}
                ]
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='요약', index=False)
            
            return {"status": "success", "path": output_path}
            
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def get_comparison_statistics(self):
        """비교 통계 정보 반환"""
        if not hasattr(self, 'last_comparison_results') or not self.last_comparison_results:
            return {
                "comparison_executed": False,
                "message": "비교가 실행되지 않았습니다."
            }
        
        summary = self.last_comparison_results.get("summary", {})
        
        return {
            "comparison_executed": True,
            "new_items": summary.get("new_items", 0),
            "deleted_items": summary.get("deleted_items", 0),
            "modified_items": summary.get("modified_items", 0),
            "unchanged_items": summary.get("unchanged_items", 0),
            "total_translation": self.last_comparison_results.get("translation_count", 0),
            "total_master": self.last_comparison_results.get("master_count", 0),
            "comparison_criteria": self.last_comparison_results.get("comparison_criteria", "알 수 없음")
        }