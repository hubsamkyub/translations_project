# tools/request_extraction_manager.py

import os
import sqlite3
import pandas as pd
from openpyxl import load_workbook, Workbook
from ui.common_components import LoadingPopup, show_message

class RequestExtractionManager:
    def __init__(self, parent_app):
        self.parent_app = parent_app

    def log(self, message):
        self.parent_app.log_message(message)

    # --- 기본 추출 로직 ---
    def run_basic_extraction(self, selected_files, db_path, conditions, mark_as_transferred, save_to_db, completion_callback):
        """기본 추출 로직 실행"""
        loading_popup = LoadingPopup(self.parent_app.root, "추출 중", "데이터 추출 준비 중...")
        
        extracted_data = []
        files_to_update = {} 

        try:
            if save_to_db:
                if os.path.exists(db_path): os.remove(db_path)
                conn = sqlite3.connect(db_path)
                cursor = conn.cursor()
                # ▼▼▼ [수정 1-1] CREATE TABLE 구문에서 EN, TH 컬럼 제거 ▼▼▼
                cursor.execute('''
                CREATE TABLE translation_requests (
                    id INTEGER PRIMARY KEY, file_name TEXT, sheet_name TEXT,
                    string_id TEXT, kr TEXT, cn TEXT, tw TEXT,
                    request_type TEXT, additional_info TEXT
                )''')

            total_files = len(selected_files)
            for i, (file_name, file_path) in enumerate(selected_files):
                loading_popup.update_message(f"파일 처리 중 ({i+1}/{total_files}): {file_name}")
                self.log(f"파일 처리 시작: {file_name}")

                wb = load_workbook(file_path, read_only=True, data_only=True)
                for sheet_name in wb.sheetnames:
                    if not sheet_name.lower().startswith("string"): continue
                    ws = wb[sheet_name]
                    headers, header_row_idx = self.parent_app._find_headers_in_worksheet(ws)
                    
                    if not headers or "#번역요청" not in headers.values():
                        continue
                    
                    req_col_idx = list(headers.keys())[list(headers.values()).index("#번역요청")]
                    
                    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
                        if not row or len(row) < req_col_idx: continue
                        request_type = row[req_col_idx - 1]
                        
                        if request_type and str(request_type).lower() in [c.lower() for c in conditions]:
                            string_id_idx = list(headers.keys())[list(headers.values()).index("STRING_ID")] if "STRING_ID" in headers.values() else None
                            string_id = row[string_id_idx - 1] if string_id_idx and len(row) >= string_id_idx else ""
                            kr = row[list(headers.keys())[list(headers.values()).index("KR")] - 1] if "KR" in headers.values() and len(row) >= list(headers.keys())[list(headers.values()).index("KR")] else ""
                            
                            # ▼▼▼ [수정 1-2] en, th 변수 추출 로직 제거 ▼▼▼
                            cn, tw = "", ""
                            if "CN" in headers.values() and len(row) >= list(headers.keys())[list(headers.values()).index("CN")]: cn = row[list(headers.keys())[list(headers.values()).index("CN")]-1]
                            if "TW" in headers.values() and len(row) >= list(headers.keys())[list(headers.values()).index("TW")]: tw = row[list(headers.keys())[list(headers.values()).index("TW")]-1]
                            
                            # ▼▼▼ [수정 1-3] append 구문에서 en, th 변수 제거 ▼▼▼
                            extracted_data.append((file_name, sheet_name, string_id, kr, cn, tw, request_type, ""))

                            if mark_as_transferred:
                                if file_path not in files_to_update: files_to_update[file_path] = {}
                                if sheet_name not in files_to_update[file_path]: files_to_update[file_path][sheet_name] = []
                                files_to_update[file_path][sheet_name].append(row_idx)
                wb.close()

            if save_to_db:
                if extracted_data:
                    # ▼▼▼ [수정 1-4] INSERT 구문에서 en, th 컬럼 및 ? 플레이스홀더 제거 ▼▼▼
                    cursor.executemany("INSERT INTO translation_requests (file_name, sheet_name, string_id, kr, cn, tw, request_type, additional_info) VALUES (?,?,?,?,?,?,?,?)", extracted_data)
                    conn.commit()
                conn.close()

                if mark_as_transferred and files_to_update:
                    loading_popup.update_message("원본 파일에 '전달' 표시 중...")
                    self.log("추출된 항목에 '전달' 표시를 시작합니다...")
                    self.parent_app._update_files_as_transferred(files_to_update, "#번역요청", "전달")
                
                loading_popup.close()
                show_message(self.parent_app.root, "info", "완료", f"기본 추출 완료: {len(extracted_data)}개 항목이 DB에 저장되었습니다.")
                self.log(f"기본 추출 완료. 총 {len(extracted_data)}개 항목 DB 저장.")
            else:
                loading_popup.close()
                if completion_callback:
                    self.parent_app.root.after(0, completion_callback, extracted_data)
                show_message(self.parent_app.root, "info", "완료", f"기본 추출 완료: {len(extracted_data)}개 항목을 찾았습니다.")
                self.log(f"기본 추출 완료. 총 {len(extracted_data)}개 항목 확인.")

        except Exception as e:
            loading_popup.close()
            show_message(self.parent_app.root, "error", "오류", f"추출 중 오류 발생: {e}")
            self.log(f"오류 발생: {e}")

    def run_compare_extraction(self, selected_files, output_db_path, compare_options, save_to_db, completion_callback):
        """비교 추출 로직 실행"""
        loading_popup = LoadingPopup(self.parent_app.root, "비교 추출 중", "비교 데이터 로딩 중...")
        try:
            comparison_cache = self._load_comparison_data(compare_options)
            
            if save_to_db:
                if os.path.exists(output_db_path): os.remove(output_db_path)
                conn = sqlite3.connect(output_db_path)
                cursor = conn.cursor()
                # ▼▼▼ [수정 2-1] CREATE TABLE 구문에서 EN, TH 컬럼 제거 ▼▼▼
                cursor.execute('''
                CREATE TABLE translation_requests (
                    id INTEGER PRIMARY KEY, file_name TEXT, sheet_name TEXT,
                    string_id TEXT, kr TEXT, cn TEXT, tw TEXT,
                    request_type TEXT, additional_info TEXT
                )''')
            
            extracted_data = []
            total_files = len(selected_files)
            extract_new = compare_options['extract_new']
            extract_modified = compare_options['extract_modified']
            apply_by_request_col = compare_options.get('apply_by_request_col', False)

            for i, (file_name, file_path) in enumerate(selected_files):
                loading_popup.update_message(f"파일 비교 중 ({i+1}/{total_files}): {file_name}")
                self.log(f"파일 비교 시작: {file_name}")
                
                wb = load_workbook(file_path, read_only=True, data_only=True)
                for sheet_name in wb.sheetnames:
                    if not sheet_name.lower().startswith("string"): continue

                    ws = wb[sheet_name]
                    headers, header_row_idx = self.parent_app._find_headers_in_worksheet(ws)
                    if not headers or "STRING_ID" not in headers.values(): continue

                    string_id_col = list(headers.keys())[list(headers.values()).index("STRING_ID")]
                    kr_col = list(headers.keys())[list(headers.values()).index("KR")] if "KR" in headers.values() else None
                    req_col_idx = list(headers.keys())[list(headers.values()).index("#번역요청")] if "#번역요청" in headers.values() else None

                    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                        if not row or len(row) < string_id_col: continue
                        
                        if apply_by_request_col:
                            if not req_col_idx or len(row) < req_col_idx: continue
                            request_val = str(row[req_col_idx - 1] or '').strip().lower()
                            if request_val not in ['신규', 'change']: continue

                        string_id = row[string_id_col - 1]
                        if not string_id: continue
                        
                        string_id_str = str(string_id)
                        source_kr = str(row[kr_col - 1]) if kr_col and len(row) >= kr_col and row[kr_col - 1] is not None else ""

                        request_type = None
                        if string_id_str not in comparison_cache:
                            if extract_new: request_type = "신규"
                        elif extract_modified:
                            if source_kr != comparison_cache[string_id_str].get('kr', ""): request_type = "변경"

                        if request_type:
                            # ▼▼▼ [수정 2-2] en, th 변수 추출 로직 제거 ▼▼▼
                            cn, tw = "", ""
                            if "CN" in headers.values() and len(row) >= list(headers.keys())[list(headers.values()).index("CN")]: cn = row[list(headers.keys())[list(headers.values()).index("CN")]-1]
                            if "TW" in headers.values() and len(row) >= list(headers.keys())[list(headers.values()).index("TW")]: tw = row[list(headers.keys())[list(headers.values()).index("TW")]-1]
                            
                            # ▼▼▼ [수정 2-3] append 구문에서 en, th 변수 제거 ▼▼▼
                            extracted_data.append((file_name, sheet_name, string_id, source_kr, cn, tw, request_type, ""))
                wb.close()

            if save_to_db:
                if extracted_data:
                    # ▼▼▼ [수정 2-4] INSERT 구문에서 en, th 컬럼 및 ? 플레이스홀더 제거 ▼▼▼
                    cursor.executemany("INSERT INTO translation_requests (id, file_name, sheet_name, string_id, kr, cn, tw, request_type, additional_info) VALUES (NULL,?,?,?,?,?,?,?,?)", extracted_data)
                    conn.commit()
                conn.close()
                loading_popup.close()
                show_message(self.parent_app.root, "info", "완료", f"비교 추출 완료: {len(extracted_data)}개 항목이 DB에 저장되었습니다.")
                self.log(f"비교 추출 완료. 총 {len(extracted_data)}개 항목 DB 저장.")
            else:
                loading_popup.close()
                if completion_callback:
                    self.parent_app.root.after(0, completion_callback, extracted_data)
                show_message(self.parent_app.root, "info", "완료", f"비교 추출 완료: {len(extracted_data)}개 항목을 찾았습니다.")
                self.log(f"비교 추출 완료. 총 {len(extracted_data)}개 항목 확인.")
            
        except Exception as e:
            loading_popup.close()
            show_message(self.parent_app.root, "error", "오류", f"비교 추출 중 오류 발생: {e}")
            self.log(f"오류 발생: {e}")
            
    def _load_comparison_data(self, compare_options):
        cache = {}
        source_type = compare_options['source_type']
        
        if source_type == "Excel":
            excel_path = compare_options['excel_path']
            sheet_name = compare_options['sheet_name']
            if not excel_path or not sheet_name: raise ValueError("비교할 엑셀 파일과 시트를 선택해야 합니다.")
            
            self.log(f"비교 데이터 로딩(Excel): {os.path.basename(excel_path)} - {sheet_name}")
            wb = load_workbook(excel_path, read_only=True, data_only=True)
            ws = wb[sheet_name]
            headers, header_row_idx = self.parent_app._find_headers_in_worksheet(ws)
            if not headers or "STRING_ID" not in headers.values():
                wb.close()
                raise ValueError("비교 엑셀에서 STRING_ID 컬럼을 찾을 수 없습니다.")

            string_id_col = list(headers.keys())[list(headers.values()).index("STRING_ID")]
            kr_col = list(headers.keys())[list(headers.values()).index("KR")] if "KR" in headers.values() else None

            for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                string_id = row[string_id_col - 1]
                if string_id:
                    kr_val = row[kr_col - 1] if kr_col and len(row) >= kr_col else ""
                    cache[str(string_id)] = {'kr': str(kr_val) if kr_val is not None else ""}
            wb.close()

        elif source_type == "DB":
            db_path = compare_options['db_path']
            if not db_path: raise ValueError("비교할 DB 파일을 선택해야 합니다.")
            self.log(f"비교 데이터 로딩(DB): {os.path.basename(db_path)}")
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='translation_requests'")
            if cursor.fetchone():
                for row in cursor.execute("SELECT string_id, kr FROM translation_requests"):
                    if row[0]: cache[str(row[0])] = {'kr': str(row[1]) if row[1] is not None else ""}
            conn.close()

        self.log(f"비교 데이터 로딩 완료: {len(cache)}개 항목")
        return cache

    # --- 엑셀 내보내기 로직 ---
    def export_to_excel(self, db_path, save_path):
        """추출된 DB의 내용을 Excel 파일로 내보냅니다."""
        self.log("엑셀로 내보내기 작업 시작...")
        try:
            conn = sqlite3.connect(db_path)
            df = pd.read_sql_query("SELECT * FROM translation_requests", conn)
            conn.close()
            
            if df.empty:
                show_message(self.parent_app.root, "info", "알림", "내보낼 데이터가 없습니다.")
                return

            df.to_excel(save_path, index=False)
            self.log(f"엑셀 내보내기 완료: {len(df)}개 항목")
            show_message(self.parent_app.root, "info", "완료", f"데이터를 엑셀로 내보냈습니다.\n파일: {save_path}")
        except Exception as e:
            self.log(f"엑셀 내보내기 오류: {e}")
            show_message(self.parent_app.root, "error", "오류", f"엑셀 내보내기 실패: {e}")