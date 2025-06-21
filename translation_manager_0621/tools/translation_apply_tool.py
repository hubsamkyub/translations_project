# tools/translate/translation_apply_tool.py (수정 후)

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import threading
import time
import sys # sys 모듈 추가

# --- 경로 문제 해결을 위한 코드 ---
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if project_root not in sys.path:
    sys.path.append(project_root)
# ---------------------------------

from ui.common_components import ScrollableCheckList, LoadingPopup
from tools.translation_apply_manager import TranslationApplyManager
import openpyxl

class TranslationApplyTool(tk.Frame):
    def __init__(self, parent, excluded_files):
        super().__init__(parent)
        self.parent = parent
        self.translation_apply_manager = TranslationApplyManager(self)
        
        # --- UI 변수 선언 ---
        # 소스 선택 관련
        self.translation_db_var = tk.StringVar()
        self.excel_source_path_var = tk.StringVar()
        self.original_folder_var = tk.StringVar()
        
        # 다중 시트 선택 관련
        self.selected_sheets_display_var = tk.StringVar(value="선택된 시트 없음")
        self.selected_sheets = []

        # 번역 적용 옵션 관련
        self.available_languages = ["KR", "EN", "CN", "TW", "TH"]
        self.apply_lang_vars = {}
        self.record_date_var = tk.BooleanVar(value=True)
        self.kr_match_check_var = tk.BooleanVar(value=True)
        self.kr_mismatch_delete_var = tk.BooleanVar(value=False)
        self.apply_smart_lookup_var = tk.BooleanVar(value=True) # [추가] 이 줄이 누락되었습니다.
        
        # 조건부 적용 옵션
        self.apply_on_new_var = tk.BooleanVar(value=True)
        self.apply_on_change_var = tk.BooleanVar(value=True)
        self.apply_on_transferred_var = tk.BooleanVar(value=False)
    
        # --- 내부 데이터 ---
        self.view_data_button = None
        self.original_files = []
        self.excluded_files = excluded_files
        self.cached_excel_path = None
        self.cached_sheet_names = []
        
        self.setup_ui()

# tools/translation_apply_tool.py

    def setup_ui(self):
        """번역 적용 탭 UI 구성 (좌/우 분할 레이아웃)"""

        # --- 상단 소스 선택 프레임 (좌/우 분할) ---
        source_selection_frame = ttk.Frame(self)
        source_selection_frame.pack(fill="x", padx=5, pady=5)
        source_selection_frame.columnconfigure(0, weight=1)
        source_selection_frame.columnconfigure(1, weight=1)

        # --- 좌측 프레임: 번역 DB 선택 ---
        db_frame = ttk.LabelFrame(source_selection_frame, text="옵션 1: 번역 DB 선택")
        db_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
        
        ttk.Label(db_frame, text="번역 DB:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        db_entry = ttk.Entry(db_frame, textvariable=self.translation_db_var)
        db_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(db_frame, text="찾아보기", command=self.select_translation_db_file).grid(row=0, column=2, padx=5, pady=5)
        db_frame.columnconfigure(1, weight=1)

        # --- 우측 프레임: 번역 엑셀 파일 선택 ---
        excel_frame = ttk.LabelFrame(source_selection_frame, text="옵션 2: 번역 엑셀 파일 선택")
        excel_frame.grid(row=0, column=1, sticky="nsew", padx=(5, 0))

        ttk.Label(excel_frame, text="엑셀 파일:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        excel_entry = ttk.Entry(excel_frame, textvariable=self.excel_source_path_var)
        excel_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(excel_frame, text="찾아보기", command=self.select_excel_source_file).grid(row=0, column=2, padx=5, pady=5)
        
        ttk.Label(excel_frame, text="시트 선택:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        selected_sheets_entry = ttk.Entry(excel_frame, textvariable=self.selected_sheets_display_var, state="readonly")
        selected_sheets_entry.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(excel_frame, text="시트 선택", command=self.open_sheet_selection_popup).grid(row=1, column=2, padx=5, pady=5)
        excel_frame.columnconfigure(1, weight=1)

        # --- 원본 파일 및 옵션 (공통 영역) ---
        original_files_frame = ttk.LabelFrame(self, text="번역을 적용할 원본 파일")
        original_files_frame.pack(fill="x", padx=5, pady=5)
        
        ttk.Label(original_files_frame, text="원본 폴더:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(original_files_frame, textvariable=self.original_folder_var).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(original_files_frame, text="찾아보기", command=self.select_original_folder).grid(row=0, column=2, padx=5, pady=5)
        ttk.Button(original_files_frame, text="파일 검색", command=self.search_original_files).grid(row=0, column=3, padx=5, pady=5)
        original_files_frame.columnconfigure(1, weight=1)
        
        files_list_frame = ttk.LabelFrame(self, text="원본 파일 목록")
        files_list_frame.pack(fill="both", expand=True, padx=5, pady=5)
        self.original_files_list = ScrollableCheckList(files_list_frame)
        self.original_files_list.pack(fill="both", expand=True, padx=5, pady=5)

        options_frame = ttk.LabelFrame(self, text="적용 옵션")
        options_frame.pack(fill="x", padx=5, pady=5)

        # --- 적용 기준 선택 (ID vs KR) ---
        self.apply_mode_var = tk.StringVar(value="id")
        self.apply_mode_var.trace_add("write", self.toggle_options_by_mode)
        
        mode_frame = ttk.Frame(options_frame)
        mode_frame.pack(fill="x", padx=5, pady=2)
        ttk.Label(mode_frame, text="적용 기준:").pack(side="left", padx=5)
        ttk.Radiobutton(mode_frame, text="ID 기반", variable=self.apply_mode_var, value="id").pack(side="left", padx=5)
        ttk.Radiobutton(mode_frame, text="KR 기반", variable=self.apply_mode_var, value="kr").pack(side="left", padx=5)

        # --- ID 기반 적용 옵션 ---
        self.id_based_options_frame = ttk.Frame(options_frame)
        self.id_based_options_frame.pack(fill="x", padx=15, pady=5)
        
        self.kr_match_check_var = tk.BooleanVar(value=True)
        self.kr_mismatch_delete_var = tk.BooleanVar(value=False)
        self.kr_overwrite_var = tk.BooleanVar(value=False)
        
        id_opt1 = ttk.Checkbutton(self.id_based_options_frame, text="KR 일치 검사 (불일치 시 건너뛰기)", variable=self.kr_match_check_var, command=self.toggle_kr_options)
        id_opt1.pack(side="left", padx=5)
        self.id_mismatch_delete_cb = ttk.Checkbutton(self.id_based_options_frame, text="└ KR 불일치 시 다국어 제거", variable=self.kr_mismatch_delete_var)
        self.id_mismatch_delete_cb.pack(side="left", padx=5)
        self.id_overwrite_cb = ttk.Checkbutton(self.id_based_options_frame, text="└ 선택 언어 덮어쓰기 (KR 일치 시)", variable=self.kr_overwrite_var)
        self.id_overwrite_cb.pack(side="left", padx=5)
        
        # --- KR 기반 적용 옵션 ---
        self.kr_based_options_frame = ttk.Frame(options_frame)
        # pack()은 toggle_options_by_mode에서 호출
        self.kr_overwrite_on_kr_mode_var = tk.BooleanVar(value=False)
        kr_opt1 = ttk.Checkbutton(self.kr_based_options_frame, text="선택 언어 덮어쓰기", variable=self.kr_overwrite_on_kr_mode_var)
        kr_opt1.pack(side="left", padx=5)
        
        # --- 공통 옵션 ---
        common_options_frame = ttk.Frame(options_frame)
        common_options_frame.pack(fill="x", pady=5)

        lang_frame = ttk.Frame(common_options_frame)
        lang_frame.pack(fill="x", padx=5, pady=2, anchor="w")
        ttk.Label(lang_frame, text="적용 언어:").pack(side="left", padx=5)
        for i, lang in enumerate(self.available_languages):
            var = tk.BooleanVar(value=True if lang in ["CN", "TW"] else False)
            self.apply_lang_vars[lang] = var
            ttk.Checkbutton(lang_frame, text=lang, variable=var).pack(side="left", padx=5)

        conditional_frame = ttk.LabelFrame(common_options_frame, text="조건부 적용")
        conditional_frame.pack(fill="x", padx=5, pady=2)
        
        cond_inner_frame = ttk.Frame(conditional_frame)
        cond_inner_frame.pack(pady=2, padx=5)
        ttk.Label(cond_inner_frame, text="#번역요청 컬럼 값이 다음과 같을 때만 적용:").pack(side="left", anchor="w")
        ttk.Checkbutton(cond_inner_frame, text="신규", variable=self.apply_on_new_var).pack(side="left", padx=(10, 5))
        ttk.Checkbutton(cond_inner_frame, text="change", variable=self.apply_on_change_var).pack(side="left", padx=5)
        ttk.Checkbutton(cond_inner_frame, text="전달", variable=self.apply_on_transferred_var).pack(side="left", padx=5)
        
        other_frame = ttk.Frame(common_options_frame)
        other_frame.pack(fill="x", padx=5, pady=2, anchor="w")
        ttk.Checkbutton(other_frame, text="번역 적용 표시", variable=self.record_date_var).pack(side="left", padx=5)
        
        action_frame = ttk.Frame(self)
        action_frame.pack(fill="x", padx=5, pady=5)
        
        ttk.Button(action_frame, text="번역 적용", command=self.apply_translation).pack(side="right", padx=5, pady=5)
        self.view_data_button = ttk.Button(action_frame, text="로드된 데이터 보기", command=self.show_loaded_data_viewer, state="disabled")
        self.view_data_button.pack(side="right", padx=5, pady=5)
        ttk.Button(action_frame, text="번역 데이터 로드", command=self.load_translation_data).pack(side="right", padx=5, pady=5)
        
        log_frame = ttk.LabelFrame(self, text="작업 로그")
        log_frame.pack(fill="both", expand=True, padx=5, pady=5)
        self.log_text = tk.Text(log_frame, wrap="word")
        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.log_text.pack(fill="both", expand=True)
        
        status_frame = ttk.Frame(self)
        status_frame.pack(fill="x", padx=5, pady=5)
        self.status_label_apply = ttk.Label(status_frame, text="대기 중...")
        self.status_label_apply.pack(side="left", padx=5)
        
        self.toggle_options_by_mode()
        self.toggle_kr_options()

    def toggle_options_by_mode(self, *args):
        """적용 기준(ID/KR)에 따라 옵션 프레임을 교체하여 보여줍니다."""
        mode = self.apply_mode_var.get()
        if mode == "id":
            self.kr_based_options_frame.pack_forget()
            self.id_based_options_frame.pack(fill="x", padx=15, pady=5)
        elif mode == "kr":
            self.id_based_options_frame.pack_forget()
            self.kr_based_options_frame.pack(fill="x", padx=15, pady=5)
        self.toggle_kr_options()

    def toggle_kr_options(self):
        """KR 일치 검사 체크박스 상태에 따라 하위 옵션 활성화/비활성화"""
        mode = self.apply_mode_var.get()
        if mode == "id":
            is_kr_check_enabled = self.kr_match_check_var.get()
            state = "normal" if is_kr_check_enabled else "disabled"
            
            # 오류가 발생한 self.kr_mismatch_cb를 올바른 변수명인 self.id_mismatch_delete_cb로 수정
            self.id_mismatch_delete_cb.config(state=state)
            self.id_overwrite_cb.config(state=state)
            
            if not is_kr_check_enabled:
                self.kr_mismatch_delete_var.set(False)
                self.kr_overwrite_var.set(False)
        else:
            # KR 기반 모드에서는 KR 일치 검사 관련 옵션 비활성화
            self.id_mismatch_delete_cb.config(state="disabled")
            self.id_overwrite_cb.config(state="disabled")
                    
# 수정 후
    def select_excel_source_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")],
            title="번역 엑셀 파일 선택", parent=self
        )
        if file_path:
            self.excel_source_path_var.set(file_path)
            self.translation_db_var.set("")
            # 다중 시트 선택 관련 변수들 초기화
            self.selected_sheets = []
            self.selected_sheets_display_var.set("선택된 시트 없음")
            # 캐시 초기화
            self.cached_excel_path = None
            self.cached_sheet_names = []
            
    def _populate_sheets(self, file_path):
        """엑셀 파일에서 시트 목록을 읽어 콤보박스를 채웁니다."""
        try:
            self.after(0, lambda: self.sheet_combobox.set("시트 목록 읽는 중..."))
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet_names = workbook.sheetnames
            
            def update_combobox():
                self.sheet_combobox['values'] = sheet_names
                if sheet_names:
                    self.sheet_combobox.set(sheet_names[0]) # 첫 번째 시트를 기본값으로
                self.sheet_combobox.config(state="readonly")

            self.after(0, update_combobox)
        except Exception as e:
            self.after(0, lambda: messagebox.showerror("오류", f"엑셀 파일 시트를 읽는 중 오류 발생: {e}", parent=self))
            self.after(0, lambda: self.sheet_combobox.set("시트 읽기 실패"))


    def load_translation_data(self):
        db_path = self.translation_db_var.get()
        excel_path = self.excel_source_path_var.get()

        if db_path:
            self.load_from_db(db_path)
        elif excel_path:
            # [수정] 선택된 시트 목록(self.selected_sheets)을 전달
            if not self.selected_sheets:
                messagebox.showwarning("경고", "'시트 선택' 버튼을 눌러 데이터를 읽어올 시트를 선택하세요.", parent=self)
                return
            self.load_from_excel(excel_path, self.selected_sheets)
        else:
            messagebox.showwarning("경고", "번역 DB 또는 엑셀 파일을 선택하세요.", parent=self)


    def open_sheet_selection_popup(self):
        """[최적화] 다중 시트 선택을 위한 팝업창을 열고, 시트 목록을 캐싱합니다."""
        excel_path = self.excel_source_path_var.get()
        if not excel_path or not os.path.exists(excel_path):
            messagebox.showwarning("파일 선택 필요", "먼저 번역 엑셀 파일을 선택해주세요.", parent=self)
            return

        all_sheets = []
        # [수정] 캐시 확인 로직 추가
        if self.cached_excel_path == excel_path and self.cached_sheet_names:
            # 캐시된 경로와 현재 경로가 같으면, 저장된 시트 목록을 사용
            all_sheets = self.cached_sheet_names
            self.log_text.insert(tk.END, "캐시된 시트 목록을 사용합니다.\n")
        else:
            # 캐시가 없으면 파일을 읽고, 결과를 캐시에 저장
            try:
                from openpyxl import load_workbook
                self.log_text.insert(tk.END, f"'{os.path.basename(excel_path)}' 파일에서 시트 목록을 읽는 중...\n")
                self.update() # 로그가 즉시 보이도록 강제 업데이트
                
                wb = load_workbook(excel_path, read_only=True)
                all_sheets = wb.sheetnames
                wb.close()
                
                # 읽어온 결과를 캐시에 저장
                self.cached_excel_path = excel_path
                self.cached_sheet_names = all_sheets
                self.log_text.insert(tk.END, "시트 목록 로드 및 캐시 저장 완료.\n")

            except Exception as e:
                messagebox.showerror("파일 오류", f"엑셀 파일을 읽는 중 오류가 발생했습니다:\n{e}", parent=self)
                return

        popup = tk.Toplevel(self)
        popup.title("시트 선택")
        popup.geometry("400x500")
        popup.transient(self)
        popup.grab_set()

        checklist = ScrollableCheckList(popup, height=350)
        checklist.pack(fill="both", expand=True, padx=10, pady=10)

        selected_set = set(self.selected_sheets)
        for sheet in all_sheets:
            checklist.add_item(sheet, checked=(sheet in selected_set))

        def on_confirm():
            self.selected_sheets = checklist.get_checked_items()
            if self.selected_sheets:
                if len(self.selected_sheets) > 3:
                    display_text = f"{len(self.selected_sheets)}개 시트 선택됨"
                else:
                    display_text = ", ".join(self.selected_sheets)
                self.selected_sheets_display_var.set(display_text)
            else:
                self.selected_sheets_display_var.set("선택된 시트 없음")
            popup.destroy()

        confirm_button = ttk.Button(popup, text="확인", command=on_confirm)
        confirm_button.pack(pady=10)

    def load_from_db(self, db_path):
        if not os.path.isfile(db_path):
            messagebox.showwarning("경고", "유효한 번역 DB 파일을 선택하세요.", parent=self)
            return
            
        self.log_text.insert(tk.END, "번역 DB 캐싱 중...\n")
        loading_popup = LoadingPopup(self, "DB 캐싱 중", "번역 데이터 캐싱 중...")
        
        def task():
            result = self.translation_apply_manager.load_translation_cache_from_db(db_path)
            self.after(0, lambda: self.process_cache_load_result(result, loading_popup))
        
        threading.Thread(target=task, daemon=True).start()

    def load_from_excel(self, file_path, sheet_names): # sheet_name -> sheet_names
        self.log_text.insert(tk.END, f"'{os.path.basename(file_path)}' 파일의 {len(sheet_names)}개 시트 캐싱 중...\n")
        loading_popup = LoadingPopup(self, "엑셀 캐싱 중", "번역 데이터 캐싱 중...")
        
        def task():
            result = self.translation_apply_manager.load_translation_cache_from_excel(file_path, sheet_names) # sheet_names 리스트 전달
            self.after(0, lambda: self.process_cache_load_result(result, loading_popup))
        
        threading.Thread(target=task, daemon=True).start()
   
   
    def select_translation_db_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("DB 파일", "*.db"), ("모든 파일", "*.*")],
            title="번역 DB 선택", parent=self
        )
        if file_path:
            self.translation_db_var.set(file_path)
            # 다른 옵션 초기화
            self.excel_source_path_var.set("") 
            self.selected_sheets = []
            self.selected_sheets_display_var.set("선택된 시트 없음")
            self.cached_excel_path = None
            self.cached_sheet_names = []
    def select_original_folder(self):
        folder = filedialog.askdirectory(title="원본 파일 폴더 선택", parent=self)
        if folder:
            self.original_folder_var.set(folder)

    def search_original_files(self):
        folder = self.original_folder_var.get()
        if not folder or not os.path.isdir(folder):
            messagebox.showwarning("경고", "유효한 폴더를 선택하세요.", parent=self)
            return
        
        self.original_files_list.clear()
        self.original_files = []
        
        for root_dir, _, files in os.walk(folder):
            for file in files:
                if file.startswith("String") and file.endswith(".xlsx") and not file.startswith("~$"):
                    if file not in self.excluded_files:
                        file_path = os.path.join(root_dir, file)
                        self.original_files.append((file, file_path))
                        self.original_files_list.add_item(file, checked=True)
        
        if not self.original_files:
            messagebox.showinfo("알림", "String으로 시작하는 엑셀 파일을 찾지 못했습니다.", parent=self)
        else:
            messagebox.showinfo("알림", f"{len(self.original_files)}개의 엑셀 파일을 찾았습니다.", parent=self)

    def load_translation_cache(self):
        db_path = self.translation_db_var.get()
        if not db_path or not os.path.isfile(db_path):
            messagebox.showwarning("경고", "유효한 번역 DB 파일을 선택하세요.", parent=self)
            return
        
        self.log_text.insert(tk.END, "번역 DB 캐싱 중...\n")
        self.update()
        
        loading_popup = LoadingPopup(self, "번역 DB 캐싱 중", "번역 데이터 캐싱 중...")
        
        def load_cache():
            try:
                result = self.translation_apply_manager.load_translation_cache(db_path)
                self.after(0, lambda: self.process_cache_load_result(result, loading_popup))
            except Exception as e:
                error_msg = str(e)
                self.after(0, lambda: [
                    loading_popup.close(),
                    self.log_text.insert(tk.END, f"캐싱 중 오류 발생: {error_msg}\n"),
                    self.status_label_apply.config(text="오류 발생"),
                    messagebox.showerror("오류", f"DB 캐싱 중 오류 발생: {error_msg}", parent=self)
                ])
                
        thread = threading.Thread(target=load_cache, daemon=True)
        thread.start()
        
    def process_cache_load_result(self, result, loading_popup):
        loading_popup.close()
        
        if result["status"] == "error":
            messagebox.showerror("오류", f"캐싱 중 오류 발생: {result['message']}", parent=self)
            self.log_text.insert(tk.END, f"캐싱 실패: {result['message']}\n")
            return
            
        self.translation_apply_manager.translation_cache = result["translation_cache"]
        self.translation_apply_manager.translation_file_cache = result["translation_file_cache"]
        self.translation_apply_manager.translation_sheet_cache = result["translation_sheet_cache"]
        self.translation_apply_manager.duplicate_ids = result["duplicate_ids"]
        
        file_count = result["file_count"]
        sheet_count = result["sheet_count"]
        id_count = result["id_count"]
        
        duplicate_count = sum(1 for ids in result["duplicate_ids"].values() if len(ids) > 1)
        if duplicate_count > 0:
            self.log_text.insert(tk.END, f"\n주의: {duplicate_count}개의 STRING_ID가 여러 파일에 중복 존재합니다.\n")
            dup_examples = [(id, files) for id, files in result["duplicate_ids"].items() if len(files) > 1][:5]
            for id, files in dup_examples:
                self.log_text.insert(tk.END, f"  - {id}: {', '.join(files)}\n")
            if len(dup_examples) < duplicate_count:
                self.log_text.insert(tk.END, f"  ... 외 {duplicate_count - len(dup_examples)}개\n")
        
        self.log_text.insert(tk.END, f"번역 DB 캐싱 완료:\n")
        self.log_text.insert(tk.END, f"- 파일별 캐시: {file_count}개 파일, {sum(len(ids) for ids in result['translation_file_cache'].values())}개 항목\n")
        self.log_text.insert(tk.END, f"- 시트별 캐시: {sheet_count}개 시트, {sum(len(ids) for ids in result['translation_sheet_cache'].values())}개 항목\n")
        self.log_text.insert(tk.END, f"- 전체 고유 STRING_ID: {id_count}개\n")
        
        self.status_label_apply.config(text=f"번역 DB 캐싱 완료 - {id_count}개 항목")
        self.view_data_button.config(state="normal")
        
        messagebox.showinfo(
            "완료", 
            f"번역 DB 캐싱 완료!\n파일 수: {file_count}개\n시트 수: {sheet_count}개\n항목 수: {id_count}개", 
            parent=self
        )


    def apply_translation(self):
        if not hasattr(self.translation_apply_manager, 'translation_cache') or not self.translation_apply_manager.translation_cache:
            messagebox.showwarning("경고", "먼저 '번역 데이터 로드'를 실행하세요.", parent=self)
            return
            
        selected_files = self.original_files_list.get_checked_items()
        if not selected_files:
            messagebox.showwarning("경고", "적용할 파일을 선택하세요.", parent=self)
            return
            
        selected_langs = [lang for lang, var in self.apply_lang_vars.items() if var.get()]
        if not selected_langs:
            messagebox.showwarning("경고", "적용할 언어를 하나 이상 선택하세요.", parent=self)
            return

        files_to_process = [item for item in self.original_files if item[0] in selected_files]
        open_files = self._check_files_are_open([path for name, path in files_to_process])
        if open_files:
            messagebox.showwarning("작업 중단", f"다음 파일이 열려 있어 작업을 시작할 수 없습니다:\n\n" + "\n".join(open_files), parent=self)
            return

        self.log_text.delete(1.0, tk.END)
        
        # 작업 시작 로그
        allowed_statuses = []
        if self.apply_on_new_var.get(): allowed_statuses.append('신규')
        if self.apply_on_change_var.get(): allowed_statuses.append('change')
        if self.apply_on_transferred_var.get(): allowed_statuses.append('전달')
        
        # 옵션 요약 출력
        mode_text = "ID 기반" if self.apply_mode_var.get() == "id" else "KR 기반"
        lang_text = ", ".join(selected_langs)
        condition_text = ", ".join(allowed_statuses) if allowed_statuses else "모든 항목"
        
        self.log_text.insert(tk.END, "="*60 + "\n")
        self.log_text.insert(tk.END, "🚀 번역 적용 작업 시작\n")
        self.log_text.insert(tk.END, f"📋 적용 모드: {mode_text}\n")
        self.log_text.insert(tk.END, f"🌍 적용 언어: {lang_text}\n")
        self.log_text.insert(tk.END, f"🎯 적용 조건: {condition_text}\n")
        self.log_text.insert(tk.END, f"📁 대상 파일: {len(files_to_process)}개\n")
        self.log_text.insert(tk.END, "="*60 + "\n\n")
        
        self.status_label_apply.config(text="작업 중...")
        self.update()
            
        loading_popup = LoadingPopup(self, "번역 적용 중", "번역 적용 준비 중...")
            
        apply_options = {
            "mode": self.apply_mode_var.get(),
            "selected_langs": selected_langs,
            "record_date": self.record_date_var.get(),
            "kr_match_check": self.kr_match_check_var.get(),
            "kr_mismatch_delete": self.kr_mismatch_delete_var.get(),
            "kr_overwrite": self.kr_overwrite_var.get(),
            "kr_overwrite_on_kr_mode": self.kr_overwrite_on_kr_mode_var.get(),
            "allowed_statuses": allowed_statuses,
        }
            
        def apply_translations_thread():
            total_results = {
                "total_updated": 0, "total_overwritten": 0, "total_kr_mismatch_skipped": 0, 
                "total_kr_mismatch_deleted": 0, "total_smart_applied": 0, 
                "total_conditional_skipped": 0,
            }
            processed_count = 0
            error_count = 0
            successful_files = []
            failed_files = []
            
            start_time = time.time()
            
            for idx, (file_name, file_path) in enumerate(files_to_process):
                self.after(0, lambda i=idx, n=file_name: [
                    loading_popup.update_progress((i / len(files_to_process)) * 100, f"파일 처리 중 ({i+1}/{len(files_to_process)}): {n}"),
                ])
                
                result = self.translation_apply_manager.apply_translation(
                    file_path,
                    apply_options
                )
                
                if result["status"] == "success":
                    processed_count += 1
                    successful_files.append(file_name)
                    for key in total_results:
                        total_results[key] += result.get(key, 0)
                else:
                    error_count += 1
                    failed_files.append((file_name, result.get("message", "알 수 없는 오류")))
            
            elapsed_time = time.time() - start_time
            
            self.after(0, lambda: self.process_translation_apply_result(
                total_results, processed_count, error_count, loading_popup, 
                successful_files, failed_files, elapsed_time)
            )

        thread = threading.Thread(target=apply_translations_thread, daemon=True)
        thread.start()

    def process_translation_apply_result(self, total_results, processed_count, error_count, loading_popup, successful_files, failed_files, elapsed_time):
        """번역 적용 스레드 완료 후 결과를 처리하고 UI에 표시합니다."""
        loading_popup.close()

        # 시간 포맷팅
        minutes = int(elapsed_time // 60)
        seconds = int(elapsed_time % 60)
        time_str = f"{minutes}분 {seconds}초" if minutes > 0 else f"{seconds}초"

        # 최종 요약 로그
        self.log_text.insert(tk.END, "\n" + "="*60 + "\n")
        self.log_text.insert(tk.END, "🎉 번역 적용 작업 완료\n")
        self.log_text.insert(tk.END, "="*60 + "\n")
        
        # 성공/실패 요약
        self.log_text.insert(tk.END, f"⏱️  소요 시간: {time_str}\n")
        self.log_text.insert(tk.END, f"✅ 성공: {processed_count}개 파일\n")
        if error_count > 0:
            self.log_text.insert(tk.END, f"❌ 실패: {error_count}개 파일\n")
        
        # 작업 통계
        total_applied = total_results["total_updated"] + total_results["total_overwritten"]
        self.log_text.insert(tk.END, f"\n📊 작업 통계:\n")
        self.log_text.insert(tk.END, f"   • 신규 적용: {total_results['total_updated']:,}개\n")
        if total_results["total_overwritten"] > 0:
            self.log_text.insert(tk.END, f"   • 덮어쓰기: {total_results['total_overwritten']:,}개\n")
        if total_results["total_conditional_skipped"] > 0:
            self.log_text.insert(tk.END, f"   • 조건 불일치로 건너뜀: {total_results['total_conditional_skipped']:,}개\n")
        if total_results["total_kr_mismatch_skipped"] > 0:
            self.log_text.insert(tk.END, f"   • KR 불일치로 건너뜀: {total_results['total_kr_mismatch_skipped']:,}개\n")
        if total_results["total_kr_mismatch_deleted"] > 0:
            self.log_text.insert(tk.END, f"   • KR 불일치로 삭제: {total_results['total_kr_mismatch_deleted']:,}개\n")
        
        self.log_text.insert(tk.END, f"\n🎯 총 적용된 번역: {total_applied:,}개\n")
        
        # 실패한 파일 상세 정보
        if failed_files:
            self.log_text.insert(tk.END, f"\n❌ 실패한 파일:\n")
            for file_name, error_msg in failed_files[:5]:  # 최대 5개까지만 표시
                self.log_text.insert(tk.END, f"   • {file_name}: {error_msg}\n")
            if len(failed_files) > 5:
                self.log_text.insert(tk.END, f"   ... 외 {len(failed_files) - 5}개\n")
        
        self.log_text.insert(tk.END, "="*60 + "\n")
        self.log_text.see(tk.END)
        
        self.status_label_apply.config(text=f"번역 적용 완료 - {total_applied:,}개 항목 적용")
        
        # 완료 메시지 박스
        completion_msg = f"번역 적용이 완료되었습니다!\n\n"
        completion_msg += f"⏱️ 소요 시간: {time_str}\n"
        completion_msg += f"✅ 성공: {processed_count}개 파일\n"
        if error_count > 0:
            completion_msg += f"❌ 실패: {error_count}개 파일\n"
        completion_msg += f"\n🎯 총 적용된 번역: {total_applied:,}개"
        
        if total_results["total_updated"] > 0:
            completion_msg += f"\n   • 신규 적용: {total_results['total_updated']:,}개"
        if total_results["total_overwritten"] > 0:
            completion_msg += f"\n   • 덮어쓰기: {total_results['total_overwritten']:,}개"
        
        messagebox.showinfo("완료", completion_msg, parent=self)




    def _check_files_are_open(self, file_paths_to_check):
        """
        주어진 파일 경로 목록을 확인하여 열려 있는 파일이 있는지 검사합니다.
        파일을 리네임하는 방식으로 잠금 상태를 확인하며, 이는 Windows 환경에서 효과적입니다.
        """
        open_files = []
        for file_path in file_paths_to_check:
            if not os.path.exists(file_path):
                continue
            try:
                # 파일을 자기 자신으로 리네임 시도. 파일이 열려있으면 OSError(PermissionError) 발생
                os.rename(file_path, file_path)
            except OSError:
                open_files.append(os.path.basename(file_path))
        return open_files
    
    # tools/translation_apply_tool.py

    def show_loaded_data_viewer(self):
        """로드된 번역 데이터를 보여주는 새 창을 엽니다."""
        if not hasattr(self.translation_apply_manager, 'translation_cache') or not self.translation_apply_manager.translation_cache:
            messagebox.showinfo("정보", "먼저 번역 데이터를 로드해주세요.", parent=self)
            return

        viewer_win = tk.Toplevel(self)
        viewer_win.title("로드된 번역 데이터 보기")
        viewer_win.geometry("1200x700")
        viewer_win.transient(self)
        viewer_win.grab_set()

        # --- 상단 검색 프레임 ---
        search_frame = ttk.Frame(viewer_win, padding="5")
        search_frame.pack(fill="x")
        
        ttk.Label(search_frame, text="STRING_ID:").pack(side="left", padx=(0, 2))
        id_search_var = tk.StringVar()
        id_search_entry = ttk.Entry(search_frame, textvariable=id_search_var, width=30)
        id_search_entry.pack(side="left", padx=(0, 10))

        ttk.Label(search_frame, text="KR:").pack(side="left", padx=(0, 2))
        kr_search_var = tk.StringVar()
        kr_search_entry = ttk.Entry(search_frame, textvariable=kr_search_var, width=40)
        kr_search_entry.pack(side="left", padx=(0, 10))

        # --- 중간 데이터 표시 프레임 (Treeview) ---
        tree_frame = ttk.Frame(viewer_win, padding="5")
        tree_frame.pack(fill="both", expand=True)

        columns = ("string_id", "kr", "en", "cn", "tw", "th", "file_name", "sheet_name")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings")
        
        # 컬럼 헤더 설정
        tree.heading("string_id", text="STRING_ID")
        tree.heading("kr", text="KR")
        tree.heading("en", text="EN")
        tree.heading("cn", text="CN")
        tree.heading("tw", text="TW")
        tree.heading("th", text="TH")
        tree.heading("file_name", text="파일명")
        tree.heading("sheet_name", text="시트명")

        # 컬럼 너비 설정
        tree.column("string_id", width=150)
        tree.column("kr", width=250)
        tree.column("en", width=200)
        tree.column("cn", width=200)
        tree.column("tw", width=200)
        tree.column("th", width=100)
        tree.column("file_name", width=150)
        tree.column("sheet_name", width=150)

        # 스크롤바 추가
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        vsb.pack(side="right", fill="y")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
        hsb.pack(side="bottom", fill="x")
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.pack(fill="both", expand=True)

        # --- 하단 상태 표시줄 ---
        status_frame = ttk.Frame(viewer_win, padding="5")
        status_frame.pack(fill="x")
        status_label = ttk.Label(status_frame, text="데이터 준비 중...")
        status_label.pack(side="left")

        # --- 데이터 처리 및 함수 ---
        # 원본 데이터 준비 (STRING_ID를 각 딕셔너리에 포함시켜 관리 용이성 증대)
        all_data = []
        for string_id, data_dict in self.translation_apply_manager.translation_cache.items():
            item = data_dict.copy()
            item['string_id'] = string_id
            all_data.append(item)

        def populate_tree(data_to_show):
            """Treeview를 주어진 데이터로 채우는 함수"""
            # 기존 데이터 삭제 (성능을 위해 보이지 않게 처리)
            tree.delete(*tree.get_children())
            
            # 새 데이터 추가
            for item in data_to_show:
                values = (
                    item.get('string_id', ''),
                    item.get('kr', ''),
                    item.get('en', ''),
                    item.get('cn', ''),
                    item.get('tw', ''),
                    item.get('th', ''),
                    item.get('file_name', ''),
                    item.get('sheet_name', '')
                )
                tree.insert("", "end", values=values)
            status_label.config(text=f"{len(data_to_show):,} / {len(all_data):,}개 항목 표시 중")

        def perform_search():
            """검색 버튼 클릭 시 필터링 수행"""
            id_query = id_search_var.get().lower().strip()
            kr_query = kr_search_var.get().lower().strip()

            if not id_query and not kr_query:
                populate_tree(all_data)
                return

            # 필터링 로직
            filtered_data = []
            for item in all_data:
                id_match = (id_query in item.get('string_id', '').lower()) if id_query else True
                kr_match = (kr_query in item.get('kr', '').lower()) if kr_query else True
                
                if id_match and kr_match:
                    filtered_data.append(item)
            
            populate_tree(filtered_data)

        def reset_search():
            """검색 조건 초기화"""
            id_search_var.set("")
            kr_search_var.set("")
            populate_tree(all_data)

        # --- 검색 버튼과 함수 연결 ---
        search_button = ttk.Button(search_frame, text="검색", command=perform_search)
        search_button.pack(side="left", padx=5)
        reset_button = ttk.Button(search_frame, text="초기화", command=reset_search)
        reset_button.pack(side="left", padx=5)
        
        # 엔터 키로 검색 실행
        id_search_entry.bind("<Return>", lambda event: perform_search())
        kr_search_entry.bind("<Return>", lambda event: perform_search())

        # --- 초기 데이터 로드 ---
        populate_tree(all_data)