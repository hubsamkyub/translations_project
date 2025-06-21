# translation_db_manager.py 수정 사항
from collections import defaultdict
import sqlite3
import os
import json
import time
import pandas as pd
import re
from openpyxl import load_workbook
import gc
from openpyxl.styles import PatternFill
from datetime import datetime

class TranslationDBManager:
    def __init__(self, parent_window=None):
        self.parent = parent_window
        self.excluded_count = 0
    
    def find_string_id_position(self, worksheet):
        for row in worksheet.iter_rows(min_row=1, max_row=6, max_col=5):
            for cell in row:
                if isinstance(cell.value, str) and "STRING_ID" in cell.value.upper():
                    return cell.column, cell.row
        return None, None

    def find_language_columns(self, worksheet, header_row, langs, language_mapping=None):
        if not header_row: return {}
        lang_cols = {}
        extended_mapping = {}
        if language_mapping:
            for alt, main in language_mapping.items():
                extended_mapping[alt.upper()] = main
        for row in worksheet.iter_rows(min_row=header_row, max_row=header_row):
            for cell in row:
                if not cell.value: continue
                header_text = str(cell.value).strip().upper()
                if header_text in langs:
                    lang_cols[header_text] = cell.column
                elif header_text in extended_mapping:
                    mapped_lang = extended_mapping[header_text]
                    if mapped_lang in langs:
                        lang_cols[mapped_lang] = cell.column
        return lang_cols

    def build_translation_db(self, excel_files, output_db_path, language_list, batch_size=2000, use_read_only=True, progress_callback=None):
        if not excel_files: return {"status": "error", "message": "번역 파일이 선택되지 않았습니다."}
        if not output_db_path: return {"status": "error", "message": "DB 파일 경로가 지정되지 않았습니다."}
        if not language_list: return {"status": "error", "message": "하나 이상의 언어를 선택하세요."}
        
        language_mapping = {"ZH": "CN"}
        if os.path.exists(output_db_path): os.remove(output_db_path)
        
        conn = None
        try:
            conn = sqlite3.connect(output_db_path)
            cursor = conn.cursor()
            cursor.execute('''
            CREATE TABLE translation_data (
                id INTEGER PRIMARY KEY, file_name TEXT, sheet_name TEXT, string_id TEXT UNIQUE,
                kr TEXT, en TEXT, cn TEXT, tw TEXT, th TEXT, status TEXT DEFAULT 'active', update_date TEXT
            )''')
            conn.commit()

            unique_data = {}
            duplicate_data_preview = defaultdict(list)
            processed_count = 0
            error_count = 0
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            for idx, (file_name, file_path) in enumerate(excel_files):
                if progress_callback:
                    progress_callback(f"파일 ({idx+1}/{len(excel_files)}) {file_name} 처리 중...", idx, len(excel_files))
                
                try:
                    gc.collect()
                    workbook = load_workbook(file_path, read_only=use_read_only, data_only=True)
                    
                    for sheet_name in workbook.sheetnames:
                        if not sheet_name.lower().startswith("string") or sheet_name.startswith("#"): continue
                        worksheet = workbook[sheet_name]
                        string_id_col, header_row = self.find_string_id_position(worksheet)
                        if not string_id_col or not header_row: continue
                        lang_cols = self.find_language_columns(worksheet, header_row, language_list, language_mapping)
                        if not lang_cols: continue
                        
                        for row_cells in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
                            if not row_cells or string_id_col - 1 >= len(row_cells): continue
                            string_id = row_cells[string_id_col - 1]
                            if not string_id: continue
                            
                            status = '비활성' if str(row_cells[0] or '').strip().startswith('#') else 'active'
                            
                            values = {"string_id": string_id}
                            for lang, col in lang_cols.items():
                                if col - 1 < len(row_cells):
                                    values[lang.lower()] = row_cells[col - 1]
                            
                            current_data_tuple = (
                                file_name, sheet_name, values.get("string_id"),
                                values.get("kr"), values.get("en"), values.get("cn"), 
                                values.get("tw"), values.get("th"), status, current_time
                            )
                            current_data_dict = {
                                'file_name': file_name, 'sheet_name': sheet_name, 'string_id': string_id,
                                'kr': values.get("kr"), 'en': values.get("en"), 'cn': values.get("cn"),
                                'tw': values.get("tw"), 'th': values.get("th"), 'status': status
                            }

                            if string_id not in unique_data and string_id not in duplicate_data_preview:
                                unique_data[string_id] = current_data_tuple
                            elif string_id in unique_data:
                                original_tuple = unique_data.pop(string_id)
                                original_dict = {
                                    'file_name': original_tuple[0], 'sheet_name': original_tuple[1], 'string_id': original_tuple[2],
                                    'kr': original_tuple[3], 'en': original_tuple[4], 'cn': original_tuple[5],
                                    'tw': original_tuple[6], 'th': original_tuple[7], 'status': original_tuple[8]
                                }
                                duplicate_data_preview[string_id].extend([original_dict, current_data_dict])
                            else:
                                duplicate_data_preview[string_id].append(current_data_dict)

                    workbook.close()
                    processed_count += 1
                except Exception as e:
                    if progress_callback: progress_callback(f"파일 처리 오류: {e}", idx + 1, len(excel_files))
                    error_count += 1
            
            batch_data_to_insert = list(unique_data.values())
            if batch_data_to_insert:
                cursor.executemany('''
                INSERT OR IGNORE INTO translation_data (file_name, sheet_name, string_id, kr, en, cn, tw, th, status, update_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', batch_data_to_insert)
            
            conn.commit()
            return {"status": "success", "processed_count": processed_count, "error_count": error_count, "total_rows": len(batch_data_to_insert), "duplicates": dict(duplicate_data_preview)}
        except Exception as e:
            return {"status": "error", "message": str(e)}
        finally:
            if conn: conn.close()
                
    def create_translation_table(self, cursor, table_name="translation_data"):
        """번역 테이블 생성 (상태 및 업데이트 날짜 포함)"""
        cursor.execute(f'''
        CREATE TABLE IF NOT EXISTS {table_name} (
            id INTEGER PRIMARY KEY,
            file_name TEXT,
            sheet_name TEXT,
            string_id TEXT UNIQUE,
            kr TEXT,
            en TEXT, 
            cn TEXT,
            tw TEXT,
            th TEXT,
            status TEXT DEFAULT 'active',
            update_date TEXT
        )
        ''')
        
        # 인덱스 생성
        cursor.execute(f'CREATE INDEX IF NOT EXISTS idx_{table_name}_string_id ON {table_name}(string_id)')
        cursor.execute(f'CREATE INDEX IF NOT EXISTS idx_{table_name}_file_sheet ON {table_name}(file_name, sheet_name)')
        cursor.execute(f'CREATE INDEX IF NOT EXISTS idx_{table_name}_status ON {table_name}(status)')

    def update_translation_db(self, excel_files, db_path, language_list, batch_size=2000, use_read_only=True, progress_callback=None, update_option="default", debug_string_id=None):
        if not excel_files: return {"status": "error", "message": "번역 파일이 선택되지 않았습니다."}
        if not db_path or not os.path.exists(db_path): return {"status": "error", "message": "유효한 DB 파일 경로를 지정하세요."}
        if not language_list: return {"status": "error", "message": "하나 이상의 언어를 선택하세요."}

        language_mapping = {"ZH": "CN"}
        conn = None
        try:
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            self._update_table_schema(cursor)
            
            # 중복 데이터와 고유 데이터 분리
            unique_excel_data, duplicate_data_preview = self._collect_and_filter_excel_data(
                excel_files, language_list, language_mapping, progress_callback, debug_string_id
            )
            
            db_data_map = self._load_existing_data(cursor, update_option)
            updates, inserts = [], []
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # 고유 데이터만으로 업데이트/삽입 목록 생성
            for key, excel_data in unique_excel_data.items():
                status = excel_data.get('status', 'active')
                
                # DB에 키가 있는지 확인
                if key in db_data_map:
                    db_row = db_data_map[key]
                    if self._is_update_needed(excel_data, db_row, language_list, update_option, status):
                        excel_data.update({'key': key, 'status': status, 'update_date': current_time})
                        updates.append(excel_data)
                else:
                    excel_data.update({'status': status, 'update_date': current_time})
                    inserts.append(excel_data)

            # 배치 처리
            updated_row_count = self._execute_batch_update(cursor, updates, language_list, update_option)
            new_row_count = self._execute_batch_insert(cursor, inserts, language_list)

            conn.commit()
            return {
                "status": "success", "processed_count": len(excel_files), "error_count": 0,
                "total_rows": updated_row_count + new_row_count, "updated_rows": updated_row_count, "new_rows": new_row_count,
                "deleted_rows": 0, "duplicates": dict(duplicate_data_preview)
            }
        except Exception as e:
            return {"status": "error", "message": str(e)}
        finally:
            if conn: conn.close()

    def _collect_and_filter_excel_data(self, excel_files, language_list, language_mapping, progress_callback, debug_string_id):
        unique_data = {}
        duplicate_data_preview = defaultdict(list)

        for idx, (file_name, file_path) in enumerate(excel_files):
            if progress_callback:
                progress_callback(f"파일 데이터 수집 중 ({idx+1}/{len(excel_files)}) {file_name}...", idx, len(excel_files))

            workbook = load_workbook(file_path, read_only=True, data_only=True)
            for sheet_name in workbook.sheetnames:
                if not sheet_name.lower().startswith('string') or sheet_name.startswith("#"):
                    continue
                worksheet = workbook[sheet_name]
                string_id_col, header_row = self.find_string_id_position(worksheet)
                if not string_id_col or not header_row: continue
                lang_cols = self.find_language_columns(worksheet, header_row, language_list, language_mapping)

                for row_cells in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
                    if not row_cells or string_id_col - 1 >= len(row_cells): continue
                    string_id = row_cells[string_id_col - 1]
                    if not string_id: continue
                    
                    status = '비활성' if str(row_cells[0] or '').strip().startswith('#') else 'active'
                    
                    excel_data = {'string_id': string_id, 'file_name': file_name, 'sheet_name': sheet_name, 'status': status}
                    for lang, col in lang_cols.items():
                        if col - 1 < len(row_cells):
                            excel_data[lang.lower()] = row_cells[col - 1]
                    
                    if string_id not in unique_data and string_id not in duplicate_data_preview:
                        unique_data[string_id] = excel_data
                    elif string_id in unique_data:
                        duplicate_data_preview[string_id].extend([unique_data.pop(string_id), excel_data])
                    else:
                        duplicate_data_preview[string_id].append(excel_data)
            workbook.close()
        return unique_data, duplicate_data_preview

    def _is_update_needed(self, excel_data, db_row, language_list, update_option, new_status):
        if db_row.get('status') != new_status: return True
        
        update_langs = language_list.copy()
        if update_option in ['default', 'kr_additional_compare'] and 'KR' in update_langs:
            update_langs.remove('KR')

        for lang in update_langs:
            lang_key = lang.lower()
            excel_val = str(excel_data.get(lang_key) or '').strip()
            db_val = str(db_row.get(lang_key) or '').strip()
            if excel_val != db_val: return True
        return False
        
    def _load_existing_data(self, cursor, update_option):
        cursor.execute("SELECT * FROM translation_data")
        db_data_map = {}
        for row in cursor.fetchall():
            row_dict = {desc[0]: value for desc, value in zip(cursor.description, row)}
            key = None
            if update_option == "kr_compare":
                if row_dict.get('kr'): key = row_dict['kr']
            elif update_option == "kr_additional_compare":
                if row_dict.get('string_id') and row_dict.get('kr'): key = (row_dict['string_id'], row_dict['kr'])
            else: key = row_dict.get('string_id')
            if key and key not in db_data_map:
                db_data_map[key] = row_dict
        return db_data_map

    def _execute_batch_update(self, cursor, updates, language_list, update_option):
        if not updates: return 0
        update_cols = [lang.lower() for lang in language_list]
        if update_option in ["default", "kr_additional_compare"] and 'kr' in update_cols:
            update_cols.remove('kr')
        if update_option == "kr_compare" and 'string_id' not in update_cols:
            update_cols.append('string_id')
        set_clause = ", ".join([f"{col} = :{col}" for col in update_cols])
        set_clause += ", status = :status, update_date = :update_date, file_name = :file_name, sheet_name = :sheet_name"
        
        where_clause = ""
        if update_option == "kr_compare": where_clause = "WHERE kr = :key"
        elif update_option == "kr_additional_compare": where_clause = "WHERE string_id = :key0 AND kr = :key1"
        else: where_clause = "WHERE string_id = :key"

        query = f"UPDATE translation_data SET {set_clause} {where_clause}"
        update_params = []
        for item in updates:
            params = item.copy()
            if update_option == "kr_additional_compare":
                params['key0'], params['key1'] = item['key']
            update_params.append(params)

        cursor.executemany(query, update_params)
        return cursor.rowcount

    def _execute_batch_insert(self, cursor, inserts, language_list):
        if not inserts: return 0
        cols = ['file_name', 'sheet_name', 'string_id', 'status', 'update_date'] + [lang.lower() for lang in language_list]
        placeholders = ", ".join([f":{col}" for col in cols])
        query = f"INSERT OR IGNORE INTO translation_data ({', '.join(cols)}) VALUES ({placeholders})"
        cursor.executemany(query, inserts)
        return cursor.rowcount
                     
    def _update_table_schema(self, cursor):
        """테이블 스키마 업데이트 (status, update_date 컬럼 추가)"""
        try:
            # 기존 테이블 구조 확인
            cursor.execute("PRAGMA table_info(translation_data)")
            columns = [row[1] for row in cursor.fetchall()]
            
            # status 컬럼이 없으면 추가
            if 'status' not in columns:
                cursor.execute("ALTER TABLE translation_data ADD COLUMN status TEXT DEFAULT 'active'")
                # 기존 데이터의 status를 'active'로 업데이트
                cursor.execute("UPDATE translation_data SET status = 'active' WHERE status IS NULL")
            
            # update_date 컬럼이 없으면 추가
            if 'update_date' not in columns:
                cursor.execute("ALTER TABLE translation_data ADD COLUMN update_date TEXT")
                # 기존 데이터의 update_date를 현재 시간으로 설정
                current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                cursor.execute("UPDATE translation_data SET update_date = ? WHERE update_date IS NULL", (current_time,))
            
            # 기존 UNIQUE 제약조건 확인 및 수정 (STRING_ID만 유일하도록)
            cursor.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='translation_data'")
            table_schema = cursor.fetchone()
            if table_schema and 'UNIQUE(file_name, sheet_name, string_id)' in table_schema[0]:
                # 기존 테이블을 백업하고 새로운 구조로 재생성
                cursor.execute('''
                CREATE TABLE translation_data_new (
                    id INTEGER PRIMARY KEY,
                    file_name TEXT,
                    sheet_name TEXT,
                    string_id TEXT UNIQUE,
                    kr TEXT,
                    en TEXT, 
                    cn TEXT,
                    tw TEXT,
                    th TEXT,
                    status TEXT DEFAULT 'active',
                    update_date TEXT
                )
                ''')
                
                # 데이터 이전 (중복 STRING_ID는 최신 것만 유지)
                cursor.execute('''
                INSERT OR REPLACE INTO translation_data_new 
                SELECT * FROM translation_data
                ''')
                
                # 기존 테이블 삭제 후 이름 변경
                cursor.execute('DROP TABLE translation_data')
                cursor.execute('ALTER TABLE translation_data_new RENAME TO translation_data')
            
            # 인덱스 생성
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_string_id ON translation_data(string_id)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_file_sheet ON translation_data(file_name, sheet_name)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_status ON translation_data(status)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_update_date ON translation_data(update_date)')
            
        except Exception as e:
            print(f"스키마 업데이트 오류: {e}")

    def _load_existing_data_by_string_id(self, cursor):
        """기존 데이터 로드 (STRING_ID만을 키로 사용)"""
        cursor.execute("""
            SELECT file_name, sheet_name, string_id, kr, en, cn, tw, th, status, update_date
            FROM translation_data
        """)
        
        existing_data = {}
        for row in cursor.fetchall():
            string_id = row[2]  # string_id만을 키로 사용
            existing_data[string_id] = {
                'file_name': row[0],
                'sheet_name': row[1],
                'kr': row[3],
                'en': row[4],
                'cn': row[5],
                'tw': row[6],
                'th': row[7],
                'status': row[8],
                'update_date': row[9]
            }
        
        return existing_data

    def _process_batch_data(self, cursor, batch_data):
        """배치 데이터 처리"""
        for operation, data in batch_data:
            try:
                if operation == 'insert':
                    cursor.execute('''
                        INSERT OR REPLACE INTO translation_data 
                        (file_name, sheet_name, string_id, kr, en, cn, tw, th, status, update_date)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        data['file_name'], data['sheet_name'], data['string_id'],
                        data.get('kr'), data.get('en'), data.get('cn'), 
                        data.get('tw'), data.get('th'), data['status'], data['update_date']
                    ))
                    
                elif operation == 'update':
                    cursor.execute('''
                        UPDATE translation_data 
                        SET file_name = ?, sheet_name = ?, kr = ?, en = ?, cn = ?, tw = ?, th = ?, status = ?, update_date = ?
                        WHERE string_id = ?
                    ''', (
                        data['file_name'], data['sheet_name'],
                        data.get('kr'), data.get('en'), data.get('cn'), 
                        data.get('tw'), data.get('th'), 
                        data['status'], data['update_date'], data['string_id']
                    ))
                    
                elif operation == 'status':
                    cursor.execute('''
                        UPDATE translation_data 
                        SET status = ?, update_date = ?
                        WHERE string_id = ?
                    ''', (
                        data['status'], data['update_date'], data['string_id']
                    ))
                    
            except Exception as e:
                print(f"배치 처리 오류: {e}")
                continue
   
    def load_translation_cache(self, db_path):
        """번역 DB를 메모리에 캐싱 (활성 상태만)"""
        try:
            # DB 연결
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            # 활성 상태의 데이터만 로드
            cursor.execute("SELECT * FROM translation_data WHERE status = 'active'")
            rows = cursor.fetchall()
            
            # 3단계 캐싱 구조 구성
            translation_cache = {}              # STRING_ID만 (3순위)
            translation_file_cache = {}         # 파일명 + STRING_ID (1순위)
            translation_sheet_cache = {}        # 시트명 + STRING_ID (2순위)
            
            # 중복 STRING_ID 추적용
            duplicate_ids = {}
            
            for row in rows:
                file_name = row["file_name"]
                sheet_name = row["sheet_name"]
                string_id = row["string_id"]
                
                # 중복 STRING_ID 추적
                if string_id not in duplicate_ids:
                    duplicate_ids[string_id] = []
                duplicate_ids[string_id].append(file_name)
                
                # 데이터 딕셔너리 생성
                data = {
                    "kr": row["kr"],
                    "en": row["en"],
                    "cn": row["cn"],
                    "tw": row["tw"],
                    "th": row["th"],
                    "file_name": file_name,
                    "sheet_name": sheet_name
                }
                
                # 1. 파일명 + STRING_ID 캐싱 (1순위)
                norm_file_name = file_name.lower()
                if norm_file_name not in translation_file_cache:
                    translation_file_cache[norm_file_name] = {}
                
                if string_id not in translation_file_cache[norm_file_name]:
                    translation_file_cache[norm_file_name][string_id] = data
                
                # 2. 시트명 + STRING_ID 캐싱 (2순위)
                norm_sheet_name = sheet_name.lower()
                if norm_sheet_name not in translation_sheet_cache:
                    translation_sheet_cache[norm_sheet_name] = {}
                
                if string_id not in translation_sheet_cache[norm_sheet_name]:
                    translation_sheet_cache[norm_sheet_name][string_id] = data
                
                # 3. STRING_ID만 캐싱 (3순위)
                translation_cache[string_id] = data
            
            conn.close()
            
            # 결과 반환
            return {
                "translation_cache": translation_cache,
                "translation_file_cache": translation_file_cache,
                "translation_sheet_cache": translation_sheet_cache,
                "duplicate_ids": duplicate_ids,
                "file_count": len(translation_file_cache),
                "sheet_count": len(translation_sheet_cache),
                "id_count": len(translation_cache)
            }
            
        except Exception as e:
            return {
                "status": "error",
                "message": str(e)
            }

    def _process_single_excel_for_update(self, file_path, file_name, language_list, language_mapping, db_data_map, update_option, updates, inserts, debug_string_id=None):
        """단일 엑셀 파일을 순회하며 업데이트/삽입할 데이터를 수집합니다. (수정된 버전)"""
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for sheet_name in workbook.sheetnames:
            if not sheet_name.lower().startswith('string') or sheet_name.startswith("#"):
                continue

            worksheet = workbook[sheet_name]
            string_id_col, header_row = self.find_string_id_position(worksheet)
            if not string_id_col or not header_row: continue

            lang_cols = self.find_language_columns(worksheet, header_row, language_list, language_mapping)
            if not lang_cols: continue

            print(f"파일 {file_name}에서 발견된 언어 컬럼: {lang_cols}")

            for row_cells in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
                if not row_cells or string_id_col - 1 >= len(row_cells): continue

                string_id = row_cells[string_id_col - 1]
                if not string_id: continue

                # 디버깅 ID와 일치하는 경우, 상세 로그
                is_debug_target = (string_id == debug_string_id)
                if is_debug_target:
                    print("\n" + "="*60)
                    print(f"DEBUG: STRING_ID '{string_id}' 추적 시작")
                    print(f"파일: {file_name}, 시트: {sheet_name}")
                    print("="*60)

                # 엑셀 데이터 추출 (표준화된 키 사용)
                excel_data = {'string_id': string_id}
                for lang, col_idx in lang_cols.items():
                    if col_idx - 1 < len(row_cells):
                        # 모든 언어 키를 소문자로 통일
                        lang_key = lang.lower()
                        excel_data[lang_key] = row_cells[col_idx - 1]

                if is_debug_target:
                    print(f"엑셀에서 추출한 데이터: {excel_data}")

                status = '비활성' if str(row_cells[0] or '').strip().startswith('#') else 'active'

                # 키 생성
                key = None
                if update_option == "kr_compare":
                    key = excel_data.get('kr')
                elif update_option == "kr_additional_compare":
                    key = (excel_data.get('string_id'), excel_data.get('kr'))
                else:
                    key = excel_data.get('string_id')

                if is_debug_target:
                    print(f"업데이트 옵션: {update_option}")
                    print(f"생성된 키: {key}")
                    print(f"DB에 키가 존재하는가: {key in db_data_map if key else False}")

                # DB에 키가 있는지 확인
                if key and key in db_data_map:
                    db_row = db_data_map[key]
                    update_needed = False

                    if is_debug_target:
                        print(f"DB에서 찾은 기존 데이터: {db_row}")

                    # 상태 변경 체크
                    if db_row.get('status') != status:
                        update_needed = True
                        if is_debug_target:
                            print(f"상태 변경 감지: DB='{db_row.get('status')}', 엑셀='{status}' -> 업데이트 필요")

                    # 언어 값 변경 체크 (상태가 동일할 때만)
                    if not update_needed:
                        # 업데이트할 언어 목록 결정
                        update_langs = language_list.copy()
                        if update_option in ['default', 'kr_additional_compare'] and 'KR' in update_langs:
                            update_langs.remove('KR')

                        if is_debug_target:
                            print(f"비교할 언어 목록: {update_langs}")

                        for lang in update_langs:
                            lang_key = lang.lower()
                            excel_val = excel_data.get(lang_key)
                            db_val = db_row.get(lang_key)

                            # 공백 제거 후 비교
                            excel_val_clean = str(excel_val or '').strip()
                            db_val_clean = str(db_val or '').strip()

                            is_different = excel_val_clean != db_val_clean

                            if is_debug_target:
                                print(f"--- [{lang}] 값 비교 ---")
                                print(f"  엑셀 원본: {repr(excel_val)}")
                                print(f"  DB 원본:   {repr(db_val)}")
                                print(f"  엑셀 정제: '{excel_val_clean}'")
                                print(f"  DB 정제:   '{db_val_clean}'")
                                print(f"  다른가?:   {is_different}")

                            if is_different:
                                update_needed = True
                                if is_debug_target:
                                    print(f"[{lang}] 값 변경 감지 -> 업데이트 필요")
                                break

                    # 업데이트 필요한 경우 목록에 추가
                    if update_needed:
                        update_payload = {'key': key}
                        update_payload.update(excel_data)
                        update_payload.update({
                            'status': status, 'update_date': current_time,
                            'file_name': file_name, 'sheet_name': sheet_name
                        })
                        updates.append(update_payload)
                        if is_debug_target:
                            print("결론: '업데이트' 목록에 추가됨")
                    elif is_debug_target:
                        print("결론: 변경사항 없음 -> 업데이트 안 함")

                else:
                    # DB에 키가 없으면 신규 삽입
                    insert_payload = excel_data.copy()
                    insert_payload.update({
                        'file_name': file_name, 'sheet_name': sheet_name,
                        'status': status, 'update_date': current_time
                    })
                    inserts.append(insert_payload)
                    if is_debug_target:
                        print(f"결론: 키 '{key}'가 DB에 없음 -> '신규 삽입' 목록에 추가됨")

                if is_debug_target:
                    print("="*60 + "\n")

        workbook.close()

        """수집된 삽입 데이터를 배치 처리합니다."""
        if not inserts: return 0

        # 모든 컬럼명을 소문자로 통일
        cols = ['file_name', 'sheet_name', 'string_id', 'status', 'update_date'] + [lang.lower() for lang in language_list]
        placeholders = ", ".join([f":{col}" for col in cols])
        query = f"INSERT OR IGNORE INTO translation_data ({', '.join(cols)}) VALUES ({placeholders})"

        cursor.executemany(query, inserts)
        return cursor.rowcount