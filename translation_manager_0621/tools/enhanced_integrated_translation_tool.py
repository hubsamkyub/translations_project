# enhanced_integrated_translation_tool.py

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import threading
import time
import sys
import pandas as pd
from datetime import datetime

# --- 경로 문제 해결을 위한 코드 ---
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if project_root not in sys.path:
    sys.path.append(project_root)
# ---------------------------------

from ui.common_components import ScrollableCheckList, LoadingPopup
from tools.enhanced_integrated_translation_manager import EnhancedIntegratedTranslationManager

class EnhancedIntegratedTranslationTool(tk.Frame):

    def __init__(self, parent, excluded_files):
        super().__init__(parent)
        self.parent = parent
        self.manager = EnhancedIntegratedTranslationManager(self)
        self.excluded_files = excluded_files
        
        # 기존 UI 변수들
        self.excel_folder_var = tk.StringVar()
        self.individual_file_var = tk.StringVar()
        self.output_excel_var = tk.StringVar()
        self.output_db_var = tk.StringVar()
        
        # [신규] 마스터 파일 관련 변수
        self.master_folder_var = tk.StringVar()
        self.master_files = []
        
        # [수정] 언어 선택 변수 (KR, CN, TW만)
        self.available_languages = ["KR", "CN", "TW"]
        self.lang_vars = {}
        
        # [신규] 비교 기준 변수
        self.comparison_criteria_var = tk.StringVar(value="file_id")
        
        # 비교 옵션 변수
        self.include_new_var = tk.BooleanVar(value=True)
        self.include_deleted_var = tk.BooleanVar(value=True)
        self.include_modified_var = tk.BooleanVar(value=True)
        
        # [신규] 특수 컬럼 필터링 옵션 변수 (사용하지 않으므로 제거 가능)
        self.use_special_filter_var = tk.BooleanVar(value=False)
        self.special_column_var = tk.StringVar()
        self.special_condition_var = tk.StringVar()
        self.special_status_var = tk.StringVar(value="특수 컬럼 감지 전")
        
        # 출력 옵션 변수
        self.export_new_var = tk.BooleanVar(value=True)
        self.export_deleted_var = tk.BooleanVar(value=True)
        self.export_modified_var = tk.BooleanVar(value=True)
        self.export_duplicates_var = tk.BooleanVar(value=True)
        self.export_special_filtered_var = tk.BooleanVar(value=True)
        self.save_db_var = tk.BooleanVar(value=False)
        
        # [신규] 비교 결과 관련 변수
        self.comparison_results = None
        self.comparison_executed = False
        self.result_summary_var = tk.StringVar(value="비교를 실행하지 않았습니다.")
        
        # 내부 데이터
        self.excel_files = []
        self.current_results = None
        self.detected_special_columns = {}
        
        self.setup_ui()

    def setup_ui(self):
        """[재구성] 4단계 프로세스에 맞는 UI 구성"""
        
        # 메인 좌우 분할 프레임
        main_paned = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        main_paned.pack(fill="both", expand=True, padx=5, pady=5)
        
        # === 왼쪽 프레임 (1~2단계: 파일 선택 및 비교 설정) ===
        left_frame = ttk.Frame(main_paned)
        main_paned.add(left_frame, weight=3)
        
        # 왼쪽 스크롤 프레임
        left_canvas = tk.Canvas(left_frame)
        left_scrollbar = ttk.Scrollbar(left_frame, orient="vertical", command=left_canvas.yview)
        left_scrollable = ttk.Frame(left_canvas)
        
        left_scrollable.bind(
            "<Configure>",
            lambda e: left_canvas.configure(scrollregion=left_canvas.bbox("all"))
        )
        
        left_canvas.create_window((0, 0), window=left_scrollable, anchor="nw")
        left_canvas.configure(yscrollcommand=left_scrollbar.set)
        
        # --- 1단계: 파일 선택 영역 ---
        self.setup_file_selection_ui(left_scrollable)
        
        # --- 2단계: 비교 설정 영역 ---
        self.setup_comparison_settings_ui(left_scrollable)
        
        # 왼쪽 스크롤바 패킹
        left_canvas.pack(side="left", fill="both", expand=True)
        left_scrollbar.pack(side="right", fill="y")
        
        # === 오른쪽 프레임 (3~4단계: 비교 실행 및 결과 적용) ===
        right_frame = ttk.Frame(main_paned)
        main_paned.add(right_frame, weight=2)
        
        # --- 3단계: 비교 실행 및 결과 확인 ---
        self.setup_comparison_execution_ui(right_frame)
        
        # --- 4단계: 결과 적용 설정 ---
        self.setup_result_application_ui(right_frame)
        
        # --- 로그 및 상태 ---
        self.setup_log_and_status_ui(right_frame)

    def setup_file_selection_ui(self, parent):
        """1단계: 파일 선택 UI"""
        file_frame = ttk.LabelFrame(parent, text="1단계: 파일 선택")
        file_frame.pack(fill="x", padx=5, pady=5)
        
        # 번역 파일 선택 (기존과 동일)
        translation_section = ttk.LabelFrame(file_frame, text="번역 파일")
        translation_section.pack(fill="x", padx=5, pady=5)
        
        # 폴더 선택
        folder_frame = ttk.Frame(translation_section)
        folder_frame.pack(fill="x", padx=5, pady=3)
        ttk.Label(folder_frame, text="번역 폴더:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(folder_frame, textvariable=self.excel_folder_var, width=40).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(folder_frame, text="찾아보기", command=self.select_excel_folder).grid(row=0, column=2, padx=5, pady=5)
        ttk.Button(folder_frame, text="파일 검색", command=self.search_excel_files).grid(row=0, column=3, padx=5, pady=5)
        folder_frame.columnconfigure(1, weight=1)
        
        # 파일 목록
        files_list_frame = ttk.LabelFrame(translation_section, text="번역 파일 목록")
        files_list_frame.pack(fill="both", expand=True, padx=5, pady=5)
        self.excel_files_list = ScrollableCheckList(files_list_frame, width=500, height=80)
        self.excel_files_list.pack(fill="both", expand=True, padx=5, pady=5)
        
        # 마스터 파일 선택 (신규)
        master_section = ttk.LabelFrame(file_frame, text="비교 기준 파일 (마스터)")
        master_section.pack(fill="x", padx=5, pady=5)
        
        # 마스터 폴더 선택
        master_folder_frame = ttk.Frame(master_section)
        master_folder_frame.pack(fill="x", padx=5, pady=3)
        ttk.Label(master_folder_frame, text="마스터 폴더:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(master_folder_frame, textvariable=self.master_folder_var, width=40).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(master_folder_frame, text="찾아보기", command=self.select_master_folder).grid(row=0, column=2, padx=5, pady=5)
        ttk.Button(master_folder_frame, text="파일 검색", command=self.search_master_files).grid(row=0, column=3, padx=5, pady=5)
        master_folder_frame.columnconfigure(1, weight=1)
        
        # 마스터 파일 목록
        master_files_list_frame = ttk.LabelFrame(master_section, text="마스터 파일 목록")
        master_files_list_frame.pack(fill="both", expand=True, padx=5, pady=5)
        self.master_files_list = ScrollableCheckList(master_files_list_frame, width=500, height=80)
        self.master_files_list.pack(fill="both", expand=True, padx=5, pady=5)

    def setup_comparison_settings_ui(self, parent):
        """2단계: 비교 설정 UI"""
        compare_frame = ttk.LabelFrame(parent, text="2단계: 비교 설정")
        compare_frame.pack(fill="x", padx=5, pady=5)
        
        # 비교할 언어 선택
        lang_frame = ttk.LabelFrame(compare_frame, text="비교할 언어")
        lang_frame.pack(fill="x", padx=5, pady=5)
        lang_checkboxes_frame = ttk.Frame(lang_frame)
        lang_checkboxes_frame.pack(fill="x", padx=5, pady=5)
        
        for i, lang in enumerate(self.available_languages):
            var = tk.BooleanVar(value=True)
            self.lang_vars[lang] = var
            ttk.Checkbutton(lang_checkboxes_frame, text=lang, variable=var).grid(row=0, column=i, padx=10, pady=5, sticky="w")
        
        # 비교 기준 선택
        criteria_frame = ttk.LabelFrame(compare_frame, text="비교 기준 우선순위")
        criteria_frame.pack(fill="x", padx=5, pady=5)
        
        self.comparison_criteria_var = tk.StringVar(value="file_id")
        criteria_options = [
            ("file_id", "파일명 + STRING_ID"),
            ("sheet_id", "시트명 + STRING_ID"),
            ("id_only", "STRING_ID만"),
            ("id_kr", "STRING_ID + KR"),
            ("kr_only", "KR만"),
            ("id_cn", "STRING_ID + CN만"),
            ("id_tw", "STRING_ID + TW만")
        ]
        
        for i, (value, text) in enumerate(criteria_options):
            row = i // 2
            col = i % 2
            ttk.Radiobutton(criteria_frame, text=text, variable=self.comparison_criteria_var, 
                        value=value).grid(row=row, column=col, padx=10, pady=2, sticky="w")
        
        # 비교 옵션
        options_frame = ttk.LabelFrame(compare_frame, text="비교 옵션")
        options_frame.pack(fill="x", padx=5, pady=5)
        options_inner_frame = ttk.Frame(options_frame)
        options_inner_frame.pack(fill="x", padx=5, pady=5)
        
        ttk.Checkbutton(options_inner_frame, text="신규 항목 포함", variable=self.include_new_var).pack(side="left", padx=10)
        ttk.Checkbutton(options_inner_frame, text="삭제된 항목 포함", variable=self.include_deleted_var).pack(side="left", padx=10)
        ttk.Checkbutton(options_inner_frame, text="변경된 항목 포함", variable=self.include_modified_var).pack(side="left", padx=10)

    def setup_comparison_execution_ui(self, parent):
        """3단계: 비교 실행 및 결과 확인 UI"""
        execution_frame = ttk.LabelFrame(parent, text="3단계: 비교 실행 및 결과 확인")
        execution_frame.pack(fill="x", padx=5, pady=5)
        
        # 비교 실행 버튼
        execute_frame = ttk.Frame(execution_frame)
        execute_frame.pack(fill="x", padx=5, pady=5)
        
        self.compare_button = ttk.Button(execute_frame, text="🔍 비교 실행", command=self.execute_comparison)
        self.compare_button.pack(side="left", padx=5, pady=5)
        
        # 결과 미리보기 버튼들
        self.preview_all_button = ttk.Button(execute_frame, text="👁️ 전체 결과 보기", 
                                            command=self.preview_all_results, state="disabled")
        self.preview_all_button.pack(side="left", padx=5, pady=5)
        
        # 결과 요약 표시
        self.result_summary_var = tk.StringVar(value="비교를 실행하지 않았습니다.")
        result_summary_label = ttk.Label(execution_frame, textvariable=self.result_summary_var, 
                                    foreground="blue", font=("Arial", 9))
        result_summary_label.pack(side="left", padx=10, pady=5)

    def setup_result_application_ui(self, parent):
        """4단계: 결과 적용 설정 UI"""
        application_frame = ttk.LabelFrame(parent, text="4단계: 결과 적용")
        application_frame.pack(fill="x", padx=5, pady=5)
        
        # 결과 필터링
        filter_frame = ttk.LabelFrame(application_frame, text="결과 필터링")
        filter_frame.pack(fill="x", padx=5, pady=5)
        
        filter_buttons_frame = ttk.Frame(filter_frame)
        filter_buttons_frame.pack(fill="x", padx=5, pady=5)
        
        self.filter_new_button = ttk.Button(filter_buttons_frame, text="신규만 보기", 
                                        command=self.preview_new_only, state="disabled")
        self.filter_new_button.pack(side="left", padx=5, pady=2)
        
        self.filter_modified_button = ttk.Button(filter_buttons_frame, text="변경만 보기", 
                                                command=self.preview_modified_only, state="disabled")
        self.filter_modified_button.pack(side="left", padx=5, pady=2)
        
        self.filter_deleted_button = ttk.Button(filter_buttons_frame, text="삭제만 보기", 
                                            command=self.preview_deleted_only, state="disabled")
        self.filter_deleted_button.pack(side="left", padx=5, pady=2)
        
        # 출력 설정
        output_frame = ttk.LabelFrame(application_frame, text="출력 설정")
        output_frame.pack(fill="x", padx=5, pady=5)
        
        # 엑셀 출력 설정
        excel_output_frame = ttk.Frame(output_frame)
        excel_output_frame.pack(fill="x", padx=5, pady=3)
        ttk.Label(excel_output_frame, text="결과 엑셀:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(excel_output_frame, textvariable=self.output_excel_var, width=30).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(excel_output_frame, text="경로 선택", command=self.select_output_excel).grid(row=0, column=2, padx=5, pady=5)
        excel_output_frame.columnconfigure(1, weight=1)
        
        # 최종 적용 버튼
        final_execute_frame = ttk.Frame(application_frame)
        final_execute_frame.pack(fill="x", padx=5, pady=5)
        
        self.export_button = ttk.Button(final_execute_frame, text="📊 결과 엑셀로 내보내기", 
                                    command=self.export_comparison_results, state="disabled")
        self.export_button.pack(side="right", padx=5, pady=5)
   
    def toggle_special_filter(self):
        """[신규] 특수 컬럼 필터링 옵션 토글"""
        if self.use_special_filter_var.get():
            self.special_config_frame.pack(fill="x", padx=15, pady=5)
        else:
            self.special_config_frame.pack_forget()

    def verify_special_column(self):
        """[개선] 사용자가 입력한 특수 컬럼명이 파일에 존재하는지 빠르게 확인"""
        special_column_name = self.special_column_var.get().strip()
        if not special_column_name:
            messagebox.showwarning("경고", "확인할 특수 컬럼명을 입력하세요.", parent=self)
            return
        
        selected_files = self.excel_files_list.get_checked_items()
        if not selected_files:
            messagebox.showwarning("경고", "먼저 번역 파일을 선택하세요.", parent=self)
            return
        
        excel_files = [(name, path) for name, path in self.excel_files if name in selected_files]
        
        self.log_text.insert(tk.END, f"=== 특수 컬럼 '{special_column_name}' 확인 시작 ===\n")
        self.log_text.insert(tk.END, f"검색 대상: {len(excel_files)}개 파일\n")
        
        found_files = []
        total_found_count = 0
        
        try:
            for file_name, file_path in excel_files:
                self.log_text.insert(tk.END, f"  검색 중: {file_name}...\n")
                self.log_text.see(tk.END)
                self.update_idletasks()
                
                result = self.quick_check_special_column(file_path, special_column_name)
                if result["found"]:
                    found_files.append(file_name)
                    total_found_count += result["count"]
                    self.log_text.insert(tk.END, f"    ✅ 발견: {result['sheets']}에서 {result['count']}개\n")
                else:
                    self.log_text.insert(tk.END, f"    ❌ 없음\n")
            
            # 결과 정리
            if found_files:
                self.special_status_var.set(f"{len(found_files)}개 파일에서 발견됨 (총 {total_found_count}개)")
                self.log_text.insert(tk.END, f"✅ 확인 완료: {len(found_files)}개 파일에서 '{special_column_name}' 컬럼 발견\n")
                self.log_text.insert(tk.END, f"총 {total_found_count}개 항목에서 해당 컬럼 존재\n")
                
                messagebox.showinfo("확인 완료", 
                    f"특수 컬럼 '{special_column_name}'이 {len(found_files)}개 파일에서 발견되었습니다.\n\n"
                    f"총 {total_found_count}개 항목에서 해당 컬럼이 존재합니다.", parent=self)
            else:
                self.special_status_var.set("지정된 컬럼을 찾을 수 없음")
                self.log_text.insert(tk.END, f"❌ '{special_column_name}' 컬럼을 찾을 수 없습니다.\n")
                messagebox.showinfo("확인 결과", 
                    f"특수 컬럼 '{special_column_name}'을 찾을 수 없습니다.\n\n"
                    "컬럼명을 정확히 입력했는지 확인해주세요.", parent=self)
        
        except Exception as e:
            self.log_text.insert(tk.END, f"❌ 확인 중 오류: {str(e)}\n")
            messagebox.showerror("오류", f"특수 컬럼 확인 중 오류 발생: {str(e)}", parent=self)

    def quick_check_special_column(self, file_path, target_column_name):
        """빠른 특수 컬럼 존재 확인 (헤더 행만 검색)"""
        from openpyxl import load_workbook
        
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            found_sheets = []
            total_count = 0
            
            target_column_clean = target_column_name.strip().lower()
            
            for sheet_name in wb.sheetnames:
                if not sheet_name.lower().startswith("string") or sheet_name.startswith("#"):
                    continue
                
                worksheet = wb[sheet_name]
                
                # STRING_ID가 있는 헤더 행 찾기 (최대 5행까지만)
                header_row = None
                for row_idx in range(1, min(6, worksheet.max_row + 1)):
                    for col_idx in range(1, min(11, worksheet.max_column + 1)):  # 최대 10개 컬럼까지만
                        cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                        if cell_value and isinstance(cell_value, str):
                            if "string_id" in cell_value.strip().lower():
                                header_row = row_idx
                                break
                    if header_row:
                        break
                
                if not header_row:
                    continue
                
                # 해당 헤더 행에서 특수 컬럼 찾기
                special_column_found = False
                for cell in worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True):
                    for cell_value in cell:
                        if cell_value and isinstance(cell_value, str):
                            cell_clean = cell_value.strip().lower()
                            if cell_clean == target_column_clean:
                                special_column_found = True
                                break
                    if special_column_found:
                        break
                
                if special_column_found:
                    found_sheets.append(sheet_name)
                    # 해당 시트의 데이터 행 수 계산 (빠른 추정)
                    data_rows = max(0, worksheet.max_row - header_row)
                    total_count += data_rows
            
            wb.close()
            
            return {
                "found": len(found_sheets) > 0,
                "sheets": found_sheets,
                "count": total_count
            }
            
        except Exception as e:
            return {"found": False, "sheets": [], "count": 0, "error": str(e)}

    def preset_special_only(self):
        """[신규] 특수 필터링만 프리셋"""
        self.use_special_filter_var.set(True)
        self.include_new_var.set(False)
        self.include_deleted_var.set(False)
        self.include_modified_var.set(False)
        self.export_new_var.set(False)
        self.export_deleted_var.set(False)
        self.export_modified_var.set(False)
        self.export_duplicates_var.set(False)
        self.export_special_filtered_var.set(True)
        self.toggle_special_filter()
        self.log_text.insert(tk.END, "프리셋 적용: 특수 필터링만\n")

    def export_special_filtered_standalone(self):
        """[신규] 특수 컬럼 필터링 데이터만 별도 엑셀로 내보내기"""
        if not self.manager.special_column_data:
            messagebox.showerror("오류", "내보낼 특수 컬럼 필터링 데이터가 없습니다.", parent=self)
            return
        
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            title="특수 컬럼 필터링 데이터 엑셀 저장",
            parent=self
        )
        if not save_path:
            return
        
        try:
            data_list = list(self.manager.special_column_data.values())
            df = pd.DataFrame(data_list)
            df.to_excel(save_path, index=False)
            
            self.log_text.insert(tk.END, f"특수 컬럼 필터링 데이터 엑셀 저장 완료: {save_path}\n")
            messagebox.showinfo("성공", f"특수 컬럼 필터링 데이터가 성공적으로 저장되었습니다:\n{save_path}", parent=self)
            
        except Exception as e:
            messagebox.showerror("저장 오류", f"파일 저장 중 오류 발생:\n{e}", parent=self)

    # [기존 메서드들은 동일하게 유지하되, 필요한 부분만 수정]
    
    def select_excel_folder(self):
        """엑셀 폴더 선택"""
        folder = filedialog.askdirectory(title="번역 엑셀 폴더 선택", parent=self)
        if folder:
            self.excel_folder_var.set(folder)

    def search_excel_files(self):
        """폴더에서 엑셀 파일 검색"""
        folder = self.excel_folder_var.get()
        if not folder or not os.path.isdir(folder):
            messagebox.showwarning("경고", "유효한 폴더를 선택하세요.", parent=self)
            return
        
        self.excel_files_list.clear()
        self.excel_files = []
        
        for root, _, files in os.walk(folder):
            for file in files:
                if file.endswith(".xlsx") and not file.startswith("~$"):
                    if file not in self.excluded_files:
                        file_name_without_ext = os.path.splitext(file)[0].lower()
                        if file_name_without_ext.startswith("string"):
                            file_path = os.path.join(root, file)
                            self.excel_files.append((file, file_path))
                            self.excel_files_list.add_item(file, checked=True)
        
        if not self.excel_files:
            messagebox.showinfo("알림", "번역 엑셀 파일을 찾지 못했습니다.", parent=self)
        else:
            self.log_text.insert(tk.END, f"번역 파일 {len(self.excel_files)}개를 찾았습니다.\n")
            messagebox.showinfo("알림", f"번역 파일 {len(self.excel_files)}개를 찾았습니다.", parent=self)

    def add_excel_files(self):
        """개별 엑셀 파일 추가"""
        file_paths = filedialog.askopenfilenames(
            title="추가할 번역 엑셀 파일 선택",
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")],
            parent=self
        )
        if not file_paths:
            return
        
        added_count = 0
        for file_path in file_paths:
            file_name = os.path.basename(file_path)
            if not any(f[1] == file_path for f in self.excel_files):
                self.excel_files.append((file_name, file_path))
                self.excel_files_list.add_item(file_name, checked=True)
                added_count += 1
        
        if added_count > 0:
            self.log_text.insert(tk.END, f"번역 파일 {added_count}개가 목록에 추가되었습니다.\n")
            self.individual_file_var.set(f"{added_count}개 파일 추가됨")

    def select_master_db(self):
        """마스터 DB 파일 선택"""
        file_path = filedialog.askopenfilename(
            filetypes=[("DB 파일", "*.db"), ("모든 파일", "*.*")],
            title="마스터 DB 파일 선택",
            parent=self
        )
        if file_path:
            self.master_db_var.set(file_path)

    def select_output_excel(self):
        """출력 엑셀 파일 경로 선택"""
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            title="비교 결과 엑셀 파일 저장",
            parent=self
        )
        if file_path:
            self.output_excel_var.set(file_path)

    def select_output_db(self):
        """출력 DB 파일 경로 선택"""
        file_path = filedialog.asksaveasfilename(
            defaultextension=".db",
            filetypes=[("DB 파일", "*.db")],
            title="결과 DB 파일 저장",
            parent=self
        )
        if file_path:
            self.output_db_var.set(file_path)

    def preset_full_comparison(self):
        """전체 비교 프리셋"""
        self.include_new_var.set(True)
        self.include_deleted_var.set(True)
        self.include_modified_var.set(True)
        self.comparison_criteria_var.set("file_id")
        for var in self.lang_vars.values():
            var.set(True)
        self.log_text.insert(tk.END, "프리셋 적용: 전체 비교 모드\n")

    def preset_new_only(self):
        """신규만 프리셋"""
        self.include_new_var.set(True)
        self.include_deleted_var.set(False)
        self.include_modified_var.set(False)
        self.comparison_criteria_var.set("file_id")
        self.log_text.insert(tk.END, "프리셋 적용: 신규 항목만\n")

    def preset_modified_only(self):
        """변경된 항목만 프리셋"""
        self.include_new_var.set(False)
        self.include_deleted_var.set(False)
        self.include_modified_var.set(True)
        self.comparison_criteria_var.set("file_id")
        self.log_text.insert(tk.END, "프리셋 적용: 변경된 항목만\n")
        
    def reset_settings(self):
        """설정 초기화"""
        # 언어 설정 초기화
        for var in self.lang_vars.values():
            var.set(True)
        
        # 비교 옵션 초기화
        self.include_new_var.set(True)
        self.include_deleted_var.set(True)
        self.include_modified_var.set(True)
        
        # 비교 기준 초기화
        self.comparison_criteria_var.set("file_id")
        
        # 결과 초기화
        self.comparison_results = None
        self.comparison_executed = False
        self.result_summary_var.set("비교를 실행하지 않았습니다.")
        
        # 버튼 상태 초기화
        self.preview_all_button.config(state="disabled")
        self.filter_new_button.config(state="disabled")
        self.filter_modified_button.config(state="disabled")
        self.filter_deleted_button.config(state="disabled")
        self.export_button.config(state="disabled")
        
        self.log_text.insert(tk.END, "설정이 초기화되었습니다.\n")


    def preview_duplicates(self):
        """중복 데이터 미리보기 (기존 유지)"""
        if not self.current_results or not self.manager.duplicate_data:
            messagebox.showinfo("정보", "표시할 중복 데이터가 없습니다.", parent=self)
            return
        
        # 중복 데이터 미리보기 창 생성 (기존 코드와 동일)
        popup = tk.Toplevel(self)
        popup.title("중복 STRING_ID 미리보기")
        popup.geometry("1200x700")
        popup.transient(self)
        popup.grab_set()
        
        # 트리뷰 생성
        tree_frame = ttk.Frame(popup, padding=10)
        tree_frame.pack(fill="both", expand=True)
        
        columns = ("string_id", "kr", "cn", "tw", "file_name", "sheet_name", "status")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings")
        
        for col in columns:
            tree.heading(col, text=col.upper())
            tree.column(col, width=150 if col != "kr" else 200)
        
        # 스크롤바
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        vsb.pack(side="right", fill="y")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
        hsb.pack(side="bottom", fill="x")
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.pack(fill="both", expand=True)
        
        # 데이터 추가
        tree.tag_configure('group', background='#E8E8E8')
        
        for string_id, items in self.manager.duplicate_data.items():
            parent_id = tree.insert("", "end", text=string_id, 
                                  values=(f"📋 {string_id} ({len(items)}개)",), 
                                  open=True, tags=('group',))
            for item in items:
                values = (
                    item.get('string_id', ''),
                    item.get('kr', ''),
                    item.get('cn', ''),
                    item.get('tw', ''),
                    item.get('file_name', ''),
                    item.get('sheet_name', ''),
                    item.get('status', '')
                )
                tree.insert(parent_id, "end", values=values)
        
        # 버튼 프레임
        button_frame = ttk.Frame(popup, padding=10)
        button_frame.pack(fill="x")
        
        ttk.Button(button_frame, text="Excel로 내보내기", 
                  command=lambda: self.export_duplicates_standalone()).pack(side="left")
        ttk.Button(button_frame, text="닫기", 
                  command=popup.destroy).pack(side="right")

    def export_duplicates_standalone(self):
        """중복 데이터만 별도 엑셀로 내보내기 (기존 유지)"""
        if not self.manager.duplicate_data:
            messagebox.showerror("오류", "내보낼 중복 데이터가 없습니다.", parent=self)
            return
        
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            title="중복 데이터 엑셀 저장",
            parent=self
        )
        if not save_path:
            return
        
        try:
            flat_list = []
            for string_id, items in self.manager.duplicate_data.items():
                for item in items:
                    flat_list.append(item)
            
            df = pd.DataFrame(flat_list)
            df.to_excel(save_path, index=False)
            
            self.log_text.insert(tk.END, f"중복 데이터 엑셀 저장 완료: {save_path}\n")
            messagebox.showinfo("성공", f"중복 데이터가 성공적으로 저장되었습니다:\n{save_path}", parent=self)
            
        except Exception as e:
            messagebox.showerror("저장 오류", f"파일 저장 중 오류 발생:\n{e}", parent=self)

    def select_master_excel(self):
        """마스터 비교용 엑셀 파일 선택"""
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")],
            title="마스터 비교용 엑셀 파일 선택",
            parent=self
        )
        if file_path:
            self.master_excel_var.set(file_path)

# enhanced_integrated_translation_tool.py에 추가할 새로운 메서드들

    def select_master_folder(self):
        """마스터 폴더 선택"""
        folder = filedialog.askdirectory(title="마스터 파일 폴더 선택", parent=self)
        if folder:
            self.master_folder_var.set(folder)

    def search_master_files(self):
        """마스터 폴더에서 엑셀 파일 검색"""
        folder = self.master_folder_var.get()
        if not folder or not os.path.isdir(folder):
            messagebox.showwarning("경고", "유효한 마스터 폴더를 선택하세요.", parent=self)
            return
        
        self.master_files_list.clear()
        self.master_files = []
        
        for root, _, files in os.walk(folder):
            for file in files:
                if file.endswith(".xlsx") and not file.startswith("~$"):
                    if file not in self.excluded_files:
                        file_name_without_ext = os.path.splitext(file)[0].lower()
                        if file_name_without_ext.startswith("string"):
                            file_path = os.path.join(root, file)
                            self.master_files.append((file, file_path))
                            self.master_files_list.add_item(file, checked=True)
        
        if not self.master_files:
            messagebox.showinfo("알림", "마스터 엑셀 파일을 찾지 못했습니다.", parent=self)
        else:
            self.log_text.insert(tk.END, f"마스터 파일 {len(self.master_files)}개를 찾았습니다.\n")
            messagebox.showinfo("알림", f"마스터 파일 {len(self.master_files)}개를 찾았습니다.", parent=self)

    def execute_comparison(self):
        """3단계: 비교 실행"""
        # 입력 검증
        selected_translation_files = self.excel_files_list.get_checked_items()
        selected_master_files = self.master_files_list.get_checked_items()
        
        if not selected_translation_files:
            messagebox.showwarning("경고", "번역 파일을 선택하세요.", parent=self)
            return
        
        if not selected_master_files:
            messagebox.showwarning("경고", "마스터 파일을 선택하세요.", parent=self)
            return
        
        selected_langs = [lang for lang, var in self.lang_vars.items() if var.get()]
        if not selected_langs:
            messagebox.showwarning("경고", "비교할 언어를 하나 이상 선택하세요.", parent=self)
            return
        
        # 비교 옵션 설정
        comparison_options = {
            "include_new": self.include_new_var.get(),
            "include_deleted": self.include_deleted_var.get(),
            "include_modified": self.include_modified_var.get(),
            "comparison_criteria": self.comparison_criteria_var.get(),
            "languages": selected_langs
        }
        
        # 파일 리스트 준비
        translation_files = [(name, path) for name, path in self.excel_files if name in selected_translation_files]
        master_files = [(name, path) for name, path in self.master_files if name in selected_master_files]
        
        self.log_text.delete(1.0, tk.END)
        self.log_text.insert(tk.END, "=== 비교 실행 시작 ===\n")
        self.log_text.insert(tk.END, f"번역 파일: {len(translation_files)}개\n")
        self.log_text.insert(tk.END, f"마스터 파일: {len(master_files)}개\n")
        self.log_text.insert(tk.END, f"비교 언어: {', '.join(selected_langs)}\n")
        self.log_text.insert(tk.END, f"비교 기준: {self.get_criteria_description()}\n\n")
        
        # UI 비활성화
        self.compare_button.config(state="disabled")
        
        # 로딩 팝업
        loading_popup = LoadingPopup(self, "비교 진행 중", "파일 비교를 진행합니다...")
        
        def progress_callback(message, current, total):
            self.after(0, lambda: [
                loading_popup.update_progress((current / total) * 100, f"{current}/{total} - {message}"),
                self.log_text.insert(tk.END, f"{message}\n"),
                self.log_text.see(tk.END)
            ])
        
        def comparison_thread():
            try:
                # 매니저를 통한 비교 실행
                results = self.manager.execute_file_comparison(
                    translation_files, master_files, comparison_options, progress_callback
                )
                
                self.after(0, lambda: self.process_comparison_result(results, loading_popup))
                
            except Exception as e:
                self.after(0, lambda: [
                    loading_popup.close(),
                    self.log_text.insert(tk.END, f"\n오류 발생: {str(e)}\n"),
                    self.compare_button.config(state="normal"),
                    messagebox.showerror("오류", f"비교 중 오류 발생: {str(e)}", parent=self)
                ])
        
        threading.Thread(target=comparison_thread, daemon=True).start()

    def process_comparison_result(self, results, loading_popup):
        """비교 결과 처리"""
        loading_popup.close()
        
        if results["status"] == "error":
            self.log_text.insert(tk.END, f"\n오류 발생: {results['message']}\n")
            self.compare_button.config(state="normal")
            messagebox.showerror("오류", f"비교 중 오류 발생: {results['message']}", parent=self)
            return
        
        # 결과 저장
        self.comparison_results = results
        self.comparison_executed = True
        
        # 결과 요약
        summary = results.get("summary", {})
        new_count = summary.get("new_items", 0)
        modified_count = summary.get("modified_items", 0)
        deleted_count = summary.get("deleted_items", 0)
        unchanged_count = summary.get("unchanged_items", 0)
        
        # 요약 텍스트 업데이트
        summary_text = f"신규 {new_count}개, 변경 {modified_count}개, 삭제 {deleted_count}개, 동일 {unchanged_count}개"
        self.result_summary_var.set(summary_text)
        
        # 버튼 활성화
        self.preview_all_button.config(state="normal")
        self.filter_new_button.config(state="normal" if new_count > 0 else "disabled")
        self.filter_modified_button.config(state="normal" if modified_count > 0 else "disabled")
        self.filter_deleted_button.config(state="normal" if deleted_count > 0 else "disabled")
        self.export_button.config(state="normal")
        self.compare_button.config(state="normal")
        
        # 로그 출력
        self.log_text.insert(tk.END, f"\n=== 비교 완료 ===\n")
        self.log_text.insert(tk.END, f"📊 결과 요약:\n")
        self.log_text.insert(tk.END, f"• 신규 항목: {new_count}개\n")
        self.log_text.insert(tk.END, f"• 변경된 항목: {modified_count}개\n")
        self.log_text.insert(tk.END, f"• 삭제된 항목: {deleted_count}개\n")
        self.log_text.insert(tk.END, f"• 동일한 항목: {unchanged_count}개\n")
        
        messagebox.showinfo("비교 완료", f"파일 비교가 완료되었습니다!\n\n{summary_text}", parent=self)

    def get_criteria_description(self):
        """비교 기준 설명 반환"""
        criteria_map = {
            "file_id": "파일명 + STRING_ID",
            "sheet_id": "시트명 + STRING_ID", 
            "id_only": "STRING_ID만",
            "id_kr": "STRING_ID + KR",
            "kr_only": "KR만",
            "id_cn": "STRING_ID + CN만",
            "id_tw": "STRING_ID + TW만"
        }
        return criteria_map.get(self.comparison_criteria_var.get(), "알 수 없음")

    def preview_all_results(self):
        """전체 비교 결과 미리보기"""
        if not self.comparison_executed or not self.comparison_results:
            messagebox.showinfo("정보", "먼저 비교를 실행해주세요.", parent=self)
            return
        
        self.show_comparison_viewer("전체 결과", None)

    def preview_new_only(self):
        """신규 항목만 미리보기"""
        if not self.comparison_executed or not self.comparison_results:
            messagebox.showinfo("정보", "먼저 비교를 실행해주세요.", parent=self)
            return
        
        self.show_comparison_viewer("신규 항목", "new_in_target")

    def preview_modified_only(self):
        """변경된 항목만 미리보기"""
        if not self.comparison_executed or not self.comparison_results:
            messagebox.showinfo("정보", "먼저 비교를 실행해주세요.", parent=self)
            return
        
        self.show_comparison_viewer("변경된 항목", "modified")

    def preview_deleted_only(self):
        """삭제된 항목만 미리보기"""
        if not self.comparison_executed or not self.comparison_results:
            messagebox.showinfo("정보", "먼저 비교를 실행해주세요.", parent=self)
            return
        
        self.show_comparison_viewer("삭제된 항목", "new_in_master")

    def show_comparison_viewer(self, title, filter_type):
        """비교 결과 뷰어 창 표시"""
        viewer_win = tk.Toplevel(self)
        viewer_win.title(f"비교 결과 - {title}")
        viewer_win.geometry("1400x800")
        viewer_win.transient(self)
        viewer_win.grab_set()
        
        # 상단 검색 프레임
        search_frame = ttk.Frame(viewer_win, padding="5")
        search_frame.pack(fill="x")
        
        ttk.Label(search_frame, text="STRING_ID:").pack(side="left", padx=(0, 2))
        id_search_var = tk.StringVar()
        id_search_entry = ttk.Entry(search_frame, textvariable=id_search_var, width=25)
        id_search_entry.pack(side="left", padx=(0, 10))
        
        ttk.Label(search_frame, text="KR:").pack(side="left", padx=(0, 2))
        kr_search_var = tk.StringVar()
        kr_search_entry = ttk.Entry(search_frame, textvariable=kr_search_var, width=30)
        kr_search_entry.pack(side="left", padx=(0, 10))
        
        ttk.Label(search_frame, text="CN:").pack(side="left", padx=(0, 2))
        cn_search_var = tk.StringVar()
        cn_search_entry = ttk.Entry(search_frame, textvariable=cn_search_var, width=25)
        cn_search_entry.pack(side="left", padx=(0, 10))
        
        ttk.Label(search_frame, text="TW:").pack(side="left", padx=(0, 2))
        tw_search_var = tk.StringVar()
        tw_search_entry = ttk.Entry(search_frame, textvariable=tw_search_var, width=25)
        tw_search_entry.pack(side="left", padx=(0, 10))
        
        # 트리뷰 프레임
        tree_frame = ttk.Frame(viewer_win, padding="5")
        tree_frame.pack(fill="both", expand=True)
        
        columns = ("category", "string_id", "kr_master", "kr_target", "cn_master", "cn_target", 
                "tw_master", "tw_target", "file_name", "sheet_name")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings")
        
        # 컬럼 헤더 설정
        headers = {
            "category": "구분",
            "string_id": "STRING_ID", 
            "kr_master": "KR(마스터)",
            "kr_target": "KR(번역)",
            "cn_master": "CN(마스터)",
            "cn_target": "CN(번역)",
            "tw_master": "TW(마스터)",
            "tw_target": "TW(번역)",
            "file_name": "파일명",
            "sheet_name": "시트명"
        }
        
        for col, header in headers.items():
            tree.heading(col, text=header)
            if col == "category":
                tree.column(col, width=80)
            elif col == "string_id":
                tree.column(col, width=150)
            elif col in ["kr_master", "kr_target"]:
                tree.column(col, width=200)
            elif col in ["cn_master", "cn_target", "tw_master", "tw_target"]:
                tree.column(col, width=150)
            else:
                tree.column(col, width=120)
        
        # 스크롤바
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        vsb.pack(side="right", fill="y")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
        hsb.pack(side="bottom", fill="x")
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.pack(fill="both", expand=True)
        
        # 하단 상태 표시줄
        status_frame = ttk.Frame(viewer_win, padding="5")
        status_frame.pack(fill="x")
        status_label = ttk.Label(status_frame, text="데이터 로딩 중...")
        status_label.pack(side="left")
        
        # 데이터 준비
        all_data = self.prepare_viewer_data(filter_type)
        
        def populate_tree(data_to_show):
            """트리뷰 데이터 채우기"""
            tree.delete(*tree.get_children())
            
            for item in data_to_show:
                values = (
                    item.get('category', ''),
                    item.get('file_name', ''),
                    item.get('sheet_name', ''),
                    item.get('string_id', ''),
                    item.get('kr_master', ''),
                    item.get('kr_target', ''),
                    item.get('cn_master', ''),
                    item.get('cn_target', ''),
                    item.get('tw_master', ''),
                    item.get('tw_target', '')
                )
                tree.insert("", "end", values=values)
            
            status_label.config(text=f"{len(data_to_show):,} / {len(all_data):,}개 항목 표시 중")
        
        def perform_search():
            """검색 실행"""
            id_query = id_search_var.get().lower().strip()
            kr_query = kr_search_var.get().lower().strip()
            cn_query = cn_search_var.get().lower().strip()
            tw_query = tw_search_var.get().lower().strip()
            
            if not any([id_query, kr_query, cn_query, tw_query]):
                populate_tree(all_data)
                return
            
            filtered_data = []
            for item in all_data:
                matches = []
                
                if id_query:
                    matches.append(id_query in item.get('string_id', '').lower())
                if kr_query:
                    kr_match = (kr_query in item.get('kr_master', '').lower() or 
                            kr_query in item.get('kr_target', '').lower())
                    matches.append(kr_match)
                if cn_query:
                    cn_match = (cn_query in item.get('cn_master', '').lower() or 
                            cn_query in item.get('cn_target', '').lower())
                    matches.append(cn_match)
                if tw_query:
                    tw_match = (tw_query in item.get('tw_master', '').lower() or 
                            tw_query in item.get('tw_target', '').lower())
                    matches.append(tw_match)
                
                if all(matches):
                    filtered_data.append(item)
            
            populate_tree(filtered_data)
        
        def reset_search():
            """검색 초기화"""
            id_search_var.set("")
            kr_search_var.set("")
            cn_search_var.set("")
            tw_search_var.set("")
            populate_tree(all_data)
        
        # 검색 버튼
        search_button = ttk.Button(search_frame, text="검색", command=perform_search)
        search_button.pack(side="left", padx=5)
        reset_button = ttk.Button(search_frame, text="초기화", command=reset_search)
        reset_button.pack(side="left", padx=5)
        
        # 엔터키 바인딩
        for entry in [id_search_entry, kr_search_entry, cn_search_entry, tw_search_entry]:
            entry.bind("<Return>", lambda event: perform_search())
        
        # 하단 버튼
        button_frame = ttk.Frame(viewer_win, padding="5")
        button_frame.pack(fill="x")
        
        ttk.Button(button_frame, text="Excel로 내보내기", 
                command=lambda: self.export_viewer_data(all_data, title)).pack(side="left")
        ttk.Button(button_frame, text="닫기", command=viewer_win.destroy).pack(side="right")
        
        # 초기 데이터 로드
        populate_tree(all_data)

    def prepare_viewer_data(self, filter_type):
        """뷰어용 데이터 준비"""
        data_list = []
        
        if not self.comparison_results:
            return data_list
        
        comparison_results = self.comparison_results.get("comparison_results", {})
        
        # 필터 타입에 따라 데이터 선택
        if filter_type is None:
            # 전체 결과
            data_sources = [
                ("신규", comparison_results.get("new_in_target", [])),
                ("변경", comparison_results.get("modified", [])),
                ("삭제", comparison_results.get("new_in_master", [])),
                ("동일", comparison_results.get("unchanged", []))
            ]
        else:
            # 특정 타입만
            category_map = {
                "new_in_target": "신규",
                "modified": "변경", 
                "new_in_master": "삭제",
                "unchanged": "동일"
            }
            category = category_map.get(filter_type, "알 수 없음")
            data_sources = [(category, comparison_results.get(filter_type, []))]
        
        # 데이터 변환
        for category, items in data_sources:
            for item in items:
                processed_item = {
                    'category': category,
                    'file_name': item.get('file_name', ''),
                    'sheet_name': item.get('sheet_name', ''),
                    'string_id': item.get('string_id', ''),
                }
                
                # 마스터/타겟 데이터 처리
                if category == "변경":
                    # 변경된 항목은 master/target 구분 데이터가 있음
                    for lang in ['kr', 'cn', 'tw']:
                        processed_item[f'{lang}_master'] = item.get(f'{lang}_master', '')
                        processed_item[f'{lang}_target'] = item.get(f'{lang}_target', '')
                else:
                    # 다른 항목들은 단일 데이터
                    for lang in ['kr', 'cn', 'tw']:
                        value = item.get(lang, '')
                        if category == "삭제":
                            processed_item[f'{lang}_master'] = value
                            processed_item[f'{lang}_target'] = ''
                        else:
                            processed_item[f'{lang}_master'] = ''
                            processed_item[f'{lang}_target'] = value
                
                data_list.append(processed_item)
        
        return data_list

    def export_viewer_data(self, data, title):
        """뷰어 데이터 엑셀로 내보내기"""
        if not data:
            messagebox.showerror("오류", "내보낼 데이터가 없습니다.", parent=self)
            return
        
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            title=f"{title} 엑셀 저장",
            parent=self
        )
        if not save_path:
            return
        
        try:
            df = pd.DataFrame(data)
            df.to_excel(save_path, index=False)
            
            self.log_text.insert(tk.END, f"{title} 엑셀 저장 완료: {save_path}\n")
            messagebox.showinfo("성공", f"{title}이 성공적으로 저장되었습니다:\n{save_path}", parent=self)
            
        except Exception as e:
            messagebox.showerror("저장 오류", f"파일 저장 중 오류 발생:\n{e}", parent=self)

    def export_comparison_results(self):
        """최종 비교 결과 엑셀로 내보내기"""
        if not self.comparison_executed or not self.comparison_results:
            messagebox.showwarning("경고", "먼저 비교를 실행해주세요.", parent=self)
            return
        
        output_path = self.output_excel_var.get()
        if not output_path:
            messagebox.showwarning("경고", "출력 엑셀 파일 경로를 지정하세요.", parent=self)
            return
        
        try:
            # 매니저를 통한 엑셀 내보내기
            export_result = self.manager.export_comparison_results_to_excel(
                output_path, self.comparison_results
            )
            
            if export_result["status"] == "success":
                self.log_text.insert(tk.END, f"✅ 비교 결과 엑셀 저장 완료: {output_path}\n")
                messagebox.showinfo("완료", f"비교 결과가 성공적으로 저장되었습니다:\n{output_path}", parent=self)
            else:
                self.log_text.insert(tk.END, f"❌ 엑셀 저장 실패: {export_result['message']}\n")
                messagebox.showerror("오류", f"엑셀 저장 실패: {export_result['message']}", parent=self)
        
        except Exception as e:
            messagebox.showerror("오류", f"결과 내보내기 중 오류 발생: {str(e)}", parent=self)

    def setup_log_and_status_ui(self, parent):
        """로그 및 상태 UI 설정"""
        # 로그 프레임
        log_frame = ttk.LabelFrame(parent, text="📋 작업 로그")
        log_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        self.log_text = tk.Text(log_frame, wrap="word", height=15)
        log_scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_scrollbar.set)
        log_scrollbar.pack(side="right", fill="y")
        self.log_text.pack(fill="both", expand=True)
        
        # 상태 표시
        status_frame = ttk.Frame(parent)
        status_frame.pack(fill="x", padx=5, pady=5)
        self.status_label = ttk.Label(status_frame, text="대기 중...")
        self.status_label.pack(side="left", padx=5)