# tools/compare_request_extractor.py (Refactored UI Only)

import tkinter as tk
from tkinter import ttk, filedialog
import os
import threading
from openpyxl import load_workbook
from ui.common_components import show_message

class CompareRequestExtractor(tk.Frame):
    def __init__(self, parent, parent_app):
        super().__init__(parent)
        self.parent_app = parent_app

        self.compare_source_type = tk.StringVar(value="Excel")
        self.compare_excel_path_var = tk.StringVar()
        self.compare_excel_sheet_var = tk.StringVar()
        self.compare_db_path_var = tk.StringVar()
        self.compare_extract_new_var = tk.BooleanVar(value=True)
        self.compare_extract_modified_var = tk.BooleanVar(value=True)
        self.apply_by_request_col_var = tk.BooleanVar(value=True)
        self.save_to_db_var = tk.BooleanVar(value=True)
        self.full_compare_results = [] 

        self.setup_ui()
        self._toggle_compare_source()
        self._toggle_db_path_entry()

    def setup_ui(self):
        compare_source_frame = ttk.LabelFrame(self, text="2. 비교 대상 선택")
        compare_source_frame.pack(fill="x", padx=5, pady=5)
        
        tk.Radiobutton(compare_source_frame, text="Excel 파일", variable=self.compare_source_type, value="Excel", command=self._toggle_compare_source).pack(anchor="w")
        self.excel_compare_frame = ttk.Frame(compare_source_frame)
        self.excel_compare_frame.pack(fill="x", padx=20)
        ttk.Entry(self.excel_compare_frame, textvariable=self.compare_excel_path_var, width=50).pack(side="left", expand=True, fill="x")
        self.compare_sheet_combo = ttk.Combobox(self.excel_compare_frame, textvariable=self.compare_excel_sheet_var, state="readonly", width=20)
        self.compare_sheet_combo.pack(side="left", padx=5)
        ttk.Button(self.excel_compare_frame, text="파일 찾기", command=self.select_compare_excel).pack(side="left")

        tk.Radiobutton(compare_source_frame, text="DB 파일", variable=self.compare_source_type, value="DB", command=self._toggle_compare_source).pack(anchor="w", pady=(5,0))
        self.db_compare_frame = ttk.Frame(compare_source_frame)
        self.db_compare_frame.pack(fill="x", padx=20)
        ttk.Entry(self.db_compare_frame, textvariable=self.compare_db_path_var, width=50).pack(side="left", expand=True, fill="x")
        ttk.Button(self.db_compare_frame, text="파일 찾기", command=self.select_compare_db).pack(side="left")

        compare_options_frame = ttk.LabelFrame(self, text="3. 비교 추출 조건")
        compare_options_frame.pack(fill="x", padx=5, pady=5)
        
        basic_compare_frame = ttk.Frame(compare_options_frame)
        basic_compare_frame.pack(fill="x", anchor="w", pady=(0, 5))
        ttk.Checkbutton(basic_compare_frame, text="신규 항목 추출", variable=self.compare_extract_new_var).pack(side="left", padx=5)
        ttk.Checkbutton(basic_compare_frame, text="변경된 항목 추출", variable=self.compare_extract_modified_var).pack(side="left", padx=5)
        
        request_filter_frame = ttk.Frame(compare_options_frame)
        request_filter_frame.pack(fill="x", anchor="w")
        ttk.Checkbutton(request_filter_frame, text="#번역요청 컬럼이 '신규' 또는 'change'인 행만 비교", variable=self.apply_by_request_col_var).pack(side="left", padx=5)

        output_frame = ttk.LabelFrame(self, text="4. 출력 설정")
        output_frame.pack(fill="x", padx=5, pady=5)

        db_option_frame = ttk.Frame(output_frame)
        db_option_frame.pack(fill="x", padx=5, pady=5)
        ttk.Checkbutton(db_option_frame, text="추출 결과를 DB 파일로 저장", variable=self.save_to_db_var, command=self._toggle_db_path_entry).pack(anchor="w")

        self.db_path_frame = ttk.Frame(output_frame)
        self.db_path_frame.pack(fill="x", padx=5, pady=5)
        ttk.Label(self.db_path_frame, text="DB 파일:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.db_entry = ttk.Entry(self.db_path_frame, textvariable=self.parent_app.output_db_var, width=50)
        self.db_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        self.db_button = ttk.Button(self.db_path_frame, text="저장 위치", command=self.parent_app.select_output_db)
        self.db_button.grid(row=0, column=2, padx=5, pady=5)
        self.db_path_frame.columnconfigure(1, weight=1)

        action_frame = ttk.Frame(self)
        action_frame.pack(fill="x", padx=5, pady=5)
        ttk.Button(action_frame, text="추출 결과 Excel로 내보내기", command=self.export_results).pack(side="right", padx=(0, 5))
        ttk.Button(action_frame, text="비교하여 추출 실행", command=self.start_compare_extraction).pack(side="right", padx=5)

        result_frame = ttk.LabelFrame(self, text="추출 결과 (DB 저장 비활성 시)")
        result_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        columns = ("file_name", "sheet_name", "string_id", "kr", "request_type")
        self.result_tree = ttk.Treeview(result_frame, columns=columns, show="headings", height=8)
        self.result_tree.heading("file_name", text="파일명")
        self.result_tree.heading("sheet_name", text="시트명")
        self.result_tree.heading("string_id", text="STRING_ID")
        self.result_tree.heading("kr", text="KR")
        self.result_tree.heading("request_type", text="요청 타입")
        for col in columns: self.result_tree.column(col, width=150)
        vsb = ttk.Scrollbar(result_frame, orient="vertical", command=self.result_tree.yview)
        hsb = ttk.Scrollbar(result_frame, orient="horizontal", command=self.result_tree.xview)
        self.result_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.result_tree.pack(fill="both", expand=True)

    def _toggle_db_path_entry(self):
        state = "normal" if self.save_to_db_var.get() else "disabled"
        self.db_entry.config(state=state)
        self.db_button.config(state=state)
                
    def _toggle_compare_source(self):
        is_excel = self.compare_source_type.get() == "Excel"
        for child in self.excel_compare_frame.winfo_children(): child.config(state="normal" if is_excel else "disabled")
        for child in self.db_compare_frame.winfo_children(): child.config(state="disabled" if is_excel else "normal")

    def select_compare_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls"), ("All files", "*.*")])
        if path:
            self.compare_excel_path_var.set(path)
            try:
                wb = load_workbook(path, read_only=True)
                self.compare_sheet_combo['values'] = wb.sheetnames
                if wb.sheetnames: self.compare_excel_sheet_var.set(wb.sheetnames[0])
                wb.close()
            except Exception as e:
                show_message(self, "error", "파일 오류", f"시트 목록을 불러올 수 없습니다: {e}")

    def select_compare_db(self):
        path = filedialog.askopenfilename(filetypes=[("DB files", "*.db"), ("All files", "*.*")])
        if path: self.compare_db_path_var.set(path)


    def start_compare_extraction(self):
        if self.parent_app._is_task_running(): return
        selected_files = self.parent_app._get_selected_files()
        if not selected_files: return

        save_to_db = self.save_to_db_var.get()
        db_path = self.parent_app.output_db_var.get()

        if save_to_db and not db_path:
            show_message(self, "warning", "경고", "저장할 DB 파일 경로를 지정하세요.")
            return
        if save_to_db and os.path.exists(db_path):
            if not show_message(self, "yesno", "확인", f"'{os.path.basename(db_path)}' 파일이 이미 존재합니다. 덮어쓰시겠습니까?"):
                return

        if not self.compare_extract_new_var.get() and not self.compare_extract_modified_var.get():
            show_message(self, "warning", "경고", "추출 조건(신규 또는 변경)을 선택하세요.")
            return

        compare_options = {
            'source_type': self.compare_source_type.get(),
            'excel_path': self.compare_excel_path_var.get(),
            'sheet_name': self.compare_excel_sheet_var.get(),
            'db_path': self.compare_db_path_var.get(),
            'extract_new': self.compare_extract_new_var.get(),
            'extract_modified': self.compare_extract_modified_var.get(),
            'apply_by_request_col': self.apply_by_request_col_var.get(),
        }
        self.parent_app.log_text.delete(1.0, tk.END)
        self.parent_app.log_message("비교 추출 작업을 시작합니다...")
        self.result_tree.delete(*self.result_tree.get_children())

        self.parent_app.extraction_thread = threading.Thread(
            target=self.parent_app.extraction_manager.run_compare_extraction,
            args=(selected_files, db_path, compare_options, save_to_db, self._display_results_in_tree)
        )
        self.parent_app.extraction_thread.daemon = True
        self.parent_app.extraction_thread.start()

    def _display_results_in_tree(self, results):
        # [수정] 전체 결과를 클래스 변수에 저장
        self.full_compare_results = results

        self.result_tree.delete(*self.result_tree.get_children())
        for row in self.full_compare_results: # results -> self.full_compare_results
            # [수정] 변경된 데이터 구조에 맞게 주석과 인덱스 수정
            # (file_name, sheet_name, string_id, kr, cn, tw, request_type, additional_info)
            display_row = (row[0], row[1], row[2], row[3], row[6]) # 기존 row[8]에서 row[6]으로 변경
            self.result_tree.insert("", "end", values=display_row)
    
    # tools/compare_request_extractor.py의 export_results 함수를 교체해주세요.
    def export_results(self):
        """추출 결과를 엑셀로 내보내기 (로직 정리 및 버그 수정)"""
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            title="번역 요청 내보내기"
        )
        if not save_path:
            return

        # 'DB 저장' 옵션에 따라 데이터 소스를 결정
        if self.save_to_db_var.get():
            db_path = self.parent_app.output_db_var.get()
            if not db_path or not os.path.exists(db_path):
                show_message(self, "warning", "경고", "먼저 추출 작업을 실행하여 DB 파일을 생성해야 합니다.")
                return
            # DB에서 직접 엑셀로 내보내기
            self.parent_app.extraction_manager.export_to_excel(db_path, save_path)

        else:
            # 'DB 저장' 옵션을 껐을 경우, 메모리에 저장된 전체 결과를 사용
            if not self.full_compare_results:
                show_message(self, "warning", "경고", "먼저 추출 작업을 실행하여 결과를 확인해야 합니다.")
                return

            try:
                import pandas as pd
                # 전체 컬럼 이름을 사용하여 DataFrame 생성
                columns = ["file_name", "sheet_name", "string_id", "kr", "cn", "tw", "request_type", "additional_info"]
                df = pd.DataFrame(self.full_compare_results, columns=columns)
                df.to_excel(save_path, index=False)
                show_message(self, "info", "완료", f"데이터를 엑셀로 내보냈습니다.\n파일: {save_path}")
            except Exception as e:
                show_message(self, "error", "엑셀 저장 오류", f"파일 저장 중 오류가 발생했습니다: {e}")
            
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            title="번역 요청 내보내기",
            parent=self
        )
        if not save_path: return

        self.parent_app.extraction_manager.export_to_excel(db_path, save_path)