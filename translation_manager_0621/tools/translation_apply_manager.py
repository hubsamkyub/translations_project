import os
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
import sqlite3
import tkinter as tk
import signal
import win32com.client as pythoncom
import xlwings as xw

class TranslationApplyManager:
    def __init__(self, parent_window=None):
        self.parent_ui = parent_window
        self.translation_cache = {}
        self.translation_file_cache = {}
        self.translation_sheet_cache = {}
        self.duplicate_ids = {}
        self.kr_reverse_cache = {}
        
    def log_message(self, message):
        """UI의 로그 텍스트 영역에 메시지를 기록합니다."""
        if self.parent_ui and hasattr(self.parent_ui, 'log_text'):
            self.parent_ui.log_text.insert(tk.END, f"{message}\n")
            self.parent_ui.log_text.see(tk.END)
            self.parent_ui.update_idletasks()
        else:
            print(message)


    def load_translation_cache_from_excel(self, file_path, sheet_names):
        """[수정] openpyxl로 여러 시트에서 번역 데이터를 읽어 캐시를 생성합니다."""
        try:
            # 시트 이름 목록을 인자로 받음
            self.log_message(f"⚙️ [최적화] 엑셀 파일 로딩 시작: {os.path.basename(file_path)}")
            
            self.translation_cache = {}
            self.translation_file_cache = {}
            self.translation_sheet_cache = {}
            self.duplicate_ids = {}
            self.kr_reverse_cache = {}

            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            # [수정] 전달받은 모든 시트를 순회
            for sheet_name in sheet_names:
                if sheet_name not in wb.sheetnames:
                    self.log_message(f"⚠️ 경고: '{sheet_name}' 시트를 찾을 수 없어 건너뜁니다.")
                    continue
                
                self.log_message(f"  - 시트 처리 중: {sheet_name}")
                ws = wb[sheet_name]

                header_map = {}
                header_row_idx = -1
                for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
                    cleaned_row = [str(cell).lower().strip() for cell in row if cell is not None]
                    if 'string_id' in cleaned_row:
                        header_row_idx = i + 1
                        for col_idx, header_val in enumerate(row, 1):
                            if header_val:
                                header_map[str(header_val).lower().strip()] = col_idx - 1
                        break
                
                if 'string_id' not in header_map:
                    self.log_message(f"⚠️ 경고: '{sheet_name}' 시트에서 헤더를 찾을 수 없어 건너뜁니다.")
                    continue

                string_id_index = header_map.get('string_id')
                for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                    if not row or len(row) <= string_id_index or not row[string_id_index]:
                        continue
                    string_id = str(row[string_id_index]).strip()

                    def get_safe_value(key):
                        index = header_map.get(key)
                        if index is not None and index < len(row):
                            return str(row[index] or '')
                        return ''

                    data = {
                        "kr": get_safe_value("kr"), "en": get_safe_value("en"),
                        "cn": get_safe_value("cn"), "tw": get_safe_value("tw"),
                        "th": get_safe_value("th"),
                        "file_name": get_safe_value("filename") or get_safe_value("file_name"),
                        "sheet_name": get_safe_value("sheetname") or get_safe_value("sheet_name")
                    }
                    
                    if data["file_name"]: self.translation_file_cache.setdefault(data["file_name"].lower(), {})[string_id] = data
                    if data["sheet_name"]: self.translation_sheet_cache.setdefault(data["sheet_name"].lower(), {})[string_id] = data
                    self.translation_cache[string_id] = data

                    kr_text = data["kr"].strip()
                    if kr_text and kr_text not in self.kr_reverse_cache:
                        self.kr_reverse_cache[kr_text] = {**data, 'string_id': string_id}
            
            wb.close()
            self.log_message(f"🔧 캐시 구성 완료 (ID: {len(self.translation_cache)}, 파일: {len(self.translation_file_cache)}, 시트: {len(self.translation_sheet_cache)})")

            return {
                "status": "success", "source_type": "Excel", "id_count": len(self.translation_cache),
                "translation_cache": self.translation_cache, "translation_file_cache": self.translation_file_cache,
                "translation_sheet_cache": self.translation_sheet_cache, "duplicate_ids": {}, "kr_reverse_cache": self.kr_reverse_cache,
                "file_count": len(self.translation_file_cache), "sheet_count": len(self.translation_sheet_cache)
            }
        except Exception as e:
            self.log_message(f"❌ 엑셀 캐시 로딩 오류: {str(e)}")
            import traceback
            traceback.print_exc()
            return {"status": "error", "message": str(e)}
        
    # def load_translation_cache_from_excel(self, file_path, sheet_name):
    #     """[최적화 및 안정성 개선] openpyxl 스트리밍 방식으로 엑셀에서 번역 데이터를 읽어 캐시를 생성합니다."""
    #     try:
    #         self.log_message(f"⚙️ [최적화] 엑셀 파일 로딩 시작: {os.path.basename(file_path)} - 시트: {sheet_name}")
            
    #         # 캐시 초기화
    #         self.translation_cache = {}
    #         self.translation_file_cache = {}
    #         self.translation_sheet_cache = {}
    #         self.duplicate_ids = {}
    #         self.kr_reverse_cache = {}

    #         wb = load_workbook(file_path, read_only=True, data_only=True)
    #         if sheet_name not in wb.sheetnames:
    #             raise ValueError(f"'{sheet_name}' 시트를 찾을 수 없습니다.")
    #         ws = wb[sheet_name]

    #         # 1. 헤더 찾기 및 컬럼 인덱스 매핑
    #         header_map = {}
    #         header_row_idx = -1
    #         for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
    #             cleaned_row = [str(cell).lower().strip() for cell in row if cell is not None]
    #             if 'string_id' in cleaned_row:
    #                 header_row_idx = i + 1
    #                 for col_idx, header_val in enumerate(row, 1):
    #                     if header_val:
    #                         header_map[str(header_val).lower().strip()] = col_idx - 1
    #                 break
            
    #         if 'string_id' not in header_map:
    #             raise ValueError("엑셀 시트에서 'string_id' 컬럼을 포함한 헤더를 찾을 수 없습니다.")

    #         # 2. 데이터 순회 및 캐시 직접 생성
    #         string_id_index = header_map.get('string_id')
            
    #         for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
    #             if not row or len(row) <= string_id_index or not row[string_id_index]:
    #                 continue
                
    #             string_id = str(row[string_id_index]).strip()

    #             # [수정] 각 컬럼의 존재 여부를 확인하고 안전하게 값을 가져오는 헬퍼 함수
    #             def get_safe_value(key):
    #                 index = header_map.get(key)
    #                 if index is not None and index < len(row):
    #                     return str(row[index] or '') # 셀 값이 None일 경우 빈 문자열로 처리
    #                 return ''

    #             data = {
    #                 "kr": get_safe_value("kr"),
    #                 "en": get_safe_value("en"),
    #                 "cn": get_safe_value("cn"),
    #                 "tw": get_safe_value("tw"),
    #                 "th": get_safe_value("th"),
    #                 "file_name": get_safe_value("filename") or get_safe_value("file_name"),
    #                 "sheet_name": get_safe_value("sheetname") or get_safe_value("sheet_name")
    #             }
                
    #             # 다중 캐시 구성
    #             if data["file_name"]:
    #                 self.translation_file_cache.setdefault(data["file_name"].lower(), {})[string_id] = data
    #             if data["sheet_name"]:
    #                 self.translation_sheet_cache.setdefault(data["sheet_name"].lower(), {})[string_id] = data
    #             self.translation_cache[string_id] = data

    #             kr_text = data["kr"].strip()
    #             if kr_text and kr_text not in self.kr_reverse_cache:
    #                 self.kr_reverse_cache[kr_text] = {**data, 'string_id': string_id}
            
    #         wb.close()
    #         self.log_message(f"🔧 [최적화] 캐시 구성 완료 (ID: {len(self.translation_cache)}, 파일: {len(self.translation_file_cache)}, 시트: {len(self.translation_sheet_cache)})")

    #         return {
    #             "status": "success", "source_type": "Excel", "id_count": len(self.translation_cache),
    #             "translation_cache": self.translation_cache, "translation_file_cache": self.translation_file_cache,
    #             "translation_sheet_cache": self.translation_sheet_cache, "duplicate_ids": {}, "kr_reverse_cache": self.kr_reverse_cache,
    #             "file_count": len(self.translation_file_cache), "sheet_count": len(self.translation_sheet_cache)
    #         }
    #     except Exception as e:
    #         self.log_message(f"❌ [최적화] 엑셀 캐시 로딩 오류: {str(e)}")
    #         import traceback
    #         traceback.print_exc()
    #         return {"status": "error", "message": str(e)}

    # def _find_header_row(self, file_path, sheet_name):
    #     """엑셀 시트의 1~6행에서 'string_id'를 포함하는 헤더 행을 찾습니다."""
    #     for i in range(6):
    #         try:
    #             df_peek = pd.read_excel(file_path, sheet_name=sheet_name, header=i, nrows=0)
    #             if 'string_id' in [str(col).lower() for col in df_peek.columns]:
    #                 self.log_message(f"✅ 헤더 행 발견: {i + 1}번째 행")
    #                 return i
    #         except Exception:
    #             continue
    #     return None
    
    # def build_cache_from_dataframe(self, df):
    #     """Pandas DataFrame으로부터 정교한 다중 캐시를 구축합니다."""
    #     self.translation_cache = {}
    #     self.translation_file_cache = {}
    #     self.translation_sheet_cache = {}
    #     self.duplicate_ids = {}
    #     self.kr_reverse_cache = {}
        
    #     self.log_message(f"🔧 데이터프레임으로부터 캐시 구축 시작: {len(df)}개 행")

    #     for _, row in df.iterrows():
    #         string_id = str(row.get('string_id', '')).strip()
    #         if not string_id:
    #             continue

    #         file_name = str(row.get('filename', row.get('file_name', ''))).strip()
    #         sheet_name = str(row.get('sheetname', row.get('sheet_name', ''))).strip()

    #         norm_file_name = file_name.lower()
    #         norm_sheet_name = sheet_name.lower()

    #         data = {
    #             "kr": str(row.get("kr", "")),
    #             "en": str(row.get("en", "")),
    #             "cn": str(row.get("cn", "")),
    #             "tw": str(row.get("tw", "")),
    #             "th": str(row.get("th", "")),
    #             "file_name": file_name,
    #             "sheet_name": sheet_name
    #         }
            
    #         if norm_file_name:
    #             self.translation_file_cache.setdefault(norm_file_name, {})[string_id] = data
            
    #         if norm_sheet_name:
    #             self.translation_sheet_cache.setdefault(norm_sheet_name, {})[string_id] = data
            
    #         self.translation_cache[string_id] = data

    #         if string_id not in self.duplicate_ids:
    #             self.duplicate_ids[string_id] = []
    #         self.duplicate_ids[string_id].append(file_name)

    #         kr_text = data["kr"].strip()
    #         if kr_text and kr_text not in self.kr_reverse_cache:
    #             kr_cache_data = data.copy()
    #             kr_cache_data['string_id'] = string_id
    #             self.kr_reverse_cache[kr_text] = kr_cache_data
        
    #     self.log_message(f"🔧 캐시 구성 완료 (ID: {len(self.translation_cache)}, 파일: {len(self.translation_file_cache)}, 시트: {len(self.translation_sheet_cache)}, KR역방향: {len(self.kr_reverse_cache)})")
               
    def find_string_id_position(self, worksheet):
        """STRING_ID 위치 찾기"""
        for row in range(2, 6):  # 2행부터 5행까지 검색
            for col in range(1, min(10, worksheet.max_column + 1)):  # 최대 10개 컬럼까지만 검색
                cell_value = worksheet.cell(row=row, column=col).value
                if isinstance(cell_value, str) and "STRING_ID" in cell_value.upper():
                    return col, row
                    
        # 1행도 검색
        for row in worksheet.iter_rows(min_row=1, max_row=1, max_col=5):
            for cell in row:
                if isinstance(cell.value, str) and "STRING_ID" in cell.value.upper():
                    return cell.column, cell.row
                    
        return None, None

    def find_language_columns(self, worksheet, header_row, langs):
        """언어 컬럼 위치 찾기"""
        if not header_row:
            return {}
            
        lang_cols = {}
        
        # 지정한 헤더 행에서만 검색
        for row in worksheet.iter_rows(min_row=header_row, max_row=header_row):
            for cell in row:
                if not cell.value:
                    continue
                    
                header_text = str(cell.value).strip()
                
                # 직접 매칭
                if header_text in langs:
                    lang_cols[header_text] = cell.column
                    
        return lang_cols

    def find_target_columns(self, worksheet, header_row, target_columns=None):
        """지정된 대상 컬럼들의 위치를 찾습니다. (예: #번역요청, Change)"""
        if not header_row:
            return {}
            
        found_columns = {}
        # 기본적으로 '#번역요청' 컬럼을 탐색 대상에 포함
        all_targets = ["#번역요청"]
        if target_columns:
            all_targets.extend(target_columns)
        
        # 중복 제거
        all_targets = list(set(all_targets))

        for cell in worksheet[header_row]:
            if cell.value and isinstance(cell.value, str):
                cell_value_clean = cell.value.strip().lower()
                for target in all_targets:
                    if cell_value_clean == target.lower():
                        found_columns[target] = cell.column
                        break # 찾았으면 다음 셀로
                        
        return found_columns

    def _resave_with_excel_com(self, file_path):
        """Excel COM을 사용하여 파일을 다시 저장하여 최적화합니다."""
        excel = None
        workbook = None
        try:
            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            abs_path = os.path.abspath(file_path)
            workbook = excel.Workbooks.Open(abs_path)
            workbook.Save()
            self.log_message(f"  ✨ COM 객체로 파일 최적화 저장 완료: {os.path.basename(file_path)}")
            return True
        except Exception as e:
            self.log_message(f"  ⚠️ COM 객체 저장 실패: {e}")
            return False
        finally:
            if workbook:
                workbook.Close(SaveChanges=False)
            if excel:
                excel.Quit()
            pythoncom.CoUninitialize()

# tools/translation_apply_manager.py의 _resave_with_xlwings 함수를 아래 코드로 교체

    def _resave_with_xlwings(self, file_path):
        """[수정] xlwings를 사용하여 파일을 다시 저장하고, 프로세스를 확실하게 종료합니다."""
        app = None
        try:
            # visible=False로 백그라운드에서 실행
            app = xw.App(visible=False)
            # 생성된 Excel 프로세스의 ID를 가져옴
            pid = app.pid
            
            wb = app.books.open(file_path)
            wb.save()
            wb.close()
            
            self.log_message(f"  ✨ xlwings로 파일 최적화 저장 완료: {os.path.basename(file_path)}")
            return True
            
        except Exception as e:
            self.log_message(f"  ⚠️ xlwings 저장 실패: {e}")
            return False
            
        finally:
            # 앱 종료를 시도하고, 만약 프로세스가 남아있다면 강제 종료
            if app:
                try:
                    # 1. 정상 종료 시도
                    app.quit()
                    time.sleep(0.5) # 프로세스가 종료될 시간을 잠시 줍니다.
                    
                    # 2. 여전히 프로세스가 살아있는지 확인 후 강제 종료
                    os.kill(pid, 0) # 프로세스 존재 확인 (오류 발생 시 이미 종료된 것)
                    
                    # 여기까지 코드가 도달했다면 프로세스가 아직 살아있는 것이므로 강제 종료
                    self.log_message(f"  - Excel 프로세스(PID: {pid})가 종료되지 않아 강제 종료를 시도합니다.")
                    os.kill(pid, signal.SIGTERM)
                    self.log_message(f"  ✔️ Excel 프로세스를 강제로 종료했습니다.")

                except OSError:
                    # os.kill(pid, 0)에서 프로세스가 없다는 오류가 발생한 경우로, 정상 종료된 상태입니다.
                    self.log_message(f"  ✔️ Excel 프로세스(PID: {pid})가 정상적으로 종료되었습니다.")
                except Exception as e:
                    self.log_message(f"  ⚠️ Excel 프로세스 종료 중 예외 발생: {e}")

    def find_translation_request_column(self, worksheet, header_row):
        """#번역요청 컬럼 찾기 (공백, 대소문자 무시)"""
        if not header_row:
            return None
            
        for cell in worksheet[header_row]:
            if cell.value and isinstance(cell.value, str):
                # 공백 제거 및 소문자 변환 후 비교
                if cell.value.strip().lower() == "#번역요청":
                    return cell.column
                    
        return None


    def apply_translation(self, file_path, options):
        """
        [수정] ID 또는 KR 기반으로 번역을 적용하고, 상세한 로그를 제공합니다.
        """
        # --- 옵션 추출 ---
        mode = options.get("mode", "id")
        selected_langs = options.get("selected_langs", [])
        record_date = options.get("record_date", True)
        # ID 모드 옵션
        kr_match_check = options.get("kr_match_check", True)
        kr_mismatch_delete = options.get("kr_mismatch_delete", False)
        kr_overwrite = options.get("kr_overwrite", False)
        # KR 모드 옵션
        kr_overwrite_on_kr_mode = options.get("kr_overwrite_on_kr_mode", False)
        
        allowed_statuses = options.get("allowed_statuses", [])
        allowed_statuses_lower = [status.lower() for status in allowed_statuses] if allowed_statuses else []

        # --- 캐시 확인 ---
        if not self.translation_cache:
            return {"status": "error", "message": "번역 캐시가 로드되지 않았습니다."}
        if mode == 'kr' and not self.kr_reverse_cache:
            return {"status": "error", "message": "KR 기반 적용을 위한 역방향 캐시가 없습니다."}

        file_name = os.path.basename(file_path)
        
        # 옵션 요약 로그
        option_summary = []
        option_summary.append(f"{mode.upper()} 기반")
        if mode == 'id' and kr_match_check:
            option_summary.append("KR일치검사")
            if kr_mismatch_delete:
                option_summary.append("불일치시삭제")
            if kr_overwrite:
                option_summary.append("덮어쓰기")
        elif mode == 'kr' and kr_overwrite_on_kr_mode:
            option_summary.append("덮어쓰기")
        
        if allowed_statuses:
            option_summary.append(f"조건:{','.join(allowed_statuses)}")
        
        self.log_message(f"📁 {file_name} 처리시작 [{' | '.join(option_summary)}]")
        
        workbook = None
        try:
            current_file_name_lower = os.path.basename(file_path).lower()
            workbook = load_workbook(file_path)

            string_sheets = [sheet for sheet in workbook.sheetnames if sheet.lower().startswith("string") and not sheet.startswith("#")]
            
            if not string_sheets:
                self.log_message(f"   ⚠️ String 시트 없음")
                return {"status": "info", "message": "파일에 String 시트가 없습니다"}

            file_modified = False
            results = {
                "total_updated": 0, "total_overwritten": 0, "total_kr_mismatch_skipped": 0,
                "total_kr_mismatch_deleted": 0, "total_conditional_skipped": 0
            }
            
            # 시트별 상세 결과 저장
            sheet_details = {}
            
            fill_green = PatternFill(start_color="DAF2D0", end_color="DAF2D0", fill_type="solid")
            fill_orange = PatternFill(start_color="FFDDC1", end_color="FFDDC1", fill_type="solid") # '덮어씀' 표시용

            for sheet_name in string_sheets:
                worksheet = workbook[sheet_name]
                string_id_col, header_row = self.find_string_id_position(worksheet)
                if not string_id_col or not header_row:
                    self.log_message(f"   ⚠️ {sheet_name}: STRING_ID 컬럼 없음")
                    continue
                
                lang_cols = self.find_language_columns(worksheet, header_row, selected_langs + ['KR'])
                request_col_idx = self.find_target_columns(worksheet, header_row, ["#번역요청"]).get("#번역요청")
                
                # 시트별 카운터
                sheet_stats = {
                    "updated": 0, "overwritten": 0, "conditional_skipped": 0,
                    "kr_mismatch_skipped": 0, "kr_mismatch_deleted": 0,
                    "total_rows": 0, "processed_rows": 0
                }
                
                # 언어별 적용 카운터
                lang_apply_count = {lang: 0 for lang in selected_langs if lang != 'KR'}
                
                # 전체 행 수 계산
                sheet_stats["total_rows"] = worksheet.max_row - header_row
                
                for row_idx in range(header_row + 1, worksheet.max_row + 1):
                    sheet_stats["processed_rows"] += 1
                    
                    # 조건부 적용 로직
                    if allowed_statuses_lower and request_col_idx:
                        request_val = str(worksheet.cell(row=row_idx, column=request_col_idx).value or '').strip().lower()
                        if request_val not in allowed_statuses_lower:
                            sheet_stats["conditional_skipped"] += 1
                            continue
                    
                    # --- 데이터 조회 로직 (ID vs KR) ---
                    trans_data = None
                    key_value = ''
                    if mode == 'id':
                        key_value = str(worksheet.cell(row=row_idx, column=string_id_col).value or '').strip()
                        if key_value:
                            trans_data = self.translation_cache.get(key_value)
                    else: # mode == 'kr'
                        if 'KR' in lang_cols:
                            key_value = str(worksheet.cell(row=row_idx, column=lang_cols['KR']).value or '').strip()
                            if key_value:
                                trans_data = self.kr_reverse_cache.get(key_value)

                    if not key_value or not trans_data:
                        continue
                    
                    row_modified_this_iteration = False
                    
                    # --- 적용 로직 ---
                    if mode == 'id' and kr_match_check:
                        current_kr_val = str(worksheet.cell(row=row_idx, column=lang_cols['KR']).value or '').strip()
                        cache_kr_val = str(trans_data.get('kr', '')).strip()
                        if current_kr_val != cache_kr_val:
                            if kr_mismatch_delete:
                                deleted_count = 0
                                for lang, col_idx in lang_cols.items():
                                    if lang != 'KR' and worksheet.cell(row=row_idx, column=col_idx).value:
                                        worksheet.cell(row=row_idx, column=col_idx).value = ""
                                        deleted_count += 1
                                        row_modified_this_iteration = True
                                if deleted_count > 0:
                                    sheet_stats["kr_mismatch_deleted"] += 1
                            else:
                                sheet_stats["kr_mismatch_skipped"] += 1
                            continue # KR 불일치 시 건너뛰기
                    
                    # 번역 적용 또는 덮어쓰기 로직
                    for lang in selected_langs:
                        if lang == 'KR': continue
                        
                        lang_lower = lang.lower()
                        col_idx = lang_cols.get(lang)
                        if not col_idx: continue
                        
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        current_val = str(cell.value or '').strip()
                        cached_val = str(trans_data.get(lang_lower, '')).strip()

                        if cached_val and current_val != cached_val:
                            should_overwrite = False
                            if mode == 'id' and kr_match_check and kr_overwrite:
                                should_overwrite = True # ID 모드, KR 일치, 덮어쓰기 옵션 켬
                            elif mode == 'kr' and kr_overwrite_on_kr_mode:
                                should_overwrite = True # KR 모드, 덮어쓰기 옵션 켬
                            
                            if should_overwrite:
                                cell.value = cached_val
                                cell.fill = fill_orange # 주황색으로 "덮어씀" 표시
                                sheet_stats["overwritten"] += 1
                                lang_apply_count[lang] += 1
                                row_modified_this_iteration = True
                            elif not should_overwrite and not current_val: # 빈 칸에만 적용
                                cell.value = cached_val
                                cell.fill = fill_green
                                sheet_stats["updated"] += 1
                                lang_apply_count[lang] += 1
                                row_modified_this_iteration = True
                    
                    if row_modified_this_iteration:
                        file_modified = True
                        if record_date and request_col_idx:
                            worksheet.cell(row=row_idx, column=request_col_idx).value = "적용"

                # 시트 처리 결과 로그
                if sheet_stats["updated"] > 0 or sheet_stats["overwritten"] > 0:
                    lang_details = []
                    for lang, count in lang_apply_count.items():
                        if count > 0:
                            lang_details.append(f"{lang}:{count}")
                    
                    log_parts = []
                    if sheet_stats["updated"] > 0:
                        log_parts.append(f"신규:{sheet_stats['updated']}")
                    if sheet_stats["overwritten"] > 0:
                        log_parts.append(f"덮어씀:{sheet_stats['overwritten']}")
                    if lang_details:
                        log_parts.append(f"[{', '.join(lang_details)}]")
                    
                    self.log_message(f"   ✅ {sheet_name}: {' | '.join(log_parts)}")
                else:
                    skip_reasons = []
                    if sheet_stats["conditional_skipped"] > 0:
                        skip_reasons.append(f"조건불일치:{sheet_stats['conditional_skipped']}")
                    if sheet_stats["kr_mismatch_skipped"] > 0:
                        skip_reasons.append(f"KR불일치:{sheet_stats['kr_mismatch_skipped']}")
                    if sheet_stats["kr_mismatch_deleted"] > 0:
                        skip_reasons.append(f"KR불일치삭제:{sheet_stats['kr_mismatch_deleted']}")
                    
                    if skip_reasons:
                        self.log_message(f"   ⚠️ {sheet_name}: 적용없음 ({' | '.join(skip_reasons)})")
                    else:
                        self.log_message(f"   ⚠️ {sheet_name}: 적용없음 (번역데이터 없음)")
                
                # 전체 결과에 누적
                for key in results:
                    if key.startswith("total_"):
                        stat_key = key[6:]  # "total_" 제거
                        results[key] += sheet_stats.get(stat_key, 0)
                
                sheet_details[sheet_name] = sheet_stats
            
            if file_modified:
                self.log_message(f"   💾 변경사항 저장 중...")
                workbook.save(file_path)
                
                # 최종 파일 요약
                summary_parts = []
                if results["total_updated"] > 0:
                    summary_parts.append(f"신규 {results['total_updated']}개")
                if results["total_overwritten"] > 0:
                    summary_parts.append(f"덮어씀 {results['total_overwritten']}개")
                
                total_applied = results["total_updated"] + results["total_overwritten"]
                self.log_message(f"   ✅ {file_name} 완료: {' | '.join(summary_parts)} (총 {total_applied}개 적용)")
            else:
                skip_summary = []
                if results["total_conditional_skipped"] > 0:
                    skip_summary.append(f"조건 {results['total_conditional_skipped']}개")
                if results["total_kr_mismatch_skipped"] > 0:
                    skip_summary.append(f"KR불일치 {results['total_kr_mismatch_skipped']}개")
                
                if skip_summary:
                    self.log_message(f"   ⚠️ {file_name} 완료: 변경없음 ({' | '.join(skip_summary)} 건너뜀)")
                else:
                    self.log_message(f"   ⚠️ {file_name} 완료: 변경없음 (번역 데이터 없음)")
            
            return {"status": "success", **results}
            
        except Exception as e:
            self.log_message(f"   ❌ {file_name} 오류: {str(e)}")
            import traceback
            traceback.print_exc()
            return {"status": "error", "message": str(e), "error_type": "processing_error"}
        finally:
            if workbook:
                workbook.close()

               
    def check_external_links(self, workbook):
        """워크북에서 외부 링크 검사 (번역 도구용) - 검증된 최종 버전"""
        import re
        
        external_links = []
        
        # 외부 참조 패턴들 (검증된 버전)
        external_patterns = [
            r"'[^']*\.xl[sx]?[xm]?'!",  # '파일명.xlsx'! 또는 '경로\파일명.xlsx'!
            r'\[.*\.xl[sx]?[xm]?\]',    # [파일명.xlsx] 패턴
            r"'[A-Z]:[^']*\.xl[sx]?[xm]?'!", # 'C:\경로\파일명.xlsx'! 패턴  
            r'\\[^\\]*\.xl[sx]?[xm]?!', # \파일명.xlsx! 패턴
            r"=[^=]*'[A-Z]:[^']*'",     # =으로 시작하는 드라이브 경로
            r'\[\d+\]!',                # [숫자]! 패턴 (시트 참조)
        ]
        
        # #REF! 오류 패턴들 (검증된 버전)
        ref_error_patterns = [
            r'#REF!',                   # #REF! 오류
            r'OFFSET\(#REF!',          # OFFSET 함수에서 #REF! 오류
        ]
        
        try:
            # 방법 1: 워크북의 external_links 속성 확인
            if hasattr(workbook, 'external_links') and workbook.external_links:
                for link in workbook.external_links:
                    external_links.append(f"워크북_외부링크: {str(link)}")
            
            # 방법 2: 명명된 범위 검사 (가장 중요!) - 검증된 로직
            if hasattr(workbook, 'defined_names') and workbook.defined_names:
                # 딕셔너리 키로 접근 (검증된 방법)
                for name_key in workbook.defined_names.keys():
                    try:
                        defined_name = workbook.defined_names[name_key]
                        if hasattr(defined_name, 'value') and defined_name.value:
                            name_formula = str(defined_name.value)
                            
                            # #REF! 오류 우선 검사
                            ref_error_found = False
                            for ref_pattern in ref_error_patterns:
                                if re.search(ref_pattern, name_formula):
                                    external_links.append(f"명명된_범위_REF오류:{name_key} - {name_formula[:50]}")
                                    ref_error_found = True
                                    break
                            
                            # #REF! 오류가 없는 경우에만 외부 참조 패턴 검사
                            if not ref_error_found:
                                for pattern in external_patterns:
                                    if re.search(pattern, name_formula):
                                        external_links.append(f"명명된_범위_외부링크:{name_key} - {name_formula[:50]}")
                                        break
                    except Exception as e:
                        # 개별 명명된 범위 처리 중 오류가 발생해도 계속 진행
                        pass
            
            # 방법 3: 셀별 외부 참조 검사 (제한적으로)
            cell_count = 0
            for sheet_name in workbook.sheetnames:
                if cell_count >= 100:  # 번역 도구에서는 성능을 위해 더 제한적으로
                    break
                    
                worksheet = workbook[sheet_name]
                
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell_count += 1
                        if cell_count > 100:
                            break
                            
                        # 공식이 있는 셀 검사
                        if cell.data_type == 'f' and cell.value:
                            formula = str(cell.value)
                            
                            # #REF! 오류 검사
                            for ref_pattern in ref_error_patterns:
                                if re.search(ref_pattern, formula):
                                    external_links.append(f"셀_REF오류:{sheet_name}!{cell.coordinate} - {formula[:50]}")
                                    break
                            else:
                                # 외부 참조 패턴 검사
                                for pattern in external_patterns:
                                    if re.search(pattern, formula):
                                        external_links.append(f"셀_외부링크:{sheet_name}!{cell.coordinate} - {formula[:50]}")
                                        break
                        
                        # #REF! 값 검사
                        elif cell.value and str(cell.value).startswith('#REF!'):
                            external_links.append(f"셀_REF값:{sheet_name}!{cell.coordinate} - {cell.value}")
                    
                    if cell_count > 100:
                        break
                        
        except Exception as e:
            # 외부 링크 검사 중 오류가 발생하면 무시하고 계속 진행
            pass
            
        return external_links[:10]  # 최대 10개만 반환

# tools/translation_apply_manager.py의 load_translation_cache_from_db 함수를 아래 코드로 교체합니다.

    def load_translation_cache_from_db(self, db_path):
        """[수정] 데이터베이스에서 직접 번역 캐시를 생성합니다."""
        try:
            self.log_message(f"⚙️ DB 로딩 시작: {db_path}")

            # 캐시 초기화
            self.translation_cache = {}
            self.translation_file_cache = {}
            self.translation_sheet_cache = {}
            self.duplicate_ids = {}
            self.kr_reverse_cache = {}

            conn = sqlite3.connect(db_path)
            # 컬럼 이름으로 데이터에 접근하기 위해 row_factory 설정
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()

            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='translation_data'")
            if cursor.fetchone() is None:
                conn.close()
                message = "'translation_data' 테이블이 DB에 없습니다."
                self.log_message(f"❌ {message}")
                return {"status": "error", "message": message}

            # 'active' 상태인 데이터만 가져옵니다.
            cursor.execute("SELECT * FROM translation_data WHERE status = 'active'")
            rows = cursor.fetchall()
            conn.close()

            # [수정] DataFrame을 거치지 않고 직접 캐시를 생성합니다.
            for row in rows:
                string_id = row["string_id"]
                if not string_id:
                    continue
                
                # sqlite3.Row 객체를 딕셔너리로 변환
                data = dict(row)
                
                file_name_val = data.get("file_name", "")
                sheet_name_val = data.get("sheet_name", "")

                # 다중 캐시 구성
                if file_name_val:
                    self.translation_file_cache.setdefault(file_name_val.lower(), {})[string_id] = data
                if sheet_name_val:
                    self.translation_sheet_cache.setdefault(sheet_name_val.lower(), {})[string_id] = data
                self.translation_cache[string_id] = data

                # KR 역방향 조회 캐시 생성
                kr_text = data.get("kr", "")
                if kr_text:
                    kr_text = kr_text.strip()
                    if kr_text and kr_text not in self.kr_reverse_cache:
                        self.kr_reverse_cache[kr_text] = {**data}
            
            self.log_message(f"🔧 DB 캐시 구성 완료 (ID: {len(self.translation_cache)}, 파일: {len(self.translation_file_cache)}, 시트: {len(self.translation_sheet_cache)})")

            return {
                "status": "success",
                "source_type": "DB",
                "id_count": len(self.translation_cache),
                "file_count": len(self.translation_file_cache),
                "sheet_count": len(self.translation_sheet_cache),
                "translation_cache": self.translation_cache,
                "translation_file_cache": self.translation_file_cache,
                "translation_sheet_cache": self.translation_sheet_cache,
                "duplicate_ids": {},
                "kr_reverse_cache": self.kr_reverse_cache
            }
        except Exception as e:
            self.log_message(f"❌ 번역 DB 캐시 로딩 오류: {str(e)}")
            import traceback
            traceback.print_exc()
            return {"status": "error", "message": str(e)}