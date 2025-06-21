# tools/workflow_manager.py

import os
import threading
from tkinter import messagebox
from openpyxl import load_workbook
import sqlite3

# 기존 매니저들을 임포트하여 로직을 재사용합니다.
from tools.translation_db_manager import TranslationDBManager
from tools.translation_apply_manager import TranslationApplyManager
from ui.common_components import LoadingPopup, show_message

class WorkflowManager:
    def __init__(self, parent_app):
        self.parent_app = parent_app # UI 업데이트 및 로그 기록을 위함
        self.db_manager = TranslationDBManager(parent_app)
        self.apply_manager = TranslationApplyManager(parent_app)

    def log(self, message):
        """UI에 로그 메시지를 전달합니다."""
        self.parent_app.log(message)


    def run_step1_build_db(self, is_update, db_path, source_folder_path, files_to_process, progress_callback, completion_callback):
        """1단계: 번역 DB 구축 또는 업데이트를 실행합니다. (UI 제어 로직 분리)"""
        
        try:
            excel_files = [(file_name, full_path) for file_name, full_path in files_to_process]
            
            if is_update:
                self.log(f"기존 DB 업데이트 시작: {db_path}")
                result = self.db_manager.update_translation_db(
                    excel_files=excel_files,
                    db_path=db_path,
                    language_list=["KR", "EN", "CN", "TW", "TH"],
                    progress_callback=progress_callback
                )
            else:
                self.log(f"새로운 DB 구축 시작: {db_path}")
                result = self.db_manager.build_translation_db(
                    excel_files=excel_files,
                    output_db_path=db_path,
                    language_list=["KR", "EN", "CN", "TW", "TH"],
                    progress_callback=progress_callback
                )
            
            # 작업 완료 후 메인 스레드에서 콜백 함수 실행
            self.parent_app.root.after(0, completion_callback, result)

        except Exception as e:
            error_result = {"status": "error", "message": str(e)}
            self.parent_app.root.after(0, completion_callback, error_result)


    def run_step2_update_db_from_excel(self, base_db_path, excel_path, sheet_names_to_process, options, completion_callback):
        """2단계: 번역된 엑셀 파일 내용으로 기준 DB를 업데이트 (다중 시트 처리 기능 추가)"""

        def task():
            total_updated_rows = 0
            try:
                from collections import defaultdict

                loading_popup.update_message("번역 파일 읽는 중...")
                self.log(f"번역 파일 로드 시작: {os.path.basename(excel_path)}")
                
                wb = load_workbook(excel_path, read_only=True, data_only=True)
                
                # [수정] 모든 처리 대상 시트의 데이터를 미리 로드
                all_excel_cache = {}
                for sheet_name in sheet_names_to_process:
                    if sheet_name not in wb.sheetnames: continue
                    
                    ws = wb[sheet_name]
                    headers, header_row_idx = self.parent_app._find_headers_in_worksheet(ws)
                    if not headers or "STRING_ID" not in headers.values(): continue
                    
                    excel_cache = defaultdict(list)
                    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                        string_id_idx = list(headers.keys())[list(headers.values()).index("STRING_ID")]
                        string_id = row[string_id_idx - 1]
                        if string_id:
                            row_data = {headers[col_idx].lower(): row[col_idx - 1] for col_idx in headers}
                            excel_cache[str(string_id)].append(row_data)
                    
                    all_excel_cache[sheet_name] = excel_cache
                wb.close()
                self.log(f"번역 파일의 {len(all_excel_cache)}개 시트 로드 완료")

                # DB 업데이트
                loading_popup.update_message("DB 업데이트 중...")
                conn = sqlite3.connect(base_db_path)
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM translation_data")
                db_rows = cursor.fetchall()
                
                # 중복 및 고유 항목 분리
                unique_entries = {}
                duplicate_entries_for_ui = {}

                for sheet_name, excel_cache in all_excel_cache.items():
                    for sid, data_list in excel_cache.items():
                        if len(data_list) > 1:
                            if sid not in duplicate_entries_for_ui:
                                duplicate_entries_for_ui[sid] = []
                            duplicate_entries_for_ui[sid].extend(data_list)
                        else:
                            if sid not in unique_entries:
                                unique_entries[sid] = data_list[0]

                self.log(f"고유 항목: {len(unique_entries)}개 / 중복 ID: {len(duplicate_entries_for_ui)}개")
                
                # 고유 항목 업데이트
                update_queries = []
                only_if_kr_matches = options.get("only_if_kr_matches", True)
                
                for db_row in db_rows:
                    string_id = db_row["string_id"]
                    if string_id in unique_entries:
                        excel_row = unique_entries[string_id]
                        if only_if_kr_matches and str(db_row["kr"] or "") != str(excel_row.get("kr", "") or ""): continue
                        
                        set_clauses = [f"{lang} = ?" for lang in ["en", "cn", "tw", "th"] if lang in excel_row and str(db_row[lang] or '') != str(excel_row[lang] or '')]
                        params = [excel_row[lang] for lang in ["en", "cn", "tw", "th"] if lang in excel_row and str(db_row[lang] or '') != str(excel_row[lang] or '')]
                        
                        if set_clauses:
                            params.append(string_id)
                            update_queries.append((f"UPDATE translation_data SET {', '.join(set_clauses)} WHERE string_id = ?", params))
                
                if update_queries:
                    cursor.execute("BEGIN TRANSACTION")
                    for query, params in update_queries:
                        cursor.execute(query, params)
                    total_updated_rows += len(update_queries)
                    conn.commit()

                conn.close()
                self.log(f"고유 항목 DB 업데이트 완료: {total_updated_rows}개 항목 업데이트됨.")
                
                final_result = {"status": "success", "updated_rows": total_updated_rows, "duplicate_entries": duplicate_entries_for_ui}
                self.parent_app.root.after(0, completion_callback, final_result)

            except Exception as e:
                error_result = {"status": "error", "message": str(e)}
                self.parent_app.root.after(0, completion_callback, error_result)
            finally:
                if 'conn' in locals() and conn: conn.close()
                if 'wb' in locals() and wb: wb.close()
                loading_popup.close()

        loading_popup = LoadingPopup(self.parent_app.root, "DB 업데이트 중", "준비 중...")
        thread = threading.Thread(target=task, daemon=True)
        thread.start()

# tools/workflow_manager.py 파일의 WorkflowManager 클래스 내부에 아래 함수를 추가하세요.

    def run_step2_apply_resolved_data(self, db_path, resolved_data, completion_callback):
        """사용자가 수정한 중복 항목들을 DB에 최종 반영합니다."""
        try:
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            update_queries = []
            for string_id, data in resolved_data.items():
                set_clauses = [f"{lang} = ?" for lang in ["en", "cn", "tw", "th"] if lang in data]
                params = [data[lang] for lang in ["en", "cn", "tw", "th"] if lang in data]
                
                if set_clauses:
                    params.append(string_id)
                    query = f"UPDATE translation_data SET {', '.join(set_clauses)} WHERE string_id = ?"
                    update_queries.append((query, params))

            if update_queries:
                cursor.execute("BEGIN TRANSACTION")
                for query, params in update_queries:
                    cursor.execute(query, params)
                conn.commit()

            conn.close()
            completion_callback({"status": "success", "updated_rows": len(update_queries)})
        except Exception as e:
            if conn: conn.close()
            completion_callback({"status": "error", "message": str(e)})
            

    def run_step3_apply_to_files(self, base_db_path, files_to_process, options, completion_callback):
        """3단계: 완성된 DB를 원본 엑셀 파일들에 최종 적용합니다."""
        
        def task():
            try:
                # 1. 번역 DB 캐시 로드
                loading_popup.update_message("기준 DB 캐시 로드 중...")
                self.log("기준 DB 캐시 로드를 시작합니다...")
                cache_result = self.apply_manager.load_translation_cache_from_db(base_db_path)
                if cache_result["status"] == "error":
                    raise Exception(f"DB 캐시 로드 실패: {cache_result['message']}")
                self.log(f"DB 캐시 로드 완료: {cache_result['id_count']}개 항목")

                # 2. 각 파일에 번역 적용
                total_files = len(files_to_process)
                total_updated_rows = 0
                processed_files_count = 0
                
                for i, (file_name, file_path) in enumerate(files_to_process):
                    loading_popup.update_message(f"번역 적용 중 ({i+1}/{total_files}): {file_name}")
                    
                    apply_result = self.apply_manager.apply_translation(
                        file_path=file_path,
                        selected_langs=options.get("selected_langs", []),
                        record_date=options.get("record_date", True),
                        kr_match_check=options.get("kr_match_check", True),
                        kr_mismatch_delete=options.get("kr_mismatch_delete", False),
                        allowed_statuses=[] # 워크플로우에서는 모든 항목 적용
                    )
                    
                    if apply_result["status"] == "success":
                        total_updated_rows += apply_result.get("total_updated", 0)
                        processed_files_count += 1
                
                final_result = {
                    "status": "success",
                    "total_updated": total_updated_rows,
                    "processed_files": processed_files_count
                }
                self.parent_app.root.after(0, completion_callback, final_result)

            except Exception as e:
                error_result = {"status": "error", "message": str(e)}
                self.parent_app.root.after(0, completion_callback, error_result)
            finally:
                loading_popup.close()

        loading_popup = LoadingPopup(self.parent_app.root, "번역 적용 중", "준비 중...")
        thread = threading.Thread(target=task, daemon=True)
        thread.start()