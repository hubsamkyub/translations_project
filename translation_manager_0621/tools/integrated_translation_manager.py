# integrated_translation_manager.py
# 이 파일을 프로젝트의 tools/ 폴더에 저장하세요

import os
import sqlite3
import pandas as pd
from collections import defaultdict
from datetime import datetime
from openpyxl import load_workbook
import gc
import threading
import time

class IntegratedTranslationManager:
    def __init__(self, parent_window=None):
        self.parent = parent_window
        self.master_data = {}  # 메모리 내 마스터 데이터
        self.target_data = {}  # 메모리 내 타겟 데이터
        self.duplicate_data = {}  # 중복 데이터 추적
        self.comparison_results = []  # 비교 결과
        
    def find_string_id_position(self, worksheet):
        """STRING_ID 컬럼 위치 찾기"""
        for row in worksheet.iter_rows(min_row=1, max_row=6, max_col=5):
            for cell in row:
                if isinstance(cell.value, str) and "STRING_ID" in cell.value.upper():
                    return cell.column, cell.row
        return None, None

    def find_language_columns(self, worksheet, header_row, langs, language_mapping=None):
        """언어 컬럼 매핑 찾기"""
        if not header_row: 
            return {}
        
        lang_cols = {}
        extended_mapping = {}
        
        if language_mapping:
            for alt, main in language_mapping.items():
                extended_mapping[alt.upper()] = main
        
        for row in worksheet.iter_rows(min_row=header_row, max_row=header_row):
            for cell in row:
                if not cell.value: 
                    continue
                header_text = str(cell.value).strip().upper()
                
                if header_text in langs:
                    lang_cols[header_text] = cell.column
                elif header_text in extended_mapping:
                    mapped_lang = extended_mapping[header_text]
                    if mapped_lang in langs:
                        lang_cols[mapped_lang] = cell.column
        
        return lang_cols

    def load_excel_data_to_memory(self, excel_files, language_list, progress_callback=None):
        """엑셀 파일들을 메모리로 로드하여 데이터와 중복 정보 반환"""
        language_mapping = {"ZH": "CN"}
        unique_data = {}
        duplicate_data = defaultdict(list)
        
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        processed_count = 0
        error_count = 0

        for idx, (file_name, file_path) in enumerate(excel_files):
            if progress_callback:
                progress_callback(f"파일 로드 중 ({idx+1}/{len(excel_files)}) {file_name}...", idx, len(excel_files))
            
            try:
                gc.collect()
                workbook = load_workbook(file_path, read_only=True, data_only=True)
                
                for sheet_name in workbook.sheetnames:
                    if not sheet_name.lower().startswith("string") or sheet_name.startswith("#"):
                        continue
                    
                    worksheet = workbook[sheet_name]
                    string_id_col, header_row = self.find_string_id_position(worksheet)
                    if not string_id_col or not header_row:
                        continue
                    
                    lang_cols = self.find_language_columns(worksheet, header_row, language_list, language_mapping)
                    if not lang_cols:
                        continue
                    
                    for row_cells in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
                        if not row_cells or string_id_col - 1 >= len(row_cells):
                            continue
                        
                        string_id = row_cells[string_id_col - 1]
                        if not string_id:
                            continue
                        
                        status = '비활성' if str(row_cells[0] or '').strip().startswith('#') else 'active'
                        
                        # 데이터 구조 생성
                        data_dict = {
                            'string_id': string_id,
                            'file_name': file_name,
                            'sheet_name': sheet_name,
                            'status': status,
                            'update_date': current_time
                        }
                        
                        # 언어 데이터 추가
                        for lang, col in lang_cols.items():
                            if col - 1 < len(row_cells):
                                data_dict[lang.lower()] = row_cells[col - 1]
                        
                        # 중복 검사 및 데이터 저장
                        if string_id not in unique_data and string_id not in duplicate_data:
                            unique_data[string_id] = data_dict
                        elif string_id in unique_data:
                            # 기존 데이터를 중복 리스트로 이동
                            original_data = unique_data.pop(string_id)
                            duplicate_data[string_id].extend([original_data, data_dict])
                        else:
                            # 이미 중복 리스트에 있으면 추가
                            duplicate_data[string_id].append(data_dict)

                workbook.close()
                processed_count += 1
                
            except Exception as e:
                if progress_callback:
                    progress_callback(f"파일 처리 오류: {e}", idx + 1, len(excel_files))
                error_count += 1
        
        return {
            "unique_data": unique_data,
            "duplicate_data": dict(duplicate_data),
            "processed_count": processed_count,
            "error_count": error_count,
            "total_unique": len(unique_data)
        }

    def load_db_data_to_memory(self, db_path):
        """기존 DB 파일을 메모리로 로드"""
        if not os.path.exists(db_path):
            return {}
        
        try:
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            cursor.execute("SELECT * FROM translation_data WHERE status = 'active'")
            rows = cursor.fetchall()
            
            # 컬럼 정보 가져오기
            column_names = [description[0] for description in cursor.description]
            
            db_data = {}
            for row in rows:
                row_dict = dict(zip(column_names, row))
                string_id = row_dict.get('string_id')
                if string_id:
                    db_data[string_id] = row_dict
            
            conn.close()
            return db_data
            
        except Exception as e:
            print(f"DB 로드 중 오류: {e}")
            return {}

    def compare_data_in_memory(self, master_data, target_data, language_list, comparison_options):
        """메모리 내 데이터 비교"""
        results = {
            "new_in_target": [],      # 타겟에만 있는 항목 (신규)
            "new_in_master": [],      # 마스터에만 있는 항목 (삭제됨)
            "modified": [],           # 변경된 항목
            "unchanged": []           # 변경되지 않은 항목
        }
        
        # 타겟에만 있는 항목 (신규)
        if comparison_options.get("include_new", True):
            for string_id, data in target_data.items():
                if string_id not in master_data:
                    results["new_in_target"].append(data)
        
        # 마스터에만 있는 항목 (삭제됨)
        if comparison_options.get("include_deleted", True):
            for string_id, data in master_data.items():
                if string_id not in target_data:
                    results["new_in_master"].append(data)
        
        # 양쪽에 모두 있는 항목 비교
        if comparison_options.get("include_modified", True):
            common_ids = set(master_data.keys()) & set(target_data.keys())
            
            for string_id in common_ids:
                master_item = master_data[string_id]
                target_item = target_data[string_id]
                
                is_changed = False
                changes_detail = {}
                
                # 언어별 비교
                for lang in language_list:
                    lang_key = lang.lower()
                    master_val = str(master_item.get(lang_key, '') or '').strip()
                    target_val = str(target_item.get(lang_key, '') or '').strip()
                    
                    if master_val != target_val:
                        is_changed = True
                        changes_detail[f'{lang_key}_master'] = master_val
                        changes_detail[f'{lang_key}_target'] = target_val
                
                if is_changed:
                    # 변경된 항목 추가
                    modified_item = target_item.copy()
                    modified_item.update(changes_detail)
                    results["modified"].append(modified_item)
                else:
                    results["unchanged"].append(target_item)
        
        return results

    def integrated_process(self, excel_files, language_list, comparison_options=None, 
                          master_db_path=None, progress_callback=None):
        """통합 프로세스: 로드 → 비교 → 결과 반환"""
        
        if comparison_options is None:
            comparison_options = {
                "include_new": True,
                "include_deleted": True, 
                "include_modified": True
            }
        
        results = {
            "status": "success",
            "load_results": None,
            "comparison_results": None,
            "duplicate_data": None,
            "master_count": 0,
            "target_count": 0
        }
        
        try:
            # 1단계: 엑셀 파일들을 메모리로 로드
            if progress_callback:
                progress_callback("엑셀 파일 로드 중...", 1, 5)
            
            load_results = self.load_excel_data_to_memory(excel_files, language_list, progress_callback)
            results["load_results"] = load_results
            results["duplicate_data"] = load_results["duplicate_data"]
            
            self.target_data = load_results["unique_data"]
            results["target_count"] = len(self.target_data)
            
            # 2단계: 마스터 DB 로드 (있는 경우)
            if master_db_path and os.path.exists(master_db_path):
                if progress_callback:
                    progress_callback("마스터 DB 로드 중...", 2, 5)
                
                self.master_data = self.load_db_data_to_memory(master_db_path)
                results["master_count"] = len(self.master_data)
            else:
                self.master_data = {}
                results["master_count"] = 0
            
            # 3단계: 데이터 비교
            if progress_callback:
                progress_callback("데이터 비교 중...", 3, 5)
            
            comparison_results = self.compare_data_in_memory(
                self.master_data, self.target_data, language_list, comparison_options
            )
            results["comparison_results"] = comparison_results
            self.comparison_results = comparison_results
            
            # 4단계: 결과 정리
            if progress_callback:
                progress_callback("결과 정리 중...", 4, 5)
            
            results["summary"] = {
                "new_items": len(comparison_results["new_in_target"]),
                "deleted_items": len(comparison_results["new_in_master"]),
                "modified_items": len(comparison_results["modified"]),
                "unchanged_items": len(comparison_results["unchanged"]),
                "duplicate_ids": len(results["duplicate_data"])
            }
            
            if progress_callback:
                progress_callback("통합 프로세스 완료!", 5, 5)
            
        except Exception as e:
            results["status"] = "error"
            results["message"] = str(e)
        
        return results

    def export_results_to_excel(self, output_path, export_options=None):
        """비교 결과를 엑셀로 내보내기"""
        if export_options is None:
            export_options = {
                "export_new": True,
                "export_deleted": True,
                "export_modified": True,
                "export_duplicates": True
            }
        
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                
                # 신규 항목 시트
                if export_options.get("export_new", True) and self.comparison_results:
                    new_items = self.comparison_results.get("new_in_target", [])
                    if new_items:
                        df_new = pd.DataFrame(new_items)
                        df_new.to_excel(writer, sheet_name='신규항목', index=False)
                
                # 삭제된 항목 시트
                if export_options.get("export_deleted", True) and self.comparison_results:
                    deleted_items = self.comparison_results.get("new_in_master", [])
                    if deleted_items:
                        df_deleted = pd.DataFrame(deleted_items)
                        df_deleted.to_excel(writer, sheet_name='삭제된항목', index=False)
                
                # 변경된 항목 시트
                if export_options.get("export_modified", True) and self.comparison_results:
                    modified_items = self.comparison_results.get("modified", [])
                    if modified_items:
                        df_modified = pd.DataFrame(modified_items)
                        df_modified.to_excel(writer, sheet_name='변경된항목', index=False)
                
                # 중복 항목 시트
                if export_options.get("export_duplicates", True) and self.duplicate_data:
                    duplicate_flat = []
                    for string_id, items in self.duplicate_data.items():
                        for item in items:
                            duplicate_flat.append(item)
                    
                    if duplicate_flat:
                        df_duplicates = pd.DataFrame(duplicate_flat)
                        df_duplicates.to_excel(writer, sheet_name='중복항목', index=False)
            
            return {"status": "success", "path": output_path}
            
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def save_to_db(self, db_path, data_source="target"):
        """메모리 데이터를 DB로 저장"""
        try:
            if os.path.exists(db_path):
                os.remove(db_path)
            
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            # 테이블 생성
            cursor.execute('''
            CREATE TABLE translation_data (
                id INTEGER PRIMARY KEY,
                file_name TEXT,
                sheet_name TEXT,
                string_id TEXT UNIQUE,
                kr TEXT,
                en TEXT,
                cn TEXT,
                tw TEXT,
                th TEXT,
                status TEXT DEFAULT 'active',
                update_date TEXT
            )
            ''')
            
            # 데이터 선택
            data_to_save = self.target_data if data_source == "target" else self.master_data
            
            # 데이터 삽입
            for string_id, data in data_to_save.items():
                cursor.execute('''
                INSERT INTO translation_data 
                (file_name, sheet_name, string_id, kr, en, cn, tw, th, status, update_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    data.get('file_name', ''),
                    data.get('sheet_name', ''),
                    data.get('string_id', ''),
                    data.get('kr', ''),
                    data.get('en', ''),
                    data.get('cn', ''),
                    data.get('tw', ''),
                    data.get('th', ''),
                    data.get('status', 'active'),
                    data.get('update_date', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                ))
            
            conn.commit()
            conn.close()
            
            return {"status": "success", "saved_count": len(data_to_save)}
            
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def get_statistics(self):
        """현재 데이터 통계 반환"""
        stats = {
            "master_data_count": len(self.master_data),
            "target_data_count": len(self.target_data),
            "duplicate_count": len(self.duplicate_data),
            "comparison_executed": bool(self.comparison_results)
        }
        
        if self.comparison_results:
            stats.update({
                "new_items": len(self.comparison_results.get("new_in_target", [])),
                "deleted_items": len(self.comparison_results.get("new_in_master", [])),
                "modified_items": len(self.comparison_results.get("modified", [])),
                "unchanged_items": len(self.comparison_results.get("unchanged", []))
            })
        
        return stats

    def clear_data(self):
        """메모리 데이터 초기화"""
        self.master_data.clear()
        self.target_data.clear()
        self.duplicate_data.clear()
        self.comparison_results.clear()
        gc.collect()