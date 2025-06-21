import tkinter as tk
from tkinter import ttk, filedialog
import os
from openpyxl import load_workbook

# 유틸리티 및 분리된 모듈 임포트
from ui.common_components import ScrollableCheckList, show_message
from tools.basic_request_extractor import BasicRequestExtractor
from tools.compare_request_extractor import CompareRequestExtractor
from tools.request_extraction_manager import RequestExtractionManager

class TranslationRequestExtractor(tk.Frame):
    def __init__(self, root):
        super().__init__(root)
        self.root = root.winfo_toplevel()
        
        # 핵심 로직 매니저 생성
        self.extraction_manager = RequestExtractionManager(self)
        
        # 공통 UI 변수
        self.folder_path_var = tk.StringVar(value="")
        self.output_db_var = tk.StringVar(value="")
        
        # 공통 데이터
        self.excel_files = []
        self.extraction_thread = None
        
        self.setup_ui()
        self.pack(fill="both", expand=True)

    def setup_ui(self):
        """메인 컨테이너 UI: 공통 부분과 탭으로 구성"""
        # 1. 소스 파일 선택 (공통 영역)
        source_files_frame = ttk.LabelFrame(self, text="1. 추출 대상 파일 선택")
        source_files_frame.pack(fill="x", padx=10, pady=5)
        
        folder_frame = ttk.Frame(source_files_frame)
        folder_frame.pack(fill="x", padx=5, pady=5)
        ttk.Label(folder_frame, text="엑셀 폴더 경로:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(folder_frame, textvariable=self.folder_path_var, width=60).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(folder_frame, text="폴더 찾기", command=self.select_excel_folder).grid(row=0, column=2, padx=5, pady=5)
        ttk.Button(folder_frame, text="파일 검색", command=self.search_excel_files).grid(row=0, column=3, padx=5, pady=5)
        folder_frame.columnconfigure(1, weight=1)
        
        self.excel_files_list = ScrollableCheckList(source_files_frame, width=700, height=120)
        self.excel_files_list.pack(fill="both", expand=True, padx=5, pady=5)

        # 2. 작업 선택 (Notebook으로 분리)
        action_notebook = ttk.Notebook(self)
        action_notebook.pack(fill="both", expand=True, padx=10, pady=5)

        # 탭 1: 기본 추출 (UI 클래스 인스턴스화)
        basic_extract_tab = ttk.Frame(action_notebook)
        self.basic_extractor = BasicRequestExtractor(basic_extract_tab, self)
        self.basic_extractor.pack(fill="both", expand=True)
        action_notebook.add(basic_extract_tab, text="기본 추출")

        # 탭 2: 비교하여 추출 (UI 클래스 인스턴스화)
        compare_extract_tab = ttk.Frame(action_notebook)
        self.compare_extractor = CompareRequestExtractor(compare_extract_tab, self)
        self.compare_extractor.pack(fill="both", expand=True)
        action_notebook.add(compare_extract_tab, text="비교하여 추출")

        # 3. 로그 및 상태 표시줄 (공통 영역)
        log_frame = ttk.LabelFrame(self, text="작업 로그")
        log_frame.pack(fill="both", expand=True, padx=10, pady=5)
        self.log_text = tk.Text(log_frame, wrap="word", height=8)
        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.log_text.pack(fill="both", expand=True)
        
        status_frame = ttk.Frame(self)
        status_frame.pack(fill="x", padx=10, pady=5)
        self.status_label = ttk.Label(status_frame, text="대기 중...")
        self.status_label.pack(side="left", padx=5)

    def select_excel_folder(self):
        folder = filedialog.askdirectory(title="엑셀 파일 폴더 선택", parent=self.root)
        if folder: self.folder_path_var.set(folder)
            
    def select_output_db(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".db", filetypes=[("DB 파일", "*.db")], title="DB 파일 저장", parent=self.root)
        if file_path: self.output_db_var.set(file_path)

    def search_excel_files(self):
        folder = self.folder_path_var.get()
        if not folder or not os.path.isdir(folder):
            show_message(self.root, "warning", "경고", "유효한 폴더를 선택하세요.")
            return
        
        self.excel_files_list.clear()
        self.excel_files = []
        self.log_text.delete(1.0, tk.END)
        self.log_message("파일 검색 중...")
        
        for root, _, files in os.walk(folder):
            for file in files:
                if file.lower().startswith("string") and file.lower().endswith(('.xlsx', '.xls')) and not file.startswith("~$"):
                    self.excel_files.append((file, os.path.join(root, file)))
        
        self.excel_files.sort()
        for file_name, _ in self.excel_files:
            self.excel_files_list.add_item(file_name, checked=True)
        self.log_message(f"{len(self.excel_files)}개의 엑셀 파일을 찾았습니다.")

    def log_message(self, message):
        self.log_text.insert(tk.END, f"{message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
        
    def _is_task_running(self):
        if self.extraction_thread and self.extraction_thread.is_alive():
            show_message(self.root, "info", "알림", "이미 다른 작업이 진행 중입니다.")
            return True
        return False

    def _get_selected_files(self):
        selected_file_names = self.excel_files_list.get_checked_items()
        if not selected_file_names:
            show_message(self.root, "warning", "경고", "추출할 파일을 선택하세요.")
            return None
        return [item for item in self.excel_files if item[0] in selected_file_names]

    def _find_headers_in_worksheet(self, ws):
        for i in range(1, 11):
            row_values = [cell.value for cell in ws[i]]
            cleaned_values = [str(v).upper().strip() for v in row_values if v is not None]
            
            if "STRING_ID" in cleaned_values:
                header_map = {idx: str(val).strip() for idx, val in enumerate(row_values, 1) if val is not None}
                return {k: v.upper() for k, v in header_map.items()}, i
        return None, -1
        
    # tools/translation_request_extractor.py
    def _update_files_as_transferred(self, files_to_update, col_name_to_find, new_value):
        # ▼▼▼ [수정] 파일 저장 전, 파일이 열려있는지 확인하는 로직 추가 ▼▼▼
        target_file_paths = list(files_to_update.keys())
        open_files = self._check_files_are_open(target_file_paths)
        if open_files:
            error_message = "다음 파일이 열려있어 '전달' 상태로 업데이트할 수 없습니다. 파일을 닫고 다시 시도해주세요:\n\n" + "\n".join(open_files)
            self.log(f"파일 업데이트 취소: 다음 파일이 열려있음 - {', '.join(open_files)}")
            show_message(self.parent_app.root, "warning", "작업 중단", error_message)
            return
        # ▲▲▲ 여기까지 추가 ▲▲▲

        for file_path, sheets in files_to_update.items():
            wb = None # finally를 위해 wb 초기화
            try:
                wb = load_workbook(file_path)
                for sheet_name, row_indices in sheets.items():
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        headers, _ = self._find_headers_in_worksheet(ws)
                        if not headers: continue

                        col_idx = None
                        for c_idx, c_name in headers.items():
                            if c_name.upper() == col_name_to_find.upper():
                                col_idx = c_idx
                                break

                        if col_idx:
                            for row_idx in row_indices:
                                ws.cell(row=row_idx, column=col_idx).value = new_value
                wb.save(file_path)
                self.log(f"'{os.path.basename(file_path)}' 파일 업데이트 완료.")
            except Exception as e:
                self.log(f"파일 업데이트 실패 '{os.path.basename(file_path)}': {e}")
            finally:
                # wb.save() 이후에는 close()가 의미 없지만, 로드 후 저장 실패 시를 대비
                if wb:
                    pass

    def _check_files_are_open(self, file_paths_to_check):
        """주어진 파일 경로 목록을 확인하여 열려 있는 파일이 있는지 검사합니다."""
        open_files = []
        for file_path in file_paths_to_check:
            if not os.path.exists(file_path):
                continue
            try:
                # 파일을 자기 자신으로 리네임 시도. 파일이 열려있으면 OSError 발생
                os.rename(file_path, file_path)
            except OSError:
                open_files.append(os.path.basename(file_path))
        return open_files