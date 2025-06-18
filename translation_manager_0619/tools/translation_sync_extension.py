# translation_sync_extension.py
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from ttkwidgets import CheckboxTreeview
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from utils.config_utils import load_config, save_config

class TranslationSyncExtension:
    def __init__(self, root):
        self.root = root
        self.root.title("번역 동기화 도구")
        self.root.geometry("800x700")
        
        # 메인 프레임
        main_frame = ttk.Frame(self.root, padding=10)
        main_frame.pack(fill="both", expand=True)
        
        # 폴더 선택 부분
        folder_frame = ttk.LabelFrame(main_frame, text="폴더 선택")
        folder_frame.pack(fill="x", padx=5, pady=5)
        
        ttk.Label(folder_frame, text="최신 데이터 폴더(A):").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.source_folder_var = tk.StringVar()
        ttk.Entry(folder_frame, textvariable=self.source_folder_var, width=50).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(folder_frame, text="찾아보기", command=self.select_source_folder).grid(row=0, column=2, padx=5, pady=5)
        
        ttk.Label(folder_frame, text="이전 번역 폴더(B):").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.target_folder_var = tk.StringVar()
        ttk.Entry(folder_frame, textvariable=self.target_folder_var, width=50).grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(folder_frame, text="찾아보기", command=self.select_target_folder).grid(row=1, column=2, padx=5, pady=5)
        
        folder_frame.columnconfigure(1, weight=1)
        
        # 옵션 설정
        options_frame = ttk.LabelFrame(main_frame, text="동기화 옵션")
        options_frame.pack(fill="x", padx=5, pady=5)
        
        self.highlight_kr_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="KR 값이 변경된 항목 하이라이트", 
                        variable=self.highlight_kr_var).grid(row=0, column=0, padx=5, pady=5, sticky="w")
        
        self.add_mark_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="변경된 항목에 '변경됨' 표시 추가", 
                        variable=self.add_mark_var).grid(row=0, column=1, padx=5, pady=5, sticky="w")
        
        # 추가 옵션
        self.copy_all_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="KR이 변경되어도 모든 번역 복사", 
                        variable=self.copy_all_var).grid(row=1, column=0, padx=5, pady=5, sticky="w")
        
        # 작업 실행 버튼
        action_frame = ttk.Frame(main_frame)
        action_frame.pack(fill="x", padx=5, pady=5)
        # action_frame에 전체 체크/해제 버튼 추가 (action_frame 정의 부분 바로 아래에 추가)
        self.check_all_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(action_frame, text="전체 선택", 
                        variable=self.check_all_var,
                        command=self.toggle_all_files).pack(side="left", padx=5, pady=5)
        
        ttk.Button(action_frame, text="매칭 파일 검색", 
                   command=self.find_matching_files).pack(side="left", padx=5, pady=5)
        ttk.Button(action_frame, text="번역 동기화", 
                   command=self.sync_translations).pack(side="right", padx=5, pady=5)
        
        # action_frame에 미리보기 버튼 추가
        ttk.Button(action_frame, text="변경사항 미리보기", 
                command=self.preview_changes).pack(side="left", padx=5, pady=5)

        
        # 파일 목록 표시 영역
        files_frame = ttk.LabelFrame(main_frame, text="매칭된 파일 목록")
        files_frame.pack(fill="x", padx=5, pady=5)

        # 기존 트리뷰 대신 CheckboxTreeview 사용
        self.files_tree = CheckboxTreeview(files_frame, columns=("source_file", "target_file"), show="tree headings", height=5)


        # 컬럼 설정
        self.files_tree.heading("source_file", text="A 파일(최신)")
        self.files_tree.heading("target_file", text="B 파일(이전 번역)")

        self.files_tree.column("source_file", width=350)
        self.files_tree.column("target_file", width=350)

        # 스크롤바 연결
        scrollbar = ttk.Scrollbar(files_frame, orient="vertical", command=self.files_tree.yview)
        self.files_tree.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        self.files_tree.pack(fill="x", expand=True, padx=5, pady=5)

        # 체크박스 이벤트 연결
        self.files_tree.bind("<<CheckboxToggled>>", self.on_checkbox_toggled)
        
        # 더블 클릭 이벤트 바인딩
        self.files_tree.bind("<Double-1>", lambda e: self.preview_changes())
        
        # 로그 표시 영역
        log_frame = ttk.LabelFrame(main_frame, text="작업 로그")
        log_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        self.log_text = tk.Text(log_frame, wrap="word")
        scrollbar_log = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar_log.set)
        
        scrollbar_log.pack(side="right", fill="y")
        self.log_text.pack(fill="both", expand=True)
        
        # 상태와 진행 표시
        status_frame = ttk.Frame(main_frame)
        status_frame.pack(fill="x", padx=5, pady=5)
        
        self.status_label = ttk.Label(status_frame, text="대기 중...")
        self.status_label.pack(side="left", padx=5)
        
        self.progress_bar = ttk.Progressbar(status_frame, length=400, mode="determinate")
        self.progress_bar.pack(side="right", fill="x", expand=True, padx=5)
        
        # 매칭된 파일 저장용
        self.matching_files = []
    
    def select_source_folder(self):
        # 현재 윈도우를 폴더 선택 대화상자의 부모로 지정
        folder = filedialog.askdirectory(title="최신 데이터 폴더(A) 선택", parent=self.root)
        if folder:
            self.source_folder_var.set(folder)
            # 폴더 선택 후 메인 창에 다시 포커스
            self.root.after(100, self.root.focus_force)
            self.root.after(100, self.root.lift)


    def select_target_folder(self):
        # 현재 윈도우를 폴더 선택 대화상자의 부모로 지정
        folder = filedialog.askdirectory(title="이전 번역 폴더(B) 선택", parent=self.root)
        if folder:
            self.target_folder_var.set(folder)
            # 폴더 선택 후 메인 창에 다시 포커스
            self.root.after(100, self.root.focus_force)
            self.root.after(100, self.root.lift)
            
            
    def find_matching_files(self):
        source_folder = self.source_folder_var.get()
        target_folder = self.target_folder_var.get()
        
        if not os.path.isdir(source_folder) or not os.path.isdir(target_folder):
            messagebox.showwarning("경고", "유효한 폴더를 선택하세요.", parent=self.top)
            return
        
        # 각 폴더에서 파일 목록 가져오기 (하위 폴더 포함)
        source_files = {}
        for root, dirs, files in os.walk(source_folder):
            for f in files:
                if f.lower().endswith(".xlsx") and f.lower().startswith("string"):
                    # 파일의 상대 경로를 키로 사용 (중복 방지)
                    rel_path = os.path.relpath(os.path.join(root, f), source_folder)
                    source_files[rel_path] = os.path.join(root, f)

        target_files = {}
        for root, dirs, files in os.walk(target_folder):
            for f in files:
                if f.lower().endswith(".xlsx") and f.lower().startswith("string"):
                    rel_path = os.path.relpath(os.path.join(root, f), target_folder)
                    target_files[rel_path] = os.path.join(root, f)
        
        # 공통 파일 찾기
        common_files = set(source_files.keys()) & set(target_files.keys())
        
        if not common_files:
            messagebox.showinfo("알림", "매칭되는 파일이 없습니다.")
            return
        
        # 트리뷰 초기화
        for item in self.files_tree.get_children():
            self.files_tree.delete(item)
        
        # 매칭 파일 저장 및 트리뷰에 표시 (파일명만 표시)
        self.matching_files = []
        for idx, filename in enumerate(sorted(common_files)):
            # 파일 경로 대신 파일명만 표시
            item_id = self.files_tree.insert("", "end", values=(filename, filename))
            self.files_tree.change_state(item_id, "checked")
            # 저장할 때는 전체 경로를 유지
            self.matching_files.append((source_files[filename], target_files[filename], item_id))
        
        self.log_text.delete(1.0, tk.END)
        self.log_text.insert(tk.END, f"{len(common_files)}개의 매칭되는 파일을 찾았습니다.\n")
        self.status_label.config(text=f"{len(common_files)}개 파일 매칭됨")
        
        # 전체 선택 체크박스 상태 업데이트
        self.check_all_var.set(True)
    
    
    #체크박스 토글 이벤트 처리
    def on_checkbox_toggled(self, event):
        """체크박스 토글 이벤트 처리"""
        # 전체 선택 체크박스 상태 업데이트
        all_items = self.files_tree.get_children()
        checked_items = self.files_tree.get_checked()
        
        all_checked = len(all_items) > 0 and len(all_items) == len(checked_items)
        self.check_all_var.set(all_checked)

    
    #전체 파일 체크/해제 토글
    def toggle_all_files(self):
        check_all = self.check_all_var.get()
        
        # 모든 항목 업데이트
        for item in self.files_tree.get_children():
            if check_all:
                self.files_tree.change_state(item, "checked")
            else:
                self.files_tree.change_state(item, "unchecked")


    def sync_translations(self):
        if not self.matching_files:
            messagebox.showwarning("경고", "먼저 매칭 파일을 검색하세요.")
            return

        # sync_translations 메서드에서 선택된 파일 필터링 부분
        selected_files = []
        for src, tgt, item_id in self.matching_files:
            if item_id in self.files_tree.get_checked():  # 체크된 항목인지 확인
                selected_files.append((src, tgt))

        # 외부 링크 및 열기 실패 파일 필터링
        valid_selected_files = []
        excluded_due_to_error = []

        for src, tgt in selected_files:
            try:
                test_wb = load_workbook(src, read_only=True)
                if getattr(test_wb, '_external_links', []):  # 외부 링크 존재 시 제외
                    excluded_due_to_error.append((src, "외부 링크 있음"))
                    test_wb.close()
                    continue
                test_wb.close()
                valid_selected_files.append((src, tgt))
            except Exception as e:
                excluded_due_to_error.append((src, f"오류: {str(e)}"))

        # 로그 초기화
        self.log_text.delete(1.0, tk.END)
        self.log_text.insert(tk.END, "번역 동기화 시작...\n")
        self.log_text.insert(tk.END, f"선택된 파일: {len(valid_selected_files)}개\n\n")

        # 제외된 파일 로그 출력
        if excluded_due_to_error:
            self.log_text.insert(tk.END, "제외된 파일 목록:\n")
            for path, reason in excluded_due_to_error:
                self.log_text.insert(tk.END, f"  {os.path.basename(path)} - {reason}\n")
            self.log_text.insert(tk.END, "\n")

        # 진행 바 설정
        self.progress_bar["maximum"] = len(valid_selected_files)
        self.progress_bar["value"] = 0

        # 통계 변수
        total_items = 0
        changed_kr_items = 0
        synced_items = 0

        # KR 변경 항목 표시용 스타일
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # 엑셀이 저장된 파일 목록
        processed_files = []

        # 각 파일 쌍에 대해 처리
        for idx, (source_path, target_path) in enumerate(valid_selected_files):
            file_name = os.path.basename(source_path)
            self.log_text.insert(tk.END, f"{file_name} 처리 중...\n")
            self.root.update()

            source_wb = None
            target_wb = None
            
            try:
                # 워크북 로드
                source_wb = load_workbook(source_path)
                target_wb = load_workbook(target_path)

                # 각 시트에 대해 처리
                for sheet_name in source_wb.sheetnames:
                    if sheet_name.startswith("String") and sheet_name in target_wb.sheetnames:
                        file_stats = self.sync_sheet(
                            source_wb[sheet_name],
                            target_wb[sheet_name],
                            yellow_fill,
                            self.highlight_kr_var.get(),
                            self.add_mark_var.get(),
                            self.copy_all_var.get()
                        )

                        total_items += file_stats["total"]
                        changed_kr_items += file_stats["changed_kr"]
                        synced_items += file_stats["synced"]

                        self.log_text.insert(tk.END, f"  시트 {sheet_name}: 총 {file_stats['total']}항목, "
                                                    f"KR 변경 {file_stats['changed_kr']}항목, "
                                                    f"동기화 {file_stats['synced']}항목\n")

                # 저장
                source_wb.save(source_path)
                processed_files.append(source_path)

            except Exception as e:
                self.log_text.insert(tk.END, f"  오류: {file_name} 처리 중 오류 발생: {e}\n")
            finally:
                # 워크북 확실히 닫기
                if source_wb:
                    try:
                        source_wb.close()
                    except:
                        pass
                if target_wb:
                    try:
                        target_wb.close()
                    except:
                        pass
                
                # 파일 핸들이 완전히 해제될 때까지 잠시 대기
                import time
                time.sleep(0.1)

            # 진행 업데이트
            self.progress_bar["value"] = idx + 1
            self.root.update()

        # 완료 메시지
        summary = (f"동기화 완료!\n"
                f"총 항목: {total_items}\n"
                f"KR 변경 항목: {changed_kr_items}\n"
                f"동기화된 항목: {synced_items}\n")

        self.log_text.insert(tk.END, summary)
        self.status_label.config(text="완료")

        # 파일이 완전히 저장되도록 안내
        if processed_files:
            self.log_text.insert(tk.END, "\n엑셀 파일 저장 확인 중...\n")
            self.log_text.insert(tk.END, "저장 완료됨.\n")

        messagebox.showinfo("완료", f"번역 동기화가 완료되었습니다.\n{summary}", parent = self.root)


    # 파일 저장을 확실히 하기 위한 새 메서드 (xlwings 방식)
    def ensure_file_saved(self, file_path):
        """단일 엑셀 파일이 완전히 저장되도록 함 (xlwings 사용)"""
        try:
            import xlwings as xw
            app = xw.App(visible=False)
            wb = app.books.open(file_path)
            wb.save()
            wb.close()
            app.quit()
        except Exception as e:
            self.log_text.insert(tk.END, f"  파일 저장 확인 중 오류: {e}\n")


    def sync_sheet(self, source_sheet, target_sheet, highlight_fill, 
                  do_highlight=True, add_mark=True, copy_all=True):
        # 헤더 위치 찾기
        string_id_col_src = None
        kr_col_src = None
        cn_col_src = None
        tw_col_src = None
        header_row_src = None
        
        string_id_col_tgt = None
        kr_col_tgt = None
        cn_col_tgt = None
        tw_col_tgt = None
        header_row_tgt = None
        
        # A 파일(source) 헤더 찾기
        for row in range(3, 6):
            for col in range(1, source_sheet.max_column + 1):
                cell_value = source_sheet.cell(row=row, column=col).value
                if cell_value == "STRING_ID":
                    string_id_col_src = col
                    header_row_src = row
                elif cell_value == "KR":
                    kr_col_src = col
                elif cell_value == "CN":
                    cn_col_src = col
                elif cell_value == "TW":
                    tw_col_src = col
        
        # B 파일(target) 헤더 찾기
        for row in range(3, 6):
            for col in range(1, target_sheet.max_column + 1):
                cell_value = target_sheet.cell(row=row, column=col).value
                if cell_value == "STRING_ID":
                    string_id_col_tgt = col
                    header_row_tgt = row
                elif cell_value == "KR":
                    kr_col_tgt = col
                elif cell_value == "CN":
                    cn_col_tgt = col
                elif cell_value == "TW":
                    tw_col_tgt = col
        
        # 필요한 컬럼이 모두 있는지 확인
        if not all([string_id_col_src, kr_col_src, cn_col_src, tw_col_src, 
                    string_id_col_tgt, kr_col_tgt, cn_col_tgt, tw_col_tgt,
                    header_row_src, header_row_tgt]):
            return {"total": 0, "changed_kr": 0, "synced": 0}
        
        # 통계용 변수
        stats = {"total": 0, "changed_kr": 0, "synced": 0}
        
        # B 파일의 데이터를 해시맵으로 변환 (빠른 검색을 위해)
        target_data = {}
        for row in range(header_row_tgt + 1, target_sheet.max_row + 1):
            string_id = target_sheet.cell(row=row, column=string_id_col_tgt).value
            if string_id:
                kr = target_sheet.cell(row=row, column=kr_col_tgt).value
                cn = target_sheet.cell(row=row, column=cn_col_tgt).value
                tw = target_sheet.cell(row=row, column=tw_col_tgt).value
                target_data[string_id] = {"kr": kr, "cn": cn, "tw": tw}
        
        # 변경 표시 컬럼 준비
        change_mark_col = None
        if add_mark:
            # 기존 #원문 컬럼 찾기
            for col in range(1, source_sheet.max_column + 1):
                cell_value = source_sheet.cell(row=header_row_src, column=col).value
                if cell_value and isinstance(cell_value, str) and cell_value.strip() == "#원문":
                    change_mark_col = col
                    break
            # #원문 컬럼이 없으면 pass (새로 추가하지 않음)
            if not change_mark_col:
                pass  # 컬럼이 없으면 변경 표시 기능 비활성화
                
        # A 파일을 순회하면서 B 파일의 데이터 동기화
        for row in range(header_row_src + 1, source_sheet.max_row + 1):
            string_id = source_sheet.cell(row=row, column=string_id_col_src).value
            if not string_id or string_id not in target_data:
                continue
            
            stats["total"] += 1
            
            # KR 값 비교
            source_kr = source_sheet.cell(row=row, column=kr_col_src).value or ""
            target_kr = target_data[string_id]["kr"] or ""
            
            kr_changed = source_kr != target_kr
            
            # CN, TW 값 옮기기
            if copy_all or not kr_changed:
                if target_data[string_id]["cn"]:
                    source_sheet.cell(row=row, column=cn_col_src).value = target_data[string_id]["cn"]
                    stats["synced"] += 1
                
                if target_data[string_id]["tw"]:
                    source_sheet.cell(row=row, column=tw_col_src).value = target_data[string_id]["tw"]
                    stats["synced"] += 1
            
            # KR 값이 변경된 경우 표시
            if kr_changed:
                stats["changed_kr"] += 1
                if do_highlight:
                    source_sheet.cell(row=row, column=kr_col_src).fill = highlight_fill
                
                # 변경 정보 표시를 위한 셀 추가
                if add_mark and change_mark_col:
                    source_sheet.cell(row=row, column=change_mark_col).value = "#변경됨"
        
        return stats
    
    
    #pandas 이용
    def preview_changes(self):
        # 선택된 항목 확인
        selected_items = self.files_tree.selection()
        if not selected_items:
            messagebox.showwarning("경고", "파일을 선택하세요.")
            return
        
        # 선택된 항목 ID 가져오기
        selected_item_id = selected_items[0]
        
        # 해당 ID를 가진 항목 찾기
        selected_file_pair = None
        for file_pair in self.matching_files:
            if file_pair[2] == selected_item_id:
                selected_file_pair = file_pair
                break
        
        if not selected_file_pair:
            messagebox.showwarning("경고", "선택한 파일 정보를 찾을 수 없습니다.")
            return
            
        source_path, target_path, _ = selected_file_pair

        # 로딩 창 표시
        loading_window = tk.Toplevel(self.root)
        loading_window.title("로딩 중...")
        loading_window.geometry("300x100")
        loading_window.transient(self.root)
        loading_window.grab_set()
        
        loading_label = ttk.Label(loading_window, text="파일 분석 중...")
        loading_label.pack(pady=10)
        
        loading_progress = ttk.Progressbar(loading_window, mode='indeterminate')
        loading_progress.pack(fill='x', padx=20, pady=10)
        loading_progress.start(10)
        
        # 로딩 창을 즉시 표시하고 업데이트
        loading_window.update_idletasks()
        self.root.update()
        
        try:
            # pandas 사용하여 빠르게 처리 (openpyxl 대신)
            import pandas as pd
            
            # 로딩 상태 업데이트
            loading_label.config(text="A 파일 로딩 중...")
            loading_window.update()
            
            # 최적화된 엑셀 파일 읽기 (필요한 시트만 로드)
            # 모든 시트 이름 확인하여 String으로 시작하는 시트만 처리
            import openpyxl
            source_wb = openpyxl.load_workbook(source_path, read_only=True)
            target_wb = openpyxl.load_workbook(target_path, read_only=True)
            
            # 공통 시트 찾기
            common_sheets = [sheet for sheet in source_wb.sheetnames 
                            if sheet in target_wb.sheetnames and sheet.startswith("String")]
            
            source_wb.close()
            target_wb.close()
            
            if not common_sheets:
                messagebox.showinfo("알림", "공통 시트가 없습니다.")
                loading_window.destroy()
                return
            
            # 시트 선택
            sheet_name = common_sheets[0]  # 기본적으로 첫 번째 시트 사용
            
            # pandas로 필요한 시트만 효율적으로 로드
            loading_label.config(text=f"시트 '{sheet_name}' 분석 중...")
            loading_window.update()
            
            # 헤더 행 찾기 (최대 10행까지만 검색)
            header_row = None
            for i in range(10):
                try:
                    temp_df = pd.read_excel(source_path, sheet_name=sheet_name, 
                                        skiprows=i, nrows=1)
                    if 'STRING_ID' in temp_df.columns:
                        header_row = i
                        break
                except:
                    continue
            
            if header_row is None:
                messagebox.showinfo("알림", "헤더 행을 찾을 수 없습니다.")
                loading_window.destroy()
                return
            
            # 효율적인 데이터 로드 (필요한 열만 선택)
            loading_label.config(text="A 파일 데이터 로드 중...")
            loading_window.update()
            source_df = pd.read_excel(source_path, sheet_name=sheet_name, 
                                    skiprows=header_row)
            
            loading_label.config(text="B 파일 데이터 로드 중...")
            loading_window.update()
            target_df = pd.read_excel(target_path, sheet_name=sheet_name, 
                                    skiprows=header_row)
            
            # 필요한 열만 추출
            source_cols = [col for col in source_df.columns if col in ['STRING_ID', 'KR', 'CN', 'TW']]
            target_cols = [col for col in target_df.columns if col in ['STRING_ID', 'KR', 'CN', 'TW']]
            
            if 'STRING_ID' not in source_cols or 'STRING_ID' not in target_cols:
                messagebox.showinfo("알림", "필수 열(STRING_ID)을 찾을 수 없습니다.")
                loading_window.destroy()
                return
            
            source_df = source_df[source_cols].fillna('')
            target_df = target_df[target_cols].fillna('')
            
            # 중복된 STRING_ID 확인 및 처리
            source_duplicated = source_df['STRING_ID'].duplicated().any()
            target_duplicated = target_df['STRING_ID'].duplicated().any()
            
            # 중복된 항목이 있다면 리스트 형태로 직접 딕셔너리 생성
            source_dict = {}
            target_dict = {}
            
            # 변경 사항 분석
            changes = []
            
            loading_label.config(text="데이터 분석 중...")
            loading_window.update()
            
            # 직접 딕셔너리 생성 (중복 처리 가능)
            for _, row in source_df.iterrows():
                string_id = row['STRING_ID']
                if string_id and string_id not in source_dict:  # 빈 값이나 중복 항목 건너뛰기
                    source_dict[string_id] = {
                        'KR': row.get('KR', ''),
                        'CN': row.get('CN', ''),
                        'TW': row.get('TW', '')
                    }
                    
            for _, row in target_df.iterrows():
                string_id = row['STRING_ID']
                if string_id and string_id not in target_dict:  # 빈 값이나 중복 항목 건너뛰기
                    target_dict[string_id] = {
                        'KR': row.get('KR', ''),
                        'CN': row.get('CN', ''),
                        'TW': row.get('TW', '')
                    }
            
            # KR 값이 변경된 항목 찾기
            for string_id, src_values in source_dict.items():
                if string_id in target_dict:
                    src_kr = src_values.get('KR', '')
                    tgt_kr = target_dict[string_id].get('KR', '')
                    
                    if src_kr != tgt_kr:
                        changes.append({
                            "string_id": string_id,
                            "source_kr": src_kr,
                            "source_cn": src_values.get('CN', ''),
                            "source_tw": src_values.get('TW', ''),
                            "target_kr": tgt_kr,
                            "target_cn": target_dict[string_id].get('CN', ''),
                            "target_tw": target_dict[string_id].get('TW', ''),
                            "status": "변경" if src_kr and tgt_kr else "추가" if src_kr else "삭제"
                        })
                else:
                    # 새로 추가된 항목
                    changes.append({
                        "string_id": string_id,
                        "source_kr": src_values.get('KR', ''),
                        "source_cn": src_values.get('CN', ''),
                        "source_tw": src_values.get('TW', ''),
                        "target_kr": '',
                        "target_cn": '',
                        "target_tw": '',
                        "status": "추가"
                    })
            
            # 삭제된 항목 확인 (최대 1000개까지만 표시하여 성능 향상)
            count = 0
            for string_id, tgt_values in target_dict.items():
                if string_id not in source_dict and count < 1000:
                    changes.append({
                        "string_id": string_id,
                        "source_kr": '',
                        "source_cn": '',
                        "source_tw": '',
                        "target_kr": tgt_values.get('KR', ''),
                        "target_cn": tgt_values.get('CN', ''),
                        "target_tw": tgt_values.get('TW', ''),
                        "status": "삭제"
                    })
                    count += 1
            
            # 변경 사항을 표시할 새 창 생성
            loading_window.destroy()
            self.show_changes_table(changes, os.path.basename(source_path), sheet_name)
            
        except Exception as e:
            import traceback
            traceback.print_exc()  # 상세 오류 출력
            messagebox.showerror("오류", f"미리보기 생성 중 오류 발생: {e}")
            loading_window.destroy()
            
    
    #변경 사항을 테이블 형식으로 표시하는 창
    def show_changes_table(self, changes, file_name, sheet_name):
        if not changes:
            messagebox.showinfo("알림", "변경된 항목이 없습니다.")
            return
        
        # 새 창 생성
        preview_window = tk.Toplevel(self.root)
        preview_window.title(f"변경 미리보기 - {file_name} ({sheet_name})")
        preview_window.geometry("1200x600")
        
        # 프레임 설정
        main_frame = ttk.Frame(preview_window, padding=10)
        main_frame.pack(fill="both", expand=True)
        
        # 헤더 정보
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill="x", pady=5)
        
        ttk.Label(header_frame, text=f"파일: {file_name}, 시트: {sheet_name}").pack(side="left")
        ttk.Label(header_frame, text=f"총 {len(changes)}개 항목").pack(side="right")
        
        # 필터링 옵션
        filter_frame = ttk.Frame(main_frame)
        filter_frame.pack(fill="x", pady=5)
        
        status_var = tk.StringVar(value="모두")
        ttk.Label(filter_frame, text="상태 필터:").pack(side="left", padx=5)
        
        status_combo = ttk.Combobox(filter_frame, textvariable=status_var, 
                                values=["모두", "변경", "추가", "삭제"], width=10)
        status_combo.pack(side="left", padx=5)
        
        search_frame = ttk.Frame(filter_frame)
        search_frame.pack(side="right")
        
        ttk.Label(search_frame, text="검색:").pack(side="left", padx=5)
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var, width=20)
        search_entry.pack(side="left", padx=5)
        
        # 테이블 생성 (트리뷰)
        table_frame = ttk.Frame(main_frame)
        table_frame.pack(fill="both", expand=True, pady=5)
        
        # 컬럼 설정
        columns = ("string_id", 
                "source_kr", "source_cn", "source_tw", 
                "target_kr", "target_cn", "target_tw", 
                "status")
        
        tree = ttk.Treeview(table_frame, columns=columns, show="headings")
        
        # 헤더 설정
        tree.heading("string_id", text="STRING_ID")
        
        # 최신 데이터 컬럼 (A 파일)
        tree.heading("source_kr", text="최신 KR")
        tree.heading("source_cn", text="최신 CN")
        tree.heading("source_tw", text="최신 TW")
        
        # 이전 번역 컬럼 (B 파일)
        tree.heading("target_kr", text="비교 KR")
        tree.heading("target_cn", text="비교 CN")
        tree.heading("target_tw", text="비교 TW")
        
        tree.heading("status", text="상태")
        
        # 컬럼 너비 설정
        tree.column("string_id", width=100)
        tree.column("source_kr", width=150)
        tree.column("source_cn", width=120)
        tree.column("source_tw", width=120)
        tree.column("target_kr", width=150)
        tree.column("target_cn", width=120)
        tree.column("target_tw", width=120)
        tree.column("status", width=80, anchor="center")
        
        # 스크롤바 추가
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side="right", fill="y")
        tree.pack(side="left", fill="both", expand=True)
        
        # 행 색상 설정
        def set_row_colors():
            for item in tree.get_children():
                status = tree.item(item, "values")[-1]
                if status == "변경":
                    tree.item(item, tags=("changed",))
                elif status == "추가":
                    tree.item(item, tags=("added",))
                elif status == "삭제":
                    tree.item(item, tags=("deleted",))
        
        tree.tag_configure("changed", background="#FFFFCC")  # 연한 노랑
        tree.tag_configure("added", background="#CCFFCC")    # 연한 녹색
        tree.tag_configure("deleted", background="#FFCCCC")  # 연한 빨강
        
        # 데이터 로드
        def load_data(filter_status="모두", search_text=""):
            # 기존 항목 삭제
            for item in tree.get_children():
                tree.delete(item)
            
            # 데이터 필터링 및 추가
            for idx, change in enumerate(changes):
                # 상태 필터링
                if filter_status != "모두" and change["status"] != filter_status:
                    continue
                
                # 검색어 필터링
                search_text = search_text.lower()
                if search_text and not (
                    search_text in change["string_id"].lower() or
                    search_text in change["source_kr"].lower() or
                    search_text in change["target_kr"].lower()
                ):
                    continue
                
                # 데이터 추가
                tree.insert("", "end", iid=idx, values=(
                    change["string_id"],
                    change["source_kr"],
                    change["source_cn"],
                    change["source_tw"],
                    change["target_kr"],
                    change["target_cn"],
                    change["target_tw"],
                    change["status"]
                ))
            
            # 행 색상 적용
            set_row_colors()
        
        # 필터링 이벤트 처리
        def apply_filters(*args):
            load_data(status_var.get(), search_var.get())
        
        status_var.trace("w", apply_filters)
        
        def on_search_keypress(event):
            # 엔터 키 눌렀을 때만 검색 적용
            if event.keysym == "Return":
                apply_filters()
        
        search_entry.bind("<KeyPress>", on_search_keypress)
        
        # 검색 버튼
        ttk.Button(search_frame, text="검색", command=apply_filters).pack(side="left")
        
        # 초기 데이터 로드
        load_data()
        
        # 버튼 프레임
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=10)
        
        # 닫기 버튼
        ttk.Button(button_frame, text="닫기", command=preview_window.destroy).pack(side="right", padx=5)
        
        # 엑셀로 내보내기 버튼
        def export_to_excel():
            # 저장할 파일 경로 선택
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel 파일", "*.xlsx")],
                initialfile=f"변경사항_{file_name}_{sheet_name}.xlsx"
            )
            
            if not save_path:
                return
            
            try:
                # 판다스로 데이터프레임 생성
                df = pd.DataFrame(changes)
                # 컬럼명 변경
                df.columns = ["STRING_ID", "최신 KR", "최신 CN", "최신 TW", "비교 KR", "비교 CN", "비교 TW", "상태"]
                # 엑셀로 저장
                df.to_excel(save_path, index=False)
                messagebox.showinfo("완료", f"엑셀 파일로 내보내기 완료: {save_path}")
            except Exception as e:
                messagebox.showerror("오류", f"엑셀 내보내기 실패: {e}")
        
        ttk.Button(button_frame, text="엑셀로 내보내기", command=export_to_excel).pack(side="left", padx=5)
    
    
    # 보조 메서드 추가
    def find_headers(self, sheet):
        """시트에서 헤더 행과 컬럼 인덱스 찾기"""
        for row in range(3, 6):  # 3행부터 5행까지 검색
            headers = {}
            for col in range(1, min(20, sheet.max_column + 1)):  # 최대 20열까지만 검색 (속도 개선)
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value in ["STRING_ID", "KR", "CN", "TW"]:
                    headers[cell_value] = col
            
            # 필수 헤더를 찾았으면 반환
            if "STRING_ID" in headers and "KR" in headers:
                return {"row": row, "columns": headers}
        
        return None

    def sheet_to_dict(self, sheet, headers):
        """시트 데이터를 딕셔너리로 변환"""
        data = {}
        header_row = headers["row"]
        columns = headers["columns"]
        
        # 데이터 행 순회 (최대 1000행까지만 - 속도 개선)
        max_rows = min(sheet.max_row, header_row + 1000)
        
        for row in range(header_row + 1, max_rows):
            string_id = sheet.cell(row=row, column=columns["STRING_ID"]).value
            if not string_id:
                continue
            
            data[string_id] = {}
            for col_name, col_idx in columns.items():
                if col_name != "STRING_ID":
                    value = sheet.cell(row=row, column=col_idx).value
                    data[string_id][col_name] = value if value is not None else ""
        
        return data


# 단독 실행 가능하도록 설정
if __name__ == "__main__":
    root = tk.Tk()
    app = TranslationSyncExtension(root)
    root.mainloop()