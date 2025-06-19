import tkinter as tk
import os
import sqlite3
import json
import re
import time
import gc
import threading
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook, Workbook
import pandas as pd

# 유틸리티 및 분리된 모듈 임포트
from ui.common_components import ScrollableCheckList, show_message
from tools.basic_request_extractor import BasicRequestExtractor
from tools.compare_request_extractor import CompareRequestExtractor
from tools.request_extraction_manager import RequestExtractionManager

# config 유틸리티 함수
def load_config():
    """설정 파일에서 설정 불러오기"""
    try:
        with open("config.json", "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def save_config(config):
    """설정 파일에 설정 저장"""
    try:
        with open("config.json", "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"설정 저장 실패: {e}")

class TranslationRequestExtractor(tk.Frame):
    def __init__(self, root):
        super().__init__(root)
        self.root = root.winfo_toplevel()
        
        # [수정] 핵심 로직 매니저를 여기서 생성합니다.
        self.extraction_manager = RequestExtractionManager(self)
        
        # 공통 UI 변수
        self.folder_path_var = tk.StringVar(value="")
        self.output_db_var = tk.StringVar(value="")
        
        # 공통 데이터
        self.excel_files = []
        self.extraction_thread = None
        
        self.setup_ui()
        self.pack(fill="both", expand=True)

    def setup_ui(self):
        """메인 컨테이너 UI: 공통 부분과 탭으로 구성"""
        # 1. 소스 파일 선택 (공통 영역)
        source_files_frame = ttk.LabelFrame(self, text="1. 추출 대상 파일 선택")
        source_files_frame.pack(fill="x", padx=10, pady=5)
        
        folder_frame = ttk.Frame(source_files_frame)
        folder_frame.pack(fill="x", padx=5, pady=5)
        ttk.Label(folder_frame, text="엑셀 폴더 경로:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(folder_frame, textvariable=self.folder_path_var, width=60).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(folder_frame, text="폴더 찾기", command=self.select_excel_folder).grid(row=0, column=2, padx=5, pady=5)
        ttk.Button(folder_frame, text="파일 검색", command=self.search_excel_files).grid(row=0, column=3, padx=5, pady=5)
        folder_frame.columnconfigure(1, weight=1)
        
        self.excel_files_list = ScrollableCheckList(source_files_frame, width=700, height=120)
        self.excel_files_list.pack(fill="both", expand=True, padx=5, pady=5)

        # 2. 작업 선택 (Notebook으로 분리)
        action_notebook = ttk.Notebook(self)
        action_notebook.pack(fill="both", expand=True, padx=10, pady=5)

        # 탭 1: 기본 추출 (UI 클래스 인스턴스화)
        basic_extract_tab = ttk.Frame(action_notebook)
        self.basic_extractor = BasicRequestExtractor(basic_extract_tab, self)
        self.basic_extractor.pack(fill="both", expand=True)
        action_notebook.add(basic_extract_tab, text="기본 추출")

        # 탭 2: 비교하여 추출 (UI 클래스 인스턴스화)
        compare_extract_tab = ttk.Frame(action_notebook)
        self.compare_extractor = CompareRequestExtractor(compare_extract_tab, self)
        self.compare_extractor.pack(fill="both", expand=True)
        action_notebook.add(compare_extract_tab, text="비교하여 추출")

        # 3. 로그 및 상태 표시줄 (공통 영역)
        log_frame = ttk.LabelFrame(self, text="작업 로그")
        log_frame.pack(fill="both", expand=True, padx=10, pady=5)
        self.log_text = tk.Text(log_frame, wrap="word", height=8)
        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.log_text.pack(fill="both", expand=True)
        
        status_frame = ttk.Frame(self)
        status_frame.pack(fill="x", padx=10, pady=5)
        self.status_label = ttk.Label(status_frame, text="대기 중...")
        self.status_label.pack(side="left", padx=5)

    def select_excel_folder(self):
        folder = filedialog.askdirectory(title="엑셀 파일 폴더 선택", parent=self.root)
        if folder: self.folder_path_var.set(folder)
            
    def select_output_db(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".db", filetypes=[("DB 파일", "*.db")], title="DB 파일 저장", parent=self.root)
        if file_path: self.output_db_var.set(file_path)

    def search_excel_files(self):
        folder = self.folder_path_var.get()
        if not folder or not os.path.isdir(folder):
            show_message(self.root, "warning", "경고", "유효한 폴더를 선택하세요.")
            return
        
        self.excel_files_list.clear()
        self.excel_files = []
        self.log_text.delete(1.0, tk.END)
        self.log_message("파일 검색 중...")
        
        for root, _, files in os.walk(folder):
            for file in files:
                if file.lower().startswith("string") and file.lower().endswith(('.xlsx', '.xls')) and not file.startswith("~$"):
                    self.excel_files.append((file, os.path.join(root, file)))
        
        self.excel_files.sort()
        for file_name, _ in self.excel_files:
            self.excel_files_list.add_item(file_name, checked=True)
        self.log_message(f"{len(self.excel_files)}개의 엑셀 파일을 찾았습니다.")

    def log_message(self, message):
        self.log_text.insert(tk.END, f"{message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def _is_task_running(self):
        if self.extraction_thread and self.extraction_thread.is_alive():
            show_message(self.root, "info", "알림", "이미 다른 작업이 진행 중입니다.")
            return True
        return False

    def _get_selected_files(self):
        selected_file_names = self.excel_files_list.get_checked_items()
        if not selected_file_names:
            show_message(self.root, "warning", "경고", "추출할 파일을 선택하세요.")
            return None
        return [item for item in self.excel_files if item[0] in selected_file_names]

    def _find_headers_in_worksheet(self, ws):
        for i in range(1, 11):
            row_values = [cell.value for cell in ws[i]]
            cleaned_values = [str(v).upper().strip() for v in row_values if v is not None]
            
            # STRING_ID는 필수, #번역요청은 있을 수도 있고 없을 수도 있음
            if "STRING_ID" in cleaned_values:
                header_map = {idx: str(val).strip() for idx, val in enumerate(row_values, 1) if val is not None}
                # 모든 헤더 이름을 소문자로 통일하여 반환
                return {k: v.upper() for k, v in header_map.items()}, i
        return None, -1
        
    def _update_files_as_transferred(self, files_to_update, col_name_to_find, new_value):
        for file_path, sheets in files_to_update.items():
            try:
                wb = load_workbook(file_path)
                for sheet_name, row_indices in sheets.items():
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        headers, _ = self._find_headers_in_worksheet(ws)
                        if not headers: continue
                        
                        col_idx = None
                        for c_idx, c_name in headers.items():
                            if c_name.upper() == col_name_to_find.upper():
                                col_idx = c_idx
                                break
                        
                        if col_idx:
                            for row_idx in row_indices:
                                ws.cell(row=row_idx, column=col_idx).value = new_value
                wb.save(file_path)
                self.log_message(f"'{os.path.basename(file_path)}' 파일 업데이트 완료.")
            except Exception as e:
                self.log_message(f"파일 업데이트 실패 '{os.path.basename(file_path)}': {e}")

    #여기까지 진행

    def _toggle_compare_source(self):
        """비교 대상 소스 타입에 따라 UI 활성화/비활성화"""
        is_excel = self.compare_source_type.get() == "Excel"
        for child in self.excel_compare_frame.winfo_children():
            child.config(state="normal" if is_excel else "disabled")
        for child in self.db_compare_frame.winfo_children():
            child.config(state="disabled" if is_excel else "normal")

    def select_compare_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls"), ("All files", "*.*")])
        if path:
            self.compare_excel_path_var.set(path)
            try:
                wb = load_workbook(path, read_only=True)
                self.compare_sheet_combo['values'] = wb.sheetnames
                if wb.sheetnames:
                    self.compare_excel_sheet_var.set(wb.sheetnames[0])
                wb.close()
            except Exception as e:
                show_message(self.root, "error", "파일 오류", f"시트 목록을 불러올 수 없습니다: {e}")

    def select_compare_db(self):
        path = filedialog.askopenfilename(filetypes=[("DB files", "*.db"), ("All files", "*.*")])
        if path:
            self.compare_db_path_var.set(path)

    # --- 로깅 함수 ---

    def _log_message_thread_safe(self, message):
        self.log_text.insert(tk.END, f"{message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
        
    def _validate_inputs(self, require_db=True, require_compare_source=False):
        if self._is_task_running(): return False

        selected_files = self._get_selected_files()
        if not selected_files: return False

        if require_db:
            if not self.output_db_var.get():
                show_message(self.root, "warning", "경고", "출력 DB 파일 경로를 지정하세요.")
                return False

        if require_compare_source:
            if self.compare_source_type.get() == "Excel":
                if not self.compare_excel_path_var.get() or not self.compare_excel_sheet_var.get():
                    show_message(self.root, "warning", "경고", "비교할 Excel 파일과 시트를 선택하세요.")
                    return False
            else: # DB
                if not self.compare_db_path_var.get():
                    show_message(self.root, "warning", "경고", "비교할 DB 파일을 선택하세요.")
                    return False
        return True

    def extract_translation_requests(self):
        """'기본 추출' 실행 함수"""
        if self._is_task_running(): return
        
        selected_files = self._get_selected_files()
        if not selected_files: return
            
        db_path = self.output_db_var.get()
        if not db_path:
            show_message(self.root, "warning", "경고", "출력 DB 파일 경로를 지정하세요.")
            return

        conditions = []
        if self.extract_new_var.get(): conditions.append("신규")
        if self.extract_change_var.get(): conditions.append("change")
        if not conditions:
            show_message(self.root, "warning", "경고", "추출할 조건을 하나 이상 선택하세요.")
            return

        if os.path.exists(db_path):
            if not show_message(self.root, "yesno", "확인", f"'{os.path.basename(db_path)}' 파일이 이미 존재합니다. 덮어쓰시겠습니까?"):
                return
        
        self.log_text.delete(1.0, tk.END)
        self.log_message("기본 추출 작업을 시작합니다...")
        
        self.extraction_thread = threading.Thread(
            target=self._extract_in_background,
            args=(selected_files, db_path, conditions, self.mark_as_transferred_var.get())
        )
        self.extraction_thread.daemon = True
        self.extraction_thread.start()
    
    def _extract_in_background(self, selected_files, db_path, conditions, mark_as_transferred):
        """백그라운드에서 기본 추출 작업 수행"""
        loading_popup = LoadingPopup(self.root, "추출 중", "데이터 추출 준비 중...")
        
        extracted_data = []
        files_to_update = {} # {file_path: {sheet_name: [row_index, ...]}}

        try:
            if os.path.exists(db_path):
                os.remove(db_path)
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            cursor.execute('''
            CREATE TABLE translation_requests (
                id INTEGER PRIMARY KEY, file_name TEXT, sheet_name TEXT,
                string_id TEXT, kr TEXT, en TEXT, cn TEXT, tw TEXT, th TEXT,
                request_type TEXT, additional_info TEXT
            )''')

            total_files = len(selected_files)
            for i, (file_name, file_path) in enumerate(selected_files):
                loading_popup.update_message(f"파일 처리 중 ({i+1}/{total_files}): {file_name}")
                self.log_message(f"파일 처리 시작: {file_name}")

                wb = load_workbook(file_path, read_only=True, data_only=True)
                for sheet_name in wb.sheetnames:
                    if not sheet_name.lower().startswith("string"): continue

                    ws = wb[sheet_name]
                    headers, header_row_idx = self._find_headers_in_dataframe(ws)
                    
                    if not headers or "#번역요청" not in headers.values():
                        self.log_message(f"  '{sheet_name}' 시트: '#번역요청' 컬럼을 찾을 수 없습니다.")
                        continue
                    
                    req_col_name = [k for k, v in headers.items() if v == "#번역요청"][0]
                    
                    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 2, values_only=True), start=header_row_idx + 2):
                        request_type = row[req_col_name - 1]
                        if request_type and str(request_type).lower() in conditions:
                            string_id = row[headers.get('string_id', 1) - 1] if 'string_id' in headers else ""
                            kr = row[headers.get('kr', 1) - 1] if 'kr' in headers else ""
                            en = row[headers.get('en', 1) - 1] if 'en' in headers else ""
                            cn = row[headers.get('cn', 1) - 1] if 'cn' in headers else ""
                            tw = row[headers.get('tw', 1) - 1] if 'tw' in headers else ""
                            th = row[headers.get('th', 1) - 1] if 'th' in headers else ""
                            
                            extracted_data.append((file_name, sheet_name, string_id, kr, en, cn, tw, th, request_type, ""))

                            if mark_as_transferred:
                                if file_path not in files_to_update:
                                    files_to_update[file_path] = {}
                                if sheet_name not in files_to_update[file_path]:
                                    files_to_update[file_path][sheet_name] = []
                                files_to_update[file_path][sheet_name].append(row_idx)

                wb.close()

            if extracted_data:
                cursor.executemany("INSERT INTO translation_requests (file_name, sheet_name, string_id, kr, en, cn, tw, th, request_type, additional_info) VALUES (?,?,?,?,?,?,?,?,?,?)", extracted_data)
                conn.commit()
            conn.close()

            if mark_as_transferred and files_to_update:
                loading_popup.update_message("원본 파일에 '전달' 표시 중...")
                self.log_message("추출된 항목에 '전달' 표시를 시작합니다...")
                req_col_header = [k for k, v in headers.items() if v == "#번역요청"][0]
                self._update_files_as_transferred(files_to_update, req_col_header, "전달")

            loading_popup.close()
            show_message(self.root, "info", "완료", f"기본 추출이 완료되었습니다. 총 {len(extracted_data)}개 항목이 추출되었습니다.")
            self.log_message(f"기본 추출 완료. 총 {len(extracted_data)}개 항목 추출.")
        except Exception as e:
            loading_popup.close()
            show_message(self.root, "error", "오류", f"추출 중 오류 발생: {e}")
            self.log_message(f"오류 발생: {e}")

    def _find_headers_in_dataframe(self, ws):
        """Worksheet에서 헤더 행과 컬럼 위치 찾기 (openpyxl 용)"""
        for i in range(1, 11): # 첫 10행 스캔
            row_values = [cell.value for cell in ws[i]]
            cleaned_values = [str(v).lower().strip() for v in row_values if v is not None]
            
            if "#번역요청" in cleaned_values and "string_id" in cleaned_values:
                header_map = {}
                for idx, val in enumerate(row_values, 1):
                    if val is not None:
                         header_map[idx] = str(val).strip()
                return header_map, i
        return None, -1

    def _update_files_as_transferred(self, files_to_update, col_name_to_find, new_value):
        """추출된 항목을 엑셀 파일에서 '전달'로 표시"""
        for file_path, sheets in files_to_update.items():
            try:
                wb = load_workbook(file_path)
                for sheet_name, row_indices in sheets.items():
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        headers, header_row_idx = self._find_headers_in_dataframe(ws)
                        
                        if not headers: continue
                        
                        col_idx = None
                        for c_idx, c_name in headers.items():
                            if c_name == col_name_to_find:
                                col_idx = c_idx
                                break
                        
                        if col_idx:
                            for row_idx in row_indices:
                                ws.cell(row=row_idx, column=col_idx).value = new_value
                wb.save(file_path)
                self.log_message(f"'{os.path.basename(file_path)}' 파일 업데이트 완료.")
            except Exception as e:
                self.log_message(f"파일 업데이트 실패 '{os.path.basename(file_path)}': {e}")
                   
    def run_compare_extract(self):
        """'비교하여 추출' 실행 함수"""
        if self._is_task_running(): return
        
        selected_files = self._get_selected_files()
        if not selected_files: return

        db_path = self.output_db_var.get()
        if not db_path:
            show_message(self.root, "warning", "경고", "출력 DB 파일 경로를 지정하세요.")
            return

        if not self.compare_extract_new_var.get() and not self.compare_extract_modified_var.get():
            show_message(self.root, "warning", "경고", "추출 조건(신규 또는 변경)을 선택하세요.")
            return

        if os.path.exists(db_path):
            if not show_message(self.root, "yesno", "확인", f"'{os.path.basename(db_path)}' 파일이 이미 존재합니다. 덮어쓰시겠습니까?"):
                return

        self.log_text.delete(1.0, tk.END)
        self.log_message("비교 추출 작업을 시작합니다...")

        self.extraction_thread = threading.Thread(
            target=self._compare_extract_in_background,
            args=(selected_files, db_path)
        )
        self.extraction_thread.daemon = True
        self.extraction_thread.start()

    def _load_comparison_data(self):
        """비교 대상 데이터를 캐시로 로드"""
        cache = {}
        source_type = self.compare_source_type.get()
        
        if source_type == "Excel":
            excel_path = self.compare_excel_path_var.get()
            sheet_name = self.compare_excel_sheet_var.get()
            if not excel_path or not sheet_name:
                raise ValueError("비교할 엑셀 파일과 시트를 선택해야 합니다.")
            
            self.log_message(f"비교 데이터 로딩(Excel): {os.path.basename(excel_path)} - {sheet_name}")
            wb = load_workbook(excel_path, read_only=True, data_only=True)
            ws = wb[sheet_name]
            headers, header_row_idx = self._find_headers_in_dataframe(ws)
            if not headers or 'string_id' not in [v.lower() for v in headers.values()]:
                wb.close()
                raise ValueError("비교 엑셀에서 STRING_ID 컬럼을 찾을 수 없습니다.")

            string_id_col = [k for k, v in headers.items() if v.lower() == 'string_id'][0]
            kr_col = [k for k, v in headers.items() if v.lower() == 'kr'][0] if 'kr' in [v.lower() for v in headers.values()] else None

            for row in ws.iter_rows(min_row=header_row_idx + 2, values_only=True):
                string_id = row[string_id_col - 1]
                if string_id:
                    kr_val = row[kr_col - 1] if kr_col else ""
                    cache[str(string_id)] = {'kr': str(kr_val) if kr_val is not None else ""}
            wb.close()

        elif source_type == "DB":
            db_path = self.compare_db_path_var.get()
            if not db_path:
                raise ValueError("비교할 DB 파일을 선택해야 합니다.")
                
            self.log_message(f"비교 데이터 로딩(DB): {os.path.basename(db_path)}")
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            # translation_data 테이블이 있는지 먼저 확인
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='translation_data'")
            if cursor.fetchone():
                for row in cursor.execute("SELECT string_id, kr FROM translation_data"):
                    if row[0]:
                        cache[str(row[0])] = {'kr': str(row[1]) if row[1] is not None else ""}
            conn.close()

        self.log_message(f"비교 데이터 로딩 완료: {len(cache)}개 항목")
        return cache

    def _compare_and_extract_in_background(self, selected_files, db_path, mark_as_transferred):
        loading_popup = LoadingPopup(self.root, "비교 추출 중", "준비 중...")
        extracted_data_for_db = []
        files_to_update = {}
        
        try:
            loading_popup.update_message("비교 데이터 로드 중...")
            compare_cache = self._load_comparison_data()
            if compare_cache is None:
                loading_popup.close()
                return

            if os.path.exists(db_path): os.remove(db_path)
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            cursor.execute('''CREATE TABLE IF NOT EXISTS translation_requests (
                              file_name TEXT, sheet_name TEXT, string_id TEXT, kr TEXT, en TEXT, 
                              cn TEXT, tw TEXT, th TEXT, request_type TEXT)''')
                              
            extract_new = self.compare_extract_new_var.get()
            extract_modified = self.compare_extract_modified_var.get()
            langs_to_compare = [lang for lang, var in self.compare_lang_vars.items() if var.get()]

            total_files = len(selected_files)
            for i, (file_name, file_path) in enumerate(selected_files):
                loading_popup.update_message(f"파일 비교 중 ({i+1}/{total_files}): {file_name}")
                try:
                    df_map = pd.read_excel(file_path, sheet_name=None, header=None, engine='openpyxl')
                    for sheet_name, df in df_map.items():
                        if not sheet_name.lower().startswith("string"): continue
                        headers, header_row_idx = self._find_headers_in_dataframe(df)
                        if not headers or "string_id" not in headers: continue
                        
                        df.columns = df.iloc[header_row_idx]
                        df_data = df.iloc[header_row_idx + 1:].reset_index(drop=True)
                        
                        for idx, row in df_data.iterrows():
                            string_id = str(row.get(headers["string_id"], "")).strip()
                            if not string_id: continue
                            
                            cached_item = compare_cache.get(string_id)
                            
                            if cached_item is None: # 신규 항목
                                if extract_new:
                                    data_tuple = self._prepare_data_for_db(headers, row, "신규")
                                    extracted_data_for_db.append(data_tuple)
                                    # ... '전달' 표시 정보 기록 ...
                            else: # 변경 항목 검사
                                if extract_modified:
                                    is_modified = False
                                    for lang in langs_to_compare:
                                        lang_lower = lang.lower()
                                        source_val = str(row.get(headers.get(lang, ''), '')).strip()
                                        cached_val = str(cached_item.get(lang_lower, '')).strip()
                                        if source_val != cached_val:
                                            is_modified = True
                                            break
                                    if is_modified:
                                        data_tuple = self._prepare_data_for_db(headers, row, "변경")
                                        extracted_data_for_db.append(data_tuple)
                                        # ... '전달' 표시 정보 기록 ...
                except Exception as e:
                    self.log_message(f"파일 비교 오류 '{file_name}': {e}")

            if extracted_data_for_db:
                cursor.executemany('INSERT INTO translation_requests VALUES (?,?,?,?,?,?,?,?,?)', extracted_data_for_db)
                conn.commit()

            # ... '전달' 표시 로직 ...
            
            loading_popup.close()
            show_message(self.root, "info", "완료", f"비교 추출 완료! 총 {len(extracted_data_for_db)}개 항목을 추출했습니다.")
        except Exception as e:
            loading_popup.close()
            show_message(self.root, "error", "오류", f"비교 추출 중 오류 발생: {e}")
        finally:
            if 'conn' in locals(): conn.close()
   
    # --- 헬퍼 함수들 ---
    def _is_task_running(self):
        if self.extraction_thread and self.extraction_thread.is_alive():
            show_message(self.root, "info", "알림", "이미 다른 작업이 진행 중입니다.")
            return True
        return False

    def _get_selected_files(self):
        selected_file_names = self.excel_files_list.get_checked_items()
        if not selected_file_names:
            show_message(self.root, "warning", "경고", "추출할 파일을 선택하세요.")
            return None
        return [item for item in self.excel_files if item[0] in selected_file_names]

    def _find_headers_in_dataframe(self, df):
        """DataFrame에서 헤더 행과 컬럼 위치 찾기 (정규화된 이름과 원본 이름 모두 반환)"""
        for i, row in df.head(10).iterrows():
            row_values = {str(v).lower().strip(): str(v) for v in row if pd.notna(v)}
            if "#번역요청" in row_values and "string_id" in row_values:
                return row_values, i # {정규화된 이름: 원본 이름} 맵 반환
        return None, -1
        
    def _prepare_data_for_db(self, headers, row, request_type):
        """DataFrame 행을 DB에 저장할 튜플 형태로 변환"""
        get_val = lambda key: row.get(headers.get(key, 'non_existing_col'))
        return (
            get_val("file_name"), get_val("sheet_name"), get_val("string_id"),
            get_val("kr"), get_val("en"), get_val("cn"), get_val("tw"), get_val("th"),
            request_type
        )

    def _update_files_as_transferred(self, files_to_update, col_name_to_find, new_value):
        """추출된 항목을 엑셀 파일에서 '전달'로 표시"""
        total_files = len(files_to_update)
        for i, (file_path, sheets) in enumerate(files_to_update.items()):
            self.log_message(f"'전달' 표시 중 ({i+1}/{total_files}): {os.path.basename(file_path)}")
            try:
                wb = load_workbook(file_path)
                for sheet_name, row_indices in sheets.items():
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        col_idx = None
                        for r in range(1, 11): # 헤더는 상위 10행 안에 있다고 가정
                            for c in range(1, ws.max_column + 1):
                                if str(ws.cell(row=r, column=c).value).strip().lower() == col_name_to_find.lower():
                                    col_idx = c
                                    break
                            if col_idx: break
                        
                        if col_idx:
                            for row_idx in row_indices:
                                ws.cell(row=row_idx, column=col_idx).value = new_value
                wb.save(file_path)
            except Exception as e:
                self.log_message(f"파일 업데이트 실패 '{os.path.basename(file_path)}': {e}")
                      
    def find_translation_request_column(self, worksheet):
        """#번역요청 컬럼 찾기 (공백 무시) - ExcelUtils 활용"""
        # ExcelFileManager에 적절한 메서드가 없으므로 커스텀 함수 유지
        # 대신 로깅 추가하여 개선
        logger.debug(f"#번역요청 컬럼 검색 시작")
        
        for row in range(2, 6):  # 2행부터 5행까지 검색
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    # 공백 제거 후 비교
                    if re.sub(r'\s+', '', cell_value).upper() == "#번역요청":
                        logger.debug(f"#번역요청 컬럼 발견: 열={col}, 행={row}")
                        return col, row
        
        logger.debug(f"#번역요청 컬럼을 찾지 못함")
        return None, None
    
    def find_string_id_position(self, worksheet):
        """STRING_ID 컬럼 위치 찾기"""
        for row in range(2, 6):  # 2행부터 5행까지 검색
            for col in range(1, min(10, worksheet.max_column + 1)):  # 최대 10개 컬럼까지만 검색
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str) and "STRING_ID" in cell_value.upper():
                    return col, row
        return None, None
    
    def reset_column_cache(self):
        """컬럼 캐시 초기화"""
        if messagebox.askyesno("확인", "컬럼 캐시를 초기화하시겠습니까?", parent=self.root):
            self.column_cache = {}
            self.save_column_cache()
            self.log_text.insert(tk.END, "컬럼 캐시를 초기화했습니다.\n")
    
    def _find_headers_in_dataframe(self, df):
        """DataFrame에서 헤더 행과 컬럼 위치 찾기"""
        for i, row in df.head(10).iterrows():
            row_values = [str(v).lower().strip() for v in row if pd.notna(v)]
            if "#번역요청" in row_values and "string_id" in row_values:
                # 헤더 맵 생성 {정규화된 이름: 원본 이름}
                header_map = {str(v).lower().strip(): str(v) for v in row if pd.notna(v)}
                return header_map, i
        return None, -1
 
    def _setup_extraction_thread(self, selected_files, db_path):
        """추출 작업용 백그라운드 스레드 설정"""
        # 진행 상태 초기화
        self.log_text.delete(1.0, tk.END)
        self.progress_bar["value"] = 0
        self.status_label.config(text="번역 요청 추출 시작...")
        
        # 취소 플래그 초기화
        self.cancel_extraction = False
        
        # 작업 스레드 생성 및 시작
        self.extraction_thread = threading.Thread(
            target=self._extract_in_background,
            args=(selected_files, db_path)
        )
        self.extraction_thread.daemon = True
        self.extraction_thread.start()
        
        # 진행 상태 모니터링
        self.root.after(100, self._check_extraction_progress)

    def _check_extraction_progress(self):
        """추출 작업 진행 상태 확인"""
        if self.extraction_thread and self.extraction_thread.is_alive():
            # 작업이 진행 중이면 상태 확인 계속
            self.root.after(100, self._check_extraction_progress)
        else:
            # 작업 완료 또는 실패
            if hasattr(self, 'extraction_thread') and self.extraction_thread:
                self.extraction_thread = None

    def _get_column_info(self, file_name, sheet_name, worksheet, extract_conditions, update_ui):
        """컬럼 정보 가져오기 (캐시 활용)"""
        condition_key = '-'.join([k for k, v in extract_conditions.items() if v])
        cache_key = f"{file_name}|{sheet_name}|{condition_key}"
        
        if cache_key in self.column_cache:
            update_ui(None, None, f"    캐시에서 컬럼 정보를 찾았습니다.\n")
            return self.column_cache[cache_key]
        
        # STRING_ID 위치 찾기
        string_id_col, header_row = self.find_string_id_position(worksheet)
        if not string_id_col or not header_row:
            return None
        
        # #번역요청 컬럼 찾기
        request_col, _ = self.find_translation_request_column(worksheet)
        if not request_col:
            return None
        
        # 필요한 언어 컬럼 찾기
        lang_cols = {}
        check_langs = ["KR"]  # KR은 항상 포함
        
        # 빈 칸 체크 조건에 따라 필요한 언어 추가
        if extract_conditions.get('en_empty'):
            check_langs.append("EN")
        if extract_conditions.get('cn_empty'):
            check_langs.append("CN")
        if extract_conditions.get('tw_empty'):
            check_langs.append("TW")
        
        # 모든 언어 컬럼을 찾음 (DB 저장용)
        all_langs = ["KR", "EN", "CN", "TW", "TH"]
        for lang in all_langs:
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=header_row, column=col).value
                if cell_value == lang:
                    lang_cols[lang] = col
                    break
        
        # 컬럼 정보 캐싱
        columns_info = {
            "string_id_col": string_id_col,
            "header_row": header_row,
            "request_col": request_col,
            "lang_cols": lang_cols,
            "extract_conditions": extract_conditions
        }
        self.column_cache[cache_key] = columns_info
        
        # 캐시 저장
        def save_cache():
            self.save_column_cache()
        self.root.after(0, save_cache)
        
        return columns_info

    def _extract_sheet_data(self, worksheet, columns_info, file_name, sheet_name, 
                        batch_data, cursor, conn, batch_size, use_read_only, extract_conditions, update_ui):
        """시트에서 데이터 추출"""
        string_id_col = columns_info["string_id_col"]
        header_row = columns_info["header_row"]
        request_col = columns_info["request_col"]
        lang_cols = columns_info["lang_cols"]
        
        sheet_rows = 0
        
        if use_read_only:
            # 읽기 전용 모드
            row_idx = 0
            for row in worksheet.iter_rows(min_row=header_row + 1):
                row_idx += 1
                if row_idx % 1000 == 0:
                    update_ui(None, None, f"    {row_idx}행 처리 중...\n", False)
                
                # 추출 조건 확인 및 데이터 처리
                if self._should_extract_row_readonly(row, request_col, lang_cols, extract_conditions, string_id_col):
                    string_id = row[string_id_col - 1].value if string_id_col - 1 < len(row) else None
                    if string_id:
                        lang_values = self._extract_lang_values_readonly(row, lang_cols)
                        
                        batch_data.append((
                            file_name, sheet_name, string_id,
                            lang_values.get("kr"), lang_values.get("en"), 
                            lang_values.get("cn"), lang_values.get("tw"), 
                            lang_values.get("th"), "번역요청", None
                        ))
                        sheet_rows += 1
                        
                        # 배치 처리
                        if len(batch_data) >= batch_size:
                            self._insert_batch_data(cursor, conn, batch_data)
                            batch_data.clear()
        else:
            # 일반 모드
            for row_num in range(header_row + 1, worksheet.max_row + 1):
                if row_num % 1000 == 0:
                    update_ui(None, None, f"    {row_num}행 처리 중...\n", False)
                
                # 추출 조건 확인 및 데이터 처리
                if self._should_extract_row_normal(worksheet, row_num, request_col, lang_cols, extract_conditions, string_id_col):
                    string_id = worksheet.cell(row=row_num, column=string_id_col).value
                    if string_id:
                        lang_values = self._extract_lang_values_normal(worksheet, row_num, lang_cols)
                        
                        batch_data.append((
                            file_name, sheet_name, string_id,
                            lang_values.get("kr"), lang_values.get("en"), 
                            lang_values.get("cn"), lang_values.get("tw"), 
                            lang_values.get("th"), "번역요청", None
                        ))
                        sheet_rows += 1
                        
                        # 배치 처리
                        if len(batch_data) >= batch_size:
                            self._insert_batch_data(cursor, conn, batch_data)
                            batch_data.clear()
        
        return sheet_rows

    def _should_extract_row_readonly(self, row, request_col, lang_cols, extract_conditions, string_id_col):
        """읽기 전용 모드에서 행 추출 조건 확인 (OR 조건)"""
        # KR 값 필터링 - 제외 조건 확인
        kr_col = lang_cols.get('KR')
        if kr_col:
            kr_value = row[kr_col - 1].value if kr_col - 1 < len(row) else None
            if self._should_exclude_kr_value(kr_value):
                return False
            
        conditions_met = []
        
        # 1. #번역요청이 "신규"인 조건 확인
        if extract_conditions.get('new_request'):
            request_value = row[request_col - 1].value if request_col - 1 < len(row) else None
            is_new_request = request_value and str(request_value).strip() == "신규"
            conditions_met.append(is_new_request)
        
        # 2. #번역요청이 "change"인 조건 확인
        if extract_conditions.get('change_request'):
            request_value = row[request_col - 1].value if request_col - 1 < len(row) else None
            is_change_request = request_value and str(request_value).strip() == "Change"
            conditions_met.append(is_change_request)
        
        # 3. 선택된 언어별 빈 칸 조건 확인
        lang_mapping = {'en_empty': 'EN', 'cn_empty': 'CN', 'tw_empty': 'TW'}
        
        for condition_key, lang in lang_mapping.items():
            if extract_conditions.get(condition_key):
                lang_col = lang_cols.get(lang)
                if lang_col:
                    lang_value = row[lang_col - 1].value if lang_col - 1 < len(row) else None
                    is_empty = not lang_value or str(lang_value).strip() == ""
                    conditions_met.append(is_empty)
        
        # OR 조건: 하나라도 만족하면 추출
        return any(conditions_met)

    def _should_extract_row_normal(self, worksheet, row_num, request_col, lang_cols, extract_conditions, string_id_col):
        """일반 모드에서 행 추출 조건 확인 (OR 조건)"""
        kr_col = lang_cols.get('KR')
        if kr_col:
            kr_value = worksheet.cell(row=row_num, column=kr_col).value
            if self._should_exclude_kr_value(kr_value):
                return False
            
        conditions_met = []
        
        # 1. #번역요청이 "신규"인 조건 확인
        if extract_conditions.get('new_request'):
            request_value = worksheet.cell(row=row_num, column=request_col).value
            is_new_request = request_value and str(request_value).strip() == "신규"
            conditions_met.append(is_new_request)
        
        # 2. #번역요청이 "change"인 조건 확인
        if extract_conditions.get('change_request'):
            request_value = worksheet.cell(row=row_num, column=request_col).value
            is_change_request = request_value and str(request_value).strip() == "Change"
            conditions_met.append(is_change_request)
        
        # 3. 선택된 언어별 빈 칸 조건 확인
        lang_mapping = {'en_empty': 'EN', 'cn_empty': 'CN', 'tw_empty': 'TW'}
        
        for condition_key, lang in lang_mapping.items():
            if extract_conditions.get(condition_key):
                lang_col = lang_cols.get(lang)
                if lang_col:
                    lang_value = worksheet.cell(row=row_num, column=lang_col).value
                    is_empty = not lang_value or str(lang_value).strip() == ""
                    conditions_met.append(is_empty)
        
        # OR 조건: 하나라도 만족하면 추출
        return any(conditions_met)

    def _should_exclude_kr_value(self, kr_value):
        """KR 값이 제외 조건에 해당하는지 확인"""
        if not kr_value:
            # 빈 값인 경우 제외
            return True
        
        kr_str = str(kr_value).strip()
        if not kr_str:
            # 공백만 있는 경우 제외
            return True
        
        # #으로 시작하는 경우 제외
        if kr_str.startswith('#'):
            return True
        
        # [@...]으로만 구성된 경우 제외 (정규식 사용)
        import re
        if re.match(r'^\[@[^\]]*\]$', kr_str):
            return True
        
        return False

    def _extract_lang_values_readonly(self, row, lang_cols):
        """읽기 전용 모드에서 언어 값 추출"""
        lang_values = {}
        # KR은 항상 추출, 나머지는 컬럼이 있는 경우만 추출
        for lang, col in lang_cols.items():
            if col - 1 < len(row):
                lang_values[lang.lower()] = row[col - 1].value
            else:
                lang_values[lang.lower()] = None
        
        # 누락된 언어는 None으로 설정
        for lang in ['kr', 'en', 'cn', 'tw', 'th']:
            if lang not in lang_values:
                lang_values[lang] = None
        
        return lang_values

    def _extract_lang_values_normal(self, worksheet, row_num, lang_cols):
        """일반 모드에서 언어 값 추출"""
        lang_values = {}
        # KR은 항상 추출, 나머지는 컬럼이 있는 경우만 추출
        for lang, col in lang_cols.items():
            lang_values[lang.lower()] = worksheet.cell(row=row_num, column=col).value
        
        # 누락된 언어는 None으로 설정
        for lang in ['kr', 'en', 'cn', 'tw', 'th']:
            if lang not in lang_values:
                lang_values[lang] = None
        
        return lang_values

    def _insert_batch_data(self, cursor, conn, batch_data):
        """배치 데이터를 DB에 삽입"""
        cursor.executemany('''
        INSERT INTO translation_requests 
        (file_name, sheet_name, string_id, kr, en, cn, tw, th, request_type, additional_info)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', batch_data)
        conn.commit()

    def _finalize_database(self, cursor, conn, update_ui):
        """DB 인덱스 생성 및 최적화"""
        update_ui(None, "인덱스 생성 중...", "\n인덱스 생성 중...\n")
        cursor.execute('CREATE INDEX idx_string_id ON translation_requests(string_id)')
        cursor.execute('CREATE INDEX idx_file_sheet ON translation_requests(file_name, sheet_name)')
        
        update_ui(None, "DB 최적화 중...", "DB 최적화 중...\n")
        cursor.execute("PRAGMA optimize")
        conn.commit()
        conn.close()

    def save_column_cache(self):
        """컬럼 캐시 저장 - FileUtils 활용"""
        try:
            FileUtils.save_json("column_cache.json", self.column_cache)
            logger.debug("컬럼 캐시 저장 완료")
        except Exception as e:
            logger.error(f"컬럼 캐시 저장 실패: {e}")
            self.log_text.insert(tk.END, f"컬럼 캐시 저장 실패: {str(e)}\n")

    def load_column_cache(self):
        """컬럼 캐시 로드 - FileUtils 활용"""
        try:
            self.column_cache = FileUtils.load_json("column_cache.json", {})
            logger.debug(f"컬럼 캐시 로드 완료: {len(self.column_cache)} 항목")
        except Exception as e:
            logger.warning(f"컬럼 캐시 로드 실패: {e}")
            self.column_cache = {}
    
    def export_to_excel(self):
        """DB를 Excel 파일로 내보내기"""
        db_path = self.output_db_var.get()
        if not db_path or not os.path.isfile(db_path):
            messagebox.showwarning("경고", "유효한 DB 파일을 선택하세요.", parent=self.root)
            return
        
        # 저장할 엑셀 파일 경로 선택
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            title="번역 요청 내보내기",
            parent=self.root
        )
        
        if not file_path:
            return
        
        self.log_text.insert(tk.END, "엑셀로 내보내기 작업 시작...\n")
        self.status_label.config(text="엑셀로 내보내기 중...")
        self.root.update()
        
        try:
            # DB 연결
            conn = sqlite3.connect(db_path)
            
            # 데이터 및 컬럼명 조회
            cursor = conn.cursor()
            cursor.execute("PRAGMA table_info(translation_requests)")
            columns = [column[1] for column in cursor.fetchall()]
            
            # 데이터 조회
            cursor.execute("SELECT * FROM translation_requests")
            rows = cursor.fetchall()
            
            if not rows:
                messagebox.showinfo("알림", "내보낼 데이터가 없습니다.", parent=self.root)
                conn.close()
                return
            
            # 판다스 데이터프레임으로 변환 (컬럼명 지정)
            df = pd.DataFrame(rows, columns=columns)
            
            # 엑셀 작성기 생성
            writer = pd.ExcelWriter(file_path, engine='openpyxl')
            
            # "String" 시트에 데이터 저장
            df.to_excel(writer, sheet_name="String", index=False)
            
            # 저장 및 닫기
            writer.close()
            
            conn.close()
            
            self.log_text.insert(tk.END, f"엑셀로 내보내기 완료: {len(rows)}개 항목\n")
            self.status_label.config(text=f"내보내기 완료 - {len(rows)}개 항목")
            
            messagebox.showinfo("완료", f"번역 요청 데이터를 엑셀로 내보냈습니다.\n파일: {file_path}", parent=self.root)
            
        except Exception as e:
            self.log_text.insert(tk.END, f"내보내기 중 오류 발생: {str(e)}\n")
            messagebox.showerror("오류", f"엑셀로 내보내기 실패: {str(e)}", parent=self.root)

    def export_to_template_excel(self):
        """템플릿 기반 Excel 내보내기"""
        db_path = self.output_db_var.get()
        if not db_path or not os.path.isfile(db_path):
            messagebox.showwarning("경고", "유효한 DB 파일을 선택하세요.", parent=self.root)
            return
        
        # 템플릿 Excel 파일 선택
        template_path = filedialog.askopenfilename(
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")],
            title="템플릿 Excel 파일 선택",
            parent=self.root
        )
        
        if not template_path:
            return
        
        # 출력할 Excel 파일 경로 선택
        output_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            title="번역 요청 템플릿 내보내기",
            parent=self.root
        )
        
        if not output_path:
            return
        
        self.log_text.insert(tk.END, "템플릿 기반 엑셀 내보내기 작업 시작...\n")
        self.status_label.config(text="템플릿 기반 내보내기 중...")
        self.root.update()
        
        try:
            # DB 연결 및 데이터 조회
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            # 데이터 조회 (시트별로 그룹화)
            cursor.execute("""
            SELECT file_name, sheet_name, string_id, kr, en, cn, tw, th, request_type
            FROM translation_requests 
            ORDER BY file_name, sheet_name, string_id
            """)
            rows = cursor.fetchall()
            conn.close()
            
            if not rows:
                messagebox.showinfo("알림", "내보낼 데이터가 없습니다.", parent=self.root)
                return
            
            # 데이터를 시트별로 그룹화
            sheets_data = {}
            for row in rows:
                file_name, sheet_name, string_id, kr, en, cn, tw, th, request_type = row
                
                if sheet_name not in sheets_data:
                    sheets_data[sheet_name] = []
                
                sheets_data[sheet_name].append({
                    'file_name': file_name,
                    'string_id': string_id,
                    'kr': kr,
                    'en': en,
                    'cn': cn,
                    'tw': tw,
                    'th': th,
                    'request_type': request_type
                })
            
            # 템플릿 기반 Excel 생성
            success_count = self._create_template_based_excel(template_path, output_path, sheets_data)
            
            self.log_text.insert(tk.END, f"템플릿 기반 내보내기 완료: {success_count}개 시트 생성\n")
            self.status_label.config(text=f"템플릿 내보내기 완료 - {success_count}개 시트")
            
            messagebox.showinfo("완료", 
                            f"템플릿 기반 Excel 내보내기가 완료되었습니다.\n"
                            f"생성된 시트: {success_count}개\n"
                            f"파일: {output_path}", parent=self.root)
            
        except Exception as e:
            self.log_text.insert(tk.END, f"템플릿 내보내기 중 오류 발생: {str(e)}\n")
            messagebox.showerror("오류", f"템플릿 기반 내보내기 실패: {str(e)}", parent=self.root)

    def _create_template_based_excel(self, template_path, output_path, sheets_data):
        """템플릿 기반 Excel 파일 생성"""
        try:
            # 템플릿 파일 로드
            template_workbook = load_workbook(template_path)
            
            # 새 워크북 생성
            output_workbook = Workbook()
            
            # 기본 시트 제거
            if 'Sheet' in output_workbook.sheetnames:
                output_workbook.remove(output_workbook['Sheet'])
            
            success_count = 0
            
            # 각 시트별로 처리
            for sheet_name, sheet_data in sheets_data.items():
                self.log_text.insert(tk.END, f"  시트 '{sheet_name}' 생성 중... ({len(sheet_data)}개 항목)\n")
                self.root.update()
                
                # 템플릿에서 해당 시트를 찾거나 첫 번째 시트를 템플릿으로 사용
                template_sheet = None
                if sheet_name in template_workbook.sheetnames:
                    template_sheet = template_workbook[sheet_name]
                else:
                    # String으로 시작하는 시트 찾기
                    string_sheets = [s for s in template_workbook.sheetnames 
                                if s.lower().startswith('string')]
                    if string_sheets:
                        template_sheet = template_workbook[string_sheets[0]]
                    else:
                        template_sheet = template_workbook.active
                
                # 새 시트 생성
                new_sheet = output_workbook.create_sheet(title=sheet_name)
                
                # 템플릿 구조 복사 및 데이터 삽입
                if self._copy_template_and_insert_data(template_sheet, new_sheet, sheet_data):
                    success_count += 1
                else:
                    self.log_text.insert(tk.END, f"    시트 '{sheet_name}' 처리 실패\n")
            
            # 템플릿 워크북 닫기
            template_workbook.close()
            
            # 결과 파일 저장
            output_workbook.save(output_path)
            output_workbook.close()
            
            return success_count
            
        except Exception as e:
            self.log_text.insert(tk.END, f"템플릿 Excel 생성 중 오류: {str(e)}\n")
            return 0

    def _copy_template_and_insert_data(self, template_sheet, new_sheet, sheet_data):
        """템플릿 구조 복사 및 데이터 삽입"""
        try:
            # 1. 템플릿 구조 복사 (값과 기본 스타일만)
            max_copy_rows = min(10, template_sheet.max_row)  # 최대 10행까지만 복사
            
            for row in range(1, max_copy_rows + 1):
                for col in range(1, template_sheet.max_column + 1):
                    source_cell = template_sheet.cell(row=row, column=col)
                    target_cell = new_sheet.cell(row=row, column=col)
                    
                    # 값만 복사 (스타일 복사 제거)
                    target_cell.value = source_cell.value
            
            # 컬럼 너비 복사
            try:
                for col in range(1, template_sheet.max_column + 1):
                    col_letter = new_sheet.cell(row=1, column=col).column_letter
                    template_col_letter = template_sheet.cell(row=1, column=col).column_letter
                    if template_col_letter in template_sheet.column_dimensions:
                        width = template_sheet.column_dimensions[template_col_letter].width
                        if width:
                            new_sheet.column_dimensions[col_letter].width = width
            except Exception as e:
                # 컬럼 너비 복사 실패해도 계속 진행
                pass
            
            # 2. 컬럼 위치 찾기
            columns_info = self._find_template_columns(template_sheet)
            if not columns_info:
                self.log_text.insert(tk.END, f"    템플릿에서 필요한 컬럼을 찾을 수 없습니다.\n")
                return False
            
            # 3. 데이터 삽입
            return self._insert_data_to_template(new_sheet, columns_info, sheet_data)
            
        except Exception as e:
            self.log_text.insert(tk.END, f"    템플릿 복사 중 오류: {str(e)}\n")
            return False

    def _find_template_columns(self, template_sheet):
        """템플릿에서 컬럼 위치 찾기"""
        columns_info = {}
        
        # 헤더 행을 찾기 위해 상위 10행 검색
        for row in range(1, min(11, template_sheet.max_row + 1)):
            found_columns = {}
            
            for col in range(1, template_sheet.max_column + 1):
                cell_value = template_sheet.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    cell_value_clean = cell_value.strip().upper()
                    
                    # 컬럼 매핑
                    if cell_value_clean == "STRING_ID":
                        found_columns["string_id"] = col
                    elif cell_value_clean == "KR":
                        found_columns["kr"] = col
                    elif cell_value_clean == "EN":
                        found_columns["en"] = col
                    elif cell_value_clean == "CN":
                        found_columns["cn"] = col
                    elif cell_value_clean == "TW":
                        found_columns["tw"] = col
                    elif cell_value_clean == "TH":
                        found_columns["th"] = col
                    elif "#번역요청" in cell_value or "번역요청" in cell_value:
                        found_columns["request"] = col
            
            # STRING_ID와 KR 컬럼이 모두 있으면 헤더 행으로 판단
            if "string_id" in found_columns and "kr" in found_columns:
                columns_info = found_columns
                columns_info["header_row"] = row
                break
        
        return columns_info if "string_id" in columns_info else None

    def _insert_data_to_template(self, worksheet, columns_info, sheet_data):
        """템플릿에 데이터 삽입"""
        try:
            header_row = columns_info["header_row"]
            data_start_row = header_row + 1
            
            # 기존 데이터가 있다면 그 다음 행부터 시작
            current_row = data_start_row
            while worksheet.cell(row=current_row, column=columns_info["string_id"]).value:
                current_row += 1
            
            # 데이터 삽입
            inserted_count = 0
            for data_item in sheet_data:
                # STRING_ID 삽입
                worksheet.cell(row=current_row, column=columns_info["string_id"]).value = data_item["string_id"]
                
                # KR 삽입
                worksheet.cell(row=current_row, column=columns_info["kr"]).value = data_item["kr"]
                
                # 다른 언어 컬럼들 삽입 (있는 경우만)
                lang_mapping = {
                    "en": data_item.get("en"),
                    "cn": data_item.get("cn"),
                    "tw": data_item.get("tw"),
                    "th": data_item.get("th")
                }
                
                for lang, value in lang_mapping.items():
                    if lang in columns_info and value:
                        worksheet.cell(row=current_row, column=columns_info[lang]).value = value
                
                # 번역요청 컬럼에 "신규" 표시 (있는 경우만)
                if "request" in columns_info:
                    worksheet.cell(row=current_row, column=columns_info["request"]).value = "신규"
                
                current_row += 1
                inserted_count += 1
            
            self.log_text.insert(tk.END, f"    {inserted_count}개 항목 삽입 완료\n")
            return True
            
        except Exception as e:
            self.log_text.insert(tk.END, f"    데이터 삽입 중 오류: {str(e)}\n")
            return False

def main():
    root = tk.Tk()
    app = TranslationRequestExtractor(root)
    root.mainloop()


if __name__ == "__main__":
    main()