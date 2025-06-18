# tools/translate/string_sync_manager.py
import tkinter as tk
import json  # 이 줄 추가
import os
import threading
import sqlite3
import time
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd

from ui.common_components import ScrollableCheckList, LoadingPopup
from tools.db_compare_manager import DBCompareManager

class StringSyncManager(tk.Frame):
    def __init__(self, parent, root):
        super().__init__(parent)
        self.root = root
        self.db_compare_manager = DBCompareManager(root)
        self.compare_results = []
        self.db_pairs = []
        self.filtered_results = []  # 필터링된 결과 저장
        self.exception_rules = []
        self.compiled_rules = []  # 이 줄 추가
        self.setup_ui()
        self.load_exception_rules()  # rules 파일명 변경 가능
        
    def setup_ui(self):
        """개선된 UI 구성 (탭 구조)"""
        menubar = tk.Menu(self.root)
        tools_menu = tk.Menu(menubar, tearoff=0)
        tools_menu.add_command(label="예외 규칙 관리", command=self.show_exception_rules_manager)
        menubar.add_cascade(label="도구", menu=tools_menu)
        self.root.config(menu=menubar)

        # 메인 탭 프레임
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=5, pady=5)
        
        # 탭 1: DB 비교 설정
        self.setup_compare_tab()
        
        # 탭 2: 결과 및 작업
        self.setup_result_tab()
        
        # 하단 상태 표시줄 (공통)
        self.setup_status_bar()

    def setup_compare_tab(self):
        """탭 1: DB 비교 설정"""
        compare_frame = ttk.Frame(self.notebook)
        self.notebook.add(compare_frame, text="DB 비교 설정")
        
        # 상단 프레임 (좌우 분할)
        top_frame = ttk.Frame(compare_frame)
        top_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # 좌측 프레임 (DB 선택) - 높이 제한
        left_frame = ttk.Frame(top_frame)
        left_frame.pack(side="left", fill="both", expand=True, padx=(0, 5))
        
        # 개별 DB 선택 (컴팩트하게)
        single_db_frame = ttk.LabelFrame(left_frame, text="개별 DB 파일 비교")
        single_db_frame.pack(fill="x", padx=5, pady=5)
        
        # 원본 DB
        ttk.Label(single_db_frame, text="원본:").grid(row=0, column=0, padx=5, pady=3, sticky="w")
        self.original_db_var = tk.StringVar()
        ttk.Entry(single_db_frame, textvariable=self.original_db_var, width=35).grid(row=0, column=1, padx=5, pady=3, sticky="ew")
        ttk.Button(single_db_frame, text="찾기", 
                command=lambda: self.select_db_file("original")).grid(row=0, column=2, padx=5, pady=3)
        
        # 비교 DB
        ttk.Label(single_db_frame, text="비교:").grid(row=1, column=0, padx=5, pady=3, sticky="w")
        self.compare_db_var = tk.StringVar()
        ttk.Entry(single_db_frame, textvariable=self.compare_db_var, width=35).grid(row=1, column=1, padx=5, pady=3, sticky="ew")
        ttk.Button(single_db_frame, text="찾기", 
                command=lambda: self.select_db_file("compare")).grid(row=1, column=2, padx=5, pady=3)
        
        single_db_frame.columnconfigure(1, weight=1)
        
        # 폴더 DB 비교 (컴팩트하게)
        folder_db_frame = ttk.LabelFrame(left_frame, text="폴더 내 DB 비교")
        folder_db_frame.pack(fill="x", padx=5, pady=5)
        
        # 원본 폴더
        ttk.Label(folder_db_frame, text="원본:").grid(row=0, column=0, padx=5, pady=3, sticky="w")
        self.original_folder_db_var = tk.StringVar()
        ttk.Entry(folder_db_frame, textvariable=self.original_folder_db_var, width=35).grid(row=0, column=1, padx=5, pady=3, sticky="ew")
        ttk.Button(folder_db_frame, text="찾기", 
                command=lambda: self.select_db_folder("original")).grid(row=0, column=2, padx=5, pady=3)
        
        # 비교 폴더
        ttk.Label(folder_db_frame, text="비교:").grid(row=1, column=0, padx=5, pady=3, sticky="w")
        self.compare_folder_db_var = tk.StringVar()
        ttk.Entry(folder_db_frame, textvariable=self.compare_folder_db_var, width=35).grid(row=1, column=1, padx=5, pady=3, sticky="ew")
        ttk.Button(folder_db_frame, text="찾기", 
                command=lambda: self.select_db_folder("compare")).grid(row=1, column=2, padx=5, pady=3)
        
        # DB 목록 보기 버튼
        ttk.Button(folder_db_frame, text="DB 목록 보기", 
                command=self.show_db_list).grid(row=2, column=1, padx=5, pady=3, sticky="e")
        
        folder_db_frame.columnconfigure(1, weight=1)
        
        # 우측 프레임 (DB 목록 + 옵션) - 높이 제한
        right_frame = ttk.Frame(top_frame)
        right_frame.pack(side="right", fill="both", expand=True, padx=(5, 0))
        
        # DB 목록 (높이 제한)
        self.db_list_frame = ttk.LabelFrame(right_frame, text="비교할 DB 목록")
        self.db_list_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        self.db_checklist = ScrollableCheckList(self.db_list_frame, width=300, height=120)  # 높이 제한
        self.db_checklist.pack(fill="both", expand=True, padx=5, pady=5)
        
        # 비교 옵션 (컴팩트하게)
        options_frame = ttk.LabelFrame(right_frame, text="비교 옵션")
        options_frame.pack(fill="x", padx=5, pady=5)

        self.changed_kr_var = tk.BooleanVar(value=True)
        self.new_items_var = tk.BooleanVar(value=True)
        self.deleted_items_var = tk.BooleanVar(value=True)
        self.apply_exception_rules_var = tk.BooleanVar(value=True)

        # 옵션들을 2열로 배치
        ttk.Checkbutton(options_frame, text="KR 변경", variable=self.changed_kr_var).grid(row=0, column=0, padx=5, pady=2, sticky="w")
        ttk.Checkbutton(options_frame, text="신규 항목", variable=self.new_items_var).grid(row=0, column=1, padx=5, pady=2, sticky="w")
        ttk.Checkbutton(options_frame, text="삭제 항목", variable=self.deleted_items_var).grid(row=1, column=0, padx=5, pady=2, sticky="w")
        ttk.Checkbutton(options_frame, text="예외 규칙", variable=self.apply_exception_rules_var).grid(row=1, column=1, padx=5, pady=2, sticky="w")
        
        # 비교 실행 버튼
        button_frame = ttk.Frame(compare_frame)
        button_frame.pack(fill="x", padx=5, pady=5)
        
        ttk.Button(button_frame, text="단일 DB 비교", 
                command=self.compare_databases).pack(side="right", padx=5)
        ttk.Button(button_frame, text="폴더 내 전체 비교", 
                command=self.compare_all_databases).pack(side="right", padx=5)
        
        # 로그 영역 (높이 제한)
        log_frame = ttk.LabelFrame(compare_frame, text="작업 로그")
        log_frame.pack(fill="x", padx=5, pady=5)
        
        log_container = ttk.Frame(log_frame)
        log_container.pack(fill="x", padx=5, pady=5)
        
        self.log_text = tk.Text(log_container, wrap="word", height=6)  # 높이 고정
        log_scrollbar = ttk.Scrollbar(log_container, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_scrollbar.set)
        
        log_scrollbar.pack(side="right", fill="y")
        self.log_text.pack(fill="x", expand=True)

    def setup_result_tab(self):
        """탭 2: 결과 및 작업"""
        result_frame = ttk.Frame(self.notebook)
        self.notebook.add(result_frame, text="결과 및 작업")
        
        # 상단: 필터링 및 통계 (높이 고정)
        top_result_frame = ttk.Frame(result_frame)
        top_result_frame.pack(fill="x", padx=5, pady=5)
        
        # 필터링 프레임 (한 줄로 컴팩트하게)
        filter_frame = ttk.LabelFrame(top_result_frame, text="결과 필터링")
        filter_frame.pack(fill="x", padx=5, pady=5)
        
        filter_row = ttk.Frame(filter_frame)
        filter_row.pack(fill="x", padx=5, pady=5)
        
        ttk.Label(filter_row, text="상태:").pack(side="left", padx=5)
        self.filter_var = tk.StringVar(value="전체")
        filter_combo = ttk.Combobox(filter_row, textvariable=self.filter_var, 
                                values=["전체", "신규", "변경됨", "삭제됨"], width=10, state="readonly")
        filter_combo.pack(side="left", padx=5)
        filter_combo.bind("<<ComboboxSelected>>", self.apply_filter)
        
        # 통계 레이블
        self.stats_label = ttk.Label(filter_row, text="비교 결과가 없습니다.")
        self.stats_label.pack(side="left", padx=20)
        
        # STRING 동기화 액션 (한 줄로 컴팩트하게)
        sync_frame = ttk.LabelFrame(top_result_frame, text="STRING 동기화 작업")
        sync_frame.pack(fill="x", padx=5, pady=5)
        
        # 옵션
        option_row = ttk.Frame(sync_frame)
        option_row.pack(fill="x", padx=5, pady=2)
        
        self.clear_translations_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(option_row, text="CN, TW 값 삭제", 
                    variable=self.clear_translations_var).pack(side="left", padx=5)
        
        # 액션 버튼들 (한 줄로)
        button_row = ttk.Frame(sync_frame)
        button_row.pack(fill="x", padx=5, pady=2)
        
        ttk.Button(button_row, text="신규→비교본", command=self.apply_new_strings).pack(side="left", padx=2)
        ttk.Button(button_row, text="신규→@_new", command=self.apply_new_strings_to_new).pack(side="left", padx=2)
        ttk.Button(button_row, text="수정→비교본", command=self.apply_modified_strings).pack(side="left", padx=2)
        ttk.Button(button_row, text="삭제→비교본", command=self.apply_deleted_strings).pack(side="left", padx=2)
        ttk.Button(button_row, text="삭제→원본", command=self.apply_reverse_new_strings).pack(side="left", padx=2)
        
        # 중간: 결과 표시 영역 (확장 가능)
        result_display_frame = ttk.LabelFrame(result_frame, text="비교 결과")
        result_display_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # 상단 버튼들
        result_button_frame = ttk.Frame(result_display_frame)
        result_button_frame.pack(fill="x", padx=5, pady=2)
        
        ttk.Button(result_button_frame, text="엑셀로 내보내기", 
                command=self.export_to_excel).pack(side="right", padx=5)
        # 디버깅 버튼 (개발 중에만 사용)
        ttk.Button(result_button_frame, text="데이터 디버깅", 
                command=self.debug_current_data).pack(side="left", padx=5)
        
        # 트리뷰 및 스크롤바
        tree_frame = ttk.Frame(result_display_frame)
        tree_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # 컬럼 순서 명확히 정의
        columns = ("file_name", "sheet_name", "type", "string_id", "kr", "original_kr")
        self.result_tree = ttk.Treeview(tree_frame, columns=columns, show="headings")
        
        # 컬럼 헤딩 설정
        # 컬럼 헤딩 설정 (더 명확하게)
        self.result_tree.heading("file_name", text="파일명")
        self.result_tree.heading("sheet_name", text="시트명")  
        self.result_tree.heading("type", text="유형")
        self.result_tree.heading("string_id", text="STRING_ID")
        self.result_tree.heading("kr", text="비교본 KR")        # 더 명확한 헤더
        self.result_tree.heading("original_kr", text="원본 KR")  # 더 명확한 헤더

        # 컬럼 너비 설정
        self.result_tree.column("file_name", width=100)
        self.result_tree.column("sheet_name", width=100)
        self.result_tree.column("type", width=60)
        self.result_tree.column("string_id", width=100)
        self.result_tree.column("kr", width=300)
        self.result_tree.column("original_kr", width=300)
        
        # 스크롤바
        scrollbar_y = ttk.Scrollbar(tree_frame, orient="vertical", command=self.result_tree.yview)
        self.result_tree.configure(yscrollcommand=scrollbar_y.set)
        
        scrollbar_x = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.result_tree.xview)
        self.result_tree.configure(xscrollcommand=scrollbar_x.set)
        
        # 배치
        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x") 
        self.result_tree.pack(fill="both", expand=True)

    def setup_status_bar(self):
        """하단 상태 표시줄"""
        status_frame = ttk.Frame(self)
        status_frame.pack(fill="x", padx=5, pady=5)
        
        self.status_label = ttk.Label(status_frame, text="대기 중...")
        self.status_label.pack(side="left", fill="x", expand=True, padx=5)
        
        self.progress_bar = ttk.Progressbar(status_frame, length=300, mode="determinate")
        self.progress_bar.pack(side="right", padx=5)

        
    def load_exception_rules(self, path="string_exception_rules.json"):
        """예외 규칙 로드 및 최적화"""
        try:
            with open(path, "r", encoding="utf-8") as f:
                rules = json.load(f)
                
            # 규칙 전처리 및 최적화
            self.exception_rules = []
            self.compiled_rules = []  # 컴파일된 규칙들
            
            for rule in rules:
                if not rule.get("enabled", True):
                    continue
                    
                # 필드명 정규화
                field = rule.get("field", "").lower()
                if not field:
                    continue
                    
                rule_type = rule.get("type", "")
                value = rule.get("value", "")
                
                # 문자열 이스케이프 미리 처리
                if rule_type in ["startswith", "endswith", "contains", "equals"]:
                    processed_value = str(value).replace("\\n", "\n").replace("\\t", "\t")
                else:
                    processed_value = value
                
                compiled_rule = {
                    "field": field,
                    "type": rule_type,
                    "value": processed_value,
                    "original_value": value,
                    "description": rule.get("description", "")
                }
                
                # 정규식 미리 컴파일
                if rule_type == "regex":
                    try:
                        import re
                        compiled_rule["compiled_pattern"] = re.compile(processed_value)
                    except Exception as e:
                        self.log_message(f"정규식 컴파일 실패: {e} (패턴: {processed_value})")
                        continue
                
                # length 타입 값 미리 변환
                elif rule_type == "length":
                    try:
                        compiled_rule["threshold"] = int(processed_value)
                    except (ValueError, TypeError):
                        self.log_message(f"length 규칙 값 변환 실패: {processed_value}")
                        continue
                
                self.compiled_rules.append(compiled_rule)
                self.exception_rules.append(rule)  # 원본도 보존
                
            self.log_message(f"예외 규칙 로드 및 최적화 완료: {len(self.compiled_rules)}개")
            
        except FileNotFoundError:
            self.exception_rules = []
            self.compiled_rules = []
            self.log_message("예외 규칙 파일이 없습니다.")
        except Exception as e:
            self.exception_rules = []
            self.compiled_rules = []
            self.log_message(f"예외 규칙 로딩 오류: {e}")


    def filter_by_exception_rules(self, results):
        """최적화된 예외 규칙 필터링"""
        if not self.compiled_rules or not results:
            return results
        
        # 필드명 매핑 캐시 생성 (첫 번째 항목 기준)
        field_mapping = {}
        first_item = results[0]
        for actual_field in first_item.keys():
            field_mapping[actual_field.lower()] = actual_field
        
        # 규칙별 함수 생성 (클로저 활용)
        rule_functions = []
        for rule in self.compiled_rules:
            field = rule["field"]
            rule_type = rule["type"]
            
            # 실제 필드명 찾기
            actual_field = field_mapping.get(field)
            if not actual_field:
                continue
            
            # 규칙 타입별 최적화된 함수 생성
            if rule_type == "startswith":
                value = rule["value"]
                rule_functions.append(lambda item, f=actual_field, v=value: 
                                    str(item.get(f, "")).startswith(v))
            
            elif rule_type == "endswith":
                value = rule["value"]
                rule_functions.append(lambda item, f=actual_field, v=value: 
                                    str(item.get(f, "")).endswith(v))
            
            elif rule_type == "contains":
                value = rule["value"]
                rule_functions.append(lambda item, f=actual_field, v=value: 
                                    v in str(item.get(f, "")))
            
            elif rule_type == "equals":
                value = rule["value"]
                rule_functions.append(lambda item, f=actual_field, v=value: 
                                    str(item.get(f, "")) == v)
            
            elif rule_type == "length":
                threshold = rule["threshold"]
                rule_functions.append(lambda item, f=actual_field, t=threshold: 
                                    len(str(item.get(f, ""))) > t)
            
            elif rule_type == "regex":
                pattern = rule["compiled_pattern"]
                rule_functions.append(lambda item, f=actual_field, p=pattern: 
                                    bool(p.search(str(item.get(f, "")))))
        
        # 실제 필터링 (최적화된 방식)
        filtered = []
        excluded_count = 0
        
        for item in results:
            excluded = False
            
            # 모든 규칙 함수 실행 (조기 종료)
            for rule_func in rule_functions:
                try:
                    if rule_func(item):
                        excluded = True
                        excluded_count += 1
                        break
                except Exception:
                    continue  # 오류 시 무시하고 계속
            
            if not excluded:
                filtered.append(item)
        
        self.log_message(f"예외 규칙 적용: {excluded_count}개 제외, {len(filtered)}개 남음")
        return filtered

    def filter_by_exception_rules_debug(self, results):
        """디버깅용 상세 로그가 있는 예외 규칙 필터링"""
        import re
        
        self.log_message("=== 디버깅 모드 예외 규칙 적용 ===")
        self.log_message(f"입력 데이터: {len(results)}개")
        self.log_message(f"적용할 규칙: {len(self.exception_rules)}개")
        
        if results:
            self.log_message(f"첫 번째 데이터: {results[0]}")
        
        filtered = []
        excluded_count = 0
        
        for idx, item in enumerate(results):
            excluded = False
            
            for rule_idx, rule in enumerate(self.exception_rules):
                if not rule.get("enabled", True):
                    continue
                    
                field = rule.get("field", "").lower()
                value = rule.get("value", "")
                rule_type = rule.get("type", "")
                
                # 필드 찾기
                field_val = None
                for key in item.keys():
                    if key.lower() == field:
                        field_val = str(item[key]) if item[key] is not None else ""
                        break
                
                if field_val is None:
                    continue
                
                self.log_message(f"항목 {idx+1}, 규칙 {rule_idx+1}: {rule_type}({field}='{field_val[:30]}...', '{value}')")
                
                # 규칙 적용
                match = False
                try:
                    if rule_type == "startswith":
                        check_value = value.replace("\\n", "\n").replace("\\t", "\t")
                        match = field_val.startswith(check_value)
                    elif rule_type == "endswith":
                        check_value = value.replace("\\n", "\n").replace("\\t", "\t")
                        match = field_val.endswith(check_value)
                    elif rule_type == "contains":
                        check_value = value.replace("\\n", "\n").replace("\\t", "\t")
                        match = check_value in field_val
                    elif rule_type == "equals":
                        check_value = value.replace("\\n", "\n").replace("\\t", "\t")
                        match = field_val == check_value
                    elif rule_type == "length":
                        threshold = int(value) if isinstance(value, str) else value
                        match = len(field_val) > threshold
                    elif rule_type == "regex":
                        match = bool(re.search(value, field_val))
                except Exception as e:
                    self.log_message(f"규칙 적용 오류: {e}")
                    continue
                
                if match:
                    excluded = True
                    excluded_count += 1
                    self.log_message(f"*** 제외: {rule.get('description', '')}")
                    break
            
            if not excluded:
                filtered.append(item)
        
        self.log_message(f"디버깅 모드 완료: {excluded_count}개 제외, {len(filtered)}개 남음")
        return filtered


    def log_message(self, message):
        """로그 메시지 추가"""
        self.log_text.insert(tk.END, f"{time.strftime('%H:%M:%S')} - {message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()


    def apply_filter(self, event=None):
        """결과 필터링 적용 (타입별 필터링만)"""
        filter_value = self.filter_var.get()
        if filter_value == "전체":
            results = self.compare_results.copy()
        else:
            results = [item for item in self.compare_results if item["type"] == filter_value]
        
        # 더 이상 여기서 예외 규칙을 적용하지 않음 (이미 적용됨)
        self.filtered_results = results

        self.update_result_display()
        self.update_statistics()



    def update_result_display(self):
        """결과 표시 업데이트 (올바른 컬럼 의미 적용)"""
        # 기존 항목 삭제
        self.result_tree.delete(*self.result_tree.get_children())
        
        # 결과 탭으로 자동 전환
        self.notebook.select(1)  # 두 번째 탭(결과 탭)으로 전환
        
        self.log_message(f"UI 업데이트: {len(self.filtered_results)}개 항목 표시")
        
        # 컬럼 순서: file_name, sheet_name, type, string_id, kr, original_kr
        for idx, item in enumerate(self.filtered_results):
            item_type = item.get("type", "")
            kr_value = item.get("kr", "")           # 비교본의 KR 값
            original_kr_value = item.get("original_kr", "")  # 원본의 KR 값
            
            # 올바른 표시 로직:
            # KR 컬럼 = 항상 비교본의 KR 값 (kr)
            # 원본 컬럼 = 항상 원본의 KR 값 (original_kr)
            
            display_kr = kr_value           # 비교본의 KR 값
            display_original_kr = original_kr_value  # 원본의 KR 값
            
            values_to_insert = (
                item.get("file_name", ""),     # 0
                item.get("sheet_name", ""),    # 1  
                item_type,                     # 2
                item.get("string_id", ""),     # 3
                display_kr,                    # 4 - 비교본 KR
                display_original_kr            # 5 - 원본 KR
            )
            
            if idx < 3:  # 처음 3개 항목만 로그
                self.log_message(f"항목 {idx+1}: 유형={item_type}, 비교본KR='{kr_value}', 원본KR='{original_kr_value}'")
            
            self.result_tree.insert("", "end", iid=idx, values=values_to_insert)
        
        self.log_message("UI 업데이트 완료")


        
    # 간단한 테스트 메서드 (임시로 추가해서 테스트)
    def test_insert_dummy_data(self):
        """더미 데이터로 트리뷰 테스트"""
        self.result_tree.delete(*self.result_tree.get_children())
        
        # 테스트 데이터
        test_data = [
            ("String.db", "String", "신규", "test_id_1", "테스트 한글 텍스트 1", ""),
            ("String.db", "String", "변경됨", "test_id_2", "테스트 한글 텍스트 2", "원본 텍스트"),
            ("String.db", "String", "삭제됨", "test_id_3", "테스트 한글 텍스트 3", "")
        ]
        
        for idx, values in enumerate(test_data):
            self.result_tree.insert("", "end", iid=idx, values=values)
        
        self.log_message("더미 데이터 삽입 완료 - KR 컬럼이 보이는지 확인하세요")
        

    def update_statistics(self):
        """통계 정보 업데이트"""
        if not self.compare_results:
            self.stats_label.config(text="비교 결과가 없습니다.")
            return
        
        # 전체 통계
        new_count = len([item for item in self.compare_results if item["type"] == "신규"])
        modified_count = len([item for item in self.compare_results if item["type"] == "변경됨"])
        deleted_count = len([item for item in self.compare_results if item["type"] == "삭제됨"])
        
        # 현재 필터 통계
        filter_value = self.filter_var.get()
        current_count = len(self.filtered_results)
        
        if filter_value == "전체":
            stats_text = f"전체: {len(self.compare_results)}개 (신규: {new_count}, 수정: {modified_count}, 삭제: {deleted_count})"
        else:
            stats_text = f"{filter_value}: {current_count}개 | 전체: {len(self.compare_results)}개"
        
        self.stats_label.config(text=stats_text)

    def select_db_file(self, db_type):
        """DB 파일 선택"""
        file_path = filedialog.askopenfilename(
            filetypes=[("DB 파일", "*.db"), ("모든 파일", "*.*")],
            title=f"{db_type.capitalize()} DB 파일 선택",
            parent=self.root
        )
        if file_path:
            if db_type == "original":
                self.original_db_var.set(file_path)
            else:
                self.compare_db_var.set(file_path)
            
            self.root.after(100, self.root.focus_force)
            self.root.after(100, self.root.lift)

    def select_db_folder(self, folder_type):
        """DB 폴더 선택"""
        folder = filedialog.askdirectory(title=f"{folder_type.capitalize()} DB 폴더 선택", parent=self.root)
        if folder:
            if folder_type == "original":
                self.original_folder_db_var.set(folder)
            else:
                self.compare_folder_db_var.set(folder)
            
            self.root.after(100, self.root.focus_force)
            self.root.after(100, self.root.lift)

    def show_db_list(self):
        """폴더 내 DB 파일 목록 표시"""
        original_folder = self.original_folder_db_var.get()
        compare_folder = self.compare_folder_db_var.get()
        
        if not original_folder or not os.path.isdir(original_folder):
            messagebox.showwarning("경고", "유효한 원본 DB 폴더를 선택하세요.", parent=self.root)
            return
        
        if not compare_folder or not os.path.isdir(compare_folder):
            messagebox.showwarning("경고", "유효한 비교 DB 폴더를 선택하세요.", parent=self.root)
            return
        
        # DB 파일 목록 가져오기
        original_dbs = {f for f in os.listdir(original_folder) if f.endswith('.db') and f.startswith('String')}
        compare_dbs = {f for f in os.listdir(compare_folder) if f.endswith('.db') and f.startswith('String')}
        
        # 공통 DB 파일만 찾기
        common_dbs = original_dbs.intersection(compare_dbs)
        
        if not common_dbs:
            messagebox.showinfo("알림", "두 폴더에 공통된 String DB 파일이 없습니다.", parent=self.root)
            return
        
        # DB 목록 업데이트
        self.db_checklist.clear()
        self.db_pairs = []
        
        for db_file in sorted(common_dbs):
            self.db_checklist.add_item(db_file, checked=True)
            self.db_pairs.append({
                'file_name': db_file,
                'original_path': os.path.join(original_folder, db_file),
                'compare_path': os.path.join(compare_folder, db_file)
            })
        
        messagebox.showinfo("알림", f"{len(common_dbs)}개의 공통 String DB 파일을 찾았습니다.", parent=self.root)

    def compare_databases(self):
        """단일 DB 파일 비교"""
        original_db_path = self.original_db_var.get()
        compare_db_path = self.compare_db_var.get()

        if not original_db_path or not os.path.isfile(original_db_path):
            messagebox.showwarning("경고", "유효한 원본 DB 파일을 선택하세요.")
            return

        if not compare_db_path or not os.path.isfile(compare_db_path):
            messagebox.showwarning("경고", "유효한 비교 DB 파일을 선택하세요.")
            return
            
        # 단일 DB 쌍을 생성하여 비교
        db_pair = {
            'file_name': os.path.basename(original_db_path),
            'original_path': original_db_path,
            'compare_path': compare_db_path
        }
        
        self._run_comparison([db_pair])

    def compare_all_databases(self):    
        """선택된 모든 DB 파일을 비교"""
        if not self.db_pairs:
            messagebox.showwarning("경고", "비교할 DB 파일 목록이 없습니다. 'DB 목록 보기'를 먼저 실행하세요.")
            return
        
        # 선택된 DB 파일 쌍 필터링
        selected_db_names = self.db_checklist.get_checked_items()
        selected_db_pairs = [pair for pair in self.db_pairs if pair['file_name'] in selected_db_names]
        
        if not selected_db_pairs:
            messagebox.showwarning("경고", "비교할 DB 파일을 선택하세요.")
            return
        
        self._run_comparison(selected_db_pairs)

    def _run_comparison(self, db_pairs):
        """DB 비교 실행"""
        # 로그 초기화
        self.log_text.delete(1.0, tk.END)
        self.log_message("DB 비교 작업 시작")
        
        # 진행 창 생성
        loading_popup = LoadingPopup(self.root, "DB 비교 중", "DB 비교 작업 준비 중...")
        
        # 작업용 스레드 함수
        def run_comparison():
            try:
                # DBCompareManager 사용하여 비교 실행
                def progress_callback(status_text, current, total):
                    self.root.after(0, lambda: loading_popup.update_progress(
                        (current / total) * 100, status_text))
                    
                result = self.db_compare_manager.compare_all_databases(
                    db_pairs, 
                    self.changed_kr_var.get(), 
                    self.new_items_var.get(), 
                    self.deleted_items_var.get(),
                    progress_callback
                )
                
                # 결과 처리 (메인 스레드에서)
                self.root.after(0, lambda: self._process_compare_results(result, loading_popup))
                
            except Exception as e:
                self.root.after(0, lambda: [
                    loading_popup.close(),
                    messagebox.showerror("오류", f"DB 비교 중 오류 발생: {str(e)}")
                ])

        # 백그라운드 스레드 실행
        thread = threading.Thread(target=run_comparison)
        thread.daemon = True
        thread.start()


    def _process_compare_results(self, result, loading_popup):
        """DB 비교 결과 처리"""
        if result["status"] != "success":
            loading_popup.close()
            messagebox.showerror("오류", result["message"])
            return
            
        # 결과 저장 (타입 매핑 제거 - DBCompareManager가 이미 올바른 타입 반환)
        original_results = result["compare_results"]
        
        # === 예외 규칙 적용 (여기가 핵심!) ===
        if self.apply_exception_rules_var.get():
            self.log_message("DB 비교 결과에 예외 규칙 적용 중...")
            filtered_results = self.filter_by_exception_rules(original_results)
            self.compare_results = filtered_results
            self.log_message(f"예외 규칙 적용 완료: {len(original_results)}개 → {len(filtered_results)}개")
        else:
            self.compare_results = original_results
            self.log_message("예외 규칙 적용하지 않음")
        
        self.filtered_results = self.compare_results.copy()
        
        # === 디버깅 로그 추가 ===
        self.log_message(f"최종 비교 결과: {len(self.compare_results)}개")
        
        # 타입별 개수 확인
        type_counts = {}
        for item in self.compare_results:
            item_type = item.get("type", "알 수 없음")
            type_counts[item_type] = type_counts.get(item_type, 0) + 1
        
        for item_type, count in type_counts.items():
            self.log_message(f"  {item_type}: {count}개")
        
        # 신규 항목 상세 로그
        new_items = [item for item in self.compare_results if item.get("type") == "신규"]
        if new_items:
            self.log_message(f"신규 항목 상세:")
            for i, item in enumerate(new_items[:5]):  # 최대 5개만 로그
                self.log_message(f"  {i+1}. {item.get('string_id', 'N/A')}: {item.get('kr', 'N/A')}")
            if len(new_items) > 5:
                self.log_message(f"  ... 외 {len(new_items) - 5}개")
        else:
            self.log_message("신규 항목이 없습니다.")
        
        # 삭제된 항목 상세 로그
        deleted_items = [item for item in self.compare_results if item.get("type") == "삭제됨"]
        if deleted_items:
            self.log_message(f"삭제된 항목 상세:")
            for i, item in enumerate(deleted_items[:5]):  # 최대 5개만 로그
                self.log_message(f"  {i+1}. {item.get('string_id', 'N/A')}: {item.get('kr', 'N/A')}")
            if len(deleted_items) > 5:
                self.log_message(f"  ... 외 {len(deleted_items) - 5}개")
        else:
            self.log_message("삭제된 항목이 없습니다.")
        
        # 비교 옵션 상태 로그
        self.log_message(f"비교 옵션 - 변경됨: {self.changed_kr_var.get()}, 신규: {self.new_items_var.get()}, 삭제됨: {self.deleted_items_var.get()}")
        # === 디버깅 로그 끝 ===
        
        # 결과 표시 업데이트
        self.update_result_display()
        self.update_statistics()
        
        # 상태 업데이트
        success_count = result["success_count"]
        total_count = result["total_count"]
        total_changes = len(self.compare_results)  # 예외 규칙 적용 후 개수로 변경
        
        self.status_label.config(
            text=f"{success_count}/{total_count}개 DB 비교 완료, {total_changes}개의 변경사항 발견"
        )
        
        self.log_message(f"DB 비교 완료: {success_count}/{total_count}개 파일, {total_changes}개 변경사항")
        
        loading_popup.close()
        
        if total_changes > 0:
            messagebox.showinfo(
                "완료", 
                f"{success_count}개의 DB를 비교했습니다.\n"
                f"총 {total_changes}개의 변경사항을 찾았습니다.", 
                parent=self.root
            )
        else:
            messagebox.showinfo("완료", "변경사항이 없습니다.", parent=self.root)
            

    def apply_new_strings(self):
        """신규 STRING 적용"""
        new_items = [item for item in self.compare_results if item["type"] == "신규"]
        if not new_items:
            messagebox.showinfo("알림", "적용할 신규 항목이 없습니다.", parent=self.root)
            return
        
        if not messagebox.askyesno("확인", f"{len(new_items)}개의 신규 항목을 적용하시겠습니까?", parent=self.root):
            return
        
        self._apply_strings(new_items, "신규", use_new_file=False)

    def apply_new_strings_to_new(self):
        """신규 STRING 적용 (@_new 파일)"""
        new_items = [item for item in self.compare_results if item["type"] == "신규"]
        if not new_items:
            messagebox.showinfo("알림", "적용할 신규 항목이 없습니다.", parent=self.root)
            return
        
        if not messagebox.askyesno("확인", f"{len(new_items)}개의 신규 항목을 @_new 파일에 적용하시겠습니까?", parent=self.root):
            return
        
        self._apply_strings(new_items, "신규 (@_new)", use_new_file=True)

    def apply_modified_strings(self):
        """수정된 STRING 적용"""
        modified_items = [item for item in self.compare_results if item["type"] == "변경됨"]
        if not modified_items:
            messagebox.showinfo("알림", "적용할 수정 항목이 없습니다.", parent=self.root)
            return
        
        if not messagebox.askyesno("확인", f"{len(modified_items)}개의 수정 항목을 적용하시겠습니까?", parent=self.root):
            return
        
        self._apply_strings(modified_items, "수정", use_new_file=False)

    def apply_deleted_strings(self):
        """삭제된 STRING 적용"""
        deleted_items = [item for item in self.compare_results if item["type"] == "삭제됨"]
        if not deleted_items:
            messagebox.showinfo("알림", "적용할 삭제 항목이 없습니다.", parent=self.root)
            return
        
        if not messagebox.askyesno("확인", f"{len(deleted_items)}개의 삭제 항목을 적용하시겠습니까?", parent=self.root):
            return
        
        self._apply_strings(deleted_items, "삭제", use_new_file=False)

    def apply_reverse_new_strings(self):
        """역방향 신규 적용 (삭제된 항목을 원본에 추가)"""
        # 삭제된 항목을 원본에 추가
        reverse_items = [item for item in self.compare_results if item["type"] == "삭제됨"]
        if not reverse_items:
            messagebox.showinfo("알림", "역방향 적용할 항목이 없습니다.", parent=self.root)
            return
        
        if not messagebox.askyesno("확인", f"{len(reverse_items)}개의 삭제된 항목을 원본에 역방향 적용하시겠습니까?", parent=self.root):
            return
        
        self._apply_reverse_strings(reverse_items)

    def _apply_strings(self, items, action_type, use_new_file=False):
        """STRING 적용 실행"""
        # 진행 창 생성
        loading_popup = LoadingPopup(self.root, f"{action_type} 적용 중", f"{action_type} 적용 준비 중...")
        
        def apply_work():
            try:
                # 파일별로 그룹화
                files_dict = {}
                for item in items:
                    file_name = item["file_name"]
                    if file_name not in files_dict:
                        files_dict[file_name] = []
                    files_dict[file_name].append(item)
                
                total_files = len(files_dict)
                processed_files = 0
                total_processed = 0
                
                # 비교 폴더에서 실제 엑셀 파일 찾기
                compare_folder = self.compare_folder_db_var.get()
                if not compare_folder and self.compare_db_var.get():
                    compare_folder = os.path.dirname(self.compare_db_var.get())
                
                if not compare_folder:
                    self.root.after(0, lambda: [
                        loading_popup.close(),
                        messagebox.showerror("오류", "비교 폴더를 찾을 수 없습니다.")
                    ])
                    return
                
                # 각 파일 처리
                for file_name, file_items in files_dict.items():
                    self.root.after(0, lambda f=file_name, c=processed_files, t=total_files: 
                                  loading_popup.update_progress((c/t)*100, f"파일 처리 중: {f}"))
                    
                    # 엑셀 파일 찾기
                    excel_path = self._find_excel_file(compare_folder, file_name, use_new_file)
                    
                    if not excel_path:
                        if use_new_file:
                            self.root.after(0, lambda f=file_name: 
                                          self.log_message(f"@_new 파일을 찾을 수 없음: {f}"))
                        continue
                    
                    # 파일 처리
                    result = self._process_excel_file(excel_path, file_items, action_type)
                    total_processed += result
                    processed_files += 1
                
                # 완료 처리
                self.root.after(0, lambda: [
                    loading_popup.close(),
                    self.log_message(f"{action_type} 적용 완료: {total_processed}개 항목 처리"),
                    messagebox.showinfo("완료", f"{action_type} 적용 완료!\n총 {total_processed}개 항목이 처리되었습니다.", parent=self.root)
                ])
                
            except Exception as e:
                self.root.after(0, lambda: [
                    loading_popup.close(),
                    self.log_message(f"{action_type} 적용 중 오류: {str(e)}"),
                    messagebox.showerror("오류", f"{action_type} 적용 중 오류 발생: {str(e)}")
                ])
        
        # 백그라운드 스레드 실행
        thread = threading.Thread(target=apply_work)
        thread.daemon = True
        thread.start()

    def _apply_reverse_strings(self, items):
        """역방향 STRING 적용 (비교본 → 원본)"""
        # 진행 창 생성
        loading_popup = LoadingPopup(self.root, "역방향 적용 중", "역방향 적용 준비 중...")
        
        def apply_work():
            try:
                # 파일별로 그룹화
                files_dict = {}
                for item in items:
                    file_name = item["file_name"]
                    if file_name not in files_dict:
                        files_dict[file_name] = []
                    files_dict[file_name].append(item)
                
                total_files = len(files_dict)
                processed_files = 0
                total_processed = 0
                
                # 원본 폴더에서 @_new 파일 찾기
                original_folder = self.original_folder_db_var.get()
                if not original_folder and self.original_db_var.get():
                    original_folder = os.path.dirname(self.original_db_var.get())
                
                if not original_folder:
                    self.root.after(0, lambda: [
                        loading_popup.close(),
                        messagebox.showerror("오류", "원본 폴더를 찾을 수 없습니다.")
                    ])
                    return
                
                # 각 파일 처리
                for file_name, file_items in files_dict.items():
                    self.root.after(0, lambda f=file_name, c=processed_files, t=total_files: 
                                  loading_popup.update_progress((c/t)*100, f"파일 처리 중: {f}"))
                    
                    # @_new 엑셀 파일 찾기
                    excel_path = self._find_excel_file(original_folder, file_name, use_new_file=True)
                    
                    if not excel_path:
                        self.root.after(0, lambda f=file_name: 
                                      self.log_message(f"원본 @_new 파일을 찾을 수 없음: {f}"))
                        continue
                    
                    # 파일 처리 (신규 추가로 처리)
                    result = self._process_excel_file(excel_path, file_items, "역방향 신규")
                    total_processed += result
                    processed_files += 1
                
                # 완료 처리
                self.root.after(0, lambda: [
                    loading_popup.close(),
                    self.log_message(f"역방향 적용 완료: {total_processed}개 항목 처리"),
                    messagebox.showinfo("완료", f"역방향 적용 완료!\n총 {total_processed}개 항목이 처리되었습니다.", parent=self.root)
                ])
                
            except Exception as e:
                self.root.after(0, lambda: [
                    loading_popup.close(),
                    self.log_message(f"역방향 적용 중 오류: {str(e)}"),
                    messagebox.showerror("오류", f"역방향 적용 중 오류 발생: {str(e)}")
                ])
        
        # 백그라운드 스레드 실행
        thread = threading.Thread(target=apply_work)
        thread.daemon = True
        thread.start()

    def _find_excel_file(self, folder, db_file_name, use_new_file=False):
        """엑셀 파일 찾기"""
        # DB 파일명에서 엑셀 파일명 생성
        base_name = db_file_name.replace('.db', '')
        if use_new_file:
            excel_name = f"{base_name}@_new.xlsx"
        else:
            excel_name = f"{base_name}.xlsx"
        
        # 폴더 내에서 엑셀 파일 찾기
        for root, dirs, files in os.walk(folder):
            # 대소문자 구분 없이 파일 찾기
            for file in files:
                if file.lower() == excel_name.lower():
                    return os.path.join(root, file)
        
        return None

    def _process_excel_file(self, excel_path, items, action_type):
        """엑셀 파일 처리"""
        try:
            workbook = load_workbook(excel_path)
            processed_count = 0
            
            # 시트별로 그룹화
            sheets_dict = {}
            for item in items:
                sheet_name = item["sheet_name"]
                if sheet_name not in sheets_dict:
                    sheets_dict[sheet_name] = []
                sheets_dict[sheet_name].append(item)
            
            # 각 시트 처리
            for sheet_name, sheet_items in sheets_dict.items():
                if sheet_name not in workbook.sheetnames:
                    continue
                
                worksheet = workbook[sheet_name]
                
                # 헤더 찾기
                headers = self._find_headers(worksheet)
                if not headers:
                    continue
                
                # 액션별 처리
                if action_type == "신규" or action_type == "신규 (@_new)" or action_type == "역방향 신규":
                    processed_count += self._add_new_strings(worksheet, sheet_items, headers)
                elif action_type == "수정":
                    processed_count += self._modify_strings(worksheet, sheet_items, headers)
                elif action_type == "삭제":
                    processed_count += self._delete_strings(worksheet, sheet_items, headers)
            
            # 파일 저장
            workbook.save(excel_path)
            workbook.close()
            return processed_count
            
        except Exception as e:
            self.log_message(f"파일 처리 오류 {os.path.basename(excel_path)}: {str(e)}")
            return 0

    def _find_headers(self, worksheet):
        """헤더 위치 찾기"""
        for row in range(2, 6):
            headers = {}
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value in ["STRING_ID", "KR", "CN", "TW", "#번역요청"]:
                    headers[cell_value] = {"col": col, "row": row}
            
            if "STRING_ID" in headers and "KR" in headers:
                return headers
        return None

    def _add_new_strings(self, worksheet, items, headers):
        """신규 STRING 추가"""
        processed = 0
        header_row = headers["STRING_ID"]["row"]
        
        # 마지막 행 찾기
        last_row = worksheet.max_row
        
        for item in items:
            last_row += 1
            
            # STRING_ID와 KR 값 추가
            worksheet.cell(row=last_row, column=headers["STRING_ID"]["col"]).value = item["string_id"]
            worksheet.cell(row=last_row, column=headers["KR"]["col"]).value = item["kr"]
            
            # #번역요청 컬럼에 "신규" 추가
            if "#번역요청" in headers:
                worksheet.cell(row=last_row, column=headers["#번역요청"]["col"]).value = "신규"
            
            processed += 1
        
        return processed

    def _modify_strings(self, worksheet, items, headers):
        """STRING 수정"""
        processed = 0
        header_row = headers["STRING_ID"]["row"]
        
        # 기존 데이터 매핑
        string_id_to_row = {}
        for row in range(header_row + 1, worksheet.max_row + 1):
            string_id = worksheet.cell(row=row, column=headers["STRING_ID"]["col"]).value
            if string_id:
                string_id_to_row[string_id] = row
        
        for item in items:
            string_id = item["string_id"]
            if string_id in string_id_to_row:
                row = string_id_to_row[string_id]
                
                # KR 값 수정
                worksheet.cell(row=row, column=headers["KR"]["col"]).value = item["kr"]
                
                # #번역요청에 "신규" 추가
                if "#번역요청" in headers:
                    worksheet.cell(row=row, column=headers["#번역요청"]["col"]).value = "신규"
                
                # CN, TW 값 삭제 (옵션)
                if self.clear_translations_var.get():
                    if "CN" in headers:
                        worksheet.cell(row=row, column=headers["CN"]["col"]).value = ""
                    if "TW" in headers:
                        worksheet.cell(row=row, column=headers["TW"]["col"]).value = ""
                
                processed += 1
        
        return processed

    def _delete_strings(self, worksheet, items, headers):
        """STRING 삭제 (A열에 # 추가)"""
        processed = 0
        header_row = headers["STRING_ID"]["row"]
        
        # 기존 데이터 매핑
        string_id_to_row = {}
        for row in range(header_row + 1, worksheet.max_row + 1):
            string_id = worksheet.cell(row=row, column=headers["STRING_ID"]["col"]).value
            if string_id:
                string_id_to_row[string_id] = row
        
        for item in items:
            string_id = item["string_id"]
            if string_id in string_id_to_row:
                row = string_id_to_row[string_id]
                
                # A열(1열)에 # 추가
                current_value = worksheet.cell(row=row, column=1).value or ""
                if not str(current_value).startswith("#"):
                    worksheet.cell(row=row, column=1).value = f"#{current_value}"
                
                processed += 1
        
        return processed
    

    def debug_current_data(self):
        """현재 데이터 상태 디버깅 (수동 호출용)"""
        self.log_message("=== 현재 데이터 상태 디버깅 ===")
        
        # compare_results 확인
        self.log_message(f"compare_results 개수: {len(self.compare_results)}")
        if self.compare_results:
            first_compare = self.compare_results[0]
            self.log_message(f"compare_results 첫 항목: {first_compare}")
        
        # filtered_results 확인  
        self.log_message(f"filtered_results 개수: {len(self.filtered_results)}")
        if self.filtered_results:
            first_filtered = self.filtered_results[0]
            self.log_message(f"filtered_results 첫 항목: {first_filtered}")
        
        # UI 트리뷰 데이터 확인
        tree_children = self.result_tree.get_children()
        self.log_message(f"UI 트리뷰 항목 개수: {len(tree_children)}")
        if tree_children:
            first_tree_item = self.result_tree.item(tree_children[0])
            self.log_message(f"UI 첫 항목 값들: {first_tree_item['values']}")
        
        # 트리뷰 컬럼 순서 확인
        tree_columns = self.result_tree["columns"]
        self.log_message(f"트리뷰 컬럼 순서: {tree_columns}")
        
        self.log_message("=== 디버깅 완료 ===")


    def export_to_excel(self):
        """비교 결과를 엑셀로 저장 (올바른 컬럼 의미 적용)"""
        if not self.filtered_results:
            messagebox.showinfo("알림", "내보낼 비교 결과가 없습니다.", parent=self.root)
            return

        # 파일 저장 경로 선택
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            title="엑셀로 저장",
            parent=self.root
        )
        if not file_path:
            return

        try:
            # 데이터 준비 (UI와 동일한 로직)
            export_data = []
            for item in self.filtered_results:
                row_data = {
                    "file_name": item.get("file_name", ""),
                    "sheet_name": item.get("sheet_name", ""),
                    "type": item.get("type", ""),
                    "string_id": item.get("string_id", ""),
                    "비교본_KR": item.get("kr", ""),           # 비교본의 KR 값
                    "원본_KR": item.get("original_kr", "")    # 원본의 KR 값
                }
                export_data.append(row_data)
            
            # DataFrame 생성 및 저장
            df = pd.DataFrame(export_data)
            df.to_excel(file_path, index=False)
            
            self.log_message(f"엑셀 내보내기 완료: {len(export_data)}개 항목")
            messagebox.showinfo("완료", f"엑셀 파일로 저장 완료!\n{file_path}", parent=self.root)
            
        except Exception as e:
            self.log_message(f"엑셀 내보내기 오류: {str(e)}")
            messagebox.showerror("오류", f"엑셀 내보내기 실패: {e}", parent=self.root)
    
    
    
    def show_exception_rules_manager(self):
        import tkinter as tk
        import tkinter.ttk as ttk
        import json

        # 팝업 생성
        win = tk.Toplevel(self.root)
        win.title("예외 규칙 관리")
        win.geometry("600x400")
        win.transient(self.root)
        win.grab_set()

        # 규칙 트리뷰
        columns = ("type", "field", "value", "enabled", "description")
        tree = ttk.Treeview(win, columns=columns, show="headings", height=10)
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=100)
        tree.pack(fill="both", expand=True, padx=5, pady=5)

        # 현재 규칙 표시
        def refresh_tree():
            tree.delete(*tree.get_children())
            for idx, rule in enumerate(self.exception_rules):
                tree.insert("", "end", iid=idx, values=(
                    rule.get("type", ""),
                    rule.get("field", ""),
                    rule.get("value", ""),
                    "O" if rule.get("enabled", True) else "X",
                    rule.get("description", "")
                ))

        refresh_tree()

        # 하단 영역: 새 규칙 추가
        frame = ttk.Frame(win)
        frame.pack(fill="x", padx=5, pady=5)
        rule_type_var = tk.StringVar(value="startswith")
        rule_field_var = tk.StringVar(value="KR")
        rule_value_var = tk.StringVar()
        rule_enabled_var = tk.BooleanVar(value=True)
        rule_desc_var = tk.StringVar()

        ttk.Label(frame, text="유형").grid(row=0, column=0)
        type_combo = ttk.Combobox(frame, textvariable=rule_type_var, width=10, values=[
            "startswith", "endswith", "contains", "equals", "length", "regex"
        ])
        type_combo.grid(row=0, column=1)
        ttk.Label(frame, text="필드").grid(row=0, column=2)
        field_combo = ttk.Combobox(frame, textvariable=rule_field_var, width=10, values=[
            "KR", "string_id", "file_name", "sheet_name"
        ])
        field_combo.grid(row=0, column=3)
        ttk.Label(frame, text="값").grid(row=0, column=4)
        ttk.Entry(frame, textvariable=rule_value_var, width=15).grid(row=0, column=5)
        ttk.Label(frame, text="설명").grid(row=0, column=6)
        ttk.Entry(frame, textvariable=rule_desc_var, width=18).grid(row=0, column=7)
        ttk.Checkbutton(frame, text="활성", variable=rule_enabled_var).grid(row=0, column=8, padx=5)

        def add_rule():
            rule = {
                "type": rule_type_var.get(),
                "field": rule_field_var.get(),
                "value": rule_value_var.get(),
                "enabled": rule_enabled_var.get(),
                "description": rule_desc_var.get()
            }
            self.exception_rules.append(rule)
            self.save_exception_rules()
            refresh_tree()
            rule_value_var.set("")
            rule_desc_var.set("")

        ttk.Button(frame, text="규칙 추가", command=add_rule).grid(row=0, column=9, padx=5)

        # 규칙 삭제/토글 버튼
        btns = ttk.Frame(win)
        btns.pack(fill="x", padx=5, pady=5)

        def delete_rule():
            sel = tree.selection()
            if not sel: return
            for iid in sel:
                del self.exception_rules[int(iid)]
            self.save_exception_rules()
            refresh_tree()

        def toggle_rule():
            sel = tree.selection()
            if not sel: return
            for iid in sel:
                idx = int(iid)
                self.exception_rules[idx]["enabled"] = not self.exception_rules[idx].get("enabled", True)
            self.save_exception_rules()
            refresh_tree()

        ttk.Button(btns, text="삭제", command=delete_rule).pack(side="left", padx=3)
        ttk.Button(btns, text="활성/비활성 전환", command=toggle_rule).pack(side="left", padx=3)
        ttk.Button(btns, text="닫기", command=win.destroy).pack(side="right", padx=3)

        # 기본값 초기화(옵션)
        def reset_rules():
            self.exception_rules = []
            self.save_exception_rules()
            refresh_tree()
            ttk.Button(btns, text="모두 초기화", command=reset_rules).pack(side="left", padx=3)

    def save_exception_rules(self, path="string_exception_rules.json"):
        """예외 규칙을 JSON 파일로 저장하고 컴파일 규칙 재생성"""
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(self.exception_rules, f, ensure_ascii=False, indent=2)
            
            # 규칙이 변경되었으므로 컴파일된 규칙도 다시 생성
            self._recompile_exception_rules()
            
            self.log_message(f"예외 규칙 저장 완료: {len(self.exception_rules)}개 규칙")
        except Exception as e:
            self.log_message(f"예외 규칙 저장 오류: {e}")

    def _recompile_exception_rules(self):
        """컴파일된 예외 규칙 재생성"""
        try:
            self.compiled_rules = []
            
            for rule in self.exception_rules:
                if not rule.get("enabled", True):
                    continue
                    
                field = rule.get("field", "").lower()
                if not field:
                    continue
                    
                rule_type = rule.get("type", "")
                value = rule.get("value", "")
                
                # 문자열 이스케이프 미리 처리
                if rule_type in ["startswith", "endswith", "contains", "equals"]:
                    processed_value = str(value).replace("\\n", "\n").replace("\\t", "\t")
                else:
                    processed_value = value
                
                compiled_rule = {
                    "field": field,
                    "type": rule_type,
                    "value": processed_value,
                    "original_value": value,
                    "description": rule.get("description", "")
                }
                
                # 정규식 미리 컴파일
                if rule_type == "regex":
                    try:
                        import re
                        compiled_rule["compiled_pattern"] = re.compile(processed_value)
                    except Exception as e:
                        self.log_message(f"정규식 컴파일 실패: {e}")
                        continue
                
                # length 타입 값 미리 변환
                elif rule_type == "length":
                    try:
                        compiled_rule["threshold"] = int(processed_value)
                    except (ValueError, TypeError):
                        self.log_message(f"length 규칙 값 변환 실패: {processed_value}")
                        continue
                
                self.compiled_rules.append(compiled_rule)
                
        except Exception as e:
            self.log_message(f"규칙 재컴파일 오류: {e}")
            self.compiled_rules = []