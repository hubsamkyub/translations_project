import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import sqlite3
import re
import time
import shutil
#import psutil
from utils.config_utils import load_config, save_config

class UniqueTextManager:
    def __init__(self, root):
        self.root = root
        self.root.title("고유 텍스트 관리")
        self.root.geometry("1600x1000")
        
        # DB 경로 설정 변수 추가
        self.db_path_var = tk.StringVar()
        # 초기값은 프로그램 폴더로 설정
        program_dir = os.path.dirname(os.path.abspath(__file__))
        default_db_path = os.path.join(program_dir, "unique_texts.db")
        self.db_path_var.set(default_db_path)
        
        # 메인 프레임 설정
        self.main_frame = ttk.Frame(self.root)
        self.main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # 설정 정보 프레임 추가
        info_frame = ttk.Frame(self.main_frame)
        info_frame.pack(fill="x", padx=10, pady=5)

        # DB 경로 설정 UI 추가
        db_path_frame = ttk.Frame(info_frame)
        db_path_frame.pack(fill="x", pady=5)

        ttk.Label(db_path_frame, text="고유 텍스트 DB 위치:").pack(side="left")
        db_path_entry = ttk.Entry(db_path_frame, textvariable=self.db_path_var, width=60)
        db_path_entry.pack(side="left", padx=5)
        ttk.Button(db_path_frame, text="찾아보기", command=self.select_db_path).pack(side="left", padx=5)
        ttk.Button(db_path_frame, text="저장", command=self.save_db_path_setting).pack(side="left", padx=5)
        ttk.Button(self.main_frame, text="엑셀로 내보내기", command=self.export_unique_texts_to_excel)\
            .pack(side="bottom", anchor="e", padx=10, pady=6)


        self.info_label = ttk.Label(info_frame, text=f"현재 DB 경로: {self.db_path_var.get()}", foreground="blue")
        self.info_label.pack(anchor="w")
        
        # 상태 표시줄 추가 (꼭 여기 추가해야 함)
        self.status_label = ttk.Label(self.main_frame, text="준비됨", anchor="w")
        self.status_label.pack(fill="x", side="bottom", padx=10, pady=2)
        
        # DB 경로 설정 불러오기
        self.load_db_path_setting()
        
        # 예외 규칙 메뉴 추가 
        self.exception_rules = []
        self.default_rules = [
            {"type": "startswith", "field": "KR", "value": "#", "enabled": True, "description": "#으로 시작하는 KR 제외"},
            {"type": "startswith", "field": "STRING_ID", "value": "cs_", "enabled": True, "description": "cs_로 시작하는 STRING_ID 제외"},
            {"type": "length", "field": "KR", "value": 100, "enabled": True, "description": "KR이 100자 넘으면 제외"}
        ]
        
        menu_bar = tk.Menu(self.root)
        self.root.config(menu=menu_bar)

        tools_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="도구", menu=tools_menu)
        tools_menu.add_command(label="예외 규칙 관리", command=self.show_exception_rules_manager)
        # 다국어 메뉴 다음에 엑셀 업데이트 메뉴 추가
        tools_menu.add_command(label="엑셀 업데이트", command=self.update_excel_with_db_data)

        # 예외 규칙 설정 로드
        self.load_exception_rules()
        
        #다국어 메뉴
        tools_menu.add_command(label="다국어 갱신", command=self.update_translations_from_excel)

        # 상단: 폴더 선택 + 최신화 버튼
        top_frame = ttk.Frame(self.main_frame)
        top_frame.pack(fill="x", padx=10, pady=5)

        # DB 폴더 변수 초기화
        self.unique_db_folder_var = tk.StringVar()
        
        # DB 폴더 경로 설정 불러오기
        self.load_db_folder_path()
        
        ttk.Label(top_frame, text="DB 폴더:").pack(side="left")
        ttk.Entry(top_frame, textvariable=self.unique_db_folder_var, width=60).pack(side="left", padx=5)
        ttk.Button(top_frame, text="찾아보기", command=self.select_unique_db_folder).pack(side="left", padx=5)
        ttk.Button(top_frame, text="📋 목록 보기", command=self.load_db_list).pack(side="left", padx=5)
        ttk.Button(top_frame, text="🔄 최신화", command=self.refresh_unique_db).pack(side="left", padx=5)
        ttk.Button(top_frame, text="🔄 DB 새로고침", command=self.reload_unique_string_db).pack(side="left", padx=5)
        
        # DB 목록 표시 프레임
        db_list_frame = ttk.LabelFrame(self.main_frame, text="대상 DB 목록")
        db_list_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        # 스크롤 가능한 캔버스 생성
        canvas = tk.Canvas(db_list_frame, height=150)
        scrollbar = ttk.Scrollbar(db_list_frame, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # 내부 체크박스를 그릴 프레임
        self.db_checks_frame = ttk.Frame(canvas)
        canvas.create_window((0, 0), window=self.db_checks_frame, anchor="nw")

        # 스크롤 동작 자동 계산
        self.db_checks_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        # 마우스 휠로 스크롤
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        self.db_checks_frame.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", _on_mousewheel))
        self.db_checks_frame.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

        self.db_check_vars = {}  # {filename: tk.BooleanVar()}

        # 중간: 테이블 뷰어와 상세 정보 패널 구성
        content_frame = ttk.Frame(self.main_frame)
        content_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        # 검색창 프레임
        search_frame = ttk.Frame(content_frame)
        search_frame.pack(fill="x", padx=5, pady=5)
        
        # STRING_ID와 KR 검색 옵션
        search_options_frame = ttk.Frame(search_frame)
        search_options_frame.pack(side="left", padx=5)

        # STRING_ID 검색
        ttk.Label(search_options_frame, text="🔍 STRING_ID:").pack(side="left", padx=5)
        self.table_id_search_var = tk.StringVar()
        id_search_entry = ttk.Entry(search_options_frame, textvariable=self.table_id_search_var, width=20)
        id_search_entry.pack(side="left", padx=5)
        
        # 수정된 코드
        ttk.Label(search_frame, text="🔍 KR 검색:").pack(side="left", padx=5)
        self.table_kr_search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=self.table_kr_search_var, width=40)
        search_entry.pack(side="left", padx=5)
        
        # 검색 옵션 프레임 (완전일치/부분일치)
        match_options_frame = ttk.Frame(search_frame)
        match_options_frame.pack(side="left", padx=10)

        self.exact_match_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(match_options_frame, text="완전 일치", variable=self.exact_match_var, 
                        command=self.filter_table_view).pack(side="left")

        # 검색 버튼 추가
        ttk.Button(match_options_frame, text="검색", command=self.filter_table_view).pack(side="left", padx=5)

        # Entry에 기본 포커스 설정
        search_entry.focus_set()

        # 검색어 변경 시 필터링 함수 연결
        self.table_kr_search_var.trace_add("write", lambda *args: self.filter_table_view())
        self.table_id_search_var.trace_add("write", lambda *args: self.filter_table_view())
        
        # 좌측: 테이블 뷰어
        left_panel = ttk.Frame(content_frame)
        left_panel.pack(side="left", fill="both", expand=True)
        
        # 테이블 뷰어 (Treeview)
        columns = ("STRING_ID", "KR", "EN", "CN", "TW")
        self.text_table = ttk.Treeview(left_panel, columns=columns, show="headings", height=20)
        
        # 각 컬럼 설정
        self.text_table.heading("STRING_ID", text="STRING_ID")
        self.text_table.heading("KR", text="KR")
        self.text_table.heading("EN", text="EN")
        self.text_table.heading("CN", text="CN")
        self.text_table.heading("TW", text="TW")
        
        # 컬럼 너비 설정
        self.text_table.column("STRING_ID", width=100)
        self.text_table.column("KR", width=200)
        self.text_table.column("EN", width=200)
        self.text_table.column("CN", width=150)
        self.text_table.column("TW", width=150)
        
        # 테이블 스크롤바
        table_scroll_y = ttk.Scrollbar(left_panel, orient="vertical", command=self.text_table.yview)
        self.text_table.configure(yscrollcommand=table_scroll_y.set)
        
        table_scroll_x = ttk.Scrollbar(left_panel, orient="horizontal", command=self.text_table.xview)
        self.text_table.configure(xscrollcommand=table_scroll_x.set)
        
        self.text_table.pack(side="left", fill="both", expand=True)
        table_scroll_y.pack(side="right", fill="y")
        table_scroll_x.pack(side="bottom", fill="x")
        
        # 테이블 항목 선택 이벤트
        self.text_table.bind("<<TreeviewSelect>>", self.on_table_row_selected)
        
        # 우측: 다국어 편집 영역
        right_panel = ttk.Frame(content_frame)
        right_panel.pack(side="right", fill="y", padx=(10, 0))
        
        # 언어별 필드 저장
        self.lang_fields = {}
        
        self.string_id_var = tk.StringVar()
        ttk.Label(right_panel, text="STRING_ID").pack(anchor="w")
        ttk.Entry(right_panel, textvariable=self.string_id_var, state="readonly", width=40).pack(anchor="w", pady=(0, 5))

        self.selected_kr_var = tk.StringVar()
        ttk.Label(right_panel, text="KR (수정 불가)").pack(anchor="w")
        ttk.Entry(right_panel, textvariable=self.selected_kr_var, state="readonly", width=40).pack(anchor="w", pady=(0, 10))

        for lang in ["EN", "CN", "TW", "JP", "DE", "FR", "TH", "PT", "ES"]:
            ttk.Label(right_panel, text=lang).pack(anchor="w")
            entry = ttk.Entry(right_panel, width=50)
            entry.pack(anchor="w", pady=2)
            self.lang_fields[lang] = entry

        # 하단: 저장/취소 버튼
        action_frame = ttk.Frame(right_panel)
        action_frame.pack(anchor="e", pady=10)

        ttk.Button(action_frame, text="💾 저장", command=self.save_translation_edits).pack(side="left", padx=5)
        ttk.Button(action_frame, text="⛔ 취소", command=self.cancel_translation_edits).pack(side="left", padx=5)
        
        # 초기 데이터 로드
        self.unique_texts = {}  # <-- 이 라인 추가
        self._is_first_load = True
        self.load_initial_data()
        
        # 추가: 검색 영역을 명시적으로 활성화
        self.load_db_list()  # DB 목록을 초기에 불러옴


    def select_db_path(self):
        """DB 파일 경로 선택"""
        file_path = filedialog.asksaveasfilename(
            title="unique_texts.db 저장 위치 선택",
            defaultextension=".db",
            filetypes=[("Database files", "*.db")],
            initialfile="unique_texts.db",
            parent=self.root
        )
        if file_path:
            self.db_path_var.set(file_path)
            self.info_label.config(text=f"현재 DB 경로: {file_path}")
            
            # 포커스 복원
            self.root.after(100, self.root.focus_force)
            self.root.after(100, self.root.lift)
            
    def save_db_path_setting(self):
        """DB 경로 설정 저장"""
        try:
            # 기존 설정 로드
            config = load_config()
            
            # db_path 설정 추가
            config["db_path"] = self.db_path_var.get()
            
            # 설정 저장
            save_config("config.json", config)
            
            messagebox.showinfo("완료", "DB 경로 설정이 저장되었습니다.", parent=self.root)
        except Exception as e:
            messagebox.showerror("오류", f"설정 저장 실패: {str(e)}", parent=self.root)


    def load_db_path_setting(self):
        """저장된 DB 경로 설정 불러오기"""
        try:            
            # 설정 로드
            config = load_config()
            
            # db_path 설정이 있으면 적용
            if "db_path" in config and config["db_path"] and os.path.exists(os.path.dirname(config["db_path"])):
                self.db_path_var.set(config["db_path"])
                self.info_label.config(text=f"현재 DB 경로: {config['db_path']}")
                return True
            else:
                return False
        except Exception as e:
            # 로드 실패 시 기본값 유지
            return False

    def load_db_folder_path(self):
        """저장된 DB 폴더 경로 설정 불러오기"""
        try:            
            # 설정 로드
            config = load_config()
            
            # db_folder_path 설정이 있으면 적용
            if "db_folder_path" in config and config["db_folder_path"] and os.path.exists(config["db_folder_path"]):
                self.unique_db_folder_var.set(config["db_folder_path"])
                self.status_label.config(text=f"DB 폴더 경로를 불러왔습니다: {config['db_folder_path']}")
                return True
            return False
        except Exception as e:
            print(f"DB 폴더 경로 로드 실패: {str(e)}")
            return False

    def update_excel_with_db_data(self):
        """unique_texts.db의 데이터로 엑셀 파일 업데이트"""
        # 1. 엑셀 파일 선택
        excel_path = filedialog.askopenfilename(
            title="업데이트할 엑셀 파일 선택",
            filetypes=[("Excel 파일", "*.xlsx")],
            parent=self.root
        )
        
        if not excel_path:
            messagebox.showwarning("경고", "파일이 선택되지 않았습니다.", parent=self.root)
            return
        
        # 포커스 복원
        self.root.after(100, self.root.focus_force)
        self.root.after(100, self.root.lift)
        
        try:
            import pandas as pd
            from openpyxl import load_workbook
            
            # 2. 데이터베이스에서 데이터 로드
            if not self.unique_texts:
                db_path = self.db_path_var.get()
                if not os.path.exists(db_path):
                    messagebox.showwarning("경고", "DB 파일이 존재하지 않습니다.", parent=self.root)
                    return
                    
                self.load_unique_string_db(db_path)
                
                if not self.unique_texts:
                    messagebox.showwarning("경고", "DB에 데이터가 없습니다.", parent=self.root)
                    return
                    
            # 3. 엑셀 파일 로드 및 #String 시트 확인
            try:
                wb = load_workbook(excel_path)
                if "#String" not in wb.sheetnames:
                    messagebox.showwarning("경고", "엑셀 파일에 #String 시트가 없습니다.", parent=self.root)
                    wb.close()
                    return
            except Exception as e:
                messagebox.showerror("오류", f"엑셀 파일 열기 실패: {str(e)}", parent=self.root)
                return
                
            # 4. 시트 데이터 분석
            sheet = wb["#String"]
            
            # 4번째 행에서 컬럼 헤더 찾기 (인덱스는 1부터 시작)
            header_row = 4
            column_headers = {}
            
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=header_row, column=col).value
                if cell_value:
                    column_headers[cell_value] = col
            
            # STRING_ID 컬럼 확인
            if "STRING_ID" not in column_headers:
                messagebox.showwarning("경고", "#String 시트의 4번째 행에 STRING_ID 컬럼이 없습니다.", parent=self.root)
                wb.close()
                return
                
            # 5. 진행 상황 창 생성
            progress_window = tk.Toplevel(self.root)
            progress_window.title("엑셀 업데이트 진행 중")
            progress_window.geometry("400x150")
            progress_window.transient(self.root)
            progress_window.grab_set()
            
            ttk.Label(progress_window, text="DB 데이터로 엑셀 파일 업데이트 중...").pack(pady=10)
            
            progress_var = tk.DoubleVar()
            progress_bar = ttk.Progressbar(progress_window, variable=progress_var, maximum=100)
            progress_bar.pack(fill="x", padx=20, pady=10)
            
            status_var = tk.StringVar(value="처리 중...")
            status_label = ttk.Label(progress_window, textvariable=status_var)
            status_label.pack(pady=10)
            
            progress_window.update()
            
            # 6. 엑셀의 기존 STRING_ID 목록 수집
            excel_string_ids = set()
            last_row = sheet.max_row
            
            for row in range(header_row + 1, last_row + 1):
                string_id = sheet.cell(row=row, column=column_headers["STRING_ID"]).value
                if string_id:
                    excel_string_ids.add(string_id)
            
            # 7. DB 데이터로 엑셀 업데이트
            updated_count = 0
            added_count = 0
            total_count = len(self.unique_texts)
            current_count = 0
            
            # DB의 모든 항목을 순회
            for kr, record in self.unique_texts.items():
                current_count += 1
                string_id = record.get("STRING_ID", "")
                
                if not string_id:
                    continue
                    
                # 진행 상황 업데이트
                progress_var.set((current_count / total_count) * 100)
                status_var.set(f"처리 중... ({current_count}/{total_count})")
                progress_window.update()
                
                # 이미 엑셀에 있는 항목인 경우 업데이트
                if string_id in excel_string_ids:
                    # 기존 행 찾기
                    for row in range(header_row + 1, last_row + 1):
                        cell_string_id = sheet.cell(row=row, column=column_headers["STRING_ID"]).value
                        if cell_string_id == string_id:
                            # 각 언어 필드 업데이트
                            for lang in ["KR", "EN", "CN", "TW", "JP", "DE", "FR", "TH", "PT", "ES"]:
                                if lang in column_headers and lang in record:
                                    sheet.cell(row=row, column=column_headers[lang]).value = record[lang]
                            updated_count += 1
                            break
                else:
                    # 엑셀에 없는 경우 새 행 추가
                    last_row += 1
                    
                    # STRING_ID 컬럼 추가
                    sheet.cell(row=last_row, column=column_headers["STRING_ID"]).value = string_id
                    
                    # 각 언어 필드 추가
                    for lang in ["KR", "EN", "CN", "TW", "JP", "DE", "FR", "TH", "PT", "ES"]:
                        if lang in column_headers and lang in record:
                            sheet.cell(row=last_row, column=column_headers[lang]).value = record[lang]
                    added_count += 1
            
            # 8. 변경사항 저장
            wb.save(excel_path)
            wb.close()
            
            # 진행 창 닫기
            progress_window.destroy()
            
            # 결과 메시지
            messagebox.showinfo(
                "엑셀 업데이트 완료", 
                f"총 {total_count}개 항목 중\n"
                f"{updated_count}개 항목이 업데이트되었고,\n"
                f"{added_count}개 항목이 새로 추가되었습니다.",
                parent=self.root
            )
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 업데이트 중 오류가 발생했습니다:\n{str(e)}", parent=self.root)
            import traceback
            traceback.print_exc()

    #초기 데이터 로드
    def load_initial_data(self):
        """초기 데이터 로드"""
        # 설정된 DB 경로에서 파일 로드
        db_path = self.db_path_var.get()
        
        if os.path.exists(db_path):
            self.load_unique_string_db(db_path)
            self.update_db_path_display()
            self.status_label.config(text=f"DB 파일 로드 완료: {db_path}")
        else:
            # DB 파일이 없는 경우 알림
            messagebox.showinfo("알림", 
                            f"지정된 경로에 DB 파일이 없습니다: {db_path}\n"
                            "DB를 최신화하거나 올바른 경로를 지정하세요.", 
                            parent=self.root)
            self.status_label.config(text=f"DB 파일이 없음: {db_path}")
        
        # DB가 로드되지 않았는지 확인
        if not self.unique_texts:
            messagebox.showinfo("알림", "데이터베이스가 비어 있습니다. DB를 최신화하거나 올바른 경로를 지정하세요.", parent=self.root)
        
        # 텍스트 박스가 비어 있더라도 filter_table_view가 한 번 실행되도록 함
        self.filter_table_view()
        
    def load_unique_string_db(self, db_path):
        """unique_texts.db 파일 로드하여 테이블 뷰어에 표시"""
        try:
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            # 테이블 존재 여부 확인
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='unique_texts'")
            if not cursor.fetchone():
                messagebox.showinfo("정보", "유효한 unique_texts 테이블이 없습니다.", parent=self.root)
                conn.close()
                return
                
            # 데이터 로드
            cursor.execute("SELECT STRING_ID, KR, EN, CN, TW, JP, DE, FR, TH, PT, ES, UpdateTime FROM unique_texts")
            columns = [desc[0] for desc in cursor.description]
            records = cursor.fetchall()
            
            # 데이터를 메모리에 저장
            self.unique_texts = {}
            for record in records:
                data = dict(zip(columns, record))
                self.unique_texts[data["KR"]] = data
            
            # 테이블 뷰어 업데이트
            self.update_table_view()
            
            conn.close()
            
            # 로드 메시지 표시
            if self._is_first_load:
                messagebox.showinfo("로드 완료", f"총 {len(self.unique_texts)}개의 고유 텍스트를 로드했습니다.", parent=self.root)
                self._is_first_load = False
        except Exception as e:
            messagebox.showerror("로드 오류", f"데이터베이스 로드 중 오류 발생: {str(e)}", parent=self.root)
            import traceback
            traceback.print_exc()

    def update_table_view(self):
        """테이블 뷰어에 데이터 표시"""
        # 테이블 초기화
        for item in self.text_table.get_children():
            self.text_table.delete(item)
        
        if not self.unique_texts:
            return
            
        # 데이터 추가
        for kr, record in self.unique_texts.items():
            self.text_table.insert("", "end", values=(
                record.get("STRING_ID", ""),
                kr,
                record.get("EN", ""),
                record.get("CN", ""),
                record.get("TW", "")
            ))

    def filter_table_view(self):
        """STRING_ID와 KR 검색어로 테이블 필터링"""
        kr_keyword = self.table_kr_search_var.get().strip().lower()
        id_keyword = self.table_id_search_var.get().strip().lower()
        exact_match = self.exact_match_var.get()
        
        # 테이블 초기화
        for item in self.text_table.get_children():
            self.text_table.delete(item)
        
        # 데이터가 없으면 상태 메시지 표시하고 리턴 (하지만 UI 요소는 계속 활성화 상태 유지)
        if not self.unique_texts:
            self.status_label.config(text="데이터가 로드되지 않았습니다. DB를 먼저 로드하세요.")
            return
            
        # 검색 조건에 맞는 항목만 표시
        for kr, record in self.unique_texts.items():
            string_id = record.get("STRING_ID", "").lower()
            
            # KR 일치 여부 확인
            kr_matches = False
            if exact_match:
                kr_matches = kr.lower() == kr_keyword if kr_keyword else True
            else:
                kr_matches = kr_keyword in kr.lower() if kr_keyword else True
            
            # STRING_ID 일치 여부 확인
            id_matches = False
            if exact_match:
                id_matches = string_id == id_keyword if id_keyword else True
            else:
                id_matches = id_keyword in string_id if id_keyword else True
            
            # 두 조건 모두 충족하는 경우만 표시
            if kr_matches and id_matches:
                self.text_table.insert("", "end", values=(
                    record.get("STRING_ID", ""),
                    kr,
                    record.get("EN", ""),
                    record.get("CN", ""),
                    record.get("TW", "")
                ))
                
    def on_table_row_selected(self, event):
        """테이블 행 선택 이벤트 핸들러"""
        selection = self.text_table.selection()
        if not selection:
            return
        
        # 선택된 행의 데이터 가져오기
        item = self.text_table.item(selection[0])
        values = item["values"]
        
        if not values or len(values) < 2:
            return
        
        kr = values[1]  # KR 컬럼 값
        record = self.unique_texts.get(kr, {})
        
        # 우측 상세 정보 패널 업데이트
        self.string_id_var.set(record.get("STRING_ID", ""))
        self.selected_kr_var.set(kr)
        
        for lang, entry in self.lang_fields.items():
            entry.delete(0, tk.END)
            entry.insert(0, record.get(lang, ""))
    
    def load_db_list(self):
        db_folder = self.unique_db_folder_var.get()
        if not db_folder or not os.path.isdir(db_folder):
            # 경고 대신 기본 경로 시도
            program_dir = os.path.dirname(os.path.abspath(__file__))
            db_folder = program_dir
            self.unique_db_folder_var.set(db_folder)
            self.save_db_folder_path(db_folder)
            return

        all_db_files = [
            os.path.join(db_folder, f)
            for f in os.listdir(db_folder)
            if f.startswith("String") and f.endswith(".db")
        ]

        if not all_db_files:
            messagebox.showinfo("정보", "DB 파일이 없습니다.", parent=self.root)
            return

        self.display_db_checkboxes(all_db_files)
        messagebox.showinfo("완료", f"총 {len(all_db_files)}개의 DB 파일을 찾았습니다.", parent=self.root)

    def display_db_checkboxes(self, db_files):
        for widget in self.db_checks_frame.winfo_children():
            widget.destroy()

        self.db_check_vars.clear()

        self.check_all_var = tk.BooleanVar()
        ttk.Checkbutton(
            self.db_checks_frame,
            text="전체 선택",
            variable=self.check_all_var,
            command=self.toggle_all_db_checks
        ).pack(anchor="w", padx=5, pady=2)

        for db_file in db_files:
            var = tk.BooleanVar(value=True)
            # ✅ 각 하위 체크박스가 변경될 때 전체 선택 상태를 갱신
            chk = ttk.Checkbutton(
                self.db_checks_frame,
                text=os.path.basename(db_file),
                variable=var,
                command=self.update_check_all_state
            )
            chk.pack(anchor="w", padx=10)
            self.db_check_vars[db_file] = var

        # 초기 상태도 검사
        self.update_check_all_state()

    def update_check_all_state(self):
        """모든 체크 상태를 보고 전체 선택 버튼 상태를 동기화"""
        all_checked = all(var.get() for var in self.db_check_vars.values())
        self.check_all_var.set(all_checked)

    def toggle_all_db_checks(self):
        state = self.check_all_var.get()
        for var in self.db_check_vars.values():
            var.set(state)

    def select_unique_db_folder(self):
        folder = filedialog.askdirectory(title="String DB 폴더 선택", parent=self.root)
        if folder:
            self.unique_db_folder_var.set(folder)
            # 선택된 폴더 경로 저장
            self.save_db_folder_path(folder)
            # 포커스 복원
            self.root.after(100, self.root.focus_force)
            self.root.after(100, self.root.lift)

    def save_db_folder_path(self, folder_path):
        """DB 폴더 경로 설정 저장"""
        try:            
            # 기존 설정 로드
            config = load_config()
            
            # db_folder_path 설정 추가
            config["db_folder_path"] = folder_path
            
            # 설정 저장
            save_config("config.json", config)
            
            self.status_label.config(text=f"DB 폴더 경로가 저장되었습니다: {folder_path}")
        except Exception as e:
            self.status_label.config(text=f"DB 폴더 경로 저장 실패: {str(e)}")


    def refresh_unique_db(self):
        # 1. 입력 유효성 검증
        db_folder = self.unique_db_folder_var.get()
        if not db_folder or not os.path.isdir(db_folder):
            messagebox.showwarning("경고", "올바른 DB 폴더를 선택하세요.", parent=self.root)
            return

        if not hasattr(self, "db_check_vars") or not self.db_check_vars:
            messagebox.showwarning("경고", "먼저 [목록 보기]를 통해 DB 목록을 불러오세요.", parent=self.root)
            return

        selected_db_files = [
            db_file for db_file, var in self.db_check_vars.items() if var.get()
        ]
        if not selected_db_files:
            messagebox.showwarning("경고", "최소 하나 이상의 DB를 선택하세요.", parent=self.root)
            return

        try:
            # 2. 선택된 DB에서 데이터 수집
            all_entries = self._collect_entries_from_dbs(selected_db_files)
            
            # 3. 기존 unique_texts.db 로딩 및 백업 (프로그램 폴더 사용)
            program_dir = os.path.dirname(os.path.abspath(__file__))
            unique_db_path = os.path.join(program_dir, "unique_texts.db")
            existing_entries = self._load_existing_entries(unique_db_path)
            
            # 4. 신규 ID 부여 및 병합
            merged_entries, updated_count = self._merge_entries(all_entries, existing_entries)
            
            # 5. 데이터베이스에 저장
            self._save_to_database(unique_db_path, merged_entries)
            
            # 6. 메모리에 반영 + UI 업데이트
            self.unique_texts = {entry["KR"]: entry for entry in merged_entries.values()}
            
            # 테이블 뷰어 업데이트
            self.update_table_view()
            
            # 제외된 항목 수 정보 추가
            excluded_count = getattr(self, 'excluded_count', 0)
            
            messagebox.showinfo(
                "완료",
                f"총 {len(self.unique_texts)}개의 고유 텍스트 중\n"
                f"{updated_count}개가 새로 추가되었습니다.\n\n"
                f"예외 규칙에 의해 {excluded_count}개 항목이 제외되었습니다.",
                parent=self.root
            )
        except Exception as e:
            messagebox.showerror("오류", f"처리 중 오류가 발생했습니다: {str(e)}", parent=self.root)
            import traceback
            traceback.print_exc()
            
    def _collect_entries_from_dbs(self, db_files):
        """선택된 DB 파일에서 항목 수집 (개선된 고유값 생성 로직)"""
        kr_entries_map = {}  # KR 텍스트를 키로 하는 항목 모음
        excluded_count = 0  # 예외 규칙으로 제외된 항목 수 추적
        
        for db_file in db_files:
            try:
                conn = sqlite3.connect(db_file)
                cursor = conn.cursor()
                
                for table in self.get_text_tables(cursor):
                    # SQL 인젝션 방지를 위한 처리
                    if not self._is_valid_table_name(table):
                        continue
                        
                    cursor.execute(f"SELECT * FROM {table}")
                    columns = [desc[0] for desc in cursor.description]
                    rows = cursor.fetchall()
                    
                    for row in rows:
                        row_data = dict(zip(columns, row))
                        kr = str(row_data.get("KR", "")).strip()
                        
                        # KR이 비어있지 않은 경우만 처리
                        if kr:
                            # 예외 규칙 검사
                            if self.should_exclude_entry(row_data):
                                excluded_count += 1
                                continue
                                
                            # 해당 KR이 처음 등장하면 리스트 초기화
                            if kr not in kr_entries_map:
                                kr_entries_map[kr] = []
                                
                            # 이 KR에 해당하는 모든 항목 수집
                            entry = {
                                "KR": kr,
                                "EN": row_data.get("EN", ""),
                                "CN": row_data.get("CN", ""),
                                "TW": row_data.get("TW", ""),
                                "JP": row_data.get("JP", ""),
                                "DE": row_data.get("DE", ""),
                                "FR": row_data.get("FR", ""),
                                "TH": row_data.get("TH", ""),
                                "PT": row_data.get("PT", ""),
                                "ES": row_data.get("ES", ""),
                            }
                            kr_entries_map[kr].append(entry)
                conn.close()
            except Exception as e:
                messagebox.showwarning("DB 오류", f"파일 '{os.path.basename(db_file)}' 처리 중 오류: {str(e)}", parent=self.root)
        
        # 각 KR별로 병합된 최선의 항목 생성
        all_merged_entries = []
        for kr, entries in kr_entries_map.items():
            merged_entry = {"KR": kr}
            
            # 지원하는 모든 언어에 대해 순회
            supported_langs = ["EN", "CN", "TW", "JP", "DE", "FR", "TH", "PT", "ES"]
            for lang in supported_langs:
                # 각 언어별로 비어있지 않은 첫 번째 값 선택
                for entry in entries:
                    if entry.get(lang, "").strip():
                        merged_entry[lang] = entry[lang]
                        break
                
                # 값을 찾지 못했다면 빈 문자열 설정
                if lang not in merged_entry:
                    merged_entry[lang] = ""
            
            all_merged_entries.append(merged_entry)
        
        # 제외된 항목 수 저장
        self.excluded_count = excluded_count
        
        return all_merged_entries

    def _is_valid_table_name(self, table_name):
        """테이블 이름 유효성 검사 (SQL 인젝션 방지)"""
        # 영문자, 숫자, 언더스코어만 허용
        return bool(re.match(r'^[a-zA-Z0-9_]+$', table_name))

    def _load_existing_entries(self, db_path):
        """기존 unique_texts.db 로딩 (백업 생성 없이)"""
        existing_entries = {}
        
        # 프로그램 폴더에 있는 unique_texts.db 사용
        program_dir = os.path.dirname(os.path.abspath(__file__))
        db_path = self.db_path_var.get()
        
        if os.path.exists(db_path):
            # 백업 생성 코드 제거됨
            
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            # 테이블이 없으면 생성
            self._create_unique_texts_table(cursor)
                
            cursor.execute("SELECT * FROM unique_texts")
            columns = [desc[0] for desc in cursor.description]
            
            for row in cursor.execute("SELECT * FROM unique_texts"):
                record = dict(zip(columns, row))
                existing_entries[record["KR"]] = record
                
            conn.close()
        
        return existing_entries


    def _create_unique_texts_table(self, cursor):
        """unique_texts 테이블 생성 쿼리"""
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS unique_texts (
                STRING_ID TEXT PRIMARY KEY, KR TEXT, EN TEXT, CN TEXT, TW TEXT,
                TH TEXT, PT TEXT, ES TEXT, DE TEXT, FR TEXT, JP TEXT, UpdateTime TEXT
            )
        """)

    def _merge_entries(self, new_entries, existing_entries):
        """신규 항목 병합 및 ID 부여 (완전 개선된 병합 로직)"""
        merged = {}  # 결과 딕셔너리
        updated_count = 0
        next_id = self.get_next_unique_id(existing_entries)
        supported_langs = ["EN", "CN", "TW", "JP", "DE", "FR", "TH", "PT", "ES"]
        
        # 기존 DB의 항목 먼저 추가 (기본 값으로 사용)
        for kr, entry in existing_entries.items():
            merged[kr] = entry.copy()
        
        # 새 항목들을 처리
        for entry in new_entries:
            kr = entry["KR"]
            
            # 이미 해당 KR이 DB에 있는 경우
            if kr in merged:
                updated = False
                
                # 각 언어별로 DB 값이 비어있고 새 항목에 값이 있는 경우만 업데이트
                for lang in supported_langs:
                    if not merged[kr].get(lang, "") and entry.get(lang, ""):
                        merged[kr][lang] = entry[lang]
                        updated = True
                        
                if updated:
                    merged[kr]["UpdateTime"] = time.strftime("%Y-%m-%d %H:%M:%S")
                    updated_count += 1
                    
            # 처음 등장하는 KR인 경우
            else:
                entry["STRING_ID"] = next_id
                entry["UpdateTime"] = time.strftime("%Y-%m-%d %H:%M:%S")
                merged[kr] = entry
                next_id = self.increment_string_id(next_id)
                updated_count += 1
        
        return merged, updated_count



    def _save_to_database(self, db_path, entries):
        """데이터베이스에 항목 저장"""
        # 프로그램 폴더에 있는 unique_texts.db 사용
        program_dir = os.path.dirname(os.path.abspath(__file__))
        db_path = self.db_path_var.get()
        
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # 테이블 재생성
        cursor.execute("DROP TABLE IF EXISTS unique_texts")
        self._create_unique_texts_table(cursor)
        
        # 데이터 삽입
        for rec in entries.values():
            cursor.execute("""
                INSERT INTO unique_texts (
                    STRING_ID, KR, EN, CN, TW, TH, PT, ES, DE, FR, JP, UpdateTime
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                rec.get("STRING_ID", ""), rec.get("KR", ""), rec.get("EN", ""), 
                rec.get("CN", ""), rec.get("TW", ""), rec.get("TH", ""), 
                rec.get("PT", ""), rec.get("ES", ""), rec.get("DE", ""), 
                rec.get("FR", ""), rec.get("JP", ""), rec.get("UpdateTime", "")
            ))
        
        conn.commit()
        conn.close()

    def get_next_unique_id(self, existing_dict):
        existing_ids = [v["STRING_ID"] for v in existing_dict.values()]
        nums = [int(i.replace("utext_", "")) for i in existing_ids if i.startswith("utext_")]
        next_num = max(nums + [0]) + 1
        return f"utext_{next_num:05}"

    def increment_string_id(self, current_id):
        prefix, num = current_id.split("_")
        return f"{prefix}_{int(num)+1:05}"

    def get_text_tables(self, cursor):
        # 단순하게 이름에 String 들어간 테이블을 탐색
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        return [r[0] for r in cursor.fetchall() if "String" in r[0]]

    def update_kr_listbox(self):
        self.kr_listbox.delete(0, tk.END)
        for kr in sorted(self.unique_texts.keys()):
            self.kr_listbox.insert(tk.END, kr)

    def filter_kr_list(self):
        keyword = self.kr_search_var.get().strip().lower()
        self.kr_listbox.delete(0, tk.END)
        
        # filtered_kr_list가 없으면 초기화
        if not hasattr(self, 'filtered_kr_list'):
            self.filtered_kr_list = []
        else:
            self.filtered_kr_list = []
            
        for kr in sorted(self.unique_texts.keys()):
            if keyword in kr.lower():
                self.kr_listbox.insert(tk.END, kr)
                self.filtered_kr_list.append(kr)

    def on_kr_selected(self, event=None):
        selection = self.kr_listbox.curselection()
        if not selection:
            return

        kr = self.kr_listbox.get(selection[0])
        record = self.unique_texts.get(kr, {})

        self.string_id_var.set(record.get("STRING_ID", ""))
        self.selected_kr_var.set(kr)

        for lang, entry in self.lang_fields.items():
            entry.delete(0, tk.END)
            entry.insert(0, record.get(lang, ""))

    def save_translation_edits(self):
        kr = self.selected_kr_var.get()
        if not kr:
            messagebox.showwarning("경고", "KR 항목이 선택되지 않았습니다.", parent=self.root)
            return

        rec = self.unique_texts.get(kr)
        if not rec:
            messagebox.showerror("오류", "선택된 KR의 레코드를 찾을 수 없습니다.", parent=self.root)
            return

        updated = False
        for lang, entry in self.lang_fields.items():
            new_val = entry.get().strip()
            if rec.get(lang, "") != new_val:
                rec[lang] = new_val
                updated = True

        if updated:
            rec["UpdateTime"] = time.strftime("%Y-%m-%d %H:%M:%S")
            messagebox.showinfo("저장 완료", "수정된 내용이 저장되었습니다.", parent=self.root)
            self.write_entry_to_db(rec)
            # 테이블 뷰어 업데이트
            self.update_table_view()
        else:
            messagebox.showinfo("알림", "변경된 내용이 없습니다.", parent=self.root)


    def write_entry_to_db(self, record):
        # 프로그램 폴더에 있는 unique_texts.db 사용
        program_dir = os.path.dirname(os.path.abspath(__file__))
        db_path = os.path.join(program_dir, "unique_texts.db")
        
        try:
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()

            cursor.execute("""
                UPDATE unique_texts SET
                    EN = ?, CN = ?, TW = ?, JP = ?, DE = ?, FR = ?,
                    TH = ?, PT = ?, ES = ?, UpdateTime = ?
                WHERE STRING_ID = ?
            """, (
                record["EN"], record["CN"], record["TW"], record["JP"], record["DE"],
                record["FR"], record["TH"], record["PT"], record["ES"],
                record["UpdateTime"], record["STRING_ID"]
            ))

            conn.commit()
            conn.close()
        except Exception as e:
            messagebox.showerror("DB 저장 오류", f"데이터베이스에 저장하는 중 오류 발생: {e}")


    def cancel_translation_edits(self):
        self.on_kr_selected()
        

    def reload_unique_string_db(self):
        """고유 텍스트 DB를 수동으로 새로고침"""
        # 사용자 확인
        result = messagebox.askyesno(
            "DB 새로고침",
            "현재 메모리에 있는 데이터를 지우고 DB에서 다시 로드하시겠습니까?\n"
            "저장하지 않은 편집 내용은 모두 사라집니다.",
            parent=self.root
        )
        
        if not result:
            return
            
        # self.db_path_var.get()을 사용하여 DB 경로 접근
        unique_db_path = self.db_path_var.get()
        
        if os.path.exists(unique_db_path):
            # 기존 데이터 초기화 후 DB 다시 로드
            self.unique_texts = None
            self.load_unique_string_db(unique_db_path)
            messagebox.showinfo("완료", "DB를 성공적으로 다시 로드했습니다.", parent=self.root)
            self.status_label.config(text=f"DB 새로고침 완료: {unique_db_path}")
        else:
            messagebox.showwarning("경고", "DB 파일을 찾을 수 없습니다.", parent=self.root)
            self.status_label.config(text="DB 파일을 찾을 수 없습니다.")
   
   
    def show_exception_rules_manager(self):
        """예외 규칙 관리 창 표시"""
        # 새 창 생성
        rules_window = tk.Toplevel(self.root)
        rules_window.title("예외 규칙 관리")
        rules_window.geometry("800x500")
        rules_window.transient(self.root)
        rules_window.grab_set()  # 모달 창으로 설정
        
        # 상단 프레임: 규칙 추가 컨트롤
        add_frame = ttk.LabelFrame(rules_window, text="새 규칙 추가")
        add_frame.pack(fill="x", padx=10, pady=10)
        
        # 규칙 유형 선택
        ttk.Label(add_frame, text="규칙 유형:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        rule_type_var = tk.StringVar(value="startswith")
        rule_type_combo = ttk.Combobox(add_frame, textvariable=rule_type_var, width=15)
        rule_type_combo['values'] = ("startswith", "endswith", "contains", "equals", "length", "regex")
        rule_type_combo.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        
        # 필드 선택
        ttk.Label(add_frame, text="적용 필드:").grid(row=0, column=2, padx=5, pady=5, sticky="w")
        field_var = tk.StringVar(value="KR")
        field_combo = ttk.Combobox(add_frame, textvariable=field_var, width=15)
        field_combo['values'] = ("KR", "STRING_ID", "EN", "CN", "TW", "JP")
        field_combo.grid(row=0, column=3, padx=5, pady=5, sticky="w")
        
        # 값 입력
        ttk.Label(add_frame, text="값:").grid(row=0, column=4, padx=5, pady=5, sticky="w")
        value_var = tk.StringVar()
        value_entry = ttk.Entry(add_frame, textvariable=value_var, width=20)
        value_entry.grid(row=0, column=5, padx=5, pady=5, sticky="w")
        
        # 설명 입력
        ttk.Label(add_frame, text="설명:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        description_var = tk.StringVar()
        description_entry = ttk.Entry(add_frame, textvariable=description_var, width=50)
        description_entry.grid(row=1, column=1, columnspan=4, padx=5, pady=5, sticky="ew")
        
        # 추가 버튼
        def add_rule():
            rule_type = rule_type_var.get()
            field = field_var.get()
            value = value_var.get()
            description = description_var.get()
            
            if not value:
                messagebox.showwarning("경고", "값을 입력하세요.", parent=rules_window)
                return
                
            # 길이 규칙의 경우 숫자 확인
            if rule_type == "length":
                try:
                    value = int(value)
                except ValueError:
                    messagebox.showwarning("경고", "길이 규칙의 값은 숫자여야 합니다.", parent=rules_window)
                    return
                    
            # 규칙 추가
            new_rule = {
                "type": rule_type,
                "field": field,
                "value": value,
                "enabled": True,
                "description": description
            }
            
            self.exception_rules.append(new_rule)
            self.save_exception_rules()
            update_rules_display()
            
            # 입력 필드 초기화
            value_var.set("")
            description_var.set("")
            
        ttk.Button(add_frame, text="규칙 추가", command=add_rule).grid(row=1, column=5, padx=5, pady=5)
        
        # 중앙 프레임: 규칙 목록
        list_frame = ttk.LabelFrame(rules_window, text="예외 규칙 목록")
        list_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # 트리뷰 생성
        columns = ("description", "type", "field", "value", "enabled")
        rules_tree = ttk.Treeview(list_frame, columns=columns, show="headings", height=10)
        
        # 각 컬럼 설정
        rules_tree.heading("description", text="설명")
        rules_tree.heading("type", text="규칙 유형")
        rules_tree.heading("field", text="적용 필드")
        rules_tree.heading("value", text="값")
        rules_tree.heading("enabled", text="활성화")
        
        # 컬럼 너비 설정
        rules_tree.column("description", width=300)
        rules_tree.column("type", width=100)
        rules_tree.column("field", width=100)
        rules_tree.column("value", width=150)
        rules_tree.column("enabled", width=80)
        
        # 트리뷰 스크롤바
        tree_scroll = ttk.Scrollbar(list_frame, orient="vertical", command=rules_tree.yview)
        rules_tree.configure(yscrollcommand=tree_scroll.set)
        
        rules_tree.pack(side="left", fill="both", expand=True)
        tree_scroll.pack(side="right", fill="y")
        
        def update_rules_display():
            # 트리뷰 초기화
            for item in rules_tree.get_children():
                rules_tree.delete(item)
                
            # 규칙 표시
            for i, rule in enumerate(self.exception_rules):
                enabled_text = "활성화" if rule["enabled"] else "비활성화"
                rules_tree.insert("", "end", values=(
                    rule.get("description", ""),
                    rule["type"],
                    rule["field"],
                    rule["value"],
                    enabled_text
                ), tags=(str(i),))
        
        # 하단 프레임: 버튼 영역
        button_frame = ttk.Frame(rules_window)
        button_frame.pack(fill="x", padx=10, pady=10)
        
        def toggle_selected_rule():
            selected = rules_tree.selection()
            if not selected:
                messagebox.showinfo("알림", "규칙을 선택하세요.", parent=rules_window)
                return
                
            # 선택된 규칙의 인덱스 찾기
            item = selected[0]
            index = int(rules_tree.item(item, "tags")[0])
            
            # 규칙 활성화 상태 변경
            self.exception_rules[index]["enabled"] = not self.exception_rules[index]["enabled"]
            self.save_exception_rules()
            update_rules_display()
        
        def delete_selected_rule():
            selected = rules_tree.selection()
            if not selected:
                messagebox.showinfo("알림", "규칙을 선택하세요.", parent=rules_window)
                return
                
            # 확인 대화상자
            if not messagebox.askyesno("확인", "선택한 규칙을 삭제하시겠습니까?", parent=rules_window):
                return
                
            # 선택된 규칙의 인덱스 찾기
            item = selected[0]
            index = int(rules_tree.item(item, "tags")[0])
            
            # 규칙 삭제
            del self.exception_rules[index]
            self.save_exception_rules()
            update_rules_display()
        
        def reset_rules():
            # 확인 대화상자
            if not messagebox.askyesno("확인", "모든 규칙을 기본값으로 초기화하시겠습니까?", parent=rules_window):
                return
                
            # 기본 규칙으로 초기화
            self.exception_rules = self.default_rules.copy()
            self.save_exception_rules()
            update_rules_display()
        
        # 버튼 추가
        ttk.Button(button_frame, text="활성화/비활성화", command=toggle_selected_rule).pack(side="left", padx=5)
        ttk.Button(button_frame, text="규칙 삭제", command=delete_selected_rule).pack(side="left", padx=5)
        ttk.Button(button_frame, text="기본값으로 초기화", command=reset_rules).pack(side="left", padx=5)
        ttk.Button(button_frame, text="창 닫기", command=rules_window.destroy).pack(side="right", padx=5)
        
        # 초기 규칙 표시
        update_rules_display()

    def load_exception_rules(self):
        """저장된 예외 규칙 로드"""
        program_dir = os.path.dirname(os.path.abspath(__file__))
        rules_path = os.path.join(program_dir, "exception_rules.json")
        
        if os.path.exists(rules_path):
            try:
                import json
                with open(rules_path, 'r', encoding='utf-8') as f:
                    self.exception_rules = json.load(f)
            except Exception as e:
                # 로드 실패 시 기본 규칙 사용
                self.exception_rules = self.default_rules.copy()
                messagebox.showwarning("경고", f"예외 규칙 로드 실패: {str(e)}", parent=self.root)
        else:
            # 파일이 없으면 기본 규칙 사용
            self.exception_rules = self.default_rules.copy()
            self.save_exception_rules()  # 기본 규칙 저장

    def save_exception_rules(self):
        """예외 규칙을 파일에 저장"""
        program_dir = os.path.dirname(os.path.abspath(__file__))
        rules_path = os.path.join(program_dir, "exception_rules.json")
        
        try:
            import json
            with open(rules_path, 'w', encoding='utf-8') as f:
                json.dump(self.exception_rules, f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showerror("오류", f"예외 규칙 저장 실패: {str(e)}", parent=self.root)

    def should_exclude_entry(self, entry):
        """예외 규칙을 적용하여 제외할지 여부 결정"""
        for rule in self.exception_rules:
            if not rule["enabled"]:
                continue
                
            field = rule["field"]
            field_value = str(entry.get(field, ""))
            
            if rule["type"] == "startswith":
                if field_value.startswith(rule["value"]):
                    return True
            elif rule["type"] == "endswith":
                if field_value.endswith(rule["value"]):
                    return True
            elif rule["type"] == "contains":
                if rule["value"] in field_value:
                    return True
            elif rule["type"] == "equals":
                if field_value == rule["value"]:
                    return True
            elif rule["type"] == "length":
                max_length = int(rule["value"])
                if len(field_value) > max_length:
                    return True
            elif rule["type"] == "regex":
                import re
                if re.match(rule["value"], field_value):
                    return True

                    
        return False


    # _collect_entries_from_dbs 메서드 수정 (기존 메서드를 대체)
    def _collect_entries_from_dbs(self, db_files):
        """선택된 DB 파일에서 항목 수집 (완전 개선된 고유값 병합 로직)"""
        kr_entries_map = {}  # KR 텍스트를 키로 하는 항목 모음
        excluded_count = 0  # 예외 규칙으로 제외된 항목 수 추적
        supported_langs = ["EN", "CN", "TW", "JP", "DE", "FR", "TH", "PT", "ES"]
        
        for db_file in db_files:
            try:
                conn = sqlite3.connect(db_file)
                cursor = conn.cursor()
                
                for table in self.get_text_tables(cursor):
                    # SQL 인젝션 방지를 위한 처리
                    if not self._is_valid_table_name(table):
                        continue
                        
                    cursor.execute(f"SELECT * FROM {table}")
                    columns = [desc[0] for desc in cursor.description]
                    rows = cursor.fetchall()
                    
                    for row in rows:
                        row_data = dict(zip(columns, row))
                        kr = str(row_data.get("KR", "")).strip()
                        
                        # KR이 비어있지 않은 경우만 처리
                        if kr:
                            # 예외 규칙 검사
                            if self.should_exclude_entry(row_data):
                                excluded_count += 1
                                continue
                                
                            # 해당 KR이 처음 등장하면 초기화
                            if kr not in kr_entries_map:
                                kr_entries_map[kr] = {lang: [] for lang in supported_langs}
                                
                            # 각 언어별로 값을 따로 수집
                            for lang in supported_langs:
                                if lang in row_data and row_data.get(lang, "").strip():
                                    kr_entries_map[kr][lang].append(row_data.get(lang, ""))
                conn.close()
            except Exception as e:
                messagebox.showwarning("DB 오류", f"파일 '{os.path.basename(db_file)}' 처리 중 오류: {str(e)}", parent=self.root)
        
        # 각 KR별로 최적의 번역 선택하여 병합
        all_merged_entries = []
        for kr, lang_values in kr_entries_map.items():
            merged_entry = {"KR": kr}
            
            # 각 언어별로 값 선택
            for lang in supported_langs:
                if lang_values[lang]:  # 해당 언어에 수집된 값이 있는 경우
                    # 첫 번째 값 사용 (필요시 다른 선택 로직 적용 가능)
                    merged_entry[lang] = lang_values[lang][0]
                else:
                    # 값이 없으면, 빈 문자열 설정
                    merged_entry[lang] = ""
            
            all_merged_entries.append(merged_entry)
        
        # 제외된 항목 수 저장
        self.excluded_count = excluded_count
        
        return all_merged_entries


    def export_unique_texts_to_excel(self):
        """현재 메모리의 유니크 텍스트 DB를 엑셀 파일로 내보내기"""
        import pandas as pd
        from tkinter import filedialog, messagebox

        if not self.unique_texts:
            messagebox.showinfo("알림", "엑셀로 내보낼 데이터가 없습니다.", parent=self.root)
            return

        # 저장 경로 선택
        file_path = filedialog.asksaveasfilename(
            title="유니크 텍스트 엑셀로 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")]
        )
        if not file_path:
            return

        try:
            # unique_texts는 KR을 key로 하는 dict임
            df = pd.DataFrame(list(self.unique_texts.values()))
            # 컬럼 순서 지정(원하는대로 정렬)
            columns = ["STRING_ID", "KR", "EN", "CN", "TW", "JP", "DE", "FR", "TH", "PT", "ES", "UpdateTime"]
            df = df[[col for col in columns if col in df.columns]]
            df.to_excel(file_path, index=False)
            messagebox.showinfo("완료", f"엑셀로 저장 완료!\n{file_path}", parent=self.root)
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 내보내기 실패: {e}", parent=self.root)



    def update_translations_from_excel(self):
        """엑셀 파일에서 다국어 데이터를 가져와 DB 업데이트"""
        # 1. 엑셀 파일 선택
        excel_path = filedialog.askopenfilename(
            title="다국어 엑셀 파일 선택",
            filetypes=[("Excel 파일", "*.xlsx *.xls")],
            parent=self.root
        )
        
        if not excel_path:
            return  # 사용자가 취소함
        
        # 포커스 복원
        self.root.after(100, self.root.focus_force)
        self.root.after(100, self.root.lift)
        
        try:
            # 2. 엑셀 파일 로드
            import pandas as pd
            
            # 엑셀 파일의 모든 시트 확인
            excel_file = pd.ExcelFile(excel_path)
            
            # 'string' 이름을 포함한 시트 찾기
            string_sheets = [sheet for sheet in excel_file.sheet_names if '#string' in sheet.lower()]
            
            if not string_sheets:
                messagebox.showwarning("경고", "엑셀 파일에 'string'이 포함된 시트가 없습니다.", parent=self.root)
                return
            
            # 첫 번째 string 시트 사용
            sheet_name = string_sheets[0]
            df = pd.read_excel(excel_path, sheet_name=sheet_name)
            
            # 3. 'KR' 컬럼 확인
            if 'KR' not in df.columns:
                messagebox.showwarning("경고", "선택한 시트에 'KR' 컬럼이 없습니다.", parent=self.root)
                return
            
            # 지원하는 언어 컬럼 목록
            supported_langs = ['EN', 'CN', 'TW', 'JP', 'DE', 'FR', 'TH', 'PT', 'ES']
            
            # 있는 언어 컬럼만 필터링
            available_langs = [lang for lang in supported_langs if lang in df.columns]
            
            if not available_langs:
                messagebox.showwarning("경고", "선택한 시트에 번역 언어 컬럼이 없습니다.", parent=self.root)
                return
            
            # 4. 데이터 로드 확인
            if not self.unique_texts:
                # DB가 로드되지 않은 경우 로드
                program_dir = os.path.dirname(os.path.abspath(__file__))
                unique_db_path = self.db_path_var.get()
                
                if os.path.exists(unique_db_path):
                    self.load_unique_string_db(unique_db_path)
                else:
                    messagebox.showwarning("경고", "unique_texts.db 파일을 찾을 수 없습니다.", parent=self.root)
                    return
            
            # 5. 다국어 업데이트 처리
            update_count = 0
            total_rows = len(df)
            matched_kr_count = 0
            
            # 진행 상황 표시 창
            progress_window = tk.Toplevel(self.root)
            progress_window.title("다국어 갱신 진행 중")
            progress_window.geometry("400x150")
            progress_window.transient(self.root)
            progress_window.grab_set()
            
            ttk.Label(progress_window, text="다국어 데이터 갱신 중...").pack(pady=10)
            
            progress_var = tk.DoubleVar()
            progress_bar = ttk.Progressbar(progress_window, variable=progress_var, maximum=100)
            progress_bar.pack(fill="x", padx=20, pady=10)
            
            status_var = tk.StringVar(value="처리 중...")
            status_label = ttk.Label(progress_window, textvariable=status_var)
            status_label.pack(pady=10)
            
            # DB 연결
            db_path = self.db_path_var.get()
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            # 각 행 처리
            for idx, row in df.iterrows():
                # 진행률 업데이트
                progress_var.set((idx / total_rows) * 100)
                status_var.set(f"처리 중... ({idx}/{total_rows})")
                progress_window.update()
                
                kr_value = str(row['KR']).strip()
                if not kr_value:
                    continue
                
                # KR 값이 DB에 있는지 확인
                if kr_value in self.unique_texts:
                    matched_kr_count += 1
                    db_record = self.unique_texts[kr_value]
                    
                    # 각 언어 확인 및 업데이트
                    updated_fields = []
                    update_values = []
                    
                    for lang in available_langs:
                        excel_value = row.get(lang, '')
                        
                        # 엑셀 값이 비어있지 않고, DB 값과 다른 경우만 업데이트
                        if pd.notna(excel_value) and excel_value and excel_value != db_record.get(lang, ''):
                            updated_fields.append(f"{lang} = ?")
                            update_values.append(str(excel_value).strip())
                    
                    # 업데이트할 필드가 있으면 쿼리 실행
                    if updated_fields:
                        update_query = f"UPDATE unique_texts SET {', '.join(updated_fields)}, UpdateTime = ? WHERE STRING_ID = ?"
                        update_values.append(time.strftime("%Y-%m-%d %H:%M:%S"))
                        update_values.append(db_record["STRING_ID"])
                        
                        cursor.execute(update_query, update_values)
                        update_count += 1
            
            # 변경사항 저장
            conn.commit()
            conn.close()
            
            # 진행 창 닫기
            progress_window.destroy()
            
            # 메모리 데이터 다시 로드
            self.load_unique_string_db(db_path)
            
            # 결과 메시지
            messagebox.showinfo(
                "다국어 갱신 완료", 
                f"총 {total_rows}개 항목 중 {matched_kr_count}개 항목이 매칭되었고,\n"
                f"{update_count}개 항목이 업데이트되었습니다.",
                parent=self.root
            )
            
        except Exception as e:
            messagebox.showerror("오류", f"다국어 갱신 중 오류가 발생했습니다:\n{str(e)}", parent=self.root)
            import traceback
            traceback.print_exc()

    def update_db_path_display(self):
        """DB 경로 정보 표시 업데이트"""
        self.info_label.config(text=f"현재 DB 경로: {self.db_path_var.get()}")
                
# 메인 실행 코드
def run_unique_text_manager(parent=None):
    """고유 텍스트 관리자 실행"""
    if parent:
        root = tk.Toplevel(parent)
    else:
        root = tk.Tk()
        
    app = UniqueTextManager(root)
    
    if not parent:
        root.mainloop()
    
    return app