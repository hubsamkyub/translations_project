import tkinter as tk
import os
import sqlite3
import json
import re
import time
import gc
import threading
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook, Workbook
import pandas as pd

# 유틸리티 모듈 임포트
from utils.config_utils import load_config, save_config
from utils.common_utils import PathUtils, FileUtils, DBUtils, show_message, logger
from utils.excel_utils import ExcelFileManager

# config 유틸리티 함수
def load_config():
    """설정 파일에서 설정 불러오기"""
    try:
        with open("config.json", "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def save_config(config):
    """설정 파일에 설정 저장"""
    try:
        with open("config.json", "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"설정 저장 실패: {e}")


class ScrollableCheckList(tk.Frame):
    def __init__(self, parent, width=300, height=150, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.config = load_config()
        
        self.original_folder_var = tk.StringVar()
        default_path = self.config.get("data_path", "")
        self.original_folder_var.set(default_path)
        
        self.canvas = tk.Canvas(self, width=width, height=height, borderwidth=0)
        self.scrollbar = tk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.inner_frame = tk.Frame(self.canvas)
        self.canvas_window = self.canvas.create_window((0, 0), window=self.inner_frame, anchor="nw")

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

        self.inner_frame.bind("<Configure>", self._on_frame_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        
        # 전체 선택/해제 변수 및 컨트롤
        self.check_all_var = tk.BooleanVar(value=True)
        self.check_all_frame = tk.Frame(self)
        self.check_all_frame.pack(fill="x", before=self.canvas)
        
        self.check_all_cb = tk.Checkbutton(
            self.check_all_frame, 
            text="전체 선택/해제", 
            variable=self.check_all_var,
            command=self._toggle_all
        )
        self.check_all_cb.pack(side="left", padx=5, pady=2)
        
        self.vars_dict = {}
        
        # 마우스 휠 이벤트 바인딩 (모든 가능한 플랫폼 이벤트 처리)
        self._bind_mouse_wheel(self.canvas)
        self._bind_mouse_wheel(self.inner_frame)
        
        # 위젯에 포커스가 있을 때만 마우스 휠 작동하도록 포커스 이벤트 추가
        self.canvas.bind("<Enter>", lambda e: self.canvas.focus_set())
        self.canvas.bind("<Leave>", lambda e: self.root.focus_set() if hasattr(self, 'root') else None)

    def _bind_mouse_wheel(self, widget):
        """여러 플랫폼에 대한 마우스 휠 이벤트 바인딩"""
        widget.bind("<MouseWheel>", self._on_mouse_wheel)       # Windows
        widget.bind("<Button-4>", self._on_mouse_wheel)         # Linux 위로 스크롤
        widget.bind("<Button-5>", self._on_mouse_wheel)         # Linux 아래로 스크롤
        widget.bind("<Button-2>", self._on_mouse_wheel)         # macOS/Linux 중간 버튼

    def _on_mouse_wheel(self, event):
        """마우스 휠 이벤트 처리 함수"""
        # Windows의 경우
        if event.num == 5 or event.delta < 0:  # 아래로 스크롤
            self.canvas.yview_scroll(1, "units")
        elif event.num == 4 or event.delta > 0:  # 위로 스크롤
            self.canvas.yview_scroll(-1, "units")
        return "break"  # 이벤트 전파 중단

    def _on_frame_configure(self, event):
        # 내부 프레임 크기가 변경되면 스크롤 영역 업데이트
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
    
    def _on_canvas_configure(self, event):
        # 캔버스 크기가 변경되면 내부 프레임의 너비 조정
        width = event.width
        self.canvas.itemconfig(self.canvas_window, width=width)
    
    def _toggle_all(self):
        # 전체 선택/해제 토글 처리
        checked = self.check_all_var.get()
        self.set_all_checked(checked)

    def add_item(self, text, checked=True):
        var = tk.BooleanVar(value=checked)
        cb = tk.Checkbutton(self.inner_frame, text=text, variable=var, anchor="w")
        cb.pack(fill="x", anchor="w")
        
        # 새로 추가된 체크박스에도 마우스 휠 이벤트 바인딩
        self._bind_mouse_wheel(cb)
        
        self.vars_dict[text] = var
        
        # 체크박스 상태 변경 시 전체 선택 상태 업데이트
        var.trace_add("write", lambda *args: self._update_check_all_state())

    def clear(self):
        for w in self.inner_frame.winfo_children():
            w.destroy()
        self.vars_dict.clear()
        self.check_all_var.set(True)  # 기본값으로 초기화

    def get_checked_items(self):
        return [t for (t, v) in self.vars_dict.items() if v.get()]

    def set_all_checked(self, checked=True):
        for var in self.vars_dict.values():
            var.set(checked)
    
    def _update_check_all_state(self):
        # 모든 항목이 선택되었는지 확인하여 전체 선택 체크박스 상태 업데이트
        if not self.vars_dict:  # 목록이 비어있으면 처리하지 않음
            return
        
        all_checked = all(var.get() for var in self.vars_dict.values())
        self.check_all_var.set(all_checked)


class TranslationRequestExtractor(tk.Frame):
    def __init__(self, root):
        super().__init__(root)
        # root가 Tk인지 Frame인지 구분 (독립 실행 vs 탭 임베드)
        if isinstance(root, tk.Tk):
            self.root = root
            self.root.title("번역 요청 추출 도구")
            self.root.geometry("1200x800")
        else:
            # 부모 프레임으로 사용
            self.root = None  # 나중에 TranslationAutomationTool의 root로 설정됨
        
        self.config = load_config()
        
        # 엑셀 파일 폴더 선택 프레임
        folder_frame = ttk.LabelFrame(self, text="엑셀 파일 폴더 선택")
        folder_frame.pack(fill="x", padx=5, pady=5)
        
        ttk.Label(folder_frame, text="폴더 경로:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.folder_path_var = tk.StringVar(value=self.config.get("excel_folder", ""))
        ttk.Entry(folder_frame, textvariable=self.folder_path_var, width=50).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(folder_frame, text="찾아보기", command=self.select_excel_folder).grid(row=0, column=2, padx=5, pady=5)
        ttk.Button(folder_frame, text="파일 검색", command=self.search_excel_files).grid(row=0, column=3, padx=5, pady=5)
        
        folder_frame.columnconfigure(1, weight=1)
        
        # 파일 목록 프레임
        files_frame = ttk.LabelFrame(self, text="엑셀 파일 목록")
        files_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        self.excel_files_list = ScrollableCheckList(files_frame, width=700, height=200)
        self.excel_files_list.pack(fill="both", expand=True, padx=5, pady=5)
        
        # DB 출력 설정 프레임
        output_frame = ttk.LabelFrame(self, text="출력 설정")
        output_frame.pack(fill="x", padx=5, pady=5)
        
        ttk.Label(output_frame, text="DB 파일:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.output_db_var = tk.StringVar(value=self.config.get("output_db", ""))
        ttk.Entry(output_frame, textvariable=self.output_db_var, width=50).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(output_frame, text="찾아보기", command=self.select_output_db).grid(row=0, column=2, padx=5, pady=5)

        # 추출 조건 프레임
        condition_frame = ttk.LabelFrame(output_frame, text="추출 조건 설정")
        condition_frame.grid(row=1, column=0, columnspan=3, sticky="ew", padx=5, pady=5)

        # 기본 조건 (항상 적용)
        basic_frame = ttk.Frame(condition_frame)
        basic_frame.pack(fill="x", padx=5, pady=2)
        
        # 신규 조건
        self.extract_new_request_var = tk.BooleanVar(value=True)
        new_request_cb = ttk.Checkbutton(basic_frame, text='#번역요청이 "신규"인 항목', variable=self.extract_new_request_var)
        new_request_cb.pack(side="left", padx=5)
        
        # change 조건 추가
        self.extract_change_request_var = tk.BooleanVar(value=False)
        change_request_cb = ttk.Checkbutton(basic_frame, text='#번역요청이 "change"인 항목', variable=self.extract_change_request_var)
        change_request_cb.pack(side="left", padx=15)

        # 추가 조건 (선택 사항)
        additional_frame = ttk.LabelFrame(condition_frame, text="추가 조건 (선택사항)")
        additional_frame.pack(fill="x", padx=5, pady=5)

        empty_frame = ttk.Frame(additional_frame)
        empty_frame.pack(fill="x", padx=5, pady=2)

        self.extract_en_empty_var = tk.BooleanVar(value=False)
        self.extract_cn_empty_var = tk.BooleanVar(value=False)  # 디폴트 해제
        self.extract_tw_empty_var = tk.BooleanVar(value=False)  # 디폴트 해제

        ttk.Checkbutton(empty_frame, text="EN이 비어있는 항목", variable=self.extract_en_empty_var).pack(side="left", padx=10, pady=2)
        ttk.Checkbutton(empty_frame, text="CN이 비어있는 항목", variable=self.extract_cn_empty_var).pack(side="left", padx=10, pady=2)
        ttk.Checkbutton(empty_frame, text="TW가 비어있는 항목", variable=self.extract_tw_empty_var).pack(side="left", padx=10, pady=2)

        
        # 성능 최적화 옵션 추가
        optimization_frame = ttk.Frame(output_frame)
        optimization_frame.grid(row=2, column=0, columnspan=3, sticky="ew", padx=5, pady=5)

        # 기존 성능 최적화 옵션 프레임 행 번호 수정
        optimization_frame = ttk.Frame(output_frame)
        optimization_frame.grid(row=2, column=0, columnspan=3, sticky="ew", padx=5, pady=5)  # row=1에서 row=2로 변경
        
        self.batch_size_var = tk.IntVar(value=100)
        ttk.Label(optimization_frame, text="일괄 처리 크기:").pack(side="left", padx=5)
        ttk.Spinbox(optimization_frame, from_=10, to=1000, increment=10, textvariable=self.batch_size_var, width=5).pack(side="left", padx=5)
        
        self.use_read_only_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(optimization_frame, text="읽기 전용 모드 사용 (빠름)", variable=self.use_read_only_var).pack(side="left", padx=15)
        
        output_frame.columnconfigure(1, weight=1)
        
        # 실행 버튼 프레임
        action_frame = ttk.Frame(self)
        action_frame.pack(fill="x", padx=5, pady=5)

        ttk.Button(action_frame, text="번역 요청 항목 추출", command=self.extract_translation_requests).pack(side="right", padx=5, pady=5)
        ttk.Button(action_frame, text="템플릿으로 내보내기", command=self.export_to_template_excel).pack(side="right", padx=5, pady=5)
        ttk.Button(action_frame, text="Excel로 내보내기", command=self.export_to_excel).pack(side="right", padx=5, pady=5)
        ttk.Button(action_frame, text="컬럼 캐시 초기화", command=self.reset_column_cache).pack(side="right", padx=5, pady=5)

        
        # 로그 표시 영역
        log_frame = ttk.LabelFrame(self, text="작업 로그")
        log_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        self.log_text = tk.Text(log_frame, wrap="word", height=10)
        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side="right", fill="y")
        self.log_text.pack(fill="both", expand=True)
        
        # 상태 표시줄
        status_frame = ttk.Frame(self)
        status_frame.pack(fill="x", padx=5, pady=5)
        
        self.status_label = ttk.Label(status_frame, text="대기 중...")
        self.status_label.pack(side="left", padx=5)
        
        self.progress_bar = ttk.Progressbar(status_frame, length=400, mode="determinate")
        self.progress_bar.pack(side="right", fill="x", expand=True, padx=5)
        
        # 내부 저장용
        self.excel_files = []
        self.column_cache = {}  # 컬럼 캐시
        self.extraction_thread = None  # 추출 작업용 스레드
        self.cancel_extraction = False  # 작업 취소 플래그
        
        # 폼 표시
        self.pack(fill="both", expand=True)
        
        # 시작 시 컬럼 캐시 로드
        self.load_column_cache()
    
    def select_excel_folder(self):
        """엑셀 파일 폴더 선택"""
        folder = filedialog.askdirectory(title="엑셀 파일 폴더 선택", parent=self.root)
        if folder:
            self.folder_path_var.set(folder)
            self.config["excel_folder"] = folder
            save_config(self.config)
            # 포커스를 다시 자동화 툴 창으로
            self.root.after(100, self.root.focus_force)
            self.root.after(100, self.root.lift)
    
    def select_output_db(self):
        """DB 파일 저장 경로 선택"""
        file_path = filedialog.asksaveasfilename(
            defaultextension=".db",
            filetypes=[("DB 파일", "*.db"), ("모든 파일", "*.*")],
            title="번역 요청 DB 파일 저장",
            parent=self.root
        )
        if file_path:
            self.output_db_var.set(file_path)
            self.config["output_db"] = file_path
            save_config(self.config)
            # 포커스를 다시 자동화 툴 창으로
            self.root.after(100, self.root.focus_force)
            self.root.after(100, self.root.lift)
    
    def search_excel_files(self):
        """엑셀 파일 검색 - common_utils 활용"""
        folder = self.folder_path_var.get()
        if not folder or not os.path.isdir(folder):
            show_message(self.root, "warning", "경고", "유효한 폴더를 선택하세요.")
            return
        
        self.excel_files_list.clear()
        self.excel_files = []
        
        self.log_text.delete(1.0, tk.END)
        self.log_text.insert(tk.END, "파일 검색 중...\n")
        self.root.update()
        
        # PathUtils 활용
        found_files = PathUtils.find_files(folder, ".xlsx", recursive=True)
        
        # String으로 시작하는 파일만 필터링
        string_files = []
        for file_path in found_files:
            file_name = os.path.basename(file_path)
            if file_name.lower().startswith("string") and not file_name.startswith("~$"):
                string_files.append((file_name, file_path))
        
        # 결과 처리
        if not string_files:
            self.log_text.insert(tk.END, "String으로 시작하는 엑셀 파일을 찾지 못했습니다.\n")
            show_message(self.root, "info", "알림", "String으로 시작하는 엑셀 파일을 찾지 못했습니다.")
        else:
            self.excel_files = string_files
            for file_name, _ in string_files:
                self.excel_files_list.add_item(file_name, checked=True)
                
            self.log_text.insert(tk.END, f"{len(string_files)}개의 엑셀 파일을 찾았습니다.\n")
            show_message(self.root, "info", "알림", f"{len(string_files)}개의 엑셀 파일을 찾았습니다.")
    
    def find_translation_request_column(self, worksheet):
        """#번역요청 컬럼 찾기 (공백 무시) - ExcelUtils 활용"""
        # ExcelFileManager에 적절한 메서드가 없으므로 커스텀 함수 유지
        # 대신 로깅 추가하여 개선
        logger.debug(f"#번역요청 컬럼 검색 시작")
        
        for row in range(2, 6):  # 2행부터 5행까지 검색
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    # 공백 제거 후 비교
                    if re.sub(r'\s+', '', cell_value).upper() == "#번역요청":
                        logger.debug(f"#번역요청 컬럼 발견: 열={col}, 행={row}")
                        return col, row
        
        logger.debug(f"#번역요청 컬럼을 찾지 못함")
        return None, None
    
    def find_string_id_position(self, worksheet):
        """STRING_ID 컬럼 위치 찾기"""
        for row in range(2, 6):  # 2행부터 5행까지 검색
            for col in range(1, min(10, worksheet.max_column + 1)):  # 최대 10개 컬럼까지만 검색
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str) and "STRING_ID" in cell_value.upper():
                    return col, row
        return None, None
    
    def reset_column_cache(self):
        """컬럼 캐시 초기화"""
        if messagebox.askyesno("확인", "컬럼 캐시를 초기화하시겠습니까?", parent=self.root):
            self.column_cache = {}
            self.save_column_cache()
            self.log_text.insert(tk.END, "컬럼 캐시를 초기화했습니다.\n")
    
    
    def extract_translation_requests(self):
        """번역 요청 항목 추출 작업 시작 (백그라운드 스레드)"""
        # 1. 중복 실행 방지
        if self.extraction_thread and self.extraction_thread.is_alive():
            show_message(self.root, "info", "알림", "이미 추출 작업이 진행 중입니다.")
            return
        
        # 2. 파일 선택 확인
        selected_files = self.excel_files_list.get_checked_items()
        if not selected_files:
            show_message(self.root, "warning", "경고", "파일을 선택하세요.")
            return
        
        # 3. DB 파일 경로 확인
        db_path = self.output_db_var.get()
        if not db_path:
            show_message(self.root, "warning", "경고", "DB 파일 경로를 지정하세요.")
            return
        
        # 4. 추출 조건 검증
        extract_conditions = {
            'new_request': self.extract_new_request_var.get(),
            'change_request': self.extract_change_request_var.get(),
            'en_empty': self.extract_en_empty_var.get(),
            'cn_empty': self.extract_cn_empty_var.get(),
            'tw_empty': self.extract_tw_empty_var.get()
        }
        
        # 최소 하나의 조건은 선택되어야 함
        if not any(extract_conditions.values()):
            show_message(self.root, "warning", "경고", "추출할 조건을 최소 하나 이상 선택하세요.")
            return
        
        # 5. 선택된 조건들 확인 및 로깅
        selected_conditions = []
        condition_labels = {
            'new_request': '#번역요청="신규"',
            'change_request': '#번역요청="change"',
            'en_empty': 'EN빈칸',
            'cn_empty': 'CN빈칸', 
            'tw_empty': 'TW빈칸'
        }
        
        for condition_key, is_selected in extract_conditions.items():
            if is_selected:
                selected_conditions.append(condition_labels[condition_key])
        
        # 6. 작업 시작 로깅
        logger.info(f"번역 요청 추출 작업 시작:")
        logger.info(f"  대상 파일: {len(selected_files)}개")
        logger.info(f"  추출 조건: {', '.join(selected_conditions)}")
        logger.info(f"  출력 DB: {db_path}")
        
        # 로그 텍스트에도 조건 표시
        self.log_text.insert(tk.END, f"추출 조건: {', '.join(selected_conditions)}\n")
        
        # 7. 기존 DB 파일 확인
        if os.path.exists(db_path):
            if not show_message(self.root, "yesno", "확인", 
                            f"'{os.path.basename(db_path)}' 파일이 이미 존재합니다.\n"
                            f"덮어쓰시겠습니까?\n\n"
                            f"추출 조건: {', '.join(selected_conditions)}"):
                return
        
        # 8. 작업 실행
        self._setup_extraction_thread(selected_files, db_path)


    def _setup_extraction_thread(self, selected_files, db_path):
        """추출 작업용 백그라운드 스레드 설정"""
        # 진행 상태 초기화
        self.log_text.delete(1.0, tk.END)
        self.progress_bar["value"] = 0
        self.status_label.config(text="번역 요청 추출 시작...")
        
        # 취소 플래그 초기화
        self.cancel_extraction = False
        
        # 작업 스레드 생성 및 시작
        self.extraction_thread = threading.Thread(
            target=self._extract_in_background,
            args=(selected_files, db_path)
        )
        self.extraction_thread.daemon = True
        self.extraction_thread.start()
        
        # 진행 상태 모니터링
        self.root.after(100, self._check_extraction_progress)

    def _check_extraction_progress(self):
        """추출 작업 진행 상태 확인"""
        if self.extraction_thread and self.extraction_thread.is_alive():
            # 작업이 진행 중이면 상태 확인 계속
            self.root.after(100, self._check_extraction_progress)
        else:
            # 작업 완료 또는 실패
            if hasattr(self, 'extraction_thread') and self.extraction_thread:
                self.extraction_thread = None


    def _extract_in_background(self, selected_files, db_path):
        """백그라운드 스레드에서 번역 요청 추출 작업 실행"""
        start_time = time.time()
        
        try:
            # DB 초기화
            if os.path.exists(db_path):
                os.remove(db_path)
            
            # DB 성능 최적화 설정으로 생성
            conn = sqlite3.connect(db_path)
            conn.execute("PRAGMA journal_mode = WAL")
            conn.execute("PRAGMA synchronous = NORMAL")
            conn.execute("PRAGMA cache_size = 10000")
            cursor = conn.cursor()
            
            # 테이블 생성
            cursor.execute('''
            CREATE TABLE translation_requests (
                id INTEGER PRIMARY KEY,
                file_name TEXT,
                sheet_name TEXT,
                string_id TEXT,
                kr TEXT,
                en TEXT,
                cn TEXT,
                tw TEXT,
                th TEXT,
                request_type TEXT,
                additional_info TEXT
            )
            ''')
            
            # 설정값 가져오기
            batch_size = self.batch_size_var.get()
            use_read_only = self.use_read_only_var.get()
            total_files = len(selected_files)
            
            # 추출 조건 옵션 가져오기
            extract_conditions = {
                'new_request': self.extract_new_request_var.get(),
                'change_request': self.extract_change_request_var.get(),
                'en_empty': self.extract_en_empty_var.get(),
                'cn_empty': self.extract_cn_empty_var.get(),
                'tw_empty': self.extract_tw_empty_var.get()
            }
            
            # UI 업데이트 함수
            def update_ui(progress_value=None, status_text=None, log_text=None, see_end=False):
                def _update():
                    if status_text:
                        self.status_label.config(text=status_text)
                    if progress_value is not None:
                        self.progress_bar["maximum"] = total_files
                        self.progress_bar["value"] = progress_value
                    if log_text:
                        self.log_text.insert(tk.END, log_text)
                        if see_end:
                            self.log_text.see(tk.END)
                    self.root.update_idletasks()
                self.root.after(0, _update)
            
            # 초기 UI 업데이트
            update_ui(0, "번역 요청 항목 추출 시작...", "작업 시작...\n", True)
            
            # 작업 통계 변수
            total_rows = 0
            processed_count = 0
            error_count = 0
            batch_data = []
            
            # 파일별 처리
            for idx, file_name in enumerate(selected_files):
                if self.cancel_extraction:
                    update_ui(None, "작업 취소됨", "\n작업이 사용자에 의해 취소되었습니다.\n", True)
                    break
                
                # 파일 경로 찾기
                file_path = next((path for name, path in self.excel_files if name == file_name), None)
                if not file_path:
                    continue
                
                update_ui(idx, f"처리 중: {file_name} ({idx+1}/{total_files})", f"\n{file_name} 처리 중...\n", True)
                
                try:
                    # 엑셀 파일 로드
                    workbook = load_workbook(file_path, read_only=use_read_only, data_only=True)
                    
                    # String 시트 찾기
                    string_sheets = [sheet for sheet in workbook.sheetnames 
                                if sheet.lower().startswith("string") and not sheet.startswith("#")]
                    
                    file_rows = 0
                    
                    # 시트별 처리
                    for sheet_name in string_sheets:
                        if self.cancel_extraction:
                            break
                        
                        update_ui(None, None, f"  시트 {sheet_name} 처리 중...\n", True)
                        worksheet = workbook[sheet_name]
                        
                        # 컬럼 정보 가져오기 (캐시 활용)
                        columns_info = self._get_column_info(file_name, sheet_name, worksheet, extract_conditions, update_ui)
                        if not columns_info:
                            update_ui(None, None, f"    필요한 컬럼을 찾을 수 없습니다.\n", True)
                            continue
                        
                        # 데이터 추출
                        sheet_rows = self._extract_sheet_data(
                            worksheet, columns_info, file_name, sheet_name, 
                            batch_data, cursor, conn, batch_size, use_read_only, extract_conditions, update_ui
                        )
                        
                        file_rows += sheet_rows
                        update_ui(None, None, f"    {sheet_rows}개 항목 추출됨\n", True)
                    
                    # 워크북 정리
                    workbook.close()
                    gc.collect()
                    
                    processed_count += 1
                    update_ui(None, None, f"  총 {file_rows}개 항목 처리 완료\n", True)
                    
                except Exception as e:
                    error_msg = str(e)
                    update_ui(None, None, f"  오류 발생: {error_msg}\n", True)
                    error_count += 1
                
                # 진행 상태 업데이트
                update_ui(idx + 1)
            
            # 남은 배치 데이터 처리
            if batch_data:
                self._insert_batch_data(cursor, conn, batch_data)
                total_rows += len(batch_data)
            
            # 최종 통계 계산
            cursor.execute("SELECT COUNT(*) FROM translation_requests")
            total_rows = cursor.fetchone()[0]
            
            # 인덱스 생성 및 최적화
            self._finalize_database(cursor, conn, update_ui)
            
            # 작업 완료 처리
            elapsed_time = time.time() - start_time
            time_str = f"{int(elapsed_time // 60)}분 {int(elapsed_time % 60)}초"
            
            completion_msg = f"\n번역 요청 추출 작업 완료!\n"
            completion_msg += f"소요 시간: {time_str}\n"
            completion_msg += f"파일 처리: {processed_count}/{total_files} (오류: {error_count})\n"
            completion_msg += f"총 {total_rows}개 항목이 추출되었습니다.\n"
            
            update_ui(total_files, f"추출 완료 - {total_rows}개 항목", completion_msg, True)
            
            # 완료 알림
            def show_completion():
                messagebox.showinfo("완료", 
                                f"번역 요청 추출 작업이 완료되었습니다.\n"
                                f"총 {total_rows}개 항목이 추출되었습니다.\n"
                                f"소요 시간: {time_str}", parent=self.root)
            self.root.after(0, show_completion)
            
        except Exception as e:
            error_msg = str(e)
            def show_error():
                self.log_text.insert(tk.END, f"\n오류 발생: {error_msg}\n")
                self.log_text.see(tk.END)
                self.status_label.config(text="오류 발생")
                messagebox.showerror("오류", f"추출 작업 중 오류 발생: {error_msg}", parent=self.root)
            self.root.after(0, show_error)

    def _get_column_info(self, file_name, sheet_name, worksheet, extract_conditions, update_ui):
        """컬럼 정보 가져오기 (캐시 활용)"""
        condition_key = '-'.join([k for k, v in extract_conditions.items() if v])
        cache_key = f"{file_name}|{sheet_name}|{condition_key}"
        
        if cache_key in self.column_cache:
            update_ui(None, None, f"    캐시에서 컬럼 정보를 찾았습니다.\n")
            return self.column_cache[cache_key]
        
        # STRING_ID 위치 찾기
        string_id_col, header_row = self.find_string_id_position(worksheet)
        if not string_id_col or not header_row:
            return None
        
        # #번역요청 컬럼 찾기
        request_col, _ = self.find_translation_request_column(worksheet)
        if not request_col:
            return None
        
        # 필요한 언어 컬럼 찾기
        lang_cols = {}
        check_langs = ["KR"]  # KR은 항상 포함
        
        # 빈 칸 체크 조건에 따라 필요한 언어 추가
        if extract_conditions.get('en_empty'):
            check_langs.append("EN")
        if extract_conditions.get('cn_empty'):
            check_langs.append("CN")
        if extract_conditions.get('tw_empty'):
            check_langs.append("TW")
        
        # 모든 언어 컬럼을 찾음 (DB 저장용)
        all_langs = ["KR", "EN", "CN", "TW", "TH"]
        for lang in all_langs:
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=header_row, column=col).value
                if cell_value == lang:
                    lang_cols[lang] = col
                    break
        
        # 컬럼 정보 캐싱
        columns_info = {
            "string_id_col": string_id_col,
            "header_row": header_row,
            "request_col": request_col,
            "lang_cols": lang_cols,
            "extract_conditions": extract_conditions
        }
        self.column_cache[cache_key] = columns_info
        
        # 캐시 저장
        def save_cache():
            self.save_column_cache()
        self.root.after(0, save_cache)
        
        return columns_info

    def _extract_sheet_data(self, worksheet, columns_info, file_name, sheet_name, 
                        batch_data, cursor, conn, batch_size, use_read_only, extract_conditions, update_ui):
        """시트에서 데이터 추출"""
        string_id_col = columns_info["string_id_col"]
        header_row = columns_info["header_row"]
        request_col = columns_info["request_col"]
        lang_cols = columns_info["lang_cols"]
        
        sheet_rows = 0
        
        if use_read_only:
            # 읽기 전용 모드
            row_idx = 0
            for row in worksheet.iter_rows(min_row=header_row + 1):
                row_idx += 1
                if row_idx % 1000 == 0:
                    update_ui(None, None, f"    {row_idx}행 처리 중...\n", False)
                
                # 추출 조건 확인 및 데이터 처리
                if self._should_extract_row_readonly(row, request_col, lang_cols, extract_conditions, string_id_col):
                    string_id = row[string_id_col - 1].value if string_id_col - 1 < len(row) else None
                    if string_id:
                        lang_values = self._extract_lang_values_readonly(row, lang_cols)
                        
                        batch_data.append((
                            file_name, sheet_name, string_id,
                            lang_values.get("kr"), lang_values.get("en"), 
                            lang_values.get("cn"), lang_values.get("tw"), 
                            lang_values.get("th"), "번역요청", None
                        ))
                        sheet_rows += 1
                        
                        # 배치 처리
                        if len(batch_data) >= batch_size:
                            self._insert_batch_data(cursor, conn, batch_data)
                            batch_data.clear()
        else:
            # 일반 모드
            for row_num in range(header_row + 1, worksheet.max_row + 1):
                if row_num % 1000 == 0:
                    update_ui(None, None, f"    {row_num}행 처리 중...\n", False)
                
                # 추출 조건 확인 및 데이터 처리
                if self._should_extract_row_normal(worksheet, row_num, request_col, lang_cols, extract_conditions, string_id_col):
                    string_id = worksheet.cell(row=row_num, column=string_id_col).value
                    if string_id:
                        lang_values = self._extract_lang_values_normal(worksheet, row_num, lang_cols)
                        
                        batch_data.append((
                            file_name, sheet_name, string_id,
                            lang_values.get("kr"), lang_values.get("en"), 
                            lang_values.get("cn"), lang_values.get("tw"), 
                            lang_values.get("th"), "번역요청", None
                        ))
                        sheet_rows += 1
                        
                        # 배치 처리
                        if len(batch_data) >= batch_size:
                            self._insert_batch_data(cursor, conn, batch_data)
                            batch_data.clear()
        
        return sheet_rows

    def _should_extract_row_readonly(self, row, request_col, lang_cols, extract_conditions, string_id_col):
        """읽기 전용 모드에서 행 추출 조건 확인 (OR 조건)"""
        # KR 값 필터링 - 제외 조건 확인
        kr_col = lang_cols.get('KR')
        if kr_col:
            kr_value = row[kr_col - 1].value if kr_col - 1 < len(row) else None
            if self._should_exclude_kr_value(kr_value):
                return False
            
        conditions_met = []
        
        # 1. #번역요청이 "신규"인 조건 확인
        if extract_conditions.get('new_request'):
            request_value = row[request_col - 1].value if request_col - 1 < len(row) else None
            is_new_request = request_value and str(request_value).strip() == "신규"
            conditions_met.append(is_new_request)
        
        # 2. #번역요청이 "change"인 조건 확인
        if extract_conditions.get('change_request'):
            request_value = row[request_col - 1].value if request_col - 1 < len(row) else None
            is_change_request = request_value and str(request_value).strip() == "Change"
            conditions_met.append(is_change_request)
        
        # 3. 선택된 언어별 빈 칸 조건 확인
        lang_mapping = {'en_empty': 'EN', 'cn_empty': 'CN', 'tw_empty': 'TW'}
        
        for condition_key, lang in lang_mapping.items():
            if extract_conditions.get(condition_key):
                lang_col = lang_cols.get(lang)
                if lang_col:
                    lang_value = row[lang_col - 1].value if lang_col - 1 < len(row) else None
                    is_empty = not lang_value or str(lang_value).strip() == ""
                    conditions_met.append(is_empty)
        
        # OR 조건: 하나라도 만족하면 추출
        return any(conditions_met)

    def _should_extract_row_normal(self, worksheet, row_num, request_col, lang_cols, extract_conditions, string_id_col):
        """일반 모드에서 행 추출 조건 확인 (OR 조건)"""
        kr_col = lang_cols.get('KR')
        if kr_col:
            kr_value = worksheet.cell(row=row_num, column=kr_col).value
            if self._should_exclude_kr_value(kr_value):
                return False
            
        conditions_met = []
        
        # 1. #번역요청이 "신규"인 조건 확인
        if extract_conditions.get('new_request'):
            request_value = worksheet.cell(row=row_num, column=request_col).value
            is_new_request = request_value and str(request_value).strip() == "신규"
            conditions_met.append(is_new_request)
        
        # 2. #번역요청이 "change"인 조건 확인
        if extract_conditions.get('change_request'):
            request_value = worksheet.cell(row=row_num, column=request_col).value
            is_change_request = request_value and str(request_value).strip() == "Change"
            conditions_met.append(is_change_request)
        
        # 3. 선택된 언어별 빈 칸 조건 확인
        lang_mapping = {'en_empty': 'EN', 'cn_empty': 'CN', 'tw_empty': 'TW'}
        
        for condition_key, lang in lang_mapping.items():
            if extract_conditions.get(condition_key):
                lang_col = lang_cols.get(lang)
                if lang_col:
                    lang_value = worksheet.cell(row=row_num, column=lang_col).value
                    is_empty = not lang_value or str(lang_value).strip() == ""
                    conditions_met.append(is_empty)
        
        # OR 조건: 하나라도 만족하면 추출
        return any(conditions_met)

    def _should_exclude_kr_value(self, kr_value):
        """KR 값이 제외 조건에 해당하는지 확인"""
        if not kr_value:
            # 빈 값인 경우 제외
            return True
        
        kr_str = str(kr_value).strip()
        if not kr_str:
            # 공백만 있는 경우 제외
            return True
        
        # #으로 시작하는 경우 제외
        if kr_str.startswith('#'):
            return True
        
        # [@...]으로만 구성된 경우 제외 (정규식 사용)
        import re
        if re.match(r'^\[@[^\]]*\]$', kr_str):
            return True
        
        return False

    def _extract_lang_values_readonly(self, row, lang_cols):
        """읽기 전용 모드에서 언어 값 추출"""
        lang_values = {}
        # KR은 항상 추출, 나머지는 컬럼이 있는 경우만 추출
        for lang, col in lang_cols.items():
            if col - 1 < len(row):
                lang_values[lang.lower()] = row[col - 1].value
            else:
                lang_values[lang.lower()] = None
        
        # 누락된 언어는 None으로 설정
        for lang in ['kr', 'en', 'cn', 'tw', 'th']:
            if lang not in lang_values:
                lang_values[lang] = None
        
        return lang_values

    def _extract_lang_values_normal(self, worksheet, row_num, lang_cols):
        """일반 모드에서 언어 값 추출"""
        lang_values = {}
        # KR은 항상 추출, 나머지는 컬럼이 있는 경우만 추출
        for lang, col in lang_cols.items():
            lang_values[lang.lower()] = worksheet.cell(row=row_num, column=col).value
        
        # 누락된 언어는 None으로 설정
        for lang in ['kr', 'en', 'cn', 'tw', 'th']:
            if lang not in lang_values:
                lang_values[lang] = None
        
        return lang_values

    def _insert_batch_data(self, cursor, conn, batch_data):
        """배치 데이터를 DB에 삽입"""
        cursor.executemany('''
        INSERT INTO translation_requests 
        (file_name, sheet_name, string_id, kr, en, cn, tw, th, request_type, additional_info)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', batch_data)
        conn.commit()

    def _finalize_database(self, cursor, conn, update_ui):
        """DB 인덱스 생성 및 최적화"""
        update_ui(None, "인덱스 생성 중...", "\n인덱스 생성 중...\n")
        cursor.execute('CREATE INDEX idx_string_id ON translation_requests(string_id)')
        cursor.execute('CREATE INDEX idx_file_sheet ON translation_requests(file_name, sheet_name)')
        
        update_ui(None, "DB 최적화 중...", "DB 최적화 중...\n")
        cursor.execute("PRAGMA optimize")
        conn.commit()
        conn.close()


    def save_column_cache(self):
        """컬럼 캐시 저장 - FileUtils 활용"""
        try:
            FileUtils.save_json("column_cache.json", self.column_cache)
            logger.debug("컬럼 캐시 저장 완료")
        except Exception as e:
            logger.error(f"컬럼 캐시 저장 실패: {e}")
            self.log_text.insert(tk.END, f"컬럼 캐시 저장 실패: {str(e)}\n")

    def load_column_cache(self):
        """컬럼 캐시 로드 - FileUtils 활용"""
        try:
            self.column_cache = FileUtils.load_json("column_cache.json", {})
            logger.debug(f"컬럼 캐시 로드 완료: {len(self.column_cache)} 항목")
        except Exception as e:
            logger.warning(f"컬럼 캐시 로드 실패: {e}")
            self.column_cache = {}
    
    def export_to_excel(self):
        """DB를 Excel 파일로 내보내기"""
        db_path = self.output_db_var.get()
        if not db_path or not os.path.isfile(db_path):
            messagebox.showwarning("경고", "유효한 DB 파일을 선택하세요.", parent=self.root)
            return
        
        # 저장할 엑셀 파일 경로 선택
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            title="번역 요청 내보내기",
            parent=self.root
        )
        
        if not file_path:
            return
        
        self.log_text.insert(tk.END, "엑셀로 내보내기 작업 시작...\n")
        self.status_label.config(text="엑셀로 내보내기 중...")
        self.root.update()
        
        try:
            # DB 연결
            conn = sqlite3.connect(db_path)
            
            # 데이터 및 컬럼명 조회
            cursor = conn.cursor()
            cursor.execute("PRAGMA table_info(translation_requests)")
            columns = [column[1] for column in cursor.fetchall()]
            
            # 데이터 조회
            cursor.execute("SELECT * FROM translation_requests")
            rows = cursor.fetchall()
            
            if not rows:
                messagebox.showinfo("알림", "내보낼 데이터가 없습니다.", parent=self.root)
                conn.close()
                return
            
            # 판다스 데이터프레임으로 변환 (컬럼명 지정)
            df = pd.DataFrame(rows, columns=columns)
            
            # 엑셀 작성기 생성
            writer = pd.ExcelWriter(file_path, engine='openpyxl')
            
            # "String" 시트에 데이터 저장
            df.to_excel(writer, sheet_name="String", index=False)
            
            # 저장 및 닫기
            writer.close()
            
            conn.close()
            
            self.log_text.insert(tk.END, f"엑셀로 내보내기 완료: {len(rows)}개 항목\n")
            self.status_label.config(text=f"내보내기 완료 - {len(rows)}개 항목")
            
            messagebox.showinfo("완료", f"번역 요청 데이터를 엑셀로 내보냈습니다.\n파일: {file_path}", parent=self.root)
            
        except Exception as e:
            self.log_text.insert(tk.END, f"내보내기 중 오류 발생: {str(e)}\n")
            messagebox.showerror("오류", f"엑셀로 내보내기 실패: {str(e)}", parent=self.root)


    def export_to_template_excel(self):
        """템플릿 기반 Excel 내보내기"""
        db_path = self.output_db_var.get()
        if not db_path or not os.path.isfile(db_path):
            messagebox.showwarning("경고", "유효한 DB 파일을 선택하세요.", parent=self.root)
            return
        
        # 템플릿 Excel 파일 선택
        template_path = filedialog.askopenfilename(
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")],
            title="템플릿 Excel 파일 선택",
            parent=self.root
        )
        
        if not template_path:
            return
        
        # 출력할 Excel 파일 경로 선택
        output_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            title="번역 요청 템플릿 내보내기",
            parent=self.root
        )
        
        if not output_path:
            return
        
        self.log_text.insert(tk.END, "템플릿 기반 엑셀 내보내기 작업 시작...\n")
        self.status_label.config(text="템플릿 기반 내보내기 중...")
        self.root.update()
        
        try:
            # DB 연결 및 데이터 조회
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            # 데이터 조회 (시트별로 그룹화)
            cursor.execute("""
            SELECT file_name, sheet_name, string_id, kr, en, cn, tw, th, request_type
            FROM translation_requests 
            ORDER BY file_name, sheet_name, string_id
            """)
            rows = cursor.fetchall()
            conn.close()
            
            if not rows:
                messagebox.showinfo("알림", "내보낼 데이터가 없습니다.", parent=self.root)
                return
            
            # 데이터를 시트별로 그룹화
            sheets_data = {}
            for row in rows:
                file_name, sheet_name, string_id, kr, en, cn, tw, th, request_type = row
                
                if sheet_name not in sheets_data:
                    sheets_data[sheet_name] = []
                
                sheets_data[sheet_name].append({
                    'file_name': file_name,
                    'string_id': string_id,
                    'kr': kr,
                    'en': en,
                    'cn': cn,
                    'tw': tw,
                    'th': th,
                    'request_type': request_type
                })
            
            # 템플릿 기반 Excel 생성
            success_count = self._create_template_based_excel(template_path, output_path, sheets_data)
            
            self.log_text.insert(tk.END, f"템플릿 기반 내보내기 완료: {success_count}개 시트 생성\n")
            self.status_label.config(text=f"템플릿 내보내기 완료 - {success_count}개 시트")
            
            messagebox.showinfo("완료", 
                            f"템플릿 기반 Excel 내보내기가 완료되었습니다.\n"
                            f"생성된 시트: {success_count}개\n"
                            f"파일: {output_path}", parent=self.root)
            
        except Exception as e:
            self.log_text.insert(tk.END, f"템플릿 내보내기 중 오류 발생: {str(e)}\n")
            messagebox.showerror("오류", f"템플릿 기반 내보내기 실패: {str(e)}", parent=self.root)

    def _create_template_based_excel(self, template_path, output_path, sheets_data):
        """템플릿 기반 Excel 파일 생성"""
        try:
            # 템플릿 파일 로드
            template_workbook = load_workbook(template_path)
            
            # 새 워크북 생성
            output_workbook = Workbook()
            
            # 기본 시트 제거
            if 'Sheet' in output_workbook.sheetnames:
                output_workbook.remove(output_workbook['Sheet'])
            
            success_count = 0
            
            # 각 시트별로 처리
            for sheet_name, sheet_data in sheets_data.items():
                self.log_text.insert(tk.END, f"  시트 '{sheet_name}' 생성 중... ({len(sheet_data)}개 항목)\n")
                self.root.update()
                
                # 템플릿에서 해당 시트를 찾거나 첫 번째 시트를 템플릿으로 사용
                template_sheet = None
                if sheet_name in template_workbook.sheetnames:
                    template_sheet = template_workbook[sheet_name]
                else:
                    # String으로 시작하는 시트 찾기
                    string_sheets = [s for s in template_workbook.sheetnames 
                                if s.lower().startswith('string')]
                    if string_sheets:
                        template_sheet = template_workbook[string_sheets[0]]
                    else:
                        template_sheet = template_workbook.active
                
                # 새 시트 생성
                new_sheet = output_workbook.create_sheet(title=sheet_name)
                
                # 템플릿 구조 복사 및 데이터 삽입
                if self._copy_template_and_insert_data(template_sheet, new_sheet, sheet_data):
                    success_count += 1
                else:
                    self.log_text.insert(tk.END, f"    시트 '{sheet_name}' 처리 실패\n")
            
            # 템플릿 워크북 닫기
            template_workbook.close()
            
            # 결과 파일 저장
            output_workbook.save(output_path)
            output_workbook.close()
            
            return success_count
            
        except Exception as e:
            self.log_text.insert(tk.END, f"템플릿 Excel 생성 중 오류: {str(e)}\n")
            return 0

    def _copy_template_and_insert_data(self, template_sheet, new_sheet, sheet_data):
        """템플릿 구조 복사 및 데이터 삽입"""
        try:
            # 1. 템플릿 구조 복사 (값과 기본 스타일만)
            max_copy_rows = min(10, template_sheet.max_row)  # 최대 10행까지만 복사
            
            for row in range(1, max_copy_rows + 1):
                for col in range(1, template_sheet.max_column + 1):
                    source_cell = template_sheet.cell(row=row, column=col)
                    target_cell = new_sheet.cell(row=row, column=col)
                    
                    # 값만 복사 (스타일 복사 제거)
                    target_cell.value = source_cell.value
            
            # 컬럼 너비 복사
            try:
                for col in range(1, template_sheet.max_column + 1):
                    col_letter = new_sheet.cell(row=1, column=col).column_letter
                    template_col_letter = template_sheet.cell(row=1, column=col).column_letter
                    if template_col_letter in template_sheet.column_dimensions:
                        width = template_sheet.column_dimensions[template_col_letter].width
                        if width:
                            new_sheet.column_dimensions[col_letter].width = width
            except Exception as e:
                # 컬럼 너비 복사 실패해도 계속 진행
                pass
            
            # 2. 컬럼 위치 찾기
            columns_info = self._find_template_columns(template_sheet)
            if not columns_info:
                self.log_text.insert(tk.END, f"    템플릿에서 필요한 컬럼을 찾을 수 없습니다.\n")
                return False
            
            # 3. 데이터 삽입
            return self._insert_data_to_template(new_sheet, columns_info, sheet_data)
            
        except Exception as e:
            self.log_text.insert(tk.END, f"    템플릿 복사 중 오류: {str(e)}\n")
            return False

    def _find_template_columns(self, template_sheet):
        """템플릿에서 컬럼 위치 찾기"""
        columns_info = {}
        
        # 헤더 행을 찾기 위해 상위 10행 검색
        for row in range(1, min(11, template_sheet.max_row + 1)):
            found_columns = {}
            
            for col in range(1, template_sheet.max_column + 1):
                cell_value = template_sheet.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    cell_value_clean = cell_value.strip().upper()
                    
                    # 컬럼 매핑
                    if cell_value_clean == "STRING_ID":
                        found_columns["string_id"] = col
                    elif cell_value_clean == "KR":
                        found_columns["kr"] = col
                    elif cell_value_clean == "EN":
                        found_columns["en"] = col
                    elif cell_value_clean == "CN":
                        found_columns["cn"] = col
                    elif cell_value_clean == "TW":
                        found_columns["tw"] = col
                    elif cell_value_clean == "TH":
                        found_columns["th"] = col
                    elif "#번역요청" in cell_value or "번역요청" in cell_value:
                        found_columns["request"] = col
            
            # STRING_ID와 KR 컬럼이 모두 있으면 헤더 행으로 판단
            if "string_id" in found_columns and "kr" in found_columns:
                columns_info = found_columns
                columns_info["header_row"] = row
                break
        
        return columns_info if "string_id" in columns_info else None

    def _insert_data_to_template(self, worksheet, columns_info, sheet_data):
        """템플릿에 데이터 삽입"""
        try:
            header_row = columns_info["header_row"]
            data_start_row = header_row + 1
            
            # 기존 데이터가 있다면 그 다음 행부터 시작
            current_row = data_start_row
            while worksheet.cell(row=current_row, column=columns_info["string_id"]).value:
                current_row += 1
            
            # 데이터 삽입
            inserted_count = 0
            for data_item in sheet_data:
                # STRING_ID 삽입
                worksheet.cell(row=current_row, column=columns_info["string_id"]).value = data_item["string_id"]
                
                # KR 삽입
                worksheet.cell(row=current_row, column=columns_info["kr"]).value = data_item["kr"]
                
                # 다른 언어 컬럼들 삽입 (있는 경우만)
                lang_mapping = {
                    "en": data_item.get("en"),
                    "cn": data_item.get("cn"),
                    "tw": data_item.get("tw"),
                    "th": data_item.get("th")
                }
                
                for lang, value in lang_mapping.items():
                    if lang in columns_info and value:
                        worksheet.cell(row=current_row, column=columns_info[lang]).value = value
                
                # 번역요청 컬럼에 "신규" 표시 (있는 경우만)
                if "request" in columns_info:
                    worksheet.cell(row=current_row, column=columns_info["request"]).value = "신규"
                
                current_row += 1
                inserted_count += 1
            
            self.log_text.insert(tk.END, f"    {inserted_count}개 항목 삽입 완료\n")
            return True
            
        except Exception as e:
            self.log_text.insert(tk.END, f"    데이터 삽입 중 오류: {str(e)}\n")
            return False

def main():
    root = tk.Tk()
    app = TranslationRequestExtractor(root)
    root.mainloop()


if __name__ == "__main__":
    main()