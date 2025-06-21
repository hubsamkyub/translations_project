# tools/translation_workflow_tool.py

import tkinter as tk
from tkinter import ttk, filedialog
import os
from ui.common_components import ScrollableCheckList, show_message
from tools.workflow_manager import WorkflowManager

class TranslationWorkflowTool(tk.Frame):
    def __init__(self, parent, root):
        super().__init__(parent)
        self.root = root
        self.workflow_manager = WorkflowManager(self)

        # UI 변수 선언
        self.source_folder_var = tk.StringVar()
        self.base_db_path_var = tk.StringVar()
        self.is_update_var = tk.BooleanVar(value=False)
        self.excel_files = []
        
        self.completed_excel_path_var = tk.StringVar()
        self.completed_excel_sheet_var = tk.StringVar()
        self.update_only_if_kr_matches_var = tk.BooleanVar(value=True)
        
        self.apply_lang_vars = {lang: tk.BooleanVar(value=False) for lang in ["EN", "CN", "TW"]}
        self.apply_lang_vars["CN"].set(True) # 기본 선택
        self.apply_lang_vars["TW"].set(True) # 기본 선택
        self.apply_kr_match_check_var = tk.BooleanVar(value=True)
        self.apply_kr_mismatch_delete_var = tk.BooleanVar(value=False)
        self.apply_record_date_var = tk.BooleanVar(value=True)

        # UI 구성
        self.setup_ui()
        self.pack(fill="both", expand=True)

    def log(self, message):
        """공통 로그 기록 함수"""
        self.log_text.insert(tk.END, f"[{self.get_timestamp()}] {message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
        
    def get_timestamp(self):
        import time
        return time.strftime("%H:%M:%S")

    def setup_ui(self):
        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill="both", expand=True)
        
        # --- 1단계: DB 구축 / 업데이트 ---
        step1_frame = ttk.LabelFrame(main_frame, text="1단계: 기준 DB 생성 또는 업데이트")
        step1_frame.pack(fill="x", padx=5, pady=5)

        # 소스 폴더 선택
        source_folder_frame = ttk.Frame(step1_frame)
        source_folder_frame.pack(fill="x", padx=5, pady=5)
        ttk.Label(source_folder_frame, text="작업 파일 폴더:").grid(row=0, column=0, sticky="w", padx=5)
        ttk.Entry(source_folder_frame, textvariable=self.source_folder_var, width=60).grid(row=0, column=1, sticky="ew", padx=5)
        ttk.Button(source_folder_frame, text="폴더 선택", command=self.select_source_folder).grid(row=0, column=2, padx=5)
        ttk.Button(source_folder_frame, text="파일 검색", command=self.search_source_files).grid(row=0, column=3, padx=5)
        source_folder_frame.columnconfigure(1, weight=1)

        # 파일 목록
        self.file_list_checklist = ScrollableCheckList(step1_frame, height=100)
        self.file_list_checklist.pack(fill="both", expand=True, padx=5, pady=5)

        # DB 경로 및 옵션
        db_frame = ttk.Frame(step1_frame)
        db_frame.pack(fill="x", padx=5, pady=5)
        ttk.Label(db_frame, text="기준 DB 파일:").grid(row=0, column=0, sticky="w", padx=5)
        ttk.Entry(db_frame, textvariable=self.base_db_path_var, width=60).grid(row=0, column=1, sticky="ew", padx=5)
        ttk.Button(db_frame, text="파일 선택/저장", command=self.select_base_db).grid(row=0, column=2, padx=5)
        ttk.Checkbutton(db_frame, text="기존 DB에 업데이트", variable=self.is_update_var).grid(row=0, column=3, padx=10)
        db_frame.columnconfigure(1, weight=1)

        # 1단계 실행 버튼
        step1_action_frame = ttk.Frame(step1_frame)
        step1_action_frame.pack(fill="x", pady=5)
        self.step1_button = ttk.Button(step1_action_frame, text="1. DB 구축/업데이트 실행", command=self.run_step1)
        self.step1_button.pack(side="right")
        self.step1_status_label = ttk.Label(step1_action_frame, text="대기 중", foreground="gray")
        self.step1_status_label.pack(side="left", padx=5)

        # --- 2단계: 번역 파일로 DB 업데이트 (UI 틀만 생성) ---
        self.step2_frame = ttk.LabelFrame(main_frame, text="2단계: 번역 파일로 DB 업데이트")
        self.step2_frame.pack(fill="x", padx=5, pady=5)
        
        # 번역 완료 파일 선택
        completed_file_frame = ttk.Frame(self.step2_frame)
        completed_file_frame.pack(fill="x", padx=5, pady=5)
        ttk.Label(completed_file_frame, text="번역 완료 파일(Excel):").grid(row=0, column=0, sticky="w", padx=5)
        ttk.Entry(completed_file_frame, textvariable=self.completed_excel_path_var, width=50).grid(row=0, column=1, sticky="ew", padx=5)
        ttk.Button(completed_file_frame, text="파일 선택", command=self.select_completed_excel).grid(row=0, column=2, padx=5)
        
        # 시트 선택
        sheet_select_frame = ttk.Frame(self.step2_frame)
        sheet_select_frame.pack(fill="x", padx=5, pady=5)
        ttk.Label(sheet_select_frame, text="적용할 시트 선택:").grid(row=0, column=0, sticky="w", padx=5)
        self.sheet_combobox = ttk.Combobox(sheet_select_frame, textvariable=self.completed_excel_sheet_var, state="readonly", width=40)
        self.sheet_combobox.grid(row=0, column=1, sticky="ew", padx=5)
        sheet_select_frame.columnconfigure(1, weight=1)
        completed_file_frame.columnconfigure(1, weight=1)

        # 업데이트 옵션
        update_option_frame = ttk.Frame(self.step2_frame)
        update_option_frame.pack(fill="x", padx=5, pady=5)
        ttk.Checkbutton(update_option_frame, text="기준 DB의 KR값과 일치할 경우에만 번역 업데이트", variable=self.update_only_if_kr_matches_var).pack(side="left", padx=5)

        # 2단계 실행 버튼
        step2_action_frame = ttk.Frame(self.step2_frame)
        step2_action_frame.pack(fill="x", pady=5)
        self.step2_button = ttk.Button(step2_action_frame, text="2. 번역 내용으로 DB 업데이트", command=self.run_step2)
        self.step2_button.pack(side="right")
        self.step2_status_label = ttk.Label(step2_action_frame, text="대기 중", foreground="gray")
        self.step2_status_label.pack(side="left", padx=5)

        self.disable_step(self.step2_frame) # 초기에는 비활성화

        # --- 3단계: 작업 파일에 번역 적용 (UI 틀만 생성) ---
        self.step3_frame = ttk.LabelFrame(main_frame, text="3단계: 작업 파일에 번역 최종 적용")
        self.step3_frame.pack(fill="x", padx=5, pady=5)
        # 적용 대상 확인
        target_info_frame = ttk.Frame(self.step3_frame)
        target_info_frame.pack(fill="x", padx=5, pady=5)
        ttk.Label(target_info_frame, text="적용 대상 폴더:").pack(side="left", padx=5)
        ttk.Entry(target_info_frame, textvariable=self.source_folder_var, state="readonly").pack(side="left", fill="x", expand=True)

        # 적용 옵션
        apply_options_frame = ttk.Frame(self.step3_frame)
        apply_options_frame.pack(fill="x", padx=5, pady=5)
        
        lang_frame = ttk.Frame(apply_options_frame)
        lang_frame.pack(anchor="w", pady=2)
        ttk.Label(lang_frame, text="적용 언어:").pack(side="left", padx=5)
        for lang, var in self.apply_lang_vars.items():
            ttk.Checkbutton(lang_frame, text=lang, variable=var).pack(side="left")

        kr_check_frame = ttk.Frame(apply_options_frame)
        kr_check_frame.pack(anchor="w", pady=2)
        ttk.Checkbutton(kr_check_frame, text="KR 일치 검사 (불일치 시 건너뛰기)", variable=self.apply_kr_match_check_var).pack(side="left", padx=5)
        ttk.Checkbutton(kr_check_frame, text="└ KR 불일치 시 다국어 제거", variable=self.apply_kr_mismatch_delete_var).pack(side="left", padx=5)

        other_option_frame = ttk.Frame(apply_options_frame)
        other_option_frame.pack(anchor="w", pady=2)
        ttk.Checkbutton(other_option_frame, text="번역 적용 표시 (#번역요청 -> '적용')", variable=self.apply_record_date_var).pack(side="left", padx=5)
        
        # 3단계 실행 버튼
        step3_action_frame = ttk.Frame(self.step3_frame)
        step3_action_frame.pack(fill="x", pady=5)
        self.step3_button = ttk.Button(step3_action_frame, text="3. 최종 번역 적용 실행", command=self.run_step3)
        self.step3_button.pack(side="right")
        self.step3_status_label = ttk.Label(step3_action_frame, text="대기 중", foreground="gray")
        self.step3_status_label.pack(side="left", padx=5)

        self.disable_step(self.step3_frame)

        # --- 로그 ---
        log_frame = ttk.LabelFrame(main_frame, text="작업 로그")
        log_frame.pack(fill="both", expand=True, padx=5, pady=5)
        self.log_text = tk.Text(log_frame, wrap="word", height=10)
        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.log_text.pack(fill="both", expand=True)

    def disable_step(self, frame):
        """프레임 내의 모든 위젯을 비활성화합니다."""
        for widget in frame.winfo_children():
            if isinstance(widget, (ttk.Button, ttk.Entry, ttk.Checkbutton, ttk.Combobox)):
                widget.config(state="disabled")
            if isinstance(widget, ttk.Frame):
                self.disable_step(widget)

    def select_source_folder(self):
        folder = filedialog.askdirectory(title="작업 파일이 있는 폴더 선택")
        if folder: self.source_folder_var.set(folder)

    def search_source_files(self):
        folder = self.source_folder_var.get()
        if not folder or not os.path.isdir(folder):
            show_message(self, "warning", "경고", "유효한 폴더를 선택하세요.")
            return
        
        self.file_list_checklist.clear()
        self.excel_files = []
        self.log("파일 검색 중...")
        
        for root, _, files in os.walk(folder):
            for file in files:
                if file.lower().startswith("string") and file.lower().endswith(('.xlsx', '.xls')) and not file.startswith("~$"):
                    full_path = os.path.join(root, file)
                    self.excel_files.append((file, full_path))
        
        self.excel_files.sort()
        for file_name, _ in self.excel_files:
            self.file_list_checklist.add_item(file_name, checked=True)
        self.log(f"{len(self.excel_files)}개의 String 엑셀 파일을 찾았습니다.")

    def select_base_db(self):
        if self.is_update_var.get():
            path = filedialog.askopenfilename(title="업데이트할 DB 파일 선택", filetypes=[("DB 파일", "*.db")])
        else:
            path = filedialog.asksaveasfilename(title="새로 저장할 DB 파일 경로", defaultextension=".db", filetypes=[("DB 파일", "*.db")])
        if path: self.base_db_path_var.set(path)

    def run_step1(self):
        is_update = self.is_update_var.get()
        db_path = self.base_db_path_var.get()
        source_folder = self.source_folder_var.get()
        selected_file_names = self.file_list_checklist.get_checked_items()

        if not db_path or not source_folder or not selected_file_names:
            show_message(self, "warning", "입력 오류", "작업 폴더, DB 파일, 대상 파일을 모두 선택해야 합니다.")
            return

        if is_update and not os.path.exists(db_path):
            show_message(self, "error", "파일 없음", "업데이트를 선택했지만 기준 DB 파일이 존재하지 않습니다.")
            return
            
        if not is_update and os.path.exists(db_path):
            if not show_message(self, "yesno", "파일 존재", "선택한 경로에 파일이 이미 존재합니다. 덮어쓰시겠습니까?"):
                return

        files_to_process = [item for item in self.excel_files if item[0] in selected_file_names]

        self.step1_status_label.config(text="처리 중...", foreground="blue")
        self.workflow_manager.run_step1_build_db(is_update, db_path, source_folder, files_to_process, self.on_step1_complete)
    
    def on_step1_complete(self, result):
        if result.get("status") == "success":
            mode_text = "업데이트" if self.is_update_var.get() else "구축"
            total_rows = result.get('total_rows', 0)
            message = f"DB {mode_text} 완료. 총 {total_rows}개 항목 처리됨."
            self.step1_status_label.config(text=message, foreground="green")
            self.log(message)
            show_message(self, "info", "1단계 완료", message)
        else:
            message = f"오류: {result.get('message', '알 수 없는 오류')}"
            self.step1_status_label.config(text="오류 발생", foreground="red")
            self.log(message)
            show_message(self, "error", "1단계 오류", message)

# tools/translation_workflow_tool.py 파일의 TranslationWorkflowTool 클래스 내부에 아래 함수들을 추가하세요.

    def enable_step(self, frame):
        """프레임 내의 모든 위젯을 활성화합니다."""
        for widget in frame.winfo_children():
            if isinstance(widget, (ttk.Button, ttk.Entry, ttk.Checkbutton, ttk.Combobox)):
                widget.config(state="normal")
            if isinstance(widget, ttk.Frame):
                self.enable_step(widget)

    def select_completed_excel(self):
        """2단계의 번역 완료 엑셀 파일을 선택하는 함수"""
        path = filedialog.askopenfilename(title="번역 완료된 엑셀 파일 선택", filetypes=[("Excel 파일", "*.xlsx;*.xls")])
        if path:
            self.completed_excel_path_var.set(path)
            self.log(f"번역 완료 파일 선택됨: {os.path.basename(path)}")
            # 엑셀 파일이 선택되면 시트 목록을 읽어 콤보박스에 채웁니다.
            self.load_sheets_from_excel(path)

    def load_sheets_from_excel(self, excel_path):
        """엑셀 파일에서 시트 목록을 읽어 콤보박스를 채웁니다."""
        try:
            self.sheet_combobox.set("시트 목록 읽는 중...")
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, read_only=True)
            string_sheets = [s for s in wb.sheetnames if s.lower().startswith("string")]
            wb.close()
            
            self.sheet_combobox['values'] = string_sheets
            if string_sheets:
                self.sheet_combobox.set(string_sheets[0])
            else:
                self.sheet_combobox.set("")
                show_message(self, "warning", "시트 없음", "선택한 파일에 'String'으로 시작하는 시트가 없습니다.")
        except Exception as e:
            self.sheet_combobox.set("시트 읽기 실패")
            show_message(self, "error", "파일 오류", f"엑셀 파일을 읽는 중 오류가 발생했습니다: {e}")

    def on_step1_complete(self, result):
        if result.get("status") == "success":
            mode_text = "업데이트" if self.is_update_var.get() else "구축"
            total_rows = result.get('total_rows', 0)
            message = f"DB {mode_text} 완료. 총 {total_rows}개 항목 처리됨."
            self.step1_status_label.config(text=message, foreground="green")
            self.log(message)
            show_message(self, "info", "1단계 완료", message)
            
            # [수정] 1단계 완료 시 2단계 활성화
            self.enable_step(self.step2_frame)
            self.log("2단계가 활성화되었습니다. 번역 완료 파일을 선택하고 DB 업데이트를 진행하세요.")
        else:
            message = f"오류: {result.get('message', '알 수 없는 오류')}"
            self.step1_status_label.config(text="오류 발생", foreground="red")
            self.log(message)
            show_message(self, "error", "1단계 오류", message)
            
    def run_step2(self):
        """2단계 DB 업데이트 로직 실행"""
        base_db_path = self.base_db_path_var.get()
        completed_excel_path = self.completed_excel_path_var.get()
        sheet_name = self.completed_excel_sheet_var.get()
        
        if not all([base_db_path, completed_excel_path, sheet_name]):
            show_message(self, "warning", "입력 오류", "기준 DB, 번역 완료 파일, 시트를 모두 선택해야 합니다.")
            return

        update_options = {
            "only_if_kr_matches": self.update_only_if_kr_matches_var.get()
        }

        self.step2_status_label.config(text="처리 중...", foreground="blue")
        self.workflow_manager.run_step2_update_db_from_excel(
            base_db_path,
            completed_excel_path,
            sheet_name,
            update_options,
            self.on_step2_complete
        )

# tools/translation_workflow_tool.py의 on_step2_complete 함수를 아래 코드로 교체

    def on_step2_complete(self, result):
        if result.get("status") == "success":
            updated_rows = result.get('updated_rows', 0)
            duplicate_entries = result.get("duplicate_entries", {})
            
            message = f"고유 항목 업데이트 완료: {updated_rows}개 항목 업데이트됨."
            self.log(message)

            if duplicate_entries:
                # 중복 항목이 있으면 번역 확인 창을 띄움
                message += f"\n\n중복된 STRING_ID {len(duplicate_entries)}개를 발견했습니다. '번역 확인' 창에서 내용을 수정하고 최종 반영해주세요."
                self.step2_status_label.config(text=f"확인 필요: 중복 {len(duplicate_entries)}개", foreground="orange")
                
                from tools.translation_verification_tool import TranslationVerificationTool
                verification_window = TranslationVerificationTool(
                    self.root,
                    duplicate_entries,
                    self.base_db_path_var.get(),
                    self.workflow_manager,
                    self.on_step2_verification_complete # 확인 창이 닫힐 때 호출될 콜백
                )
                verification_window.wait_window()
            else:
                # 중복 항목이 없으면 바로 3단계 활성화
                self.step2_status_label.config(text=message, foreground="green")
                self.enable_step(self.step3_frame)
                self.log("중복 항목이 없어 3단계가 활성화되었습니다.")
                show_message(self, "info", "2단계 완료", message)
        else:
            message = f"오류: {result.get('message', '알 수 없는 오류')}"
            self.step2_status_label.config(text="오류 발생", foreground="red")
            self.log(message)
            show_message(self, "error", "2단계 오류", message)

    def on_step2_verification_complete(self, result):
        """번역 확인 창에서 최종 DB 반영 후 호출되는 콜백 함수"""
        if result.get("status") == "success":
            updated_rows = result.get("updated_rows", 0)
            message = f"중복 항목 수정 및 DB 반영 완료: {updated_rows}개 항목."
            self.step2_status_label.config(text="DB 업데이트 완료", foreground="green")
            self.log(message)
            self.enable_step(self.step3_frame)
            self.log("모든 DB 업데이트가 완료되어 3단계가 활성화되었습니다.")

# tools/translation_workflow_tool.py 파일의 TranslationWorkflowTool 클래스 내부에 아래 함수들을 추가하세요.

    def run_step3(self):
        """3단계 최종 번역 적용 실행"""
        base_db_path = self.base_db_path_var.get()
        selected_file_names = self.file_list_checklist.get_checked_items()
        
        if not base_db_path or not os.path.exists(base_db_path):
            show_message(self, "warning", "입력 오류", "1, 2단계를 거쳐 생성된 유효한 기준 DB 파일이 필요합니다.")
            return

        if not selected_file_names:
            show_message(self, "warning", "입력 오류", "번역을 적용할 파일을 목록에서 선택해야 합니다.")
            return
            
        selected_langs = [lang for lang, var in self.apply_lang_vars.items() if var.get()]
        if not selected_langs:
            show_message(self, "warning", "경고", "적용할 언어를 하나 이상 선택하세요.")
            return

        apply_options = {
            "selected_langs": selected_langs,
            "record_date": self.apply_record_date_var.get(),
            "kr_match_check": self.apply_kr_match_check_var.get(),
            "kr_mismatch_delete": self.apply_kr_mismatch_delete_var.get()
        }
        
        files_to_process = [item for item in self.excel_files if item[0] in selected_file_names]

        self.step3_status_label.config(text="처리 중...", foreground="blue")
        self.workflow_manager.run_step3_apply_to_files(
            base_db_path,
            files_to_process,
            apply_options,
            self.on_step3_complete
        )

    def on_step3_complete(self, result):
        if result.get("status") == "success":
            total_updated = result.get('total_updated', 0)
            message = f"번역 적용 완료. 총 {total_updated}개 항목 업데이트됨."
            self.step3_status_label.config(text=message, foreground="green")
            self.log(message)
            show_message(self, "info", "워크플로우 완료", f"모든 작업이 완료되었습니다.\n{message}")
        else:
            message = f"오류: {result.get('message', '알 수 없는 오류')}"
            self.step3_status_label.config(text="오류 발생", foreground="red")
            self.log(message)
            show_message(self, "error", "3단계 오류", message)