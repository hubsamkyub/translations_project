import os
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
import sqlite3
import tkinter as tk
import win32com.client as pythoncom
import xlwings as xw

class TranslationApplyManager:
    def __init__(self, parent_window=None):
        self.parent_ui = parent_window
        self.translation_cache = {}
        self.translation_file_cache = {}
        self.translation_sheet_cache = {}
        self.duplicate_ids = {}
        self.kr_reverse_cache = {}
        
    def log_message(self, message):
        """UI의 로그 텍스트 영역에 메시지를 기록합니다."""
        if self.parent_ui and hasattr(self.parent_ui, 'log_text'):
            self.parent_ui.log_text.insert(tk.END, f"{message}\n")
            self.parent_ui.log_text.see(tk.END)
            self.parent_ui.update_idletasks()
        else:
            print(message)

    def load_translation_cache_from_excel(self, file_path, sheet_name):
        """엑셀 파일의 특정 시트에서 번역 데이터를 읽어 캐시를 생성합니다."""
        try:
            self.log_message(f"⚙️ 엑셀 파일 로딩 시작: {os.path.basename(file_path)} - 시트: {sheet_name}")
            header_row_index = self._find_header_row(file_path, sheet_name)
            if header_row_index is None:
                message = "지정한 시트의 1~6행에서 'STRING_ID' 컬럼을 찾을 수 없습니다."
                self.log_message(f"❌ {message}")
                return {"status": "error", "message": message}
            
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row_index, dtype=str)
            df.fillna('', inplace=True)

            df.columns = [str(col).lower() for col in df.columns]
            
            if 'string_id' not in df.columns:
                message = "엑셀 시트에 'string_id' 컬럼이 없습니다."
                self.log_message(f"❌ {message}")
                return {"status": "error", "message": message}

            self.build_cache_from_dataframe(df)

            # <<< [수정] file_count, sheet_count 키 추가 >>>
            return {
                "status": "success",
                "source_type": "Excel",
                "translation_cache": self.translation_cache,
                "translation_file_cache": self.translation_file_cache,
                "translation_sheet_cache": self.translation_sheet_cache,
                "duplicate_ids": self.duplicate_ids,
                "kr_reverse_cache": self.kr_reverse_cache,
                "file_count": len(self.translation_file_cache),
                "sheet_count": len(self.translation_sheet_cache),
                "id_count": len(self.translation_cache)
            }
        except Exception as e:
            self.log_message(f"❌ 엑셀 캐시 로딩 오류: {str(e)}")
            return {"status": "error", "message": str(e)}
        
    def _find_header_row(self, file_path, sheet_name):
        """엑셀 시트의 1~6행에서 'string_id'를 포함하는 헤더 행을 찾습니다."""
        for i in range(6):
            try:
                df_peek = pd.read_excel(file_path, sheet_name=sheet_name, header=i, nrows=0)
                if 'string_id' in [str(col).lower() for col in df_peek.columns]:
                    self.log_message(f"✅ 헤더 행 발견: {i + 1}번째 행")
                    return i
            except Exception:
                continue
        return None
    
    def build_cache_from_dataframe(self, df):
        """Pandas DataFrame으로부터 정교한 다중 캐시를 구축합니다."""
        self.translation_cache = {}
        self.translation_file_cache = {}
        self.translation_sheet_cache = {}
        self.duplicate_ids = {}
        self.kr_reverse_cache = {}
        
        self.log_message(f"🔧 데이터프레임으로부터 캐시 구축 시작: {len(df)}개 행")

        for _, row in df.iterrows():
            string_id = str(row.get('string_id', '')).strip()
            if not string_id:
                continue

            file_name = str(row.get('filename', row.get('file_name', ''))).strip()
            sheet_name = str(row.get('sheetname', row.get('sheet_name', ''))).strip()

            norm_file_name = file_name.lower()
            norm_sheet_name = sheet_name.lower()

            data = {
                "kr": str(row.get("kr", "")),
                "en": str(row.get("en", "")),
                "cn": str(row.get("cn", "")),
                "tw": str(row.get("tw", "")),
                "th": str(row.get("th", "")),
                "file_name": file_name,
                "sheet_name": sheet_name
            }
            
            if norm_file_name:
                self.translation_file_cache.setdefault(norm_file_name, {})[string_id] = data
            
            if norm_sheet_name:
                self.translation_sheet_cache.setdefault(norm_sheet_name, {})[string_id] = data
            
            self.translation_cache[string_id] = data

            if string_id not in self.duplicate_ids:
                self.duplicate_ids[string_id] = []
            self.duplicate_ids[string_id].append(file_name)

            kr_text = data["kr"].strip()
            if kr_text and kr_text not in self.kr_reverse_cache:
                kr_cache_data = data.copy()
                kr_cache_data['string_id'] = string_id
                self.kr_reverse_cache[kr_text] = kr_cache_data
        
        self.log_message(f"🔧 캐시 구성 완료 (ID: {len(self.translation_cache)}, 파일: {len(self.translation_file_cache)}, 시트: {len(self.translation_sheet_cache)}, KR역방향: {len(self.kr_reverse_cache)})")
        
             
    def find_string_id_position(self, worksheet):
        """STRING_ID 위치 찾기"""
        for row in range(2, 6):  # 2행부터 5행까지 검색
            for col in range(1, min(10, worksheet.max_column + 1)):  # 최대 10개 컬럼까지만 검색
                cell_value = worksheet.cell(row=row, column=col).value
                if isinstance(cell_value, str) and "STRING_ID" in cell_value.upper():
                    return col, row
                    
        # 1행도 검색
        for row in worksheet.iter_rows(min_row=1, max_row=1, max_col=5):
            for cell in row:
                if isinstance(cell.value, str) and "STRING_ID" in cell.value.upper():
                    return cell.column, cell.row
                    
        return None, None

    def find_language_columns(self, worksheet, header_row, langs):
        """언어 컬럼 위치 찾기"""
        if not header_row:
            return {}
            
        lang_cols = {}
        
        # 지정한 헤더 행에서만 검색
        for row in worksheet.iter_rows(min_row=header_row, max_row=header_row):
            for cell in row:
                if not cell.value:
                    continue
                    
                header_text = str(cell.value).strip()
                
                # 직접 매칭
                if header_text in langs:
                    lang_cols[header_text] = cell.column
                    
        return lang_cols

    def find_target_columns(self, worksheet, header_row, target_columns=None):
        """지정된 대상 컬럼들의 위치를 찾습니다. (예: #번역요청, Change)"""
        if not header_row:
            return {}
            
        found_columns = {}
        # 기본적으로 '#번역요청' 컬럼을 탐색 대상에 포함
        all_targets = ["#번역요청"]
        if target_columns:
            all_targets.extend(target_columns)
        
        # 중복 제거
        all_targets = list(set(all_targets))

        for cell in worksheet[header_row]:
            if cell.value and isinstance(cell.value, str):
                cell_value_clean = cell.value.strip().lower()
                for target in all_targets:
                    if cell_value_clean == target.lower():
                        found_columns[target] = cell.column
                        break # 찾았으면 다음 셀로
                        
        return found_columns

    def _resave_with_excel_com(self, file_path):
        """Excel COM을 사용하여 파일을 다시 저장하여 최적화합니다."""
        excel = None
        workbook = None
        try:
            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            abs_path = os.path.abspath(file_path)
            workbook = excel.Workbooks.Open(abs_path)
            workbook.Save()
            self.log_message(f"  ✨ COM 객체로 파일 최적화 저장 완료: {os.path.basename(file_path)}")
            return True
        except Exception as e:
            self.log_message(f"  ⚠️ COM 객체 저장 실패: {e}")
            return False
        finally:
            if workbook:
                workbook.Close(SaveChanges=False)
            if excel:
                excel.Quit()
            pythoncom.CoUninitialize()

    def _resave_with_xlwings(self, file_path):
        """xlwings를 사용하여 파일을 다시 저장하여 최적화합니다."""
        # xw.App()을 컨텍스트 관리자로 사용해 Excel 프로세스가 확실히 종료되도록 보장
        with xw.App(visible=False) as app:
            try:
                wb = app.books.open(file_path)
                wb.save()  # 현재 파일에 덮어쓰기
                wb.close()
                self.log_message(f"  ✨ xlwings로 파일 최적화 저장 완료: {os.path.basename(file_path)}")
                return True
            except Exception as e:
                self.log_message(f"  ⚠️ xlwings 저장 실패: {e}")
                # 오류 발생 시에도 앱이 정상적으로 종료됩니다.
                return False
            
    def find_translation_request_column(self, worksheet, header_row):
        """#번역요청 컬럼 찾기 (공백, 대소문자 무시)"""
        if not header_row:
            return None
            
        for cell in worksheet[header_row]:
            if cell.value and isinstance(cell.value, str):
                # 공백 제거 및 소문자 변환 후 비교
                if cell.value.strip().lower() == "#번역요청":
                    return cell.column
                    
        return None


# tools/translation_apply_manager.py 의 apply_translation 함수를 아래 코드로 교체

    def apply_translation(self, file_path, selected_langs, record_date=True, kr_match_check=False, kr_mismatch_delete=False, allowed_statuses=None, smart_translation=True):
        """
        파일에 번역 적용 (기능 개선 최종 버전)
        - [수정] allowed_statuses를 받아 조건부 적용 로직 변경
        """
        if not self.translation_cache:
            return {"status": "error", "message": "번역 캐시가 로드되지 않았습니다.", "error_type": "cache_not_loaded"}

        file_name = os.path.basename(file_path)
        self.log_message(f"📁 파일 처리 시작: {file_name}")
        
        workbook = None
        try:
            current_file_name = os.path.basename(file_path).lower()
            workbook = load_workbook(file_path)

            string_sheets = [sheet for sheet in workbook.sheetnames if sheet.lower().startswith("string") and not sheet.startswith("#")]
            
            if not string_sheets:
                return {"status": "info", "message": "파일에 String 시트가 없습니다"}

            file_modified = False
            total_updated = 0
            total_kr_mismatch_skipped = 0
            total_kr_mismatch_deleted = 0
            total_conditional_skipped = 0
            
            fill_green = PatternFill(start_color="DAF2D0", end_color="DAF2D0", fill_type="solid")
            
            # [수정] 소문자로 비교하기 위해 미리 변환
            allowed_statuses_lower = [status.lower() for status in allowed_statuses] if allowed_statuses else []

            for sheet_name in string_sheets:
                worksheet = workbook[sheet_name]
                string_id_col, header_row = self.find_string_id_position(worksheet)
                if not string_id_col or not header_row:
                    continue
                
                lang_cols = self.find_language_columns(worksheet, header_row, selected_langs + ['KR'])
                
                request_col_idx = None
                # [수정] allowed_statuses가 있을 때만 #번역요청 컬럼을 찾음
                if allowed_statuses_lower:
                    request_col_idx = self.find_target_columns(worksheet, header_row, ["#번역요청"]).get("#번역요청")
                    if not request_col_idx:
                        self.log_message(f"  - {sheet_name}: '#번역요청' 컬럼을 찾을 수 없어 조건부 적용을 건너뜁니다.")

                apply_cols = {}
                if record_date:
                    apply_cols = self.find_target_columns(worksheet, header_row, ["#번역요청", "신규", "change"])

                sheet_updated_count = 0
                
                for row_idx in range(header_row + 1, worksheet.max_row + 1):
                    # --- [수정] 조건부 적용 로직 ---
                    if allowed_statuses_lower and request_col_idx:
                        request_val = str(worksheet.cell(row=row_idx, column=request_col_idx).value or '').strip().lower()
                        if request_val not in allowed_statuses_lower:
                            total_conditional_skipped += 1
                            continue
                    # --- 로직 수정 끝 ---

                    string_id = str(worksheet.cell(row=row_idx, column=string_id_col).value or '').strip()
                    if not string_id:
                        continue
                        
                    trans_data = self.translation_file_cache.get(current_file_name, {}).get(string_id) or \
                                 self.translation_sheet_cache.get(sheet_name.lower(), {}).get(string_id) or \
                                 self.translation_cache.get(string_id)

                    if not trans_data:
                        continue
                    
                    row_modified_this_iteration = False
                    
                    kr_mismatched = False
                    if kr_match_check and 'KR' in lang_cols:
                        current_kr_value = str(worksheet.cell(row=row_idx, column=lang_cols['KR']).value or '').strip()
                        cache_kr_value = str(trans_data.get('kr', '') or '').strip()
                        if current_kr_value != cache_kr_value:
                            kr_mismatched = True

                    if kr_mismatched:
                        if kr_mismatch_delete:
                            for lang in selected_langs:
                                if lang == 'KR': continue
                                if lang in lang_cols:
                                    if worksheet.cell(row=row_idx, column=lang_cols[lang]).value != "":
                                        worksheet.cell(row=row_idx, column=lang_cols[lang]).value = ""
                                        row_modified_this_iteration = True
                            if row_modified_this_iteration:
                                total_kr_mismatch_deleted += 1
                        else:
                            total_kr_mismatch_skipped += 1
                            continue
                    else:
                        for lang in selected_langs:
                            if lang in lang_cols and trans_data.get(lang.lower()):
                                current_value = worksheet.cell(row=row_idx, column=lang_cols[lang]).value
                                trans_value = trans_data[lang.lower()]
                                if trans_value and str(current_value) != str(trans_value):
                                    worksheet.cell(row=row_idx, column=lang_cols[lang]).value = trans_value
                                    worksheet.cell(row=row_idx, column=lang_cols[lang]).fill = fill_green
                                    row_modified_this_iteration = True
                    
                    if row_modified_this_iteration:
                        sheet_updated_count += 1
                        file_modified = True
                        if record_date and apply_cols:
                            for col_num in apply_cols.values():
                                worksheet.cell(row=row_idx, column=col_num).value = "적용"
                
                if sheet_updated_count > 0 or total_kr_mismatch_deleted > 0:
                     self.log_message(f"  - {sheet_name}: {sheet_updated_count}개 행에 변경사항 적용 완료")
                total_updated += sheet_updated_count
            
            if file_modified:
                self.log_message(f"  💾 openpyxl로 파일 저장 중...")
                workbook.save(file_path)
                self._resave_with_xlwings(file_path)
            
            return {
                "status": "success",
                "total_updated": total_updated,
                "kr_mismatch_skipped": total_kr_mismatch_skipped,
                "kr_mismatch_deleted": total_kr_mismatch_deleted,
                "conditional_skipped": total_conditional_skipped,
            }
            
        except Exception as e:
            self.log_message(f"❌ 파일 처리 중 오류 발생: {file_name} - {str(e)}")
            import traceback
            traceback.print_exc()
            return {"status": "error", "message": str(e), "error_type": "processing_error"}
        finally:
            if workbook:
                workbook.close()
                self.log_message(f"  ✔️ 파일 핸들 해제 완료: {file_name}")
                
        
    def check_external_links(self, workbook):
        """워크북에서 외부 링크 검사 (번역 도구용) - 검증된 최종 버전"""
        import re
        
        external_links = []
        
        # 외부 참조 패턴들 (검증된 버전)
        external_patterns = [
            r"'[^']*\.xl[sx]?[xm]?'!",  # '파일명.xlsx'! 또는 '경로\파일명.xlsx'!
            r'\[.*\.xl[sx]?[xm]?\]',    # [파일명.xlsx] 패턴
            r"'[A-Z]:[^']*\.xl[sx]?[xm]?'!", # 'C:\경로\파일명.xlsx'! 패턴  
            r'\\[^\\]*\.xl[sx]?[xm]?!', # \파일명.xlsx! 패턴
            r"=[^=]*'[A-Z]:[^']*'",     # =으로 시작하는 드라이브 경로
            r'\[\d+\]!',                # [숫자]! 패턴 (시트 참조)
        ]
        
        # #REF! 오류 패턴들 (검증된 버전)
        ref_error_patterns = [
            r'#REF!',                   # #REF! 오류
            r'OFFSET\(#REF!',          # OFFSET 함수에서 #REF! 오류
        ]
        
        try:
            # 방법 1: 워크북의 external_links 속성 확인
            if hasattr(workbook, 'external_links') and workbook.external_links:
                for link in workbook.external_links:
                    external_links.append(f"워크북_외부링크: {str(link)}")
            
            # 방법 2: 명명된 범위 검사 (가장 중요!) - 검증된 로직
            if hasattr(workbook, 'defined_names') and workbook.defined_names:
                # 딕셔너리 키로 접근 (검증된 방법)
                for name_key in workbook.defined_names.keys():
                    try:
                        defined_name = workbook.defined_names[name_key]
                        if hasattr(defined_name, 'value') and defined_name.value:
                            name_formula = str(defined_name.value)
                            
                            # #REF! 오류 우선 검사
                            ref_error_found = False
                            for ref_pattern in ref_error_patterns:
                                if re.search(ref_pattern, name_formula):
                                    external_links.append(f"명명된_범위_REF오류:{name_key} - {name_formula[:50]}")
                                    ref_error_found = True
                                    break
                            
                            # #REF! 오류가 없는 경우에만 외부 참조 패턴 검사
                            if not ref_error_found:
                                for pattern in external_patterns:
                                    if re.search(pattern, name_formula):
                                        external_links.append(f"명명된_범위_외부링크:{name_key} - {name_formula[:50]}")
                                        break
                    except Exception as e:
                        # 개별 명명된 범위 처리 중 오류가 발생해도 계속 진행
                        pass
            
            # 방법 3: 셀별 외부 참조 검사 (제한적으로)
            cell_count = 0
            for sheet_name in workbook.sheetnames:
                if cell_count >= 100:  # 번역 도구에서는 성능을 위해 더 제한적으로
                    break
                    
                worksheet = workbook[sheet_name]
                
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell_count += 1
                        if cell_count > 100:
                            break
                            
                        # 공식이 있는 셀 검사
                        if cell.data_type == 'f' and cell.value:
                            formula = str(cell.value)
                            
                            # #REF! 오류 검사
                            for ref_pattern in ref_error_patterns:
                                if re.search(ref_pattern, formula):
                                    external_links.append(f"셀_REF오류:{sheet_name}!{cell.coordinate} - {formula[:50]}")
                                    break
                            else:
                                # 외부 참조 패턴 검사
                                for pattern in external_patterns:
                                    if re.search(pattern, formula):
                                        external_links.append(f"셀_외부링크:{sheet_name}!{cell.coordinate} - {formula[:50]}")
                                        break
                        
                        # #REF! 값 검사
                        elif cell.value and str(cell.value).startswith('#REF!'):
                            external_links.append(f"셀_REF값:{sheet_name}!{cell.coordinate} - {cell.value}")
                    
                    if cell_count > 100:
                        break
                        
        except Exception as e:
            # 외부 링크 검사 중 오류가 발생하면 무시하고 계속 진행
            pass
            
        return external_links[:10]  # 최대 10개만 반환

    def load_translation_cache_from_db(self, db_path):
        """데이터베이스에서 번역 데이터를 읽어 캐시를 생성합니다."""
        try:
            self.log_message(f"⚙️ DB 로딩 시작: {db_path}")
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='translation_data'")
            if cursor.fetchone() is None:
                message = "'translation_data' 테이블이 DB에 없습니다."
                self.log_message(f"❌ {message}")
                conn.close()
                return {"status": "error", "message": message}

            query = "SELECT * FROM translation_data"
            df = pd.read_sql_query(query, conn)
            conn.close()
            
            self.build_cache_from_dataframe(df)
            
            # <<< [수정] file_count, sheet_count 키 추가 >>>
            return {
                "status": "success",
                "source_type": "DB",
                "translation_cache": self.translation_cache,
                "translation_file_cache": self.translation_file_cache,
                "translation_sheet_cache": self.translation_sheet_cache,
                "duplicate_ids": self.duplicate_ids,
                "kr_reverse_cache": self.kr_reverse_cache,
                "file_count": len(self.translation_file_cache),
                "sheet_count": len(self.translation_sheet_cache),
                "id_count": len(self.translation_cache)
            }
        except Exception as e:
            self.log_message(f"❌ 번역 DB 캐시 로딩 오류: {str(e)}")
            return {"status": "error", "message": str(e)}