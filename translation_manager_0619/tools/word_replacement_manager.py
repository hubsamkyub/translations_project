# tools/translate/word_replacement_manager.py
import tkinter as tk
import gc
import os
import re
import sqlite3
import threading
import time
import uuid
import pythoncom
import win32com.client
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from datetime import datetime
from ui.common_components import ScrollableCheckList, LoadingPopup

class WordReplacementManager(tk.Frame):
    def __init__(self, parent, root):
        super().__init__(parent)
        self.root = root
        
        # 한글 포함 패턴 정규표현식 ([@...] 형식에서 한글이 포함된 것만)
        self.pattern = re.compile(r'\[@([^\]]*[가-힣][^\]]*)\]')
        
        # 치환 통계
        self.replacement_stats = {
            'total_found': 0,
            'existing_replaced': 0,
            'new_created': 0,
            'files_processed': 0
        }
        
        # 내부 데이터 초기화
        self.excel_files = []
        self.replacement_results = []
        
        # UI 설정
        self.setup_ui()

    def safe_close_workbook(self, workbook):
        """워크북 안전 종료"""
        try:
            if workbook:
                workbook.close()
                del workbook  # 명시적 삭제
        except:
            pass
        finally:
            gc.collect()  # 강제 가비지 컬렉션

    def check_file_in_use(self, file_path, max_wait=5):
        """파일이 다른 프로세스에서 사용 중인지 확인"""
        import time
        
        for attempt in range(max_wait):
            try:
                # 파일 열기 테스트
                with open(file_path, 'r+b') as f:
                    pass
                return False  # 사용 중이 아님
            except (IOError, OSError):
                if attempt < max_wait - 1:
                    self.log_message(f"⏳ 파일 사용 대기 중... ({attempt + 1}/{max_wait})")
                    time.sleep(1)
                else:
                    return True  # 사용 중
        return True
    
    def setup_ui(self):
        """UI 구성"""
        # 메인 컨테이너
        main_container = ttk.Frame(self)
        main_container.pack(fill="both", expand=True, padx=10, pady=10)
        
        # 파일 선택 프레임
        file_frame = ttk.LabelFrame(main_container, text="엑셀 파일 선택")
        file_frame.pack(fill="x", padx=5, pady=5)
        
        # 폴더 선택
        folder_select_frame = ttk.Frame(file_frame)
        folder_select_frame.pack(fill="x", padx=5, pady=5)
        
        ttk.Label(folder_select_frame, text="폴더 선택:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.excel_folder_var = tk.StringVar()
        ttk.Entry(folder_select_frame, textvariable=self.excel_folder_var, width=50).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(folder_select_frame, text="폴더 찾아보기", 
                  command=self.select_excel_folder).grid(row=0, column=2, padx=5, pady=5)
        ttk.Button(folder_select_frame, text="폴더 검색", 
                  command=self.search_excel_files).grid(row=0, column=3, padx=5, pady=5)
        
        folder_select_frame.columnconfigure(1, weight=1)
        
        # 개별 파일 선택
        file_select_frame = ttk.Frame(file_frame)
        file_select_frame.pack(fill="x", padx=5, pady=5)
        
        ttk.Label(file_select_frame, text="파일 선택:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.individual_file_var = tk.StringVar()
        ttk.Entry(file_select_frame, textvariable=self.individual_file_var, width=50).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(file_select_frame, text="파일 찾아보기", 
                  command=self.select_individual_file).grid(row=0, column=2, padx=5, pady=5)
        ttk.Button(file_select_frame, text="파일 추가", 
                  command=self.add_individual_file).grid(row=0, column=3, padx=5, pady=5)
        
        file_select_frame.columnconfigure(1, weight=1)
        
        # 파일 목록
        files_list_frame = ttk.LabelFrame(main_container, text="처리할 엑셀 파일 목록")
        files_list_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        try:
            self.excel_files_list = ScrollableCheckList(files_list_frame, width=700, height=120)
            self.excel_files_list.pack(fill="both", expand=True, padx=5, pady=5)
        except Exception as e:
            # ScrollableCheckList를 사용할 수 없는 경우 대체 위젯 사용
            print(f"ScrollableCheckList 오류: {e}")
            self.excel_files_list = tk.Listbox(files_list_frame, height=8)
            self.excel_files_list.pack(fill="both", expand=True, padx=5, pady=5)
        
        # DB 설정 프레임
        db_frame = ttk.LabelFrame(main_container, text="DB 설정")
        db_frame.pack(fill="x", padx=5, pady=5)
        
        db_select_frame = ttk.Frame(db_frame)
        db_select_frame.pack(fill="x", padx=5, pady=5)
        
        ttk.Label(db_select_frame, text="String DB:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.db_path_var = tk.StringVar()
        ttk.Entry(db_select_frame, textvariable=self.db_path_var, width=50).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(db_select_frame, text="찾아보기", 
                  command=self.select_db_file).grid(row=0, column=2, padx=5, pady=5)
        
        db_select_frame.columnconfigure(1, weight=1)

        # 신규 텍스트 저장 파일 설정
        new_file_frame = ttk.Frame(db_frame)
        new_file_frame.pack(fill="x", padx=5, pady=5)
        
        ttk.Label(new_file_frame, text="신규 텍스트 저장:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.new_text_file_var = tk.StringVar()
        ttk.Entry(new_file_frame, textvariable=self.new_text_file_var, width=50).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(new_file_frame, text="찾아보기", 
                  command=self.select_new_text_file).grid(row=0, column=2, padx=5, pady=5)
        
        new_file_frame.columnconfigure(1, weight=1)
        
        # 치환 옵션 프레임
        options_frame = ttk.LabelFrame(main_container, text="치환 옵션")
        options_frame.pack(fill="x", padx=5, pady=5)
        
        options_row1 = ttk.Frame(options_frame)
        options_row1.pack(fill="x", padx=5, pady=5)
        
        # 새 ID 접두사 설정
        ttk.Label(options_row1, text="새 ID 접두사:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.id_prefix_var = tk.StringVar(value="string_change_text")
        ttk.Entry(options_row1, textvariable=self.id_prefix_var, width=20).grid(row=0, column=1, padx=5, pady=5, sticky="w")
        
        options_row2 = ttk.Frame(options_frame)
        options_row2.pack(fill="x", padx=5, pady=5)
        
        # Excel 자동 저장 옵션
        self.excel_auto_save_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_row2, text="Excel 자동 저장 (용량 문제 해결)", 
                       variable=self.excel_auto_save_var).grid(row=0, column=0, padx=5, pady=5, sticky="w")
        
        # 안전 모드 옵션
        self.safe_mode_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_row2, text="안전 모드 (파일 손상 방지)", 
                       variable=self.safe_mode_var).grid(row=0, column=1, padx=20, pady=5, sticky="w")
        
        # 치환 일괄 적용 옵션
        bulk_apply_frame = ttk.LabelFrame(main_container, text="치환 일괄 적용 (KR 값을 다른 언어 컬럼에 복사)")
        bulk_apply_frame.pack(fill="x", padx=5, pady=5)
        
        ttk.Label(bulk_apply_frame, text="적용할 언어:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        
        language_frame = ttk.Frame(bulk_apply_frame)
        language_frame.grid(row=0, column=1, columnspan=7, padx=5, pady=5, sticky="w")
        
        self.bulk_languages = ["EN", "CN", "TW", "TH", "PT", "ES", "DE", "FR"]
        self.bulk_lang_vars = {}
        
        for i, lang in enumerate(self.bulk_languages):
            var = tk.BooleanVar(value=False)  # 기본은 체크 안됨
            self.bulk_lang_vars[lang] = var
            ttk.Checkbutton(language_frame, text=lang, variable=var).grid(
                row=0, column=i, padx=5, sticky="w")
        
        # 실행 버튼
        action_frame = ttk.Frame(main_container)
        action_frame.pack(fill="x", padx=5, pady=5)
        
        ttk.Button(action_frame, text="미리보기", 
                  command=self.preview_replacements).pack(side="right", padx=5, pady=5)
        ttk.Button(action_frame, text="치환 실행", 
                  command=self.execute_replacements).pack(side="right", padx=5, pady=5)
        
        # 결과 표시 영역
        result_frame = ttk.LabelFrame(main_container, text="치환 결과")
        result_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # 트리뷰로 결과 표시
        tree_frame = ttk.Frame(result_frame)
        tree_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        columns = ("file_name", "sheet_name", "cell", "original_text", "korean_text", "string_id", "status")
        self.result_tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=8)
        
        # 컬럼 설정
        self.result_tree.heading("file_name", text="파일명")
        self.result_tree.heading("sheet_name", text="시트명")
        self.result_tree.heading("cell", text="셀")
        self.result_tree.heading("original_text", text="원본 패턴")
        self.result_tree.heading("korean_text", text="한글 텍스트")
        self.result_tree.heading("string_id", text="STRING_ID")
        self.result_tree.heading("status", text="상태")
        
        self.result_tree.column("file_name", width=120)
        self.result_tree.column("sheet_name", width=100)
        self.result_tree.column("cell", width=60)
        self.result_tree.column("original_text", width=150)
        self.result_tree.column("korean_text", width=100)
        self.result_tree.column("string_id", width=150)
        self.result_tree.column("status", width=80)
        
        # 스크롤바
        scrollbar_y = ttk.Scrollbar(tree_frame, orient="vertical", command=self.result_tree.yview)
        self.result_tree.configure(yscrollcommand=scrollbar_y.set)
        
        scrollbar_x = ttk.Scrollbar(result_frame, orient="horizontal", command=self.result_tree.xview)
        self.result_tree.configure(xscrollcommand=scrollbar_x.set)
        
        # 배치
        scrollbar_y.pack(side="right", fill="y")
        self.result_tree.pack(side="left", fill="both", expand=True)
        scrollbar_x.pack(side="bottom", fill="x")
        
        # 상태 표시줄
        status_frame = ttk.Frame(main_container)
        status_frame.pack(fill="x", padx=5, pady=5)
        
        self.status_label = ttk.Label(status_frame, text="대기 중...")
        self.status_label.pack(side="left", fill="x", expand=True, padx=5)
        
        self.progress_bar = ttk.Progressbar(status_frame, length=300, mode="determinate")
        self.progress_bar.pack(side="right", padx=5)
        
        # 로그 영역
        log_frame = ttk.LabelFrame(main_container, text="작업 로그")
        log_frame.pack(fill="x", padx=5, pady=5)
        
        log_container = ttk.Frame(log_frame)
        log_container.pack(fill="both", expand=True, padx=5, pady=5)
        
        self.log_text = tk.Text(log_container, wrap="word", height=6)
        log_scrollbar = ttk.Scrollbar(log_container, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_scrollbar.set)
        
        log_scrollbar.pack(side="right", fill="y")
        self.log_text.pack(side="left", fill="both", expand=True)
        
        # 초기 메시지
        self.log_message("단어 치환 기능이 준비되었습니다.")
        print("UI 설정 완료")

    def log_message(self, message):
        """로그 메시지 추가"""
        self.log_text.insert(tk.END, f"{time.strftime('%H:%M:%S')} - {message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def select_excel_folder(self):
        """엑셀 폴더 선택"""
        folder = filedialog.askdirectory(title="엑셀 파일 폴더 선택", parent=self.root)
        if folder:
            self.excel_folder_var.set(folder)
            self.root.after(100, self.root.focus_force)
            self.root.after(100, self.root.lift)

    def select_individual_file(self):
        """개별 엑셀 파일 선택"""
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")],
            title="엑셀 파일 선택",
            parent=self.root
        )
        if file_path:
            self.individual_file_var.set(file_path)
            self.root.after(100, self.root.focus_force)
            self.root.after(100, self.root.lift)

    def add_individual_file(self):
        """개별 파일을 목록에 추가"""
        file_path = self.individual_file_var.get()
        if not file_path or not os.path.isfile(file_path):
            messagebox.showwarning("경고", "유효한 파일을 선택하세요.", parent=self.root)
            return
        
        file_name = os.path.basename(file_path)
        
        # 중복 확인
        existing_files = [name for name, _ in self.excel_files]
        if file_name in existing_files:
            messagebox.showinfo("알림", "이미 목록에 있는 파일입니다.", parent=self.root)
            return
        
        # 파일 추가
        self.excel_files.append((file_name, file_path))
        
        # 파일 목록에 추가
        if hasattr(self.excel_files_list, 'add_item'):
            self.excel_files_list.add_item(file_name, checked=True)
        else:
            # 일반 Listbox인 경우
            self.excel_files_list.insert(tk.END, file_name)
        
        self.log_message(f"파일 추가: {file_name}")
        messagebox.showinfo("알림", f"파일이 추가되었습니다: {file_name}", parent=self.root)

    def select_db_file(self):
        """DB 파일 선택"""
        file_path = filedialog.askopenfilename(
            filetypes=[("DB 파일", "*.db"), ("모든 파일", "*.*")],
            title="String DB 파일 선택",
            parent=self.root
        )
        if file_path:
            self.db_path_var.set(file_path)
            self.root.after(100, self.root.focus_force)
            self.root.after(100, self.root.lift)

    def select_new_text_file(self):
        """신규 텍스트 저장할 엑셀 파일 선택"""
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")],
            title="신규 텍스트 저장할 엑셀 파일 선택",
            parent=self.root
        )
        if file_path:
            self.new_text_file_var.set(file_path)
            self.root.after(100, self.root.focus_force)
            self.root.after(100, self.root.lift)

    def search_excel_files(self):
        """엑셀 파일 검색 (String으로 시작하는 파일만)"""
        folder = self.excel_folder_var.get()
        if not folder or not os.path.isdir(folder):
            messagebox.showwarning("경고", "유효한 폴더를 선택하세요.", parent=self.root)
            return
        
        # 파일 목록 초기화
        if hasattr(self.excel_files_list, 'clear'):
            self.excel_files_list.clear()
        else:
            # 일반 Listbox인 경우
            self.excel_files_list.delete(0, tk.END)
        
        self.excel_files = []
        
        # 폴더와 하위 폴더 검색 (String으로 시작하는 파일만)
        for root, _, files in os.walk(folder):
            for file in files:
                if (file.endswith(".xlsx") and not file.startswith("~") and 
                    file.lower().startswith("string")):
                    file_path = os.path.join(root, file)
                    self.excel_files.append((file, file_path))
                    
                    # 파일 목록에 추가
                    if hasattr(self.excel_files_list, 'add_item'):
                        self.excel_files_list.add_item(file, checked=True)
                    else:
                        # 일반 Listbox인 경우
                        self.excel_files_list.insert(tk.END, file)
        
        if not self.excel_files:
            messagebox.showinfo("알림", "String으로 시작하는 엑셀 파일을 찾지 못했습니다.", parent=self.root)
        else:
            messagebox.showinfo("알림", f"{len(self.excel_files)}개의 String 엑셀 파일을 찾았습니다.", parent=self.root)
            self.log_message(f"{len(self.excel_files)}개 String 엑셀 파일 검색 완료")

    def _get_selected_files(self):
        """선택된 파일 목록 가져오기"""
        if hasattr(self.excel_files_list, 'get_checked_items'):
            return self.excel_files_list.get_checked_items()
        else:
            # 일반 Listbox인 경우 - 모든 파일 반환
            return [file for file, _ in self.excel_files]

    def find_korean_patterns(self, text):
        """텍스트에서 한글 포함 패턴 찾기"""
        if not isinstance(text, str):
            return []
        
        matches = self.pattern.findall(text)
        return [(match, f"[@{match}]") for match in matches]

    def get_string_id_by_kr(self, db_path, kr_text):
        """KR 값으로 STRING_ID 찾기"""
        conn = None
        try:
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            # String DB의 경우 여러 시트에서 검색
            tables = self._get_db_tables(cursor)
            
            for table in tables:
                try:
                    cursor.execute(f"SELECT STRING_ID FROM {table} WHERE KR = ? LIMIT 1", (kr_text,))
                    result = cursor.fetchone()
                    if result:
                        return result[0]
                except:
                    continue
            
            return None
            
        except Exception as e:
            self.log_message(f"DB 검색 오류: {e}")
            return None
        finally:
            if conn:
                try:
                    conn.close()
                except:
                    pass

    def _get_db_tables(self, cursor):
        """DB에서 String 관련 테이블 목록 가져오기"""
        try:
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            all_tables = [row[0] for row in cursor.fetchall()]
            
            # String으로 시작하는 테이블 우선, 그 외는 나중에
            string_tables = [t for t in all_tables if t.startswith('String')]
            other_tables = [t for t in all_tables if not t.startswith('String') and 'STRING_ID' in self._get_table_columns(cursor, t)]
            
            return string_tables + other_tables
            
        except:
            return []

    def _get_table_columns(self, cursor, table_name):
        """테이블의 컬럼 목록 가져오기"""
        try:
            cursor.execute(f"PRAGMA table_info({table_name})")
            return [row[1] for row in cursor.fetchall()]
        except:
            return []

    def generate_new_string_id_from_excel(self, excel_path, prefix="string_change_text"):
        """새로운 STRING_ID 생성 - timestamp 방식 (파일 열지 않음)"""
        try:
            self.log_message(f"🔧 ID 생성 방식: timestamp")
            
            # 타임스탬프 방식: prefix_YYYYMMDD_HHMMSS_XXX
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            random_suffix = str(uuid.uuid4().hex)[:3]
            new_id = f"{prefix}_{timestamp}_{random_suffix}"
            
            self.log_message(f"✅ 타임스탬프 ID 생성: {new_id}")
            return new_id
                
        except Exception as e:
            self.log_message(f"❌ ID 생성 오류: {e}")
            # 폴백: 더 간단한 시간 기반 ID
            fallback_id = f"{prefix}_{int(time.time())}"
            self.log_message(f"🔄 대체 ID 생성: {fallback_id}")
            return fallback_id

    def excel_auto_save(self, file_path):
        """Excel로 파일을 열고 다시 저장하여 구조 정상화"""
        try:
            self.log_message(f"🔄 Excel 자동 저장 시작: {os.path.basename(file_path)}")
            
            # COM 초기화
            pythoncom.CoInitialize()
            
            # Excel 애플리케이션 시작
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False  # 백그라운드 실행
            excel.DisplayAlerts = False  # 알림 끄기
            
            try:
                # 파일 열기
                workbook = excel.Workbooks.Open(os.path.abspath(file_path))
                
                # 잠깐 대기
                time.sleep(0.5)
                
                # 저장
                workbook.Save()
                
                # 닫기
                workbook.Close(SaveChanges=True)
                
                self.log_message(f"✅ Excel 자동 저장 완료")
                return True
                
            except Exception as e:
                self.log_message(f"❌ Excel 자동 저장 실패: {e}")
                return False
                
            finally:
                # Excel 종료
                try:
                    excel.Quit()
                except:
                    pass
                
                # COM 해제
                try:
                    pythoncom.CoUninitialize()
                except:
                    pass
                    
        except Exception as e:
            self.log_message(f"❌ Excel COM 초기화 실패: {e}")
            return False

    def excel_auto_save_conditional(self, file_path):
        """조건부 Excel 자동 저장 - 용량 문제 해결"""
        # 사용자 옵션 확인
        if not getattr(self, 'excel_auto_save_var', None) or not self.excel_auto_save_var.get():
            self.log_message(f"📄 Excel 자동 저장 비활성화: {os.path.basename(file_path)}")
            return True
            
        try:
            # 파일 크기 확인
            original_size = os.path.getsize(file_path)
            
            # 1MB 이상 파일만 Excel 자동 저장 적용
            if original_size > 1024 * 1024:
                self.log_message(f"🔧 Excel 자동 저장 실행: {os.path.basename(file_path)}")
                return self.excel_auto_save_simple(file_path)
            else:
                self.log_message(f"📄 소용량 파일, Excel 자동 저장 생략: {os.path.basename(file_path)}")
                return True
                
        except Exception as e:
            self.log_message(f"⚠️ 파일 크기 확인 실패: {e}")
            return True  # 오류 시에도 계속 진행

    def excel_auto_save_simple(self, file_path):
        """간소화된 Excel 자동 저장 (용량 정상화)"""
        try:
            # COM 초기화
            pythoncom.CoInitialize()
            
            # Excel 애플리케이션 시작
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.ScreenUpdating = False
            
            try:
                # 파일 열기
                workbook = excel.Workbooks.Open(os.path.abspath(file_path))
                
                # 간단히 저장만 (용량 정상화)
                workbook.Save()
                workbook.Close(SaveChanges=False)
                
                self.log_message(f"    🔧 Excel 자동 저장 완료")
                return True
                
            except Exception as e:
                self.log_message(f"    ❌ Excel 저장 실패: {e}")
                return False
                
            finally:
                try:
                    excel.Quit()
                except:
                    pass
                    
                try:
                    pythoncom.CoUninitialize()
                except:
                    pass
                    
        except Exception as e:
            self.log_message(f"    ❌ Excel COM 실패: {e}")
            return False

    def add_new_string_to_excel_safe(self, excel_path, string_id, kr_text):
        """안전한 신규 STRING 추가 - 파일 손상 방지"""
        import shutil
        import tempfile
        
        # 백업 생성
        backup_path = None
        try:
            # 파일 존재 확인
            if not os.path.exists(excel_path):
                self.log_message(f"❌ 신규 스트링 파일이 존재하지 않음: {excel_path}")
                return False
            
            # 임시 백업 파일 생성
            backup_dir = tempfile.gettempdir()
            backup_filename = f"backup_{os.path.basename(excel_path)}_{int(time.time())}"
            backup_path = os.path.join(backup_dir, backup_filename)
            shutil.copy2(excel_path, backup_path)
            
            # 메인 작업
            success = self.add_new_string_to_excel(excel_path, string_id, kr_text)
            
            if success:
                # 파일 무결성 검증
                try:
                    test_workbook = load_workbook(excel_path, read_only=True)
                    
                    # 기본 구조 검증
                    has_string_sheet = any(sheet.lower().startswith("string") for sheet in test_workbook.sheetnames)
                    if not has_string_sheet:
                        raise Exception("String 시트를 찾을 수 없음")
                    
                    test_workbook.close()
                    
                    # 성공하면 백업 삭제
                    if backup_path and os.path.exists(backup_path):
                        os.remove(backup_path)
                    
                    self.log_message(f"✅ 안전 추가 완료: {string_id}")
                    return True
                    
                except Exception as verify_error:
                    # 파일이 손상되었으면 백업으로 복구
                    self.log_message(f"⚠️ 파일 손상 감지, 백업으로 복구 중: {verify_error}")
                    if backup_path and os.path.exists(backup_path):
                        shutil.copy2(backup_path, excel_path)
                        os.remove(backup_path)
                        self.log_message(f"🔄 백업 복구 완료")
                    return False
            else:
                # 작업 실패하면 백업으로 복구
                self.log_message(f"⚠️ 추가 작업 실패, 백업으로 복구 중")
                if backup_path and os.path.exists(backup_path):
                    shutil.copy2(backup_path, excel_path)
                    os.remove(backup_path)
                return False
                
        except Exception as e:
            self.log_message(f"❌ 안전 처리 중 오류: {e}")
            # 오류 시 백업으로 복구
            if backup_path and os.path.exists(backup_path):
                try:
                    shutil.copy2(backup_path, excel_path)
                    os.remove(backup_path)
                    self.log_message(f"🔄 오류 후 백업 복구 완료")
                except Exception as restore_error:
                    self.log_message(f"❌ 백업 복구도 실패: {restore_error}")
            return False

    def add_new_string_to_excel(self, excel_path, string_id, kr_text):
        """신규 STRING 추가 - 안전한 openpyxl 전용 (Excel 자동 저장 제외)"""
        try:
            # 번역 적용과 동일한 방식으로 안정적 로드
            workbook = load_workbook(excel_path, data_only=False, keep_vba=True)
            
            # String 시트 찾기
            target_sheet = None
            for sheet_name in workbook.sheetnames:
                if sheet_name.lower().startswith("string"):
                    target_sheet = workbook[sheet_name]
                    break
            
            if not target_sheet:
                workbook.close()
                return False
            
            # 헤더 찾기
            string_id_col = None
            kr_col = None
            
            for col in range(1, 21):
                for row in range(1, 6):
                    cell_value = target_sheet.cell(row=row, column=col).value
                    if cell_value:
                        header = str(cell_value).strip().upper()
                        if header == "STRING_ID":
                            string_id_col = col
                        elif header == "KR":
                            kr_col = col
            
            if not string_id_col or not kr_col:
                workbook.close()
                return False
            
            # 마지막 행 찾기 (더 효율적인 방법)
            last_row = 1  # 헤더 행부터 시작
            for row_cells in target_sheet.iter_rows(min_row=2, min_col=string_id_col, max_col=string_id_col):
                if row_cells[0].value and str(row_cells[0].value).strip():
                    last_row = row_cells[0].row
            
            new_row = last_row + 1
            
            # 데이터 추가
            target_sheet.cell(row=new_row, column=string_id_col, value=string_id)
            target_sheet.cell(row=new_row, column=kr_col, value=kr_text)
            
            # openpyxl만으로 저장 (신규 스트링 파일은 Excel 자동 저장 안함)
            workbook.save(excel_path)
            workbook.close()
            
            self.log_message(f"✅ 신규 텍스트 추가 (안전 모드): {string_id}")
            return True
            
        except Exception as e:
            self.log_message(f"❌ 신규 텍스트 추가 실패: {e}")
            return False

    def get_existing_string_id_from_new_file(self, excel_path, kr_text):
        """신규 텍스트 파일에서 동일한 한글이 있는지 검색"""
        workbook = None
        try:
            self.log_message(f"🔍 신규 텍스트 파일에서 중복 검색: '{kr_text}'")
            
            workbook = load_workbook(excel_path, read_only=True)
            
            for sheet_name in workbook.sheetnames:
                if sheet_name.lower().startswith("string"):
                    worksheet = workbook[sheet_name]
                    headers = self._find_excel_headers(worksheet)
                    
                    if "STRING_ID" in headers and "KR" in headers:
                        string_id_col = headers["STRING_ID"]
                        kr_col = headers["KR"]
                        
                        for row in range(1, worksheet.max_row + 1):
                            try:
                                kr_value = worksheet.cell(row=row, column=kr_col).value
                                if kr_value and str(kr_value).strip() == kr_text:
                                    string_id_value = worksheet.cell(row=row, column=string_id_col).value
                                    if string_id_value:
                                        self.log_message(f"✅ 기존 ID 발견 (신규 파일): {string_id_value}")
                                        workbook.close()
                                        return str(string_id_value).strip()
                            except:
                                continue
            
            workbook.close()
            self.log_message(f"❌ 신규 파일에서 중복 없음")
            return None
            
        except Exception as e:
            self.log_message(f"❌ 신규 파일 검색 오류: {e}")
            if workbook:
                try:
                    workbook.close()
                except:
                    pass
            return None

    def get_string_id_by_kr_enhanced(self, db_path, kr_text, new_text_file):
        """향상된 STRING_ID 검색 (DB + 신규 파일 모두 검색)"""
        
        # 1. 먼저 기존 DB에서 검색
        existing_id = self.get_string_id_by_kr(db_path, kr_text)
        if existing_id:
            self.log_message(f"✅ 기존 DB에서 ID 발견: {existing_id}")
            return existing_id
        
        # 2. 신규 텍스트 파일에서 검색
        new_file_id = self.get_existing_string_id_from_new_file(new_text_file, kr_text)
        if new_file_id:
            self.log_message(f"✅ 신규 파일에서 ID 발견: {new_file_id}")
            return new_file_id
        
        # 3. 둘 다 없으면 None 반환
        self.log_message(f"❌ 어디서도 찾을 수 없음: '{kr_text}'")
        return None

    def _find_excel_headers(self, worksheet):
        """엑셀 시트에서 헤더 위치 찾기 (개선된 버전)"""
        headers = {}
        
        try:
            # 처음 5행에서 헤더 찾기
            for row in range(1, 6):
                for col in range(1, min(worksheet.max_column + 1, 20)):  # 최대 20컬럼까지만 검색
                    try:
                        cell_value = worksheet.cell(row=row, column=col).value
                        if cell_value:
                            header_text = str(cell_value).strip().upper()
                            
                            # 대소문자 구분 없이 헤더 매칭
                            if header_text in ["STRING_ID", "KR", "EN", "CN", "TW", "TH", "PT", "ES", "DE", "FR", "#번역요청"]:
                                headers[header_text] = col
                                self.log_message(f"헤더 발견: {header_text} = 컬럼 {col}")
                    except:
                        continue
                
                # 필요한 헤더가 모두 찾아지면 중단
                if "STRING_ID" in headers and "KR" in headers:
                    break
            
            return headers
        except Exception as e:
            self.log_message(f"헤더 찾기 오류: {e}")
            return {}

    def preview_replacements(self):
        """치환 미리보기"""
        if not self._validate_inputs():
            return
            
        self.log_text.delete(1.0, tk.END)
        self.log_message("치환 미리보기 시작")
        
        # 결과 테이블 초기화
        self.result_tree.delete(*self.result_tree.get_children())
        self.replacement_results = []
        
        selected_files = self._get_selected_files()
        db_path = self.db_path_var.get()
        
        # 진행 창 생성
        try:
            loading_popup = LoadingPopup(self.root, "미리보기 중", "패턴 분석 중...")
        except:
            # LoadingPopup을 사용할 수 없는 경우 상태 라벨로 대체
            self.status_label.config(text="미리보기 진행 중...")
            loading_popup = None
        
        def preview_work():
            try:
                total_files = len(selected_files)
                
                for idx, file_name in enumerate(selected_files):
                    file_path = next((path for name, path in self.excel_files if name == file_name), None)
                    if not file_path:
                        continue
                    
                    if loading_popup:
                        self.root.after(0, lambda i=idx, f=file_name: loading_popup.update_progress(
                            (i / total_files) * 100, f"분석 중: {f}"))
                    else:
                        self.root.after(0, lambda i=idx, t=total_files: 
                                       self.status_label.config(text=f"미리보기 진행 중... ({i+1}/{t})"))
                    
                    self._analyze_file_patterns(file_path, db_path, preview_only=True)
                
                # 결과 표시 (메인 스레드에서)
                self.root.after(0, lambda: self._display_preview_results(loading_popup))
                
            except Exception as e:
                self.root.after(0, lambda: [
                    loading_popup.close() if loading_popup else None,
                    self.status_label.config(text="미리보기 오류 발생"),
                    messagebox.showerror("오류", f"미리보기 중 오류: {str(e)}")
                ])
        
        # 백그라운드 스레드 실행
        thread = threading.Thread(target=preview_work)
        thread.daemon = True
        thread.start()

    def execute_replacements(self):
        """치환 실행 - 고성능 openpyxl 전용 버전"""
        if not self._validate_inputs():
            return
            
        if not messagebox.askyesno("확인", "치환을 실행하시겠습니까?", parent=self.root):
            return
        
        self.log_text.delete(1.0, tk.END)
        self.log_message("🚀 치환 작업 시작!")
        
        # 통계 초기화
        self.replacement_stats = {
            'total_found': 0,
            'existing_replaced': 0,
            'new_created': 0,
            'files_processed': 0
        }
        
        selected_files = self._get_selected_files()
        db_path = self.db_path_var.get()
        
        # 시작 전 가비지 컬렉션
        gc.collect()
        
        # 진행 창 생성
        try:
            loading_popup = LoadingPopup(self.root, "치환 실행 중", "고성능 처리 준비 중...")
        except:
            self.status_label.config(text="치환 실행 중...")
            loading_popup = None
        
        def replacement_work():
            try:
                total_files = len(selected_files)
                self.log_message(f"📁 총 {total_files}개 파일 처리 예정")
                
                for idx, file_name in enumerate(selected_files):
                    file_path = next((path for name, path in self.excel_files if name == file_name), None)
                    if not file_path:
                        continue
                    
                    # 진행 상태 업데이트
                    if loading_popup:
                        self.root.after(0, lambda i=idx, f=file_name: loading_popup.update_progress(
                            (i / total_files) * 100, f"처리 중 ({i+1}/{total_files}): {f}"))
                    else:
                        self.root.after(0, lambda i=idx, t=total_files: 
                                        self.status_label.config(text=f"치환 실행 중... ({i+1}/{t})"))
                    
                    # 파일 처리
                    if self._process_file_replacement(file_path, db_path):
                        self.replacement_stats['files_processed'] += 1
                    
                    # 각 파일 처리 후 가비지 컬렉션
                    gc.collect()
                
                # 완료 처리 (메인 스레드에서)
                self.root.after(0, lambda: self._finalize_replacement(loading_popup))
                
            except Exception as e:
                self.root.after(0, lambda: [
                    loading_popup.close() if loading_popup else None,
                    self.status_label.config(text="치환 실행 오류 발생"),
                    messagebox.showerror("오류", f"치환 중 오류: {str(e)}")
                ])
            finally:
                # 작업 완료 후 최종 가비지 컬렉션
                gc.collect()
        
        # 백그라운드 스레드 실행
        thread = threading.Thread(target=replacement_work)
        thread.daemon = True
        thread.start()

    def _validate_inputs(self):
        """입력값 검증"""
        selected_files = self._get_selected_files()
        if not selected_files:
            messagebox.showwarning("경고", "처리할 파일을 선택하세요.", parent=self.root)
            return False
        
        db_path = self.db_path_var.get()
        if not db_path or not os.path.isfile(db_path):
            messagebox.showwarning("경고", "유효한 DB 파일을 선택하세요.", parent=self.root)
            return False
        
        new_text_file = self.new_text_file_var.get()
        if not new_text_file or not os.path.isfile(new_text_file):
            messagebox.showwarning("경고", "신규 텍스트를 저장할 엑셀 파일을 선택하세요.", parent=self.root)
            return False
        
        return True

    def _analyze_file_patterns(self, file_path, db_path, preview_only=True):
        """파일에서 패턴 분석 - String으로 시작하는 시트만"""
        workbook = None
        new_text_file = self.new_text_file_var.get()
        
        try:
            workbook = load_workbook(file_path, read_only=True)
            file_name = os.path.basename(file_path)
            
            selected_bulk_langs = [lang for lang, var in self.bulk_lang_vars.items() if var.get()]
            
            # String으로 시작하는 시트만 처리
            for sheet_name in workbook.sheetnames:
                if not sheet_name.lower().startswith("string"):
                    continue
                
                worksheet = workbook[sheet_name]
                
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            patterns = self.find_korean_patterns(cell.value)
                            
                            if patterns:
                                # 치환 시뮬레이션
                                simulated_text = cell.value
                                
                                for korean_text, full_pattern in patterns:
                                    # 향상된 검색 사용
                                    existing_id = self.get_string_id_by_kr_enhanced(db_path, korean_text, new_text_file)
                                    
                                    if existing_id:
                                        status = "기존 ID"
                                        string_id = existing_id
                                    else:
                                        status = "새 ID 생성"
                                        if preview_only:
                                            string_id = f"{self.id_prefix_var.get()}_{time.strftime('%Y%m%d_%H%M%S')}_{self.replacement_stats['new_created'] + 1:03d}"
                                        else:
                                            string_id = self.generate_new_string_id_from_excel(new_text_file, self.id_prefix_var.get())
                                    
                                    # 치환 시뮬레이션
                                    simulated_text = simulated_text.replace(full_pattern, f"[@{string_id}]")
                                    
                                    result = {
                                        'file_name': file_name,
                                        'sheet_name': sheet_name,
                                        'cell': f"{cell.column_letter}{cell.row}",
                                        'original_text': full_pattern,
                                        'korean_text': korean_text,
                                        'string_id': string_id,
                                        'status': status
                                    }
                                    
                                    self.replacement_results.append(result)
                                    self.replacement_stats['total_found'] += 1
                                    
                                    if existing_id:
                                        self.replacement_stats['existing_replaced'] += 1
                                    else:
                                        self.replacement_stats['new_created'] += 1
                                
                                # 일괄 적용 미리보기
                                if selected_bulk_langs:
                                    bulk_result = {
                                        'file_name': file_name,
                                        'sheet_name': sheet_name,
                                        'cell': f"{cell.column_letter}{cell.row}",
                                        'original_text': "[일괄복사 예정]",
                                        'korean_text': simulated_text,
                                        'string_id': f"→ {', '.join(selected_bulk_langs)}",
                                        'status': "일괄 적용"
                                    }
                                    self.replacement_results.append(bulk_result)
        
        except Exception as e:
            self.log_message(f"❌ 파일 분석 오류: {e}")
        finally:
            if workbook:
                try:
                    workbook.close()
                except:
                    pass
            gc.collect()

    def _process_file_replacement(self, file_path, db_path):
        """치환 처리 - String으로 시작하는 시트만"""
        new_text_file = self.new_text_file_var.get()
        
        try:
            # openpyxl 처리
            workbook = load_workbook(file_path)
            modified = False
            
            file_name = os.path.basename(file_path)
            selected_bulk_langs = [lang for lang, var in self.bulk_lang_vars.items() if var.get()]
            
            self.log_message(f"🔄 파일 처리 시작: {file_name}")
            
            # String으로 시작하는 시트만 처리
            for sheet_name in workbook.sheetnames:
                if not sheet_name.lower().startswith("string"):
                    continue
                
                worksheet = workbook[sheet_name]
                
                # 헤더 찾기
                headers = {}
                for col in range(1, 21):
                    for row in range(1, 6):
                        cell_value = worksheet.cell(row=row, column=col).value
                        if cell_value:
                            header = str(cell_value).strip().upper()
                            if header in ["STRING_ID", "KR", "EN", "CN", "TW", "TH"]:
                                headers[header] = col
                
                # 셀 처리
                for row in range(2, worksheet.max_row + 1):
                    for col in range(1, min(worksheet.max_column + 1, 15)):
                        cell = worksheet.cell(row=row, column=col)
                        
                        if cell.value and isinstance(cell.value, str):
                            patterns = self.find_korean_patterns(cell.value)
                            
                            if patterns:
                                original_text = cell.value
                                new_text = original_text
                                
                                for korean_text, full_pattern in patterns:
                                    existing_id = self.get_string_id_by_kr(db_path, korean_text)
                                    
                                    if existing_id:
                                        string_id = existing_id
                                    else:
                                        string_id = self.generate_new_string_id_from_excel(new_text_file, self.id_prefix_var.get())
                                        # 안전 모드 체크
                                        if getattr(self, 'safe_mode_var', None) and self.safe_mode_var.get():
                                            self.add_new_string_to_excel_safe(new_text_file, string_id, korean_text)
                                        else:
                                            self.add_new_string_to_excel(new_text_file, string_id, korean_text)
                                    
                                    new_text = new_text.replace(full_pattern, f"[@{string_id}]")
                                
                                if new_text != original_text:
                                    cell.value = new_text
                                    modified = True
                                    
                                    # 일괄 적용
                                    for lang in selected_bulk_langs:
                                        if lang in headers:
                                            worksheet.cell(row=row, column=headers[lang]).value = new_text
            
            if modified:
                # openpyxl 저장
                workbook.save(file_path)
                workbook.close()
                
                # Excel로 한 번 더 저장
                if self.excel_auto_save(file_path):
                    self.log_message(f"✅ 파일 처리 완료 (Excel 정상화): {file_name}")
                else:
                    self.log_message(f"⚠️ 파일 처리 완료 (Excel 정상화 실패): {file_name}")
                
                return True
            else:
                workbook.close()
                self.log_message(f"ℹ️ 변경사항 없음: {file_name}")
                return True
            
        except Exception as e:
            self.log_message(f"❌ 파일 처리 오류: {e}")
            return False

    def _display_preview_results(self, loading_popup):
        """미리보기 결과 표시"""
        if loading_popup:
            loading_popup.close()
        
        # 결과 트리뷰에 표시
        for idx, result in enumerate(self.replacement_results):
            self.result_tree.insert("", "end", iid=idx, values=(
                result['file_name'],
                result['sheet_name'],
                result['cell'],
                result['original_text'],
                result['korean_text'],
                result['string_id'],
                result['status']
            ))
        
        # 통계 표시
        total = self.replacement_stats['total_found']
        existing = self.replacement_stats['existing_replaced']
        new = self.replacement_stats['new_created']
        
        self.status_label.config(text=f"미리보기 완료 - 총 {total}개 패턴 (기존: {existing}, 신규: {new})")
        
        self.log_message(f"🔍 미리보기 완료: 총 {total}개 패턴 발견")
        self.log_message(f"📋 기존 ID 사용 예정: {existing}개")
        self.log_message(f"✨ 신규 ID 생성 예정: {new}개")
        self.log_message(f"💾 신규 텍스트 저장 예정: {os.path.basename(self.new_text_file_var.get())}")
        
        # 치환 일괄 적용 정보
        selected_bulk_langs = [lang for lang, var in self.bulk_lang_vars.items() if var.get()]
        if selected_bulk_langs:
            self.log_message(f"🌐 일괄 적용 예정 언어: {', '.join(selected_bulk_langs)}")
        
        # 빠른 완료 메시지
        if total > 0:
            messagebox.showinfo("미리보기 완료", 
                f"🔍 패턴 분석 완료!\n\n"
                f"• 발견된 패턴: {total}개\n"
                f"• 기존 ID 활용: {existing}개\n"
                f"• 신규 ID 생성: {new}개", 
                parent=self.root)
        else:
            messagebox.showinfo("미리보기 완료", "치환할 패턴을 찾지 못했습니다.", parent=self.root)

    def _finalize_replacement(self, loading_popup):
        """치환 작업 완료 처리 + 신규 스트링 파일 검증"""
        if loading_popup:
            loading_popup.close()
        
        stats = self.replacement_stats
        
        self.status_label.config(text=f"치환 완료 - {stats['files_processed']}개 파일 처리")
        
        self.log_message(f"🎉 치환 작업 완료!")
        self.log_message(f"📂 처리된 파일: {stats['files_processed']}개")
        self.log_message(f"🔄 총 치환된 패턴: {stats['total_found']}개")
        self.log_message(f"📋 기존 ID 사용: {stats['existing_replaced']}개")
        self.log_message(f"✨ 신규 ID 생성: {stats['new_created']}개")
        
        # 신규 스트링 파일 종합 검증
        new_text_file = self.new_text_file_var.get()
        file_status = "정상"
        
        if new_text_file and os.path.exists(new_text_file) and stats['new_created'] > 0:
            try:
                self.log_message(f"🔍 신규 스트링 파일 검증 중...")
                
                # 파일 열기 테스트
                test_workbook = load_workbook(new_text_file, read_only=True)
                
                # String 시트 존재 확인
                string_sheets = [sheet for sheet in test_workbook.sheetnames if sheet.lower().startswith("string")]
                if not string_sheets:
                    raise Exception("String 시트가 없음")
                
                # 첫 번째 String 시트에서 데이터 확인
                first_sheet = test_workbook[string_sheets[0]]
                if first_sheet.max_row < 2:
                    raise Exception("데이터가 없음")
                
                test_workbook.close()
                
                self.log_message(f"✅ 신규 스트링 파일 검증 완료: {os.path.basename(new_text_file)}")
                self.log_message(f"   📊 {len(string_sheets)}개 String 시트, {first_sheet.max_row-1}개 데이터 행")
                
            except Exception as e:
                file_status = "손상 가능"
                self.log_message(f"⚠️ 신규 스트링 파일 문제 감지: {e}")
                
                # 사용자에게 경고
                messagebox.showwarning("파일 검증 실패", 
                    f"⚠️ 신규 스트링 파일에 문제가 있을 수 있습니다!\n\n"
                    f"파일: {os.path.basename(new_text_file)}\n"
                    f"문제: {str(e)}\n\n"
                    f"임시 폴더의 백업 파일을 확인하시거나,\n"
                    f"작업을 다시 실행해 주세요.", 
                    parent=self.root)
        
        # 치환 일괄 적용 정보
        selected_bulk_langs = [lang for lang, var in self.bulk_lang_vars.items() if var.get()]
        if selected_bulk_langs:
            self.log_message(f"🌐 일괄 적용된 언어: {', '.join(selected_bulk_langs)}")
        
        # 모드 정보
        safe_mode = getattr(self, 'safe_mode_var', None) and self.safe_mode_var.get()
        excel_save = getattr(self, 'excel_auto_save_var', None) and self.excel_auto_save_var.get()
        
        completion_msg = (
            f"단어 치환이 완료되었습니다! 🎉\n\n"
            f"📊 처리 통계:\n"
            f"• 처리된 파일: {stats['files_processed']}개\n"
            f"• 총 치환된 패턴: {stats['total_found']}개\n"
            f"• 기존 ID 사용: {stats['existing_replaced']}개\n"
            f"• 신규 ID 생성: {stats['new_created']}개\n"
            f"• 신규 스트링 파일: {os.path.basename(self.new_text_file_var.get())} ({file_status})"
        )
        
        if selected_bulk_langs:
            completion_msg += f"\n• 일괄 적용 언어: {', '.join(selected_bulk_langs)}"
            
        if safe_mode or excel_save:
            modes = []
            if safe_mode:
                modes.append("안전 모드")
            if excel_save:
                modes.append("Excel 자동 저장")
            completion_msg += f"\n• 사용된 옵션: {', '.join(modes)}"
        
        messagebox.showinfo("완료", completion_msg, parent=self.root)